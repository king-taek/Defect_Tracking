"""Excel 결과 출력 (문서 Section 8.7).

선택된 기준 사진들에 대해 기준 layer 사진과 비교 layer 매칭 결과를 깔끔한 Excel 로 출력한다.
포함 정보: LOT명, wafer ID, 기준/비교 layer, col_row_x_y 위치, 허용 오차, 매칭 여부,
이미지 썸네일, 원본 경로(클릭하면 원본을 여는 링크, 무수정).

저장 경로는 반드시 assert_output_safe 게이트를 통과해야 하며, 원본 폴더 내부면 차단된다.
원본 이미지는 read-only 로만 읽는다(썸네일 캐시를 통해).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import BaseDefectMatches
from app.safety import assert_output_safe
from app.thumbnails import ThumbnailCache

# 다크/네온 느낌과 어울리는 보고서 색 (Excel 은 밝은 배경이 가독성 좋아 절제해 사용)
_NAVY = "FF1B2A4A"
_NEON = "FF1E90FF"
_LIGHT = "FFF2F6FF"
_MATCH = "FF1F7A1F"
_NOMATCH = "FFB00020"
_GREY = "FF6B7280"

_THIN = Side(style="thin", color="FFB8C4DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_IMG_COL_WIDTH = 30  # Excel 폭 단위
_IMG_PX = 190  # 썸네일 픽셀
_IMG_ROW_HEIGHT = 150  # 포인트


def _set_cell(ws, row, col, value, *, bold=False, color=None, fill=None,
              size=10, align="left", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color or "FF1B2A4A", size=size)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    cell.border = _BORDER
    return cell


def _place_image(ws, thumb_cache, rec, row, col) -> None:
    """(row,col) 셀에 rec 썸네일을 앵커한다. 실패 시 안내 텍스트."""
    col_letter = get_column_letter(col)
    cell = ws.cell(row=row, column=col)
    cell.border = _BORDER
    thumb = thumb_cache.get_full_thumbnail(rec.image_path, max_size=_IMG_PX)
    if thumb is not None:
        try:
            xl = XLImage(str(thumb))
            xl.anchor = f"{col_letter}{row}"
            ws.add_image(xl)
            return
        except (OSError, ValueError):
            cell.value = "(이미지 로드 실패)"
            return
    cell.value = "(이미지 없음)"


def _place_path_link(ws, row, col, image_path) -> None:
    """(row,col) 셀에 클릭하면 원본을 여는 링크를 건다(표시 텍스트는 고정 문구).

    원본을 통째로 심으면(고화질) 파일이 수백 MB~GB 로 불어나므로, 대신 항상
    최신 원본을 가리키는 가벼운 링크로 연결한다. 파일명은 '정보' 행에 이미 있어
    링크 텍스트는 용도가 드러나는 고정 문구로 둔다.
    """
    p = Path(image_path)
    cell = ws.cell(row=row, column=col, value="원본 사진 열기")
    cell.font = Font(color=_NEON, size=8, underline="single")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _BORDER
    try:
        cell.hyperlink = p.resolve().as_uri()
    except (OSError, ValueError):
        pass  # 링크 실패해도 안내 텍스트는 남는다


def export_excel(
    output_path: str | Path,
    *,
    lot_name: str,
    base_layer: str,
    compare_layers: list[str],
    tolerance: float,
    selected: list[BaseDefectMatches],
    thumb_cache: ThumbnailCache,
    source_roots: Iterable[str | Path],
    progress: Optional[Callable[[int, int], None]] = None,
    layer_order: Optional[list[str]] = None,
) -> Path:
    """선택된 기준 defect 들의 비교 결과를 Excel 로 저장한다.

    Args:
        layer_order: LOT 의 원래 layer 순서(폴더 스캔 순서). 주어지면 기준 layer 를
            맨 왼쪽에 고정하지 않고 이 순서대로 열을 배치한다. None 이면 기존처럼
            기준이 첫 데이터 열에 온다.

    Returns:
        저장된 파일의 절대 경로.

    Raises:
        OriginalProtectionError: 출력 경로가 원본 폴더 내부일 때.
    """
    out = assert_output_safe(output_path, source_roots)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "비교 결과"
    ws.sheet_view.showGridLines = False

    # 열: A=라벨, B=기준, C..=비교 layer. 각 블록이 '자기' 기준/비교 layer 로 스스로
    # 라벨링하므로(여러 layer 를 한 번에 담아도 섞이지 않게), 열 수는 블록별 비교 layer
    # 최댓값으로 잡는다.
    max_cmp = max((len(item.results) for item in selected), default=len(compare_layers))
    n_data_cols = 1 + max_cmp  # 기준 + 비교(최대)
    n_cols = n_data_cols + 1   # + 라벨열(A)

    ws.column_dimensions["A"].width = 16
    for ci in range(n_data_cols):
        ws.column_dimensions[get_column_letter(2 + ci)].width = _IMG_COL_WIDTH

    # ---- 보고서 헤더 ----
    r = 1
    title = ws.cell(row=r, column=1, value="Defect Tracker 비교 결과 보고서")
    title.font = Font(bold=True, color="FFFFFFFF", size=14)
    title.fill = PatternFill("solid", fgColor=_NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    ws.row_dimensions[r].height = 26
    r += 1

    # 담긴 항목들의 실제 기준 layer(혼합일 수 있음)를 헤더에 표기.
    base_layers_present: list[str] = []
    for item in selected:
        bl = item.base.layer or base_layer
        if bl not in base_layers_present:
            base_layers_present.append(bl)
    if len(base_layers_present) <= 1:
        base_desc = base_layers_present[0] if base_layers_present else base_layer
    else:
        base_desc = "혼합(기준 없이): " + ", ".join(base_layers_present)

    meta = (
        f"LOT: {lot_name}    기준 Layer: {base_desc}    "
        f"허용 오차: {tolerance:g}    생성: {datetime.now():%Y-%m-%d %H:%M}    "
        f"※ 각 블록의 'Layer' 행이 실제 기준/비교 layer 를 표시"
    )
    mc = ws.cell(row=r, column=1, value=meta)
    mc.font = Font(color="FFFFFFFF", size=10)
    mc.fill = PatternFill("solid", fgColor=_NEON)
    mc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    ws.row_dimensions[r].height = 20
    r += 2

    # 상단 컬럼 헤더 행은 두지 않는다 — 블록마다 'Layer' 행으로 이미 표기하므로 중복이다(항목 1).

    # ---- 각 기준 defect 블록 (블록마다 '자기' 기준/비교 layer 로 표기) ----
    _layer_pos = {name: i for i, name in enumerate(layer_order or [])}
    _total = len(selected)
    for idx, item in enumerate(selected, start=1):
        if progress is not None:
            progress(idx, _total)
        base = item.base
        base_layer_name = base.layer or base_layer
        results = list(item.results)

        # 열 배치 — 기준을 맨 앞에 고정하지 않고 LOT 의 원래 layer 순서를 따른다.
        # entries[ci] 가 데이터 열 2+ci 에 놓인다. mr 이 None 이면 기준 열.
        entries: list = [None] + results
        if layer_order:
            def _col_key(mr) -> int:
                name = base_layer_name if mr is None else mr.compare_layer
                # 목록에 없는 layer 는 뒤로(안정 정렬이라 기존 상대 순서 유지)
                return _layer_pos.get(name, len(layer_order))
            entries.sort(key=_col_key)

        # 블록 제목 — 기준 layer 를 함께 표기(여러 layer 혼합 대비).
        _set_cell(
            ws, r, 1,
            f"#{idx}", bold=True, color="FFFFFFFF", fill=_GREY, align="center",
        )
        head = (
            f"기준 {base_layer_name}   wafer {base.wafer_id}   "
            f"die ({base.col},{base.row})   pos {base.position_key}   "
            f"[{base.source.value}]"
        )
        _set_cell(ws, r, 2, head, bold=True, fill=_LIGHT, wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_cols)
        ws.row_dimensions[r].height = 18
        r += 1

        # Layer 이름 행 — 이 블록의 실제 기준/비교 layer(전역 기준에 종속되지 않음).
        _set_cell(ws, r, 1, "Layer", bold=True, align="center", fill=_LIGHT)
        for ci, mr in enumerate(entries):
            if mr is None:
                _set_cell(
                    ws, r, 2 + ci, "★ " + base_layer_name + " (기준)",
                    bold=True, color="FFFFFFFF", fill=_NEON, align="center",
                )
            else:
                _set_cell(
                    ws, r, 2 + ci, mr.compare_layer,
                    bold=True, color="FFFFFFFF", fill=_NAVY, align="center",
                )
        for ci in range(len(entries), n_data_cols):
            _set_cell(ws, r, 2 + ci, "", fill=_LIGHT, align="center")
        ws.row_dimensions[r].height = 18
        r += 1

        # 이미지 행
        _set_cell(ws, r, 1, "이미지", bold=True, align="center", fill=_LIGHT)
        ws.row_dimensions[r].height = _IMG_ROW_HEIGHT
        for ci, mr in enumerate(entries):
            if mr is None:
                _place_image(ws, thumb_cache, base, r, 2 + ci)
            elif mr.matched is not None:
                _place_image(ws, thumb_cache, mr.matched, r, 2 + ci)
            else:
                _set_cell(ws, r, 2 + ci, "매칭 없음", color=_NOMATCH, align="center")
        for ci in range(len(entries), n_data_cols):
            ws.cell(row=r, column=2 + ci).border = _BORDER
        r += 1

        # 상세 정보 행
        _set_cell(ws, r, 1, "정보", bold=True, align="center", fill=_LIGHT)
        for ci, mr in enumerate(entries):
            if mr is None:
                base_lines = [
                    "기준 ★", f"위치 {base.position_key}", Path(base.image_path).name
                ]
                extra = getattr(getattr(item, "base_cluster", None), "extra_count", 0) or 0
                if extra:
                    base_lines.append(f"+{extra} 근접중복")
                _set_cell(ws, r, 2 + ci, "\n".join(base_lines), color=_NEON, wrap=True, size=9)
                continue
            rec = mr.matched
            if rec is not None:
                dist = mr.distance
                lines = [
                    f"매칭 O (거리 {dist:.1f})" if dist is not None else "매칭 O",
                    f"위치 {rec.position_key}",
                    Path(rec.image_path).name,
                ]
                color = _MATCH
            else:
                lines = ["매칭 X"]
                color = _NOMATCH
            _set_cell(ws, r, 2 + ci, "\n".join(lines), color=color, wrap=True, size=9)
        for ci in range(len(entries), n_data_cols):
            ws.cell(row=r, column=2 + ci).border = _BORDER
        ws.row_dimensions[r].height = 60
        r += 1

        # 원본 경로 행 — 클릭하면 원본을 바로 연다(기준 + 각 매칭된 비교 defect 별로).
        _set_cell(ws, r, 1, "원본경로", bold=True, align="center", fill=_LIGHT)
        for ci, mr in enumerate(entries):
            if mr is None:
                _place_path_link(ws, r, 2 + ci, base.image_path)
            elif mr.matched is not None:
                _place_path_link(ws, r, 2 + ci, mr.matched.image_path)
            else:
                ws.cell(row=r, column=2 + ci).border = _BORDER
        for ci in range(len(entries), n_data_cols):
            ws.cell(row=r, column=2 + ci).border = _BORDER
        ws.row_dimensions[r].height = 24
        r += 1
        r += 1  # 블록 간 간격

    wb.save(out)
    return out
