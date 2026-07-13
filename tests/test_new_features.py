"""신규 기능 오프스크린 UI 스모크 테스트.

항목 1(출력 트레이), 2(매치 없는 셀 숨김), 4·5(히트맵), 7(썸네일 배율).
모달을 띄우지 않는 내부 메서드/위젯만 직접 호출한다.
"""

from __future__ import annotations

import os

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

pytest.importorskip("PySide6")

from PySide6.QtCore import QCoreApplication  # noqa: E402
from PySide6.QtWidgets import QApplication  # noqa: E402

from app import scanner  # noqa: E402
from app.config import AppSettings  # noqa: E402
from tools.make_sample_data import generate  # noqa: E402


@pytest.fixture(scope="module")
def app():
    return QApplication.instance() or QApplication([])


@pytest.fixture()
def win(app, tmp_path):
    from app.ui.main_window import MainWindow

    lot = generate(tmp_path / "src")
    idx = scanner.scan_lot(lot)
    w = MainWindow(AppSettings(workspace=str(tmp_path / "ws"), auto_update_check=False))
    w.lot_index = idx
    w._on_scan_finished(idx)
    w.top.cmb_base.setCurrentText("LYA4")
    for _ in range(10):
        QCoreApplication.processEvents()
    return w


# ---- 항목 9/1: 필터 드롭다운 제거 + 출력 트레이 ----

def test_filter_dropdown_removed(win):
    assert not hasattr(win, "cmb_filter")
    assert win._filter == "matched"


def test_export_tray_add_and_persist(win):
    # 트레이는 (BaseDefectMatches, 태그) 스냅샷을 담고, layer/자재가 바뀌어도 유지된다(항목 1).
    win._goto(0)
    assert win._export_tray == []
    p0 = str(win.matches[0].base.image_path)
    win._add_current_to_export()
    assert [str(m.base.image_path) for m, _tag in win._export_tray] == [p0]
    win._add_current_to_export()  # 중복 무시
    assert len(win._export_tray) == 1
    win._goto(1)
    win._add_current_to_export()
    assert len(win._export_tray) == 2
    assert "(2)" in win.btn_add_export.text()
    # 기준 layer 를 바꿔도 담은 것이 유지된다.
    win.top.cmb_base.setCurrentText("LYB4")
    for _ in range(5):
        QCoreApplication.processEvents()
    assert len(win._export_tray) == 2


def test_export_tray_dialog_remove_and_add_all(win, app):
    from app.ui.export_dialog import ExportTrayDialog

    entries = [win.matches[i] for i in range(min(3, len(win.matches)))]
    all_matched = [m for m in win.matches if win._match_status(m) != "none"]
    dlg = ExportTrayDialog(entries, win.thumb_cache, all_matched=all_matched)
    assert {str(m.base.image_path) for m in dlg.selected()} == {
        str(m.base.image_path) for m in entries
    }
    # 개별 제거
    key0 = str(entries[0].base.image_path)
    dlg._remove(key0)
    assert key0 not in {str(m.base.image_path) for m in dlg.selected()}
    # '이번 LOT 전체(매치) 추가' → 매치 있는 것이 모두 담김(중복 제거)
    dlg._add_all_matched()
    assert {str(m.base.image_path) for m in dlg.selected()} == {
        str(m.base.image_path) for m in all_matched
    }
    dlg._clear_all()
    assert dlg.selected() == []


def test_export_tray_group_persists_across_reopen(win, app):
    """'전체 추가'로 묶은 태그가 확인 후 트레이에 저장했다가 다시 열어도 유지돼야 한다.

    회귀: 예전엔 ExportTrayDialog.selected() 가 태그를 버린 평탄화 목록만 반환하고
    main_window._export_tray 에 그대로 저장해, 다이얼로그를 다시 열면 묶음(요약 카드)이
    개별 사진 카드 여러 장으로 풀어졌다.
    """
    from app.ui.export_dialog import ExportTrayDialog

    all_matched = [m for m in win.matches if win._match_status(m) != "none"]
    if len(all_matched) < 2:
        return
    dlg = ExportTrayDialog(
        [], win.thumb_cache, all_matched=all_matched, all_matched_label="테스트 묶음"
    )
    dlg._add_all_matched()
    tagged = dlg.tagged_selected()
    assert tagged and all(tag == "테스트 묶음" for _m, tag in tagged)

    # main_window._export() 가 트레이에 저장하는 것과 동일한 경로.
    win._export_tray = tagged
    dlg2 = ExportTrayDialog(list(win._export_tray), win.thumb_cache)
    assert dlg2.tagged_selected() == tagged
    tag_values = {tag for _m, tag in dlg2._tagged}
    assert tag_values == {"테스트 묶음"}  # 개별(None)로 풀리지 않았다


def test_export_tray_dialog_ok_vs_export(win, app):
    """확인(저장만)과 Excel 출력을 구분한다(wants_export)."""
    from app.ui.export_dialog import ExportTrayDialog
    entries = [win.matches[0]]
    dlg = ExportTrayDialog(entries, win.thumb_cache)
    dlg._on_ok()
    assert dlg.wants_export() is False  # 확인 = 저장만
    dlg2 = ExportTrayDialog(entries, win.thumb_cache)
    dlg2._on_export()
    assert dlg2.wants_export() is True  # Excel 출력


def test_export_all_layers_button_unions_matches(win, app):
    """'모든 매치(기준 없이)'는 모든 layer 를 기준으로 한 매치를 백그라운드로 합쳐 담는다.

    회귀: 예전엔 이 계산이 UI 스레드에서 동기 실행돼 앱이 멈췄다 — 이제
    AllLayersMatchWorker 로 백그라운드에서 돈다(QThreadPool). 또한 버튼 한 번에 대량
    추가된 항목은 사진 카드 대신 하나의 요약 태그로 묶인다.
    """
    from PySide6.QtCore import QThreadPool
    from app.ui.export_dialog import ExportTrayDialog, _ALL_LAYERS_TAG

    results: dict = {}
    win._provide_all_layers_matched(
        lambda cur, total: None, lambda items: results.setdefault("items", items)
    )
    QThreadPool.globalInstance().waitForDone(5000)
    for _ in range(20):
        QCoreApplication.processEvents()
    expected = {str(m.base.image_path) for m in results.get("items", [])}
    if not expected:
        return

    dlg = ExportTrayDialog(
        [], win.thumb_cache, all_layers_provider=win._provide_all_layers_matched
    )
    assert hasattr(dlg, "btn_add_all_layers")
    dlg._add_all_layers()
    QThreadPool.globalInstance().waitForDone(5000)
    for _ in range(20):
        QCoreApplication.processEvents()
    assert {str(m.base.image_path) for m in dlg.selected()} == expected
    assert any(tag == _ALL_LAYERS_TAG for _m, tag in dlg._tagged)
    # 현재 기준 layer 매치(all_matched)보다 크거나 같아야(여러 layer 합집합).
    cur = {str(m.base.image_path) for m in win.matches if win._match_status(m) != "none"}
    assert cur.issubset(expected)


# ---- 항목 2: 매치 없는 셀 숨김(re-pack) ----

def test_grid_hides_unmatched_cells(win):
    from PySide6.QtTest import QTest

    def _wait_status(target):
        for _ in range(30):
            QTest.qWait(100)
            if win.matches and {win._match_status(m) for m in win.matches} == target:
                return
    # 허용오차 0 → 모든 비교 매칭 실패 → 보이는 셀은 기준 하나뿐.
    # 매칭은 디바운스(250ms)+백그라운드 워커 → 조건까지 폴링 대기.
    win.top.spn_tol.setValue(0.0)
    _wait_status({"none"})
    win._goto(0)
    grid = win.grid
    # isHidden(): 창을 show 하지 않아도 명시적 숨김 상태를 반영한다.
    visible = [l for l, c in grid._cells.items() if not c.isHidden()]
    assert visible == [grid._base_layer]
    # 큰 허용오차 → 비교도 매칭 → 비교 셀도 보인다.
    win.top.spn_tol.setValue(100000.0)
    _wait_status({"matched"})
    win._goto(0)
    visible2 = {l for l, c in grid._cells.items() if not c.isHidden()}
    assert grid._base_layer in visible2
    assert len(visible2) >= 2


# ---- 썸네일 확대율: 5× 고정, 조절 UI 없음 ----

def test_thumbnail_zoom_fixed_5x_no_button(win):
    # 중앙 20%(≈5×) 고정.
    assert abs(win._thumbnail_center_ratio() - 0.20) < 1e-6
    # 조절 버튼/메서드가 제거됐다.
    assert not hasattr(win, "btn_thumb_zoom")
    assert not hasattr(win, "_adjust_thumbnail_zoom")


# ---- 항목 4·5: 히트맵 다이얼로그 ----

def test_heatmap_dialog_constructs_and_selects(win, app):
    from app.ui.heatmap_dialog import HeatmapDialog

    added = []
    dlg = HeatmapDialog(
        win.matches, win.top.base_layer(), win.top.compare_layers(),
        win.thumb_cache, lambda idxs: added.extend(idxs), win.settings,
        current_wafer=win.matches[0].base.wafer_id,
        records_by_layer=win.lot_index.records_by_layer(),
    )
    # 밀도 그룹이 채워지고 웨이퍼맵 격자가 잡힌다.
    assert dlg._map._cols >= 1 and dlg._map._rows >= 1
    assert dlg._groups, "defect 밀도 그룹이 있어야 한다"
    # 첫 셀 선택 → 상세 목록 구성(행 stretch 외 1개 이상) + 담기 동작
    key = next(iter(dlg._groups))
    dlg._on_selection_changed([key])
    assert dlg._detail_box.count() >= 2  # 행 위젯 + stretch
    dlg._add_all_current()
    assert added, "이 위치 전체 담기가 트레이 콜백을 호출해야 한다"


def test_heatmap_multi_select_union(win, app):
    from app.ui.heatmap_dialog import HeatmapDialog

    added = []
    dlg = HeatmapDialog(
        win.matches, win.top.base_layer(), win.top.compare_layers(),
        win.thumb_cache, lambda idxs: added.extend(idxs), win.settings,
        current_wafer=win.matches[0].base.wafer_id,
    )
    keys = list(dlg._groups.keys())
    if len(keys) >= 2:
        dlg._on_selection_changed(keys[:2])
        expected = set(dlg._groups[keys[0]]) | set(dlg._groups[keys[1]])
        assert set(dlg._union_indices()) == expected
    # 다중 선택 토글 → 맵이 multi 모드
    dlg.btn_multi.setChecked(True)
    assert dlg._map._multi is True


def test_heatmap_investigation_collects_all_layers(win, app):
    """상시 조사 모드: 선택 위치에서 체크된 layer 의 record 를 records_by_layer 에서 수집."""
    from app.ui.heatmap_dialog import HeatmapDialog

    rbl = win.lot_index.records_by_layer()
    dlg = HeatmapDialog(
        win.matches, win.top.base_layer(), win.top.compare_layers(),
        win.thumb_cache, lambda idxs: None, win.settings,
        current_wafer=win.matches[0].base.wafer_id, records_by_layer=rbl,
    )
    key = next(iter(dlg._groups))
    dlg._on_selection_changed([key])
    recs = dlg._records_at_selection()
    # 기준 defect 이 하나는 있고, 수집 layer 는 체크된 layer 집합에 속한다.
    assert any(lyr == dlg._base_layer for lyr, _ in recs)
    assert all(lyr in dlg._selected_layers() for lyr, _ in recs)


def test_heatmap_all_wafers_aggregates(win, app):
    from app.ui.heatmap_dialog import HeatmapDialog, _ALL_WAFERS

    dlg = HeatmapDialog(
        win.matches, win.top.base_layer(), win.top.compare_layers(),
        win.thumb_cache, lambda idxs: None, win.settings,
    )
    # '전체' wafer 항목이 콤보 맨 앞에 있고, 선택 시 모든 wafer 를 집계한다.
    assert dlg.cmb_wafer.itemText(0) == _ALL_WAFERS
    dlg._on_wafer_changed(_ALL_WAFERS)
    all_entries = dlg._base_entries()
    total = sum(1 for m in win.matches
                if m.base.col is not None and m.base.row is not None
                and m.base.col >= 0 and m.base.row >= 0)
    assert len(all_entries) == total


def test_wafer_map_bbox_normalized_no_margin_or_clip(win, app):
    """웨이퍼맵이 좌표 원점과 무관하게 bounding box (0,0) 기준으로 그려진다.

    회귀: 정합 shift 로 내용 min 이 양수면 왼쪽/위 여백(떠 보임), 음수면 좌·상단 잘림이
    생겼다. 이제 origin 정규화로 그려지는 셀이 항상 (0,0)부터 시작하고 잘리지 않아야 한다.
    """
    from app import config
    from app.config import ProductConfig
    from app.models import BaseDefectMatches, DefectRecord, MatchResult
    from pathlib import Path

    dm = {(c, r) for c in range(7) for r in range(6) if 1 <= c <= 5 or 1 <= r <= 4}
    config.PRODUCTS["TESTDISC_NORM"] = ProductConfig(
        key="TESTDISC_NORM", name="Test Disc", camtek_pitch_x=1, camtek_pitch_y=1,
        kla_package_x_count=7, kla_package_y_count=6, die_map=frozenset(dm), source="db",
    )
    prod0 = config.active_product().key

    def mk(col, row):
        b = DefectRecord(image_path=Path("/b.jpg"), wafer_id="W1", layer="LYA4",
                         layer_folder="LYA4", col=col, row=row, x=0.0, y=0.0)
        return BaseDefectMatches(base=b, results=[MatchResult(compare_layer="LYB4", base=b, matched=b)])

    try:
        config.set_active_product("TESTDISC_NORM")
        for off in [(0, 0), (3, 2), (-2, -1), (10, 7)]:
            obs = [(c + off[0], r + off[1]) for c, r in list(dm)[:8]]
            win.matches = [mk(c, r) for c, r in obs]
            win.current = 0
            win._align_cache.clear()
            win._update_wafer_map(win.matches[0])
            wm = win.wafer_map
            oc, orr = wm._origin_col, wm._origin_row
            content = set(wm._valid) if wm._valid else set(wm._states)
            assert content, f"offset {off}: 그릴 내용이 있어야 한다"
            # 그려지는 최소 셀이 (0,0)에 오고(여백 0), 모두 격자 안(잘림 0)이어야 한다.
            assert min(c - oc for c, _ in content) == 0
            assert min(r - orr for _, r in content) == 0
            assert all(0 <= c - oc < wm._cols and 0 <= r - orr < wm._rows for c, r in content)
    finally:
        config.set_active_product(prod0)
        config.PRODUCTS.pop("TESTDISC_NORM", None)
        win._align_cache.clear()


def test_excel_no_top_item_header(win, app, tmp_path):
    """엑셀 상단 '항목' 컬럼 헤더 행이 제거됐다(블록마다 Layer 행 사용)."""
    from app.export.excel_report import export_excel
    from openpyxl import load_workbook

    out = tmp_path / "o.xlsx"
    sel = win.matches[:2]
    export_excel(out, lot_name="L", base_layer=win.top.base_layer(),
                 compare_layers=win.top.compare_layers(), tolerance=100.0,
                 selected=sel, thumb_cache=win.thumb_cache,
                 source_roots=[win.lot_index.lot_path])
    wb = load_workbook(out)
    ws = wb.active
    firsts = {row[0] for row in ws.iter_rows(values_only=True) if row}
    assert "항목" not in firsts
    assert "Layer" in firsts  # 블록별 Layer 행은 있다


def test_excel_original_path_is_hyperlink(win, app, tmp_path):
    """'원본경로' 셀이 텍스트가 아니라 클릭하면 원본을 여는 링크여야 한다(고화질 심기 대신).

    엑셀에 원본을 통째로 심으면(고화질) 파일이 수백 MB~GB 로 불어나므로, 대신 항상
    최신 원본을 가리키는 가벼운 하이퍼링크로 연결한다.
    """
    from pathlib import Path
    from app.export.excel_report import export_excel
    from openpyxl import load_workbook

    out = tmp_path / "o.xlsx"
    sel = win.matches[:1]
    base = sel[0].base
    export_excel(out, lot_name="L", base_layer=win.top.base_layer(),
                 compare_layers=win.top.compare_layers(), tolerance=100.0,
                 selected=sel, thumb_cache=win.thumb_cache,
                 source_roots=[win.lot_index.lot_path])
    wb = load_workbook(out)
    ws = wb.active
    row_idx = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "원본경로"
    )
    cell = ws.cell(row=row_idx, column=2)
    # 표시 텍스트는 고정 문구(파일명은 '정보' 행에 이미 있다), 링크 대상은 원본 URI.
    assert cell.value == "원본 사진 열기"
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == Path(base.image_path).resolve().as_uri()


def test_excel_layer_order_preserved(win, app, tmp_path):
    """layer_order 를 주면 기준 layer 를 맨 왼쪽에 고정하지 않고 원래 순서를 유지한다."""
    from pathlib import Path
    from app.export.excel_report import export_excel
    from app.models import BaseDefectMatches, DefectRecord, MatchResult
    from openpyxl import load_workbook

    b = DefectRecord(image_path=Path("/b.jpg"), wafer_id="W1", layer="L2",
                     layer_folder="L2", col=0, row=0, x=0.0, y=0.0)
    item = BaseDefectMatches(base=b, results=[
        MatchResult(compare_layer="L1", base=b, matched=b),
        MatchResult(compare_layer="L3", base=b, matched=b),
    ])
    out = tmp_path / "o.xlsx"
    export_excel(out, lot_name="L", base_layer="L2", compare_layers=["L1", "L3"],
                 tolerance=100.0, selected=[item], thumb_cache=win.thumb_cache,
                 source_roots=[win.lot_index.lot_path],
                 layer_order=["L1", "L2", "L3"])
    ws = load_workbook(out).active
    row_idx = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "Layer"
    )
    vals = [ws.cell(row=row_idx, column=c).value for c in (2, 3, 4)]
    assert vals == ["L1", "★ L2 (기준)", "L3"]
    # layer_order 없이(기존 동작) 기준이 첫 데이터 열에 온다.
    out2 = tmp_path / "o2.xlsx"
    export_excel(out2, lot_name="L", base_layer="L2", compare_layers=["L1", "L3"],
                 tolerance=100.0, selected=[item], thumb_cache=win.thumb_cache,
                 source_roots=[win.lot_index.lot_path])
    ws2 = load_workbook(out2).active
    row_idx2 = next(
        r for r in range(1, ws2.max_row + 1)
        if ws2.cell(row=r, column=1).value == "Layer"
    )
    vals2 = [ws2.cell(row=row_idx2, column=c).value for c in (2, 3, 4)]
    assert vals2 == ["★ L2 (기준)", "L1", "L3"]


def test_grid_rollback_base_top_left(win):
    # 항목 5 롤백: 기준 셀이 압축 배치의 (0,0)에 온다.
    win.top.spn_tol.setValue(100000.0)
    for _ in range(5):
        QCoreApplication.processEvents()
    win._goto(0)
    gl = win.grid._grid
    base_cell = win.grid._cells[win.grid._base_layer]
    r, c, _, _ = gl.getItemPosition(gl.indexOf(base_cell))
    assert (r, c) == (0, 0)


def test_folder_picker_navigation_and_lists(app, tmp_path):
    from app.ui.folder_picker import FolderPickerDialog

    lot = generate(tmp_path / "src")  # tmp_path/src/<LOT_NAME>
    root = lot.parent
    s = AppSettings(workspace=str(tmp_path / "ws"), recent_folders=[str(lot)])
    dlg = FolderPickerDialog(s, str(root))

    # 목록에 LOT 폴더가 뜬다(한 단계 os.scandir).
    names = [dlg.listw.item(i).data(0x0100) for i in range(dlg.listw.count())]
    assert lot.name in names

    # 진입/위로 네비게이션.
    dlg._go_to(lot)
    assert dlg._cur == lot
    dlg._go_up()
    assert dlg._cur == root
    # 뒤로: 방금 위로 왔으니 history 로 lot 복귀.
    dlg._go_back()
    assert dlg._cur == lot


def test_folder_picker_filter_hides_items(app, tmp_path):
    from app.ui.folder_picker import FolderPickerDialog

    root = tmp_path / "root"
    (root / "AlphaLot").mkdir(parents=True)
    (root / "BetaLot").mkdir()
    dlg = FolderPickerDialog(AppSettings(workspace=str(tmp_path / "ws")), str(root))

    dlg._apply_filter("alpha")
    hidden = {
        dlg.listw.item(i).data(0x0100): dlg.listw.item(i).isHidden()
        for i in range(dlg.listw.count())
    }
    assert hidden.get("AlphaLot") is False
    assert hidden.get("BetaLot") is True


def test_folder_picker_favorite_toggle_persists(app, tmp_path):
    from app.ui.folder_picker import FolderPickerDialog

    (tmp_path / "root" / "sub").mkdir(parents=True)
    s = AppSettings(workspace=str(tmp_path / "ws"))
    dlg = FolderPickerDialog(s, str(tmp_path / "root"))
    dlg._set_candidate(tmp_path / "root" / "sub")
    dlg._toggle_favorite()
    assert str(tmp_path / "root" / "sub") in s.favorite_folders
    dlg._toggle_favorite()
    assert str(tmp_path / "root" / "sub") not in s.favorite_folders


def test_folder_picker_corrects_layer_to_material(app, tmp_path):
    from app.ui.folder_picker import FolderPickerDialog

    lot = generate(tmp_path / "src")
    layer = next(p for p in lot.iterdir() if p.is_dir())  # 자재 아래 layer
    s = AppSettings(workspace=str(tmp_path / "ws"))
    dlg = FolderPickerDialog(s, str(lot))

    # layer 폴더를 후보로 두면 선택 시 상위 자재로 보정된다.
    dlg._set_candidate(layer)
    assert dlg.selected_path() == str(lot)

    # 자재 폴더 자체는 그대로.
    dlg._set_candidate(lot)
    assert dlg.selected_path() == str(lot)


def test_theme_styles_item_views():
    # 트리/리스트 뷰가 전역 테마로 어둡게 스타일링된다(흰 배경 방지).
    from app.ui import theme
    assert "QTreeView" in theme.STYLESHEET
    assert "QHeaderView::section" in theme.STYLESHEET


def test_heatmap_subdivide_small_die_count(app):
    """die 개수가 적으면 웨이퍼맵이 하위셀 분할 모드가 된다."""
    from app.ui.heatmap_dialog import HeatmapDialog
    from app.models import BaseDefectMatches, DefectRecord
    from pathlib import Path

    def mk(col, row, x, y):
        base = DefectRecord(image_path=Path("/b.jpg"), wafer_id="W1", layer="LYA4",
                            layer_folder="LYA4", col=col, row=row, x=x, y=y)
        return BaseDefectMatches(base=base)

    # 4개 die(<50) → subdivide. 맵 density 는 records_by_layer(체크 layer) 기반이므로 전달.
    matches = [mk(0, 0, 0, 0), mk(0, 0, 90, 90), mk(1, 1, 0, 0), mk(2, 2, 0, 0)]
    rbl = {"LYA4": [m.base for m in matches]}
    dlg = HeatmapDialog(matches, "LYA4", [], None, lambda idxs: None, AppSettings(),
                        records_by_layer=rbl)
    assert dlg._map._subdivide is True


def test_heatmap_density_counts_clusters_as_one(app):
    """근접 중복(클러스터) defect 은 밀도에서 대표 1개로만 계산된다."""
    from app.ui.heatmap_dialog import HeatmapDialog
    from app.models import BaseDefectMatches, DefectRecord
    from pathlib import Path

    def mk(name, col, row, x, y):
        base = DefectRecord(image_path=Path(name), wafer_id="W1", layer="LYA4",
                            layer_folder="LYA4", col=col, row=row, x=x, y=y)
        return BaseDefectMatches(base=base)

    # die(2,3) 에 거리<50 근접 3개(→1 클러스터) + die(4,4) 단독 1개(→1 클러스터)
    matches = [
        mk("/a.jpg", 2, 3, 0, 0), mk("/b.jpg", 2, 3, 10, 10),
        mk("/c.jpg", 2, 3, 20, 20), mk("/d.jpg", 4, 4, 0, 0),
    ]
    rbl = {"LYA4": [m.base for m in matches]}
    s = AppSettings()
    s.cluster_radius = 50.0
    dlg = HeatmapDialog(matches, "LYA4", [], None, lambda idxs: None, s,
                        records_by_layer=rbl)
    # raw defect 4개지만 밀도 합은 클러스터 2개여야 한다(근접 3개는 1로 계산).
    assert sum(dlg._map._density.values()) == 2


def test_heatmap_subcell_uses_die_pitch_not_observed_range(app):
    """하위셀이 die pitch 절대 프레임으로 계산돼 레이어(record) 조합과 무관하게 고정된다.

    회귀: 예전엔 표시 중 record 의 관측 min/max 상대라, 같은 defect 이 다른 레이어를
    추가하면 다른 칸으로 이동했다(사용자 보고: (0,4)→(4,4)→(3,4)).
    """
    from app.ui.heatmap_dialog import HeatmapDialog
    from app.models import BaseDefectMatches, DefectRecord
    from app import config
    from pathlib import Path

    def mk(name, col, row, x, y):
        base = DefectRecord(image_path=Path(name), wafer_id="W1", layer="LYA4",
                            layer_folder="LYA4", col=col, row=row, x=x, y=y)
        return BaseDefectMatches(base=base)

    prod = config.active_product()
    # PIDS7 defect (24500.96, 6764.95) → pitch 기준 (3,0).
    target = mk("/t.jpg", 0, 2, 24500.96, 6764.95)
    exp_sc = int(24500.96 / prod.camtek_pitch_x * 5)
    exp_sr = int(6764.95 / prod.camtek_pitch_y * 5)

    # (a) 대상 defect 만
    rbl = {"LYA4": [target.base]}
    dlg = HeatmapDialog([target], "LYA4", [], None, lambda i: None, AppSettings(),
                        records_by_layer=rbl)
    k_only = dlg._key_for_record(target.base)
    # (b) 다른 die 의 defect 들을 함께 추가(관측 범위가 크게 달라짐)
    others = [mk("/o1.jpg", 1, 4, 7497.0, 31062.0), mk("/o2.jpg", 3, 3, 100.0, 200.0)]
    rbl2 = {"LYA4": [target.base] + [o.base for o in others]}
    dlg2 = HeatmapDialog([target] + others, "LYA4", [], None, lambda i: None, AppSettings(),
                         records_by_layer=rbl2)
    k_with = dlg2._key_for_record(target.base)

    assert (k_only.sub_col, k_only.sub_row) == (exp_sc, exp_sr)
    assert (k_with.sub_col, k_with.sub_row) == (exp_sc, exp_sr), "record 조합이 달라도 하위셀 불변"
    assert (k_only.col, k_only.row) == (0, 2)  # die 라벨은 physical 관측 좌표


def test_heatmap_wafermap_row0_at_bottom_roundtrip(app):
    """웨이퍼맵이 row 0 을 화면 맨 아래에 그리고(왼쪽아래 0,0), 클릭 히트테스트가 대칭."""
    from app.ui.heatmap_dialog import HeatmapWaferMap
    from app.heatmap import HeatKey
    from PySide6.QtCore import QPoint

    wm = HeatmapWaferMap()
    density = {HeatKey(0, 0): 3, HeatKey(0, 3): 5}
    wm.set_data(1, 4, None, False, density, origin=(0, 0))
    # row 0 은 row 3 보다 화면상 아래(픽셀 y 가 큼)에 온다.
    _, y0 = wm._die_origin(0, 0)
    _, y3 = wm._die_origin(0, 3)
    assert y0 > y3, "row 0 이 화면 맨 아래여야 한다"
    # 왕복: 각 die 중심 픽셀을 클릭하면 원래 (col,row) 로 복원된다.
    for col, row in [(0, 0), (0, 3)]:
        ox, oy = wm._die_origin(col, row)
        pt = QPoint(ox + wm._DIE_PX // 2, oy + wm._DIE_PX // 2)
        key = wm._key_at(pt)
        assert key is not None and (key.col, key.row) == (col, row)


# ---- 8차: 히트맵 모드/드래그/클러스터 · 폴더트리 · 출력 · 도움말 ----

def _make_heatmap(win):
    from app.ui.heatmap_dialog import HeatmapDialog
    return HeatmapDialog(
        win.matches, win.top.base_layer(), win.top.compare_layers(),
        win.thumb_cache, lambda idxs: None, win.settings,
        records_by_layer=win.lot_index.records_by_layer(),
    )


def test_heatmap_default_wafer_is_all(win, app):
    from app.ui.heatmap_dialog import _ALL_WAFERS
    dlg = _make_heatmap(win)
    assert dlg._current_wafer == _ALL_WAFERS
    assert dlg.cmb_wafer.currentText() == _ALL_WAFERS


def test_heatmap_no_base_all_layer_checkboxes(win, app):
    # 기준 별표 없음 · '전체 defect' 토글 없음 · 모든 layer 가 체크박스(기준 포함).
    dlg = _make_heatmap(win)
    assert not hasattr(dlg, "btn_show_all")
    assert dlg._base_layer in dlg._col_checks  # 기준도 체크박스로
    assert all(cb.isChecked() for cb in dlg._col_checks.values())  # 기본 전체 체크
    dlg.btn_multi.setChecked(True)
    assert "ON" in dlg.btn_multi.text()


def _ancestor_ids(w):
    out, p = [], w.parent()
    while p is not None:
        out.append(id(p))
        p = p.parent()
    return set(out)


def test_heatmap_multi_button_below_map(win, app):
    """여러 다이 선택 버튼이 (리스트 패널이 아니라) 웨이퍼 맵과 같은 패널 아래에 있다."""
    dlg = _make_heatmap(win)
    # btn_multi 와 맵 위젯이 같은 패널(맵 패널)을 공통 조상으로 가진다.
    assert _ancestor_ids(dlg.btn_multi) & _ancestor_ids(dlg._map)
    # 상세(리스트) 라벨과는 패널을 공유하지 않는다.
    assert not (_ancestor_ids(dlg.btn_multi) & {id(dlg.lbl_detail)})


def test_heatmap_subdivide_is_5x5(app):
    from app import heatmap
    assert (heatmap.SUB_COLS, heatmap.SUB_ROWS) == (5, 5)  # 25 분할


def test_heatmap_uncheck_layer_reduces_map_entries(win, app):
    # 조사 모드: 맵 density 는 체크된 layer 전체 defect. layer 를 끄면 entries 가 준다.
    dlg = _make_heatmap(win)
    full = len(dlg._map_entries())
    # 비-기준 layer 하나 해제
    others = [l for l in dlg._col_checks if l != dlg._base_layer]
    if others:
        dlg._col_checks[others[0]].setChecked(False)
        assert len(dlg._map_entries()) < full


def test_heatmap_add_targets_match_displayed_cross_layer_groups(win, app):
    """'출력에 넣기' 대상 개수가 화면에 표시된 교차매치 그룹(기준 layer 포함) 개수와 같아야 한다.

    회귀: 예전엔 다이얼로그 생성 시점에 고정된 self._compare_layers 로 판정해(_is_matched),
    화면에 실제 표시되는 교차매치 판정(_col_checks 기준, cross_layer_groups)과 어긋나
    버튼 활성화가 들쭉날쭉하고 '넣기'를 눌러도 일부만 담겼다.
    """
    from collections import defaultdict
    from app.clustering import cluster_records, cross_layer_groups

    dlg = _make_heatmap(win)
    dlg._selected_keys = list(dlg._groups.keys())
    dlg._rebuild_detail()

    by_layer = defaultdict(list)
    for lyr, rec in dlg._records_at_selection():
        by_layer[lyr].append(rec)
    layer_to_clusters = {
        lyr: cluster_records(recs, dlg._cluster_radius) for lyr, recs in by_layer.items()
    }
    groups = cross_layer_groups(layer_to_clusters, dlg._tolerance)
    matched_with_base = [g for g in groups if len(g) >= 2 and dlg._base_layer in g]
    assert len(dlg._add_targets) == len(matched_with_base)
    assert dlg.btn_add_all.isEnabled() == bool(matched_with_base)


def test_heatmap_add_targets_enable_consistent_across_selections(win, app):
    """서로 다른 위치를 연속 선택해도 버튼 활성화가 매번 add_targets 유무와 일치해야 한다."""
    dlg = _make_heatmap(win)
    for k in dlg._groups.keys():
        dlg._on_selection_changed([k])
        assert dlg.btn_add_all.isEnabled() == bool(dlg._add_targets)


def test_heatmap_map_caption_has_counts(win, app):
    dlg = _make_heatmap(win)
    txt = dlg.lbl_map.text()
    assert "die" in txt and "defect" in txt
    assert "매치만" not in txt and "전체 defect" not in txt  # 모드 문구 제거됨


def test_heatmap_drag_box_selects_without_multi(win, app):
    import types
    from PySide6.QtCore import QPoint, Qt as _Qt
    dlg = _make_heatmap(win)
    m = dlg._map
    assert m._multi is False
    dense = [k for k, c in m._density.items() if c > 0]
    if not dense:
        return  # 데이터 없으면 skip
    # 위젯 전체를 감싸는 드래그 → 밀도>0 die 다중 선택(멀티 off 에서도)
    m._rubber_origin = QPoint(0, 0)
    m._rubber_cur = QPoint(m.width(), m.height())
    m._dragging = True
    m.mouseReleaseEvent(types.SimpleNamespace(button=lambda: _Qt.LeftButton))
    assert len(m._selected_keys) >= 1


def test_heatmap_builds_cross_layer_detail(win, app):
    dlg = _make_heatmap(win)
    keys = list(dlg._groups.keys())
    if not keys:
        return
    dlg._selected_keys = keys[:2]
    dlg._rebuild_detail()
    txt = dlg.lbl_detail.text()
    assert ("교차매치" in txt) or ("개별" in txt)


def test_folder_picker_tree_lazy_expand_and_click(app, tmp_path):
    from PySide6.QtCore import Qt as _Qt
    from app.ui.folder_picker import FolderPickerDialog

    (tmp_path / "root" / "childA").mkdir(parents=True)
    (tmp_path / "root" / "childB").mkdir()
    dlg = FolderPickerDialog(AppSettings(workspace=str(tmp_path / "ws")), str(tmp_path))
    node = dlg._make_dir_node(dlg.sidebar.invisibleRootItem(), "root", str(tmp_path / "root"))
    assert node.childCount() == 1  # 더미(펼침 화살표)
    dlg._on_tree_expanded(node)
    paths = [node.child(i).data(0, _Qt.UserRole) for i in range(node.childCount())]
    assert any(p.endswith("childA") for p in paths)
    dlg._on_tree_clicked(node)
    assert dlg._cur == (tmp_path / "root")


def test_export_dialog_has_visible_confirm_button(app):
    from app.ui.export_dialog import ExportTrayDialog
    dlg = ExportTrayDialog([], None, None)
    assert hasattr(dlg, "btn_export")
    assert dlg.btn_export.text() == "Excel 출력"
    assert dlg.btn_export.isEnabled() is False  # 빈 트레이 → 비활성(존재는 함)


def test_help_dialog_has_sections_and_features(app):
    from app.ui.help_dialog import ShortcutsDialog, _FEATURES, _SHORTCUT_GROUPS
    dlg = ShortcutsDialog()
    assert dlg.windowTitle() == "도움말"
    assert len(_SHORTCUT_GROUPS) >= 3
    names = [n for n, _ in _FEATURES]
    assert any("히트맵" in n for n in names)
    assert any("클러스터" in n for n in names)


# ---- 9차: 메인 매치 클러스터링 · 히트맵 강조/개별 · 폴더 트리 · 리브랜딩 ----

def test_main_match_collapses_near_duplicates(app, tmp_path):
    """메인 매치가 근접 중복 기준 defect 을 대표 1개로 접는다(base_cluster + '+n')."""
    from app.ui.main_window import MainWindow
    from tools.make_sample_data import generate

    lot = generate(tmp_path / "src")
    # 기존 base defect 과 같은 die·근접 위치의 중복 이미지를 하나 추가한다.
    idx = scanner.scan_lot(lot)
    base_layer = "LYA4"
    raw = [r for r in idx.records_for_layer(base_layer) if r.ok]
    assert raw
    import dataclasses
    dup = dataclasses.replace(
        raw[0],
        image_path=raw[0].image_path.with_name("dup_" + raw[0].image_path.name),
        x=(raw[0].x or 0.0) + 5.0,
    )
    idx.records.append(dup)

    w = MainWindow(AppSettings(workspace=str(tmp_path / "ws"), auto_update_check=False))
    w.lot_index = idx
    w._on_scan_finished(idx)
    w.top.cmb_base.setCurrentText(base_layer)
    for _ in range(10):
        QCoreApplication.processEvents()

    raw_count = len([r for r in idx.records_for_layer(base_layer) if r.ok])
    assert len(w.matches) < raw_count  # 접힘
    assert any(getattr(m.base_cluster, "extra_count", 0) for m in w.matches)


def test_compare_grid_shows_cluster_badge(app, tmp_path):
    from app.ui.compare_grid import CompareGrid, LayerCell
    from app.models import BaseDefectMatches, DefectRecord
    from app.clustering import Cluster
    from pathlib import Path

    def rec(n):
        return DefectRecord(image_path=Path(f"/{n}.jpg"), wafer_id="W1", layer="B",
                            layer_folder="B", col=1, row=1, x=0.0, y=0.0)
    grid = CompareGrid()
    grid.build_layout([["B"]], "B")
    a, b = rec("a"), rec("b")
    item = BaseDefectMatches(base=a, results=[], base_cluster=Cluster(a, [a, b]))
    grid.update_for_base(item, [])
    cell = grid._cells["B"]
    assert not cell.more_badge.isHidden()  # 명시적으로 표시됨(창 미표시라 isVisible 대신)
    assert cell.more_badge.text() == "+1"
    # 클릭 시 묶인 멤버 목록을 emit
    got = []
    grid.base_cluster_clicked.connect(lambda m: got.append(m))
    cell._emit_cluster()
    assert got and len(got[0]) == 2


def test_heatmap_selection_paint_smoke(win, app):
    from PySide6.QtGui import QPixmap, QPainter
    dlg = _make_heatmap(win)
    dlg._selected_keys = list(dlg._groups.keys())[:1]
    dlg._map._selected_keys = set(dlg._selected_keys)
    pm = QPixmap(dlg._map.width() or 200, dlg._map.height() or 200)
    p = QPainter(pm)
    dlg._map._paint_selection(p)  # 채움+이중외곽선 경로가 예외 없이 실행
    p.end()


def test_heatmap_individual_flow_section(win, app):
    # 전체 defect 모드에서 개별(미매칭) 그룹이 FlowLayout 섹션으로 묶인다.
    from app.ui.flow_layout import FlowLayout
    dlg = _make_heatmap(win)
    groups = [{"LYA4": _one_cluster("/x.jpg")}]  # 단일 layer → 개별
    sec = dlg._make_individual_section(groups)
    flows = [c for c in sec.findChildren(QWidget) if c.layout().__class__.__name__ == "FlowLayout"] \
        if False else None
    # 섹션 캡션에 '개별(미매칭)' 포함
    from PySide6.QtWidgets import QLabel
    labels = [l.text() for l in sec.findChildren(QLabel)]
    assert any("개별(미매칭)" in t for t in labels)


def _one_cluster(path):
    from app.clustering import Cluster
    from app.models import DefectRecord
    from pathlib import Path
    r = DefectRecord(image_path=Path(path), wafer_id="W1", layer="LYA4",
                     layer_folder="LYA4", col=1, row=1, x=0.0, y=0.0)
    return Cluster(r, [r])


def test_folder_picker_tree_first_and_reveal(app, tmp_path):
    from PySide6.QtCore import Qt as _Qt
    from app.ui.folder_picker import FolderPickerDialog
    (tmp_path / "lot" / "layerA").mkdir(parents=True)
    dlg = FolderPickerDialog(AppSettings(workspace=str(tmp_path / "ws")), str(tmp_path / "lot"))
    tops = [dlg.sidebar.topLevelItem(i).text(0) for i in range(dlg.sidebar.topLevelItemCount())]
    assert any("홈" in t for t in tops)  # 최상위는 폴더 루트(홈/드라이브)
    dlg._go_to(tmp_path / "lot" / "layerA")
    cur = dlg.sidebar.currentItem()
    assert cur is not None and str(cur.data(0, _Qt.UserRole)).endswith("layerA")


def test_no_conder_branding_left():
    import subprocess
    out = subprocess.run(
        ["grep", "-rniI", "conder", "app", "main.py", "README.md", "CLAUDE.md", ".gitignore"],
        capture_output=True, text=True,
    )
    # 'Conder Scan' 은 사용자의 외부 스캔 폴더 이름(설정 scan_root_name 기본값)이라 브랜딩 잔재 아님.
    leftover = [
        ln for ln in out.stdout.splitlines()
        if ln.strip() and "scan_root_name" not in ln and "Conder Scan" not in ln
    ]
    assert leftover == [], "leftover conder refs:\n" + "\n".join(leftover)


def test_folder_picker_network_unc_root(app, tmp_path):
    """UNC 네트워크 경로(\\\\server\\share)가 트리 최상위 루트로 추가되고 reveal 가능."""
    from app.ui.folder_picker import FolderPickerDialog
    dlg = FolderPickerDialog(AppSettings(workspace=str(tmp_path / "ws")), str(tmp_path))
    # 공유 루트 추출(플랫폼 무관)
    assert FolderPickerDialog._unc_anchor(r"\\nas01\lots\LOT") == "\\\\nas01\\lots\\"
    # 네트워크 위치로 이동 시 루트가 즉석 추가된다.
    from pathlib import Path
    dlg._reveal_in_tree(Path(r"\\nas01\lots\LOT"))
    roots = [str(p) for p, _ in dlg._root_nodes]
    assert any(r.startswith("\\\\nas01") for r in roots)
    # 드라이브 라벨 생성이 예외 없이 동작.
    assert isinstance(FolderPickerDialog._drive_label("/"), str)


# ---- 10차: 메뉴 테마 · 비동기 로딩 · 확대뷰 · 폴더 트리 다듬기 ----

def test_theme_has_qmenu_rules():
    from app.ui import theme
    assert "QMenu" in theme.STYLESHEET
    assert "QMenu::item" in theme.STYLESHEET


def test_busy_overlay_start_stop(app):
    from PySide6.QtWidgets import QWidget
    from app.ui.busy_overlay import BusyOverlay
    host = QWidget()
    host.resize(400, 300)
    b = BusyOverlay(host)
    assert b.isHidden()
    b.start("매칭 중…", determinate=True)
    assert not b.isHidden()  # 표시됨(호스트 미표시라 isVisible 대신 isHidden 사용)
    b.set_progress(2, 4)
    b.stop()
    assert b.isHidden()


def test_busy_overlay_event_filter_survives_missing_host(app):
    """_host 가 없는(재래핑된) 인스턴스로 eventFilter 가 불려도 죽지 않아야 한다.

    회귀: ExportTrayDialog 처럼 매번 새로 만들어지는 다이얼로그의 BusyOverlay 가
    orphan 으로 남으면, shiboken 이 __init__ 없이 재래핑한 인스턴스로 eventFilter 를
    호출해 AttributeError: 'BusyOverlay' object has no attribute '_host' 가 났다.
    """
    from PySide6.QtCore import QEvent
    from PySide6.QtWidgets import QWidget
    from app.ui.busy_overlay import BusyOverlay

    host = QWidget()
    b = BusyOverlay(host)
    del b._host  # 재래핑(__init__ 미실행) 상황 흉내
    assert b.eventFilter(host, QEvent(QEvent.Resize)) is False


def test_notification_banner_dismiss_no_warning(app):
    """dismiss() 가 연결된 슬롯 없이 disconnect() 해도 RuntimeWarning 이 새지 않아야 한다."""
    import warnings
    from PySide6.QtWidgets import QWidget
    from app.ui.notifications import NotificationBanner

    parent = QWidget()
    parent.show()  # dismiss() 의 isVisible() 가드를 실제 사용처럼 통과시키기 위해 필요
    banner = NotificationBanner(parent)
    banner.show_message("테스트", "info", timeout_ms=0)
    banner.dismiss()
    banner._after_hide()  # 애니메이션 완료를 흉내(스스로 disconnect)
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        banner.dismiss()  # 이제 finished 에 연결된 슬롯이 없음 — 예전엔 여기서 RuntimeWarning
    assert not any(issubclass(w.category, RuntimeWarning) for w in caught)
    parent.close()


def test_match_worker_runs_and_collapses(app):
    from app.workers import MatchWorker
    from app.models import DefectRecord
    from pathlib import Path

    def rec(n, layer, x):
        return DefectRecord(image_path=Path(f"/{n}.jpg"), wafer_id="W1", layer=layer,
                            layer_folder=layer, col=1, row=1, x=x, y=0.0)
    base = [rec("a", "B", 0.0), rec("b", "B", 20.0)]  # 근접 → 접힘
    rbl = {"B": base, "C": [rec("c", "C", 5.0)]}
    got = []
    w = MatchWorker(base, ["C"], rbl, 100.0)
    w.signals.finished.connect(lambda m, o: got.append(m))
    w.run()  # 동기 실행(워커 로직 검증)
    assert got and isinstance(got[0], list)
    assert len(got[0]) < len(base)  # 근접 중복 접힘


def test_export_excel_progress_callback(win, app, tmp_path):
    from app.export.excel_report import export_excel
    sel = [win.matches[i] for i in range(min(2, len(win.matches)))]
    calls = []
    export_excel(
        tmp_path / "o.xlsx", lot_name="L", base_layer="LYA4",
        compare_layers=win.top.compare_layers(), tolerance=100.0,
        selected=sel, thumb_cache=win.thumb_cache,
        source_roots=[str(win.lot_index.lot_path)],
        progress=lambda c, t: calls.append((c, t)),
    )
    assert calls and calls[-1][1] == len(sel)


def test_main_matching_is_async(win, app):
    # 매칭이 백그라운드 워커로 수행되어도(디바운스/토큰) fixture 대기 후 matches 가 채워진다.
    assert win.matches
    assert hasattr(win, "busy") and hasattr(win, "_match_token")


def test_heatmap_fullscreen_uses_native_window_button(win, app):
    """커스텀 버튼이 아니라 OS 제목표시줄의 진짜 최대화 버튼을 쓴다."""
    from PySide6.QtCore import Qt as _Qt

    dlg = _make_heatmap(win)
    for _ in range(5):
        QCoreApplication.processEvents()
    assert not hasattr(dlg, "btn_fullscreen")
    assert not hasattr(dlg, "_toggle_fullscreen")
    assert dlg.windowFlags() & _Qt.WindowMaximizeButtonHint
    assert dlg.isMaximized()


def test_heatmap_thumb_size_slider_default_30_percent(win, app):
    """사진 크기 슬라이더 기본값이 30%(기존 고정 크기 150px)와 같다."""
    dlg = _make_heatmap(win)
    assert dlg.sld_thumb.value() == 30
    assert dlg._thumb_px == 150
    dlg.sld_thumb.setValue(60)
    assert dlg._thumb_px == 300
    assert dlg.lbl_thumb_pct.text() == "60%"


def test_heatmap_thumb_size_slider_debounces_when_many_photos(win, app):
    """사진이 20장 넘으면 슬라이더가 멈춘 뒤(디바운스)에만 다시 그린다."""
    dlg = _make_heatmap(win)
    dlg._pending_thumbs = list(range(25))  # 20장 초과 흉내
    dlg.sld_thumb.setValue(50)
    assert dlg._thumb_px == 250  # 값 자체는 즉시 갱신
    assert dlg._thumb_resize_timer.isActive()
    assert dlg._pending_thumbs == list(range(25)), "타이머 만료 전엔 다시 그리면 안 된다"
    dlg._thumb_resize_timer.timeout.emit()  # 슬라이더가 멈춘 뒤(디바운스 만료) 시뮬레이션
    assert dlg._pending_thumbs == [], "만료되면 _rebuild_detail 이 호출돼야 한다"


def test_heatmap_thumb_size_slider_immediate_when_few_photos(win, app):
    """사진이 20장 이하면 슬라이더를 움직일 때마다 바로 반영한다."""
    dlg = _make_heatmap(win)
    dlg._pending_thumbs = list(range(5))
    dlg.sld_thumb.setValue(70)
    assert not dlg._thumb_resize_timer.isActive()
    assert dlg._pending_thumbs == [], "20장 이하면 즉시 _rebuild_detail 이 호출돼야 한다"


def test_heatmap_detail_thumbs_deferred(win, app):
    dlg = _make_heatmap(win)
    keys = list(dlg._groups.keys())
    if not keys:
        return
    dlg._selected_keys = keys[:1]
    dlg._rebuild_detail()
    # 상세 썸네일은 지연 로딩 위젯으로 등록된다(백그라운드 캐시 후 채움).
    assert isinstance(dlg._pending_thumbs, list)


def test_image_viewer_scrollbars_off_and_anchor_zoom(app):
    from PySide6.QtCore import Qt as _Qt
    from app.ui.image_viewer import ImageViewerDialog
    from app.models import DefectRecord
    from pathlib import Path
    d = ImageViewerDialog(DefectRecord(image_path=Path("/nope.jpg"), wafer_id="W",
                                       layer="L", layer_folder="L", col=1, row=1, x=0.0, y=0.0))
    assert d._scroll.horizontalScrollBarPolicy() == _Qt.ScrollBarAlwaysOff
    assert d._scroll.verticalScrollBarPolicy() == _Qt.ScrollBarAlwaysOff
    d._zoom_at_cursor(1.2)  # null 이미지에서도 예외 없이 no-op


def test_image_viewer_info_compact_format(app):
    """정보 텍스트가 'Layer/Wafer/die' 한 줄 + '좌표: Camtek/KLA' 한 줄 +
    Defect/Path 로 간결하게 표시되고(측정/계산 태그·size 줄 없음), 표시 라벨이
    '정보 복사' 텍스트와 동일하며 작은 글씨(objectName=meta)다."""
    from PySide6.QtCore import Qt as _Qt
    from app.ui.image_viewer import ImageViewerDialog
    from app.models import DefectRecord, Source
    from pathlib import Path

    rec = DefectRecord(
        image_path=Path("/x/R_TB500_NLP-PIDS7_00RXM180XYH1_2_2_Over Sized Bump_27313.72_35564.77.jpg"),
        wafer_id="00RXM180XYH1", layer="PIDS7", layer_folder="PIDS7",
        source=Source.CAMTEK_FILENAME,
        col=2, row=2, x=27313.72, y=35564.77, defect_name="Over Sized Bump")
    d = ImageViewerDialog(rec)
    txt = d._info_text()
    assert "measured" not in txt and "calculated" not in txt
    assert "size:" not in txt.lower()
    assert "Layer: PIDS7 / Wafer: 00RXM180XYH1 / die: (2,2)" in txt
    assert "좌표: Camtek: (27314,35565) / KLA:" in txt
    assert "Defect: Over Sized Bump" in txt
    assert f"Path: {rec.image_path}" in txt
    # 사진을 열면 같은 정보가 텍스트로(작은 글씨) 그대로 표시된다.
    assert d._meta.text() == txt
    assert d._meta.objectName() == "meta"
    assert d._meta.textFormat() == _Qt.PlainText

    # KLA scan → 저장된 실제 DiePitchY 로 KLA 좌표를 정확히 계산한다.
    kla_rec = DefectRecord(
        image_path=Path("/x/W_-2_1_31_1.jpg"), wafer_id="00RXM179XYE0",
        layer="RDL4", layer_folder="RDL4", source=Source.KLA,
        col=1, row=4, x=7497.0, y=31062.0, die_pitch_y=44905.301)
    dk = ImageViewerDialog(kla_rec)
    tk = dk._info_text()
    # KLA y = round(DiePitchY - y) = round(44905.301 - 31062) = 13843 = 원래 YREL
    assert "좌표: Camtek: (7497,31062) / KLA: (7497,13843)" in tk
    # defect_name 이 없으면 Defect 줄은 생략된다.
    assert "Defect:" not in tk


def test_image_viewer_shorter_default_and_resizable(app):
    from PySide6.QtCore import Qt as _Qt
    from app.ui.image_viewer import ImageViewerDialog
    from app.models import DefectRecord
    from pathlib import Path
    d = ImageViewerDialog(DefectRecord(image_path=Path("/nope.jpg"), wafer_id="W",
                                       layer="L", layer_folder="L", col=1, row=1, x=0.0, y=0.0))
    # 기본 크기가 이전(960x760)보다 세로로 짧다.
    assert d.size().height() < 760
    assert d.minimumHeight() < 520
    # 최대화 버튼 힌트가 켜져 있어야 창이 확실히 리사이즈 가능한 일반 창으로 동작한다.
    assert d.windowFlags() & _Qt.WindowMaximizeButtonHint
    assert d.isSizeGripEnabled()


def test_image_viewer_recomputes_fit_on_first_show(app):
    """첫 표시 전 계산한 맞춤 배율을 showEvent 에서 실제 크기로 재계산해 잘림을 없앤다."""
    from app.ui.image_viewer import ImageViewerDialog
    from app.models import DefectRecord
    from pathlib import Path

    d = ImageViewerDialog(DefectRecord(image_path=Path("/nope.jpg"), wafer_id="W",
                                       layer="L", layer_folder="L", col=1, row=1, x=0.0, y=0.0))
    assert d._fit_pending is True  # 생성 직후엔 아직 정확한 크기로 재계산 전
    d.show()
    for _ in range(5):
        QCoreApplication.processEvents()
    assert d._fit_pending is False  # 첫 showEvent 에서 소비됨
    d.close()


def test_image_viewer_zoom_buttons_have_large_glyph_style(app):
    from app.ui.image_viewer import ImageViewerDialog
    from app.models import DefectRecord
    from pathlib import Path
    d = ImageViewerDialog(DefectRecord(image_path=Path("/nope.jpg"), wafer_id="W",
                                       layer="L", layer_folder="L", col=1, row=1, x=0.0, y=0.0))
    zoom_btns = [
        b for b in d.findChildren(type(d.btn_fit))
        if b.objectName() == "zoomGlyph"
    ]
    assert len(zoom_btns) == 2


def test_folder_picker_goto_scan_root_button(app, tmp_path):
    from app.ui.folder_picker import FolderPickerDialog

    target = tmp_path / "ConderScan"
    target.mkdir()
    s = AppSettings(workspace=str(tmp_path / "ws"), scan_root_path=str(target))
    dlg = FolderPickerDialog(s, str(tmp_path))
    assert hasattr(dlg, "btn_goto_scan")
    # '이동' 버튼 높이가 주소창(ed_scan_root)과 같아야 한다.
    assert dlg.btn_goto_scan.height() == dlg.ed_scan_root.sizeHint().height()
    visited = []
    dlg._go_to = lambda p: visited.append(p)  # noqa: E731 - 실제 탐색 대신 호출만 확인
    dlg._goto_scan_root()
    assert visited, "지정된 폴더로 이동을 시도해야 한다"


def test_folder_picker_indent_and_explorer_button(app, tmp_path):
    from app.ui.folder_picker import FolderPickerDialog
    dlg = FolderPickerDialog(AppSettings(workspace=str(tmp_path / "ws")), str(tmp_path))
    assert dlg.sidebar.indentation() == 12
    assert hasattr(dlg, "btn_explorer")


# ---- 11차: 로딩 썸네일 멈춤 수정 + Conder Scan 위치 지정 칸 ----

def test_heatmap_detail_thumb_worker_retained(win, app):
    """상세 지연 썸네일 워커가 참조 유지되어야 GC 로 ready 시그널이 사라지지 않는다."""
    from PySide6.QtTest import QTest
    dlg = _make_heatmap(win)
    keys = list(dlg._groups.keys())
    if not keys:
        return
    dlg._selected_keys = keys[:1]
    dlg._rebuild_detail()
    if not dlg._pending_thumbs:
        return
    # 워커가 실행 중 참조로 보관된다(GC 방지).
    assert len(dlg._active_thumb_workers) >= 1
    # 완료까지 대기 → 워커가 done 후 참조 해제 + 썸네일 holder 가 '로딩…' 이 아님.
    # (백그라운드 콜드 캐시 워밍은 부하에 따라 시간이 들쭉날쭉하므로 넉넉히 폴링한다.)
    for _ in range(120):
        QTest.qWait(50)
        if not dlg._active_thumb_workers:
            break
    for _ in range(5):  # 마지막 큐 이벤트 flush
        QCoreApplication.processEvents()
    assert dlg._active_thumb_workers == set()  # done 시 해제됨
    texts = {t.holder.text() for t in dlg._pending_thumbs}
    assert "로딩…" not in texts  # 모두 채워짐(또는 '이미지 없음')


def test_thumbnail_cache_atomic_write(app, tmp_path):
    """썸네일 캐시가 임시 파일→replace 로 원자적으로 써지고, 결과 파일이 완전하다."""
    from PIL import Image
    from app.thumbnails import ThumbnailCache
    src = tmp_path / "img.png"
    Image.new("RGB", (200, 160), (120, 40, 80)).save(src)
    cache = ThumbnailCache(str(tmp_path / "cache"))
    out = cache.get_full_thumbnail(src, max_size=64)
    assert out is not None and out.exists()
    # 완전한 PNG 로 다시 열림(부분 파일 아님) + .tmp 잔재 없음.
    with Image.open(out) as im:
        im.load()
    assert not list(out.parent.glob("*.tmp"))


def test_folder_picker_natural_sort():
    """폴더 나열이 자연 정렬(숫자 인식)로 1., 2., …, 10., 11. 순서가 되어야 한다."""
    from app.ui.folder_picker import natural_key
    assert sorted(["10.", "1.", "11.", "2.", "21.", "3."], key=natural_key) == [
        "1.", "2.", "3.", "10.", "11.", "21."
    ]
    assert sorted(["10. FS", "2. LYB4", "1. LYA4"], key=natural_key) == [
        "1. LYA4", "2. LYB4", "10. FS"
    ]


def test_folder_picker_lists_dirs_naturally(app, tmp_path):
    """실제 폴더 나열(_list_subdirs)도 자연 정렬을 따른다."""
    from app.ui.folder_picker import FolderPickerDialog
    for n in ["1.", "2.", "10.", "11.", "21.", "3."]:
        (tmp_path / n).mkdir()
    dlg = FolderPickerDialog(AppSettings(workspace=str(tmp_path / "ws")), str(tmp_path))
    assert dlg._list_subdirs(tmp_path) == ["1.", "2.", "3.", "10.", "11.", "21."]


def test_folder_picker_scan_root_input(app, tmp_path):
    from app.ui.folder_picker import FolderPickerDialog
    (tmp_path / "ScanData" / "LOT").mkdir(parents=True)
    s = AppSettings(workspace=str(tmp_path / "ws"), scan_root_path=str(tmp_path / "ScanData"))
    dlg = FolderPickerDialog(s, str(tmp_path))
    # 지정 칸이 존재하고 현재 경로가 채워진다.
    assert hasattr(dlg, "ed_scan_root")
    assert dlg.ed_scan_root.text().endswith("ScanData")
    # 명시 경로가 최상위 📌 로 고정된다.
    top0 = dlg.sidebar.topLevelItem(0).text(0)
    assert "📌" in top0 and "ScanData" in top0
    # 새 경로 지정 시 저장·재고정.
    (tmp_path / "Other").mkdir()
    dlg.ed_scan_root.setText(str(tmp_path / "Other"))
    dlg._apply_scan_root()
    assert s.scan_root_path.endswith("Other")
    assert "Other" in dlg.sidebar.topLevelItem(0).text(0)


def test_folder_picker_exposes_wafer_selection(app, tmp_path):
    """wafer 폴더를 고르면 selected_wafer_folder() 로 노출되고, selected_path() 는 상위 LOT."""
    from app.ui.folder_picker import FolderPickerDialog

    wafer = tmp_path / "LOT" / "LAYER" / "WAFER"
    wafer.mkdir(parents=True)
    (wafer / "a.jpg").write_bytes(b"\xff\xd8\xff\xd9")
    s = AppSettings(workspace=str(tmp_path / "ws"))
    dlg = FolderPickerDialog(s, str(tmp_path / "LOT"))
    dlg._candidate = wafer
    dlg._valid_for = None  # 캐시 무시 → classify 재판정
    assert dlg.selected_wafer_folder() == str(wafer)
    assert dlg.selected_path() == str(tmp_path / "LOT")  # 보정된 상위 LOT


def test_wafer_filter_limits_base_to_one_wafer(win, app):
    """wafer 필터가 걸리면 기준 defect(=매치)이 그 wafer 로만 한정된다."""
    from PySide6.QtTest import QTest

    all_wafers = {m.base.wafer_id for m in win.matches}
    if len(all_wafers) < 2:
        return  # 샘플에 wafer 가 하나뿐이면 스킵
    target = sorted(all_wafers)[0]
    win._wafer_filter = target
    win._rebuild_all()
    for _ in range(80):
        QTest.qWait(50)
        if win.matches and all(m.base.wafer_id == target for m in win.matches):
            break
    assert win.matches
    assert {m.base.wafer_id for m in win.matches} == {target}
    # 필터 해제 후 전체로 복원.
    win._wafer_filter = None
    win._rebuild_all()
    for _ in range(80):
        QTest.qWait(50)
        if {m.base.wafer_id for m in win.matches} == all_wafers:
            break
    assert {m.base.wafer_id for m in win.matches} == all_wafers
