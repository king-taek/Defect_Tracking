"""디바이스 DB 로더 및 제품 프로파일 병합 테스트."""

from __future__ import annotations

import openpyxl

from app import config
from app.device_db import load_device_db


def _make_db(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TESTDEV"
    ws["A1"] = "Package Info"
    ws["A2"], ws["B2"] = "X", 3
    ws["A3"], ws["B3"] = "Y", 3
    ws["A4"], ws["B4"] = "X1", 10.0
    ws["A5"], ws["B5"] = "Y1", 12.0
    ws["A6"] = "Map"
    # plus 모양 (die 가 있는 칸에 0)
    #  · 0 ·
    #  0 0 0
    #  · 0 ·
    ws["B7"] = 0
    ws["A8"], ws["B8"], ws["C8"] = 0, 0, 0
    ws["B9"] = 0
    # Package Info 없는 시트는 무시되어야 함
    ws2 = wb.create_sheet("NotADevice")
    ws2["A1"] = "hello"
    wb.save(path)


def test_load_device_db(tmp_path):
    p = tmp_path / "AOIDeviceDB.xlsx"
    _make_db(p)
    profs = load_device_db(p)
    assert set(profs.keys()) == {"TESTDEV"}  # Package Info 없는 시트 제외
    d = profs["TESTDEV"]
    assert d.x_count == 3 and d.y_count == 3
    assert d.pitch_x == 10000.0 and d.pitch_y == 12000.0
    assert d.die_map == frozenset({(1, 0), (0, 1), (1, 1), (2, 1), (1, 2)})


def test_register_devices_merges_into_products(tmp_path):
    p = tmp_path / "AOIDeviceDB.xlsx"
    _make_db(p)
    profs = load_device_db(p)
    try:
        config.register_devices(profs)
        assert "TESTDEV" in config.PRODUCTS
        cfg = config.PRODUCTS["TESTDEV"]
        assert cfg.kla_package_x_count == 3 and cfg.kla_package_y_count == 3
        assert cfg.source == "db"
        assert (1, 1) in cfg.die_map
        config.set_active_product("TESTDEV")
        assert config.active_product().die_map == profs["TESTDEV"].die_map
        assert config.kla_zero_x() == 1  # 3 // 2
    finally:
        config.set_active_product(config.DEFAULT_PRODUCT)
        config.PRODUCTS.pop("TESTDEV", None)
