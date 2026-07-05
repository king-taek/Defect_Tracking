# =============================================================================
# Defect Tracker — 단일 파일 배포본 (AUTO-GENERATED — 편집 금지)
#
# 이 파일은 `app/` + `main.py` 에서 자동 생성된 산출물입니다. 소스의 진실은 모듈식
# 소스이며, 이 파일을 직접 고치지 마세요. 재생성:
#     python tools/build_single_file.py
# 버전: 1.33.70   (실행: python defect_tracker.py / 의존성 설치: python bootstrap.py)
# =============================================================================


from __future__ import annotations


# --- 의존성 프리앰블: 무거운 import 전에 친절한 안내(원본 main.py UX 보존) ---
import importlib.util as _importlib_util
import sys


def _require_dependencies() -> None:
    _req = {"PySide6": "PySide6", "PIL": "Pillow", "openpyxl": "openpyxl"}
    _missing = [name for mod, name in _req.items() if _importlib_util.find_spec(mod) is None]
    if _missing:
        sys.stderr.write(
            "필요한 라이브러리가 설치되어 있지 않습니다: " + ", ".join(_missing) + "\n"
            "다음 명령으로 설치하세요:\n    python bootstrap.py\n"
            "또는:\n    pip install PySide6 Pillow openpyxl\n"
        )
        raise SystemExit(1)


_require_dependencies()


import hashlib
import importlib.util
import io
import json
import logging
import math
import os
import re
import shutil
import statistics
import subprocess
import sys
import tempfile
import threading
import zipfile
from collections import Counter, OrderedDict, defaultdict
from concurrent.futures import ThreadPoolExecutor
from dataclasses import asdict, dataclass, field
from datetime import datetime
from enum import Enum
from logging.handlers import RotatingFileHandler
from pathlib import Path, Path as _Path
from typing import Any, Callable, Iterable, Optional
from urllib.request import Request, urlopen

from PIL import Image
from PySide6.QtCore import QDir, QEasingCurve, QEvent, QEventLoop, QMargins, QObject, QParallelAnimationGroup, QPoint, QPropertyAnimation, QRect, QRectF, QRunnable, QSize, QStorageInfo, QThreadPool, QTimer, QUrl, Qt, Signal, Slot
from PySide6.QtGui import QColor, QDesktopServices, QFont, QGuiApplication, QImage, QImageReader, QKeySequence, QPainter, QPen, QPixmap, QShortcut
from PySide6.QtWidgets import QAbstractSpinBox, QApplication, QCheckBox, QComboBox, QDialog, QDialogButtonBox, QDoubleSpinBox, QFileDialog, QFormLayout, QFrame, QGraphicsOpacityEffect, QGridLayout, QHBoxLayout, QLabel, QLayout, QLayoutItem, QLineEdit, QListWidget, QListWidgetItem, QMainWindow, QMenu, QMessageBox, QProgressBar, QPushButton, QScrollArea, QSizePolicy, QSplashScreen, QSplitter, QTreeWidget, QTreeWidgetItem, QVBoxLayout, QWidget
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


__version__ = "1.33.70"


# 모듈 맵 (위상순서, leaf → top):
#   app/config.py                     — 전역 상수 및 사용자 설정 관리.
#   app/device_db.py                  — 외부 디바이스 DB(AOIDeviceDB.xlsx) 로더 — 제품 일반화.
#   app/heatmap.py                    — 히트맵 순수 로직 — defect 밀도 격자 구성과 die 하위셀(subcell) 분할.
#   app/logging_config.py             — 애플리케이션 로깅 설정 (관측성).
#   app/models.py                     — 도메인 데이터 모델.
#   app/safety.py                     — 원본 보호 게이트 (문서 Section 1 — 절대 안전 규칙).
#   app/ui/flow_layout.py             — 줄바꿈(Flow) 레이아웃.
#   app/ui/image_loader.py            — 비동기 이미지 로더 (문서 Section 10 — UI 멈춤 최소화).
#   app/ui/theme.py                   — 다크 + 파란 네온 테마 (문서 Section 9).
#   app/wafermap_align.py             — 웨이퍼 맵 die 정합(alignment) — 관측 die 와 디바이스 DB die_map 의 원점 맞추기.
#   app/clustering.py                 — defect 근접 클러스터링 + layer 간 교차 매칭 (순수 로직, UI 무관).
#   app/diagnostics.py                — 좌표 추출 실패 진단 리포트(개발용) — 단일 markdown 파일로 관리.
#   app/layout.py                     — Layer 폴더명 정규화 및 비교 그리드 배치 (문서 Section 8.2 / 8.4).
#   app/matcher.py                    — 매칭 엔진 (문서 Section 8.3 + 원본 AOI 도구 Module_Compare 알고리즘).
#   app/parsers/camtek_filename.py    — Camtek 파일명에서 직접 좌표 추출 (문서 Section 6 / 13.3.4 + 실제 AOI 도구 스키마).
#   app/parsers/camtek_ini.py         — ColorImageGrabingInfo.ini 기반 Camtek 좌표 산출 (문서 Section 13.3).
#   app/parsers/kla_info.py           — KLA info(.001) 기반 좌표 변환 (문서 Section 13.2 및 KLA 변환 보고서).
#   app/thumbnails.py                 — 썸네일 생성/캐시 (문서 Section 8.6, 10).
#   app/ui/busy_overlay.py            — 로딩(작업 중) 오버레이 — 부모 위 반투명 막 + 중앙 카드(부드러운 스피너·메시지·진행바).
#   app/ui/controls.py                — 좌측 사이드바 컨트롤 및 하단 탐색 바 (문서 Section 8.1, 8.3, 8.5).
#   app/ui/help_dialog.py             — 도움말 다이얼로그 — 단축키 + 기능 안내(섹션 구성, 스크롤).
#   app/ui/image_viewer.py            — 원본 확대 뷰어 (read-only) — defect review 사용성.
#   app/ui/notifications.py           — 비차단 알림 배너(토스트) — 매끄러운 오류/안내 처리 (문서 Section 9 / 사용성).
#   app/ui/settings_dialog.py         — 설정 다이얼로그 — 작업공간/출력 폴더/기본 허용오차/업데이트 확인 (사용성).
#   app/ui/splash.py                  — 시작 스플래시 — 무거운 MainWindow 임포트/구성 전에 즉시 피드백을 준다.
#   app/ui/wafer_map.py               — 웨이퍼 맵 네비게이터 — 현재 wafer 의 die 격자를 매칭 상태로 색칠하고,
#   app/ui/widgets.py                 — 재사용 위젯 및 애니메이션 헬퍼 (문서 Section 8.6, 9).
#   app/updater.py                    — 자동 업데이트 — 메인 브랜치를 가져와 적용.
#   app/export/excel_report.py        — Excel 결과 출력 (문서 Section 8.7).
#   app/scanner.py                    — LOT 폴더 스캔 및 인덱스 구축 (문서 Section 4, 8.1, 8.2).
#   app/ui/cluster_view.py            — 클러스터 defect 표시용 공유 위젯 (히트맵·메인 매치 공통).
#   app/ui/compare_grid.py            — Layer 비교 그리드 (문서 Section 8.4).
#   app/ui/export_dialog.py           — 결과 출력 트레이 다이얼로그 (문서 Section 8.7 재설계).
#   app/ui/nomatch_gallery.py         — 미매칭(기준 layer 의 defect 과 어떤 비교 layer 와도 매칭 안 된) 사진 갤러리.
#   app/ui/thumbnail_strip.py         — 상단 기준 썸네일 스트립 (문서 Section 8.6).
#   app/ui/folder_picker.py           — 자재(LOT) 폴더 선택 다이얼로그 — 브레드크럼 + 한 단계 목록 + 사이드바.
#   app/ui/heatmap_dialog.py          — Defect 히트맵 팝업 (항목 4·5).
#   app/workers.py                    — 백그라운드 작업 (문서 Section 10 성능 요구사항).
#   app/ui/main_window.py             — 메인 윈도우 — 전체 workflow 조립 (문서 Section 8 전체).
#   main.py                           — 진입점


# =============================================================================
# app/config.py   [#0]
# =============================================================================
"""전역 상수 및 사용자 설정 관리.

좌표 변환 상수(pitch, package count)와 매칭 기본값을 한 곳에 모아 두어
DEVA 외 제품으로 확장할 때 이 파일만 수정하면 되도록 한다.

사용자 설정(마지막 LOT 경로, tolerance, 출력 폴더 등)은 output workspace 내부의
JSON 파일에 저장한다. 원본 폴더에는 어떤 것도 쓰지 않는다.
"""




# ---------------------------------------------------------------------------
# 제품 프로파일 (좌표 변환 상수 묶음) — 제품별 확장은 PRODUCTS 에 추가만 하면 된다.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ProductConfig:
    """제품별 좌표 변환/패키지 상수 묶음 (문서 Section 13)."""

    key: str
    name: str
    camtek_pitch_x: float
    camtek_pitch_y: float
    kla_package_x_count: int
    kla_package_y_count: int
    camtek_col_offset: int = 2
    camtek_row_base: int = 7
    die_map: frozenset = field(default_factory=frozenset)  # 존재하는 (col,row) (비면 사각 전체)
    source: str = "builtin"  # builtin / db

    def kla_zero_x(self) -> int:
        return self.kla_package_x_count // 2

    def kla_zero_y(self) -> int:
        return self.kla_package_y_count // 2


def register_devices(profiles: dict) -> None:
    """외부 디바이스 DB(DeviceProfile 사전)를 PRODUCTS 에 병합한다.

    DB 가 제공하지 않는 Camtek INI 변환 상수(col_offset/row_base)는 기본값을 쓴다.
    """
    for key, prof in profiles.items():
        PRODUCTS[key] = ProductConfig(
            key=key,
            name=getattr(prof, "name", key),
            camtek_pitch_x=getattr(prof, "pitch_x", 0.0) or _DEFAULT_CFG.camtek_pitch_x,
            camtek_pitch_y=getattr(prof, "pitch_y", 0.0) or _DEFAULT_CFG.camtek_pitch_y,
            kla_package_x_count=prof.x_count,
            kla_package_y_count=prof.y_count,
            die_map=prof.die_map,
            source="db",
        )


PRODUCTS: dict[str, ProductConfig] = {
    "DEVAINT": ProductConfig(
        key="DEVAINT",
        name="DEVA Live",
        # AOIDeviceDB.xlsx "DEVA Live" 시트 실측값(정답 도구 원본 기준).
        # "DEVA"(다른 시트, Y=5/23-die)과 혼동하지 말 것 — 실제 운영 KLARF
        # SampleTestPlan(YINDEX -3~+2, 30쌍)과 일치하는 건 "DEVA Live"(Y=6/30-die).
        camtek_pitch_x=37170.0,
        camtek_pitch_y=44830.0,
        camtek_col_offset=2,
        camtek_row_base=7,
        kla_package_x_count=7,
        kla_package_y_count=6,
    ),
}
DEFAULT_PRODUCT = "DEVAINT"
_active_product = DEFAULT_PRODUCT


def active_product() -> ProductConfig:
    return PRODUCTS.get(_active_product, PRODUCTS[DEFAULT_PRODUCT])


def set_active_product(key: str) -> None:
    """활성 제품 프로파일을 바꾼다(다음 스캔부터 좌표 변환에 적용)."""
    global _active_product
    if key in PRODUCTS:
        _active_product = key


def ensure_die_map_product() -> None:
    """활성 제품에 die_map 이 없으면(빌트인 폴백) 같은 패키지 크기의 DB 제품으로 승격한다.

    DB 자동 로드 직후 호출한다. 빌트인 제품(die_map 비어 사각 표시)만 있을 때에도, DB 에
    같은 패키지(X×Y) 크기의 die_map 이 있으면 그 DB 제품을 활성화해 웨이퍼맵이 기본적으로
    실제 die 모양으로 뜨게 한다. 사용자가 이미 die_map 있는 제품을 고른 경우엔 건드리지 않는다.
    """
    cur = active_product()
    if cur.die_map:
        return
    for key, prod in PRODUCTS.items():
        if (
            prod.source == "db"
            and prod.die_map
            and prod.kla_package_x_count == cur.kla_package_x_count
            and prod.kla_package_y_count == cur.kla_package_y_count
        ):
            set_active_product(key)
            return


def match_product_for_path(lot_path) -> tuple[str | None, int]:
    """자재(LOT) 경로 구성요소에서 등록 제품을 자동 인식한다.

    자재명·상위(device 등) 폴더명을 영숫자 소문자 토큰으로 정규화하고, 등록 제품의
    key/name 토큰이 그 안에 부분 문자열로 나타나면 매칭으로 본다. 가장 긴(구체적인)
    제품 토큰을 우선한다. 반환은 (제품 key 또는 None, 점수=매칭 토큰 길이).
    """
    from pathlib import Path


    p = Path(lot_path)
    parts = [p.name] + [par.name for par in list(p.parents)[:3]]
    path_tokens = [_canon_token(s) for s in parts if s]
    path_tokens = [t for t in path_tokens if t]

    best_key: str | None = None
    best_score = 0
    for key, prod in PRODUCTS.items():
        for cand in (key, prod.name):
            ct = _canon_token(cand)
            if len(ct) < 4:  # 너무 짧은 토큰은 오매칭 위험 → 제외
                continue
            if any(ct in t for t in path_tokens):
                if len(ct) > best_score:
                    best_key, best_score = key, len(ct)
    return best_key, best_score


# ---- 하위호환 상수 (기본 제품 값) — 샘플데이터/기존 테스트가 참조 ----
_DEFAULT_CFG = PRODUCTS[DEFAULT_PRODUCT]
CAMTEK_PITCH_X = _DEFAULT_CFG.camtek_pitch_x
CAMTEK_PITCH_Y = _DEFAULT_CFG.camtek_pitch_y
CAMTEK_COL_OFFSET = _DEFAULT_CFG.camtek_col_offset
CAMTEK_ROW_BASE = _DEFAULT_CFG.camtek_row_base
KLA_PACKAGE_X_COUNT = _DEFAULT_CFG.kla_package_x_count
KLA_PACKAGE_Y_COUNT = _DEFAULT_CFG.kla_package_y_count


def kla_zero_x() -> int:
    return active_product().kla_zero_x()


def kla_zero_y() -> int:
    return active_product().kla_zero_y()


# ---------------------------------------------------------------------------
# 애플리케이션 / 자동 업데이트
# ---------------------------------------------------------------------------

APP_NAME = "Defect Tracker"

# 제작 크레딧(UI 곳곳 표기용 단일 출처)
CREDITS = "Designed by JinHan Kim, Developed by HyunTaek Lim"


def dev_mode(settings=None) -> bool:
    """개발자 모드 여부. 환경변수 DEFECT_TRACKER_DEV 또는 설정(AppSettings.dev_mode)로 켠다.

    - 환경변수가 참(1/true/on/yes/y)이면 **설정보다 우선**해 항상 켜진다(배포/디버그용).
    - 환경변수가 없으면 전달된 ``settings.dev_mode`` 값을 따른다(설정 창 토글로 저장).
    켜지면 파일 로그 생성·진단 리포트·설정의 로그 경로 노출이 활성화된다.
    """
    val = os.environ.get("DEFECT_TRACKER_DEV", "").strip().lower()
    if val in ("1", "true", "on", "yes", "y"):
        return True
    if settings is not None:
        return bool(getattr(settings, "dev_mode", False))
    return False

# 자동 업데이트 대상 저장소(메인 브랜치를 가져와 적용)
UPDATE_OWNER = "king-taek"
UPDATE_REPO = "defect_tracking"
UPDATE_BRANCH = "main"


# ---------------------------------------------------------------------------
# 매칭 / UI 기본값
# ---------------------------------------------------------------------------

# 기본 허용 오차 (Section 8.3) - 좌표 단위는 파일/INI 값 그대로 사용
DEFAULT_TOLERANCE = 100.0

# defect 근접 클러스터링 거리(같은 die 안에서 이 값 미만이면 하나로 묶음). 설정에서 조절.
DEFAULT_CLUSTER_RADIUS = 50.0

# 상단 썸네일은 사진 중앙 일부 구간만 확대 (Section 8.6) - 중앙 비율
# 0.20 = 중앙 20% 를 잘라 ≈5× 확대(고정).
THUMBNAIL_CENTER_RATIO = 0.20
THUMBNAIL_SIZE = 140

# Layer 비교 그리드 기본 배치 (Section 8.4). 2열 그리드, 위에서 아래로.
DEFAULT_LAYER_GRID = [
    ["LYA4", "LYB4"],
    ["LYA3", "LYB3"],
    ["LYA2", "LYB2"],
    ["LYA1", "FS"],
]

# 지원 이미지 확장자
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}


# ---------------------------------------------------------------------------
# Output workspace (2차 보호: 결과/캐시는 원본 밖에만 생성 - Section 1.1)
# ---------------------------------------------------------------------------

def default_workspace() -> Path:
    """원본 폴더 밖의 기본 출력 작업공간 경로."""
    base = os.environ.get("LOCALAPPDATA") or os.environ.get("XDG_DATA_HOME")
    if not base:
        base = str(Path.home())
    return Path(base) / "DefectTracker"


def bundled_device_db_path() -> Path | None:
    """앱과 함께 배포되는 기본 디바이스 DB(data/AOIDeviceDB.xlsx) 경로.

    사용자가 별도 DB 경로를 지정하지 않았을 때 자동 로드용으로 쓴다(원본은 read-only).
    - 소스 실행: 리포지토리 루트의 data/AOIDeviceDB.xlsx
    - PyInstaller onefile: sys._MEIPASS/data/AOIDeviceDB.xlsx
    찾지 못하면 None.
    """
    import sys

    candidates: list[Path] = []
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidates.append(Path(meipass) / "data" / "AOIDeviceDB.xlsx")
    # config.py -> app/ -> repo root
    # 단일 파일 배포: 파일 자기 디렉터리 및 CWD 기준으로 data/ 탐색(없으면 미로드).
    _here = Path(__file__).resolve().parent
    candidates.append(_here / "data" / "AOIDeviceDB.xlsx")
    candidates.append(Path.cwd() / "data" / "AOIDeviceDB.xlsx")
    for c in candidates:
        try:
            if c.is_file():
                return c
        except OSError:
            continue
    return None


def default_log_dir() -> str:
    """로그 전용 기본 경로(비어 있으면 workspace/logs 를 씀).

    Windows 배포 환경에서만 지정된 고정 경로를 쓰고, 그 외(테스트/CI 등)에서는
    빈 문자열로 두어 기존 workspace/logs 폴백을 그대로 따른다.
    """
    if os.name == "nt":
        return r"C:\Users\304236\Desktop\Defect_Tracking-main\log"
    return ""


@dataclass
class AppSettings:
    """사용자 설정. output workspace 내 settings.json에 저장된다."""

    workspace: str = field(default_factory=lambda: str(default_workspace()))
    last_lot_folder: str = ""
    output_folder: str = ""  # 비어 있으면 workspace/exports 사용
    tolerance: float = DEFAULT_TOLERANCE
    cluster_radius: float = DEFAULT_CLUSTER_RADIUS  # defect 근접 클러스터링 거리(설정에서 조절)
    base_layer: str = ""
    compare_layers: list[str] = field(default_factory=list)
    recent_folders: list[str] = field(default_factory=list)  # 최근 연 자재 폴더(최대 5)
    favorite_folders: list[str] = field(default_factory=list)  # 즐겨찾기(고정) 상위 폴더
    scan_root_name: str = "Conder Scan"  # 폴더 트리 최상위 고정: 이 이름의 폴더가 있는 드라이브
    scan_root_path: str = ""  # 명시 지정한 스캔 데이터 폴더(있으면 최상위 📌 고정, 비면 이름 탐지)
    product: str = DEFAULT_PRODUCT  # 활성 제품 프로파일(좌표 변환 상수)
    device_db_path: str = ""  # 외부 AOIDeviceDB.xlsx 경로(비면 번들 DB 자동 로드)
    thumbnail_center_ratio: float = THUMBNAIL_CENTER_RATIO  # 상단 썸네일 중앙 crop 비율(확대율)
    ui_font_size: str = "normal"  # 전체 UI 글자 크기: normal(보통) / large(크게)
    heatmap_layout: int = 0  # 히트맵 팝업 레이아웃 프리셋 인덱스(마지막 선택 기억)
    log_dir: str = field(default_factory=default_log_dir)  # 비어 있으면 workspace/logs 사용
    window_geometry: str = ""  # "x,y,w,h" — 모니터 환경별 창 크기/위치 기억(최대화 해제 시 복원)
    window_maximized: bool = True  # 시작 시 최대화(기본). 사용자가 해제하면 False 로 저장
    sidebar_width: int = 240  # 좌측 사이드바 폭(스플리터) 기억
    auto_update_check: bool = True  # 시작 시 백그라운드 업데이트 확인
    update_token: str = ""  # (선택) 비공개 저장소용 GitHub 토큰. public 이면 빈값.
    dev_mode: bool = False  # 개발자 모드(파일 로그·진단·로그 경로 UI). 설정 창에서 토글.

    # ---- 경로 헬퍼 -------------------------------------------------------
    @property
    def workspace_path(self) -> Path:
        return Path(self.workspace)

    @property
    def cache_path(self) -> Path:
        return self.workspace_path / "cache"

    @property
    def exports_path(self) -> Path:
        if self.output_folder:
            return Path(self.output_folder)
        return self.workspace_path / "exports"

    @property
    def log_dir_path(self) -> Path:
        if self.log_dir:
            return Path(self.log_dir)
        return self.workspace_path / "logs"

    def settings_file(self) -> Path:
        return self.workspace_path / "settings.json"

    # ---- 저장/로드 -------------------------------------------------------
    def ensure_workspace(self) -> None:
        self.workspace_path.mkdir(parents=True, exist_ok=True)
        self.cache_path.mkdir(parents=True, exist_ok=True)

    def save(self) -> None:
        """설정을 원자적으로 저장한다(임시파일 작성 후 교체 → 크래시 시 손상 방지)."""
        self.ensure_workspace()
        data = asdict(self)
        target = self.settings_file()
        tmp = target.with_name(target.name + ".tmp")
        tmp.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        os.replace(tmp, target)  # 같은 디렉터리 내 원자적 교체

    @classmethod
    def load(cls, workspace: Path | None = None) -> "AppSettings":
        ws = workspace or default_workspace()
        settings_file = ws / "settings.json"
        if settings_file.exists():
            try:
                raw: dict[str, Any] = json.loads(settings_file.read_text(encoding="utf-8"))
                known = {f for f in cls.__dataclass_fields__}  # type: ignore[attr-defined]
                return cls(**{k: v for k, v in raw.items() if k in known})
            except (json.JSONDecodeError, TypeError, ValueError) as exc:
                logging.getLogger("defect_tracker.config").warning(
                    "settings.json 을 읽지 못해 기본값을 사용합니다: %s", exc
                )
        return cls(workspace=str(ws))


# =============================================================================
# app/device_db.py   [#1]
# =============================================================================
"""외부 디바이스 DB(AOIDeviceDB.xlsx) 로더 — 제품 일반화.

원본 자료 구조: 시트 1개 = 디바이스 1개. 각 시트에 다음이 있다.
  Package Info
    X  | <package X die 개수>
    Y  | <package Y die 개수>
    X1 | <die pitch X (mm 근사)>
    Y1 | <die pitch Y (mm 근사)>
  Map
    <Y개 행 × X개 열의 격자. die 가 존재하는 칸에 값(0)이 있고, 없는 칸은 빈칸>

이 정보로 제품별 package count·pitch·die 배치(웨이퍼 모양)를 구성한다.
원본/DB 파일은 read-only 로만 읽는다.
"""



device_db__log = logging.getLogger("defect_tracker.device_db")

_PKG_KEYS = ("X", "Y", "X1", "Y1")


@dataclass(frozen=True)
class DeviceProfile:
    """디바이스 1개의 패키지/배치 정보."""

    key: str
    name: str
    x_count: int
    y_count: int
    pitch_x: float  # die pitch X (단위는 좌표계와 동일하게 ×1000 환산)
    pitch_y: float
    die_map: frozenset = field(default_factory=frozenset)  # 존재하는 (col,row) 집합


def _to_number(v) -> float | None:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_sheet(title: str, rows: list[tuple]) -> DeviceProfile | None:
    """한 시트(rows: values_only 행 목록)에서 DeviceProfile 을 만든다. 실패 시 None."""
    pkg: dict[str, float] = {}
    map_start: int | None = None
    for i, r in enumerate(rows):
        if not r:
            continue
        c0 = "" if r[0] is None else str(r[0]).strip()
        if c0 in _PKG_KEYS and len(r) > 1:
            num = _to_number(r[1])
            if num is not None:
                pkg[c0] = num
        elif c0.lower() == "map":
            map_start = i + 1

    if "X" not in pkg or "Y" not in pkg:
        return None
    x_count = int(round(pkg["X"]))
    y_count = int(round(pkg["Y"]))
    if x_count <= 0 or y_count <= 0 or x_count > 200 or y_count > 200:
        return None
    pitch_x = pkg.get("X1", 0.0) * 1000.0
    pitch_y = pkg.get("Y1", 0.0) * 1000.0

    die_map: set[tuple[int, int]] = set()
    if map_start is not None:
        for ri in range(y_count):
            idx = map_start + ri
            row = rows[idx] if idx < len(rows) else ()
            for ci in range(x_count):
                v = row[ci] if row and ci < len(row) else None
                if v is not None and str(v).strip() != "":
                    die_map.add((ci, ri))

    return DeviceProfile(
        key=title,
        name=title,
        x_count=x_count,
        y_count=y_count,
        pitch_x=pitch_x,
        pitch_y=pitch_y,
        die_map=frozenset(die_map),
    )


def load_device_db(path: str | Path) -> dict[str, DeviceProfile]:
    """AOIDeviceDB.xlsx 를 읽어 {device_name: DeviceProfile} 를 반환한다.

    Package Info(X/Y) 가 없는 시트는 건너뛴다. 파일/시트 오류는 로깅 후 건너뛴다.
    """
    import openpyxl  # 지연 임포트(시작 비용 절감)

    out: dict[str, DeviceProfile] = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        device_db__log.warning("디바이스 DB 로드 실패(%s): %s", path, exc)
        return out
    try:
        for ws in wb.worksheets:
            try:
                rows = list(ws.iter_rows(values_only=True))
                prof = _parse_sheet(ws.title, rows)
            except Exception as exc:  # noqa: BLE001
                device_db__log.warning("시트 파싱 실패(%s): %s", ws.title, exc)
                prof = None
            if prof is not None:
                out[prof.key] = prof
    finally:
        wb.close()
    device_db__log.info("디바이스 DB 로드: %d개 디바이스 (%s)", len(out), path)
    return out


# =============================================================================
# app/heatmap.py   [#2]
# =============================================================================
"""히트맵 순수 로직 — defect 밀도 격자 구성과 die 하위셀(subcell) 분할.

UI(팝업)와 분리해 단위 테스트로 검증한다. 웨이퍼맵에 defect 위치를 표시/클릭하기 위한
집계와, die 개수가 적을 때(50개 미만) 각 die 를 5×5(25) 하위셀로 나눠 die 내부 local
좌표(x,y)로 defect 을 구분 배치하는 매핑을 제공한다(항목 4·5).
"""



# die 하위셀 격자(항목 5): die 가 클 때 die 내부를 5열×5행(25칸)으로 나눈다.
SUB_COLS = 5
SUB_ROWS = 5
# die 개수가 이 값 미만이면 하위셀 분할을 적용한다.
SUBDIVIDE_THRESHOLD = 50


def should_subdivide(die_count: int) -> bool:
    """die 개수가 임계값 미만이면 하위셀 분할을 적용한다."""
    return 0 < die_count < SUBDIVIDE_THRESHOLD


def _bucket(value: float, lo: float, hi: float, n: int) -> int:
    """value 를 [lo,hi] 범위에서 n 개 구간 중 하나(0..n-1)로 버킷화한다."""
    if hi <= lo:
        return 0
    frac = (value - lo) / (hi - lo)
    b = int(frac * n)
    return max(0, min(n - 1, b))


def local_ranges(records) -> tuple[tuple[float, float], tuple[float, float]]:
    """defect record 들의 die 내부 local 좌표(x,y) 관측 범위((xmin,xmax),(ymin,ymax)).

    좌표가 없는 record 는 건너뛴다. 비어 있으면 ((0,1),(0,1)).
    """
    xs = [r.x for r in records if getattr(r, "x", None) is not None]
    ys = [r.y for r in records if getattr(r, "y", None) is not None]
    if not xs or not ys:
        return (0.0, 1.0), (0.0, 1.0)
    return (min(xs), max(xs)), (min(ys), max(ys))


def subcell_of(
    x: float,
    y: float,
    x_range: tuple[float, float],
    y_range: tuple[float, float],
) -> tuple[int, int]:
    """die 내부 local 좌표(x,y)를 하위셀 좌표(sub_col, sub_row)로 매핑한다.

    y 는 위(작은 값)→아래로 증가하도록 행 인덱스를 매긴다(화면 좌표계).
    """
    sc = _bucket(x, x_range[0], x_range[1], SUB_COLS)
    sr = _bucket(y, y_range[0], y_range[1], SUB_ROWS)
    return sc, sr


@dataclass(frozen=True)
class HeatKey:
    """히트맵 셀 키 — die (col,row) + (하위셀 sub_col/sub_row, 미분할이면 -1)."""

    col: int
    row: int
    sub_col: int = -1
    sub_row: int = -1

    @property
    def subdivided(self) -> bool:
        return self.sub_col >= 0 and self.sub_row >= 0


def group_defects(
    entries: list[tuple[int, object]],
    subdivide: bool,
    x_range: tuple[float, float] | None = None,
    y_range: tuple[float, float] | None = None,
) -> dict[HeatKey, list[int]]:
    """(index, DefectRecord) 목록을 히트맵 셀(HeatKey)별 index 리스트로 집계한다.

    subdivide=True 이면 die 내부를 하위셀로 나눠 키에 (sub_col,sub_row)를 포함한다.
    좌표(col/row) 가 없는 record 는 제외한다.
    """
    out: dict[HeatKey, list[int]] = {}
    for idx, rec in entries:
        col = getattr(rec, "col", None)
        row = getattr(rec, "row", None)
        if col is None or row is None:
            continue
        if subdivide and x_range is not None and y_range is not None \
                and getattr(rec, "x", None) is not None \
                and getattr(rec, "y", None) is not None:
            sc, sr = subcell_of(rec.x, rec.y, x_range, y_range)
            key = HeatKey(int(col), int(row), sc, sr)
        else:
            key = HeatKey(int(col), int(row))
        out.setdefault(key, []).append(idx)
    return out


# =============================================================================
# app/logging_config.py   [#3]
# =============================================================================
"""애플리케이션 로깅 설정 (관측성).

실데이터 환경(네트워크 경로, 다양한 자재 폴더)에서 발생하는 문제를 사후에 진단할 수
있도록 파일 로그를 남긴다. 로그는 항상 지정된 로그 디렉터리(원본 폴더 밖)에만 기록한다.

  - 콘솔: WARNING 이상(개발/터미널 실행 시)
  - 파일: DEBUG 이상, `<log_dir>/defect_tracker.log` (회전: 25MB×5)

원본 보호 원칙에 따라 로그 파일은 절대 원본 LOT 폴더에 쓰지 않는다.
"""



_CONFIGURED = False
_ROOT_NAME = "defect_tracker"
logging_config__MAX_LOG_SIZE = 25 * 1024 * 1024  # 25 MB


def get_logger(name: str = _ROOT_NAME) -> logging.Logger:
    """`defect_tracker.*` 네임스페이스 로거를 반환한다."""
    if name == _ROOT_NAME or name.startswith(_ROOT_NAME + "."):
        return logging.getLogger(name)
    return logging.getLogger(f"{_ROOT_NAME}.{name}")


def setup_logging(log_dir: Optional[Path] = None, level: int = logging.DEBUG) -> logging.Logger:
    """루트 `defect_tracker` 로거를 1회 구성한다(중복 핸들러 방지).

    log_dir 이 주어지고 쓰기 가능하면 회전 파일 핸들러를 그 디렉터리에 바로 추가한다
    (하위에 별도 "logs" 폴더를 만들지 않음 — 호출자가 최종 로그 디렉터리를 넘긴다).
    파일 핸들러 설치에 실패해도(권한 등) 콘솔 로깅은 유지하며 예외를 던지지 않는다.
    """
    global _CONFIGURED
    logger = logging.getLogger(_ROOT_NAME)
    logger.setLevel(level)
    logger.propagate = False

    if not _CONFIGURED:
        console = logging.StreamHandler(sys.stderr)
        console.setLevel(logging.WARNING)
        console.setFormatter(logging.Formatter("%(levelname)s [%(name)s] %(message)s"))
        logger.addHandler(console)
        _CONFIGURED = True

    if log_dir is not None:
        _attach_file_handler(logger, Path(log_dir))

    return logger


def _has_file_handler(logger: logging.Logger) -> bool:
    return any(isinstance(h, RotatingFileHandler) for h in logger.handlers)


def _attach_file_handler(logger: logging.Logger, log_dir: Path) -> None:
    if _has_file_handler(logger):
        return
    try:
        log_dir.mkdir(parents=True, exist_ok=True)
        handler = RotatingFileHandler(
            log_dir / "defect_tracker.log",
            maxBytes=logging_config__MAX_LOG_SIZE,
            backupCount=5,
            encoding="utf-8",
        )
        handler.setLevel(logging.DEBUG)
        handler.setFormatter(
            logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")
        )
        logger.addHandler(handler)
        logger.info("로깅 시작 (log_dir=%s)", log_dir)
    except OSError:
        # 파일 로깅 실패는 치명적이지 않다(콘솔 로깅 유지).
        logger.warning("로그 파일 핸들러를 설치하지 못했습니다: %s", log_dir)


# =============================================================================
# app/models.py   [#4]
# =============================================================================
"""도메인 데이터 모델.

좌표 변환 결과는 col_row_x_y 형식(문서 Section 13)으로 정규화되어 DefectRecord 에 담긴다.
모든 모델은 원본 파일 경로만 보관하고 원본을 수정하지 않는다.
"""




class Source(str, Enum):
    """defect 이미지의 좌표 출처."""

    CAMTEK_FILENAME = "Camtek(파일명)"
    CAMTEK_INI = "Camtek(INI)"
    KLA = "KLA"
    UNKNOWN = "Unknown"


class ParseStatus(str, Enum):
    """좌표 파싱 상태 (문서 Section 13의 예외 표기와 대응)."""

    OK = "OK"
    NOT_FOUND = "NOT_FOUND"
    INFO_FILE_NOT_FOUND = "INFO_FILE_NOT_FOUND"
    INVALID_INFO = "INVALID_INFO"


@dataclass
class DefectRecord:
    """단일 defect 이미지와 그 위치 정보.

    col/row 는 die 위치, x/y 는 die 내부 local 좌표(문서 Section 13.1).
    파싱 실패 시 status 로 사유를 보관하고 col/row/x/y 는 None 이다.
    """

    image_path: Path
    wafer_id: str
    layer: str  # 정규화된 layer 토큰 (예: LYA4, LYC3)
    layer_folder: str  # 원본 폴더명 (예: "2. LYC3_재리뷰")
    source: Source = Source.UNKNOWN
    status: ParseStatus = ParseStatus.OK
    col: Optional[int] = None
    row: Optional[int] = None
    x: Optional[float] = None
    y: Optional[float] = None
    defect_name: str = ""
    note: str = ""  # 파싱 관련 부가 정보
    diag: dict = field(default_factory=dict)  # 진단용 컨텍스트(폴더 파일 목록, info 내용 등)
    dx_size: Optional[float] = None  # defect 크기 X (파일명에 있을 때)
    dy_size: Optional[float] = None  # defect 크기 Y
    d_area: Optional[float] = None  # defect 면적

    @property
    def ok(self) -> bool:
        return (
            self.status == ParseStatus.OK
            and self.col is not None
            and self.row is not None
            and self.x is not None
            and self.y is not None
        )

    @property
    def position_key(self) -> str:
        """col_row_x_y 형식의 정수 위치 키 (문서 최종 출력 형식)."""
        if not self.ok:
            return self.status.value
        return f"{self.col}_{self.row}_{round(self.x)}_{round(self.y)}"

    @property
    def die_key(self) -> tuple[int, int]:
        return (int(self.col), int(self.row))  # type: ignore[arg-type]

    @property
    def die_key_full(self) -> tuple[str, int, int]:
        """매칭 인덱스 키: (wafer_id, col, row)."""
        return (self.wafer_id, int(self.col), int(self.row))  # type: ignore[arg-type]

    def distance_to(self, other: "DefectRecord") -> Optional[float]:
        """같은 die 내 local 좌표 유클리드 거리. 둘 중 하나라도 좌표 없으면 None."""
        if not self.ok or not other.ok:
            return None
        return math.hypot(self.x - other.x, self.y - other.y)  # type: ignore[operator]


@dataclass
class LayerInfo:
    """LOT 하위 layer 폴더 하나."""

    folder_name: str  # 원본 폴더명 (예: "1. LYA4_재리뷰")
    canonical: str  # 정규화 토큰 (예: LYA4)
    path: Path
    is_re_review: bool = False  # _재리뷰 여부(레벨 ≥1)
    re_review_level: int = 0  # 0 없음 / 1 재리뷰 / 2 재재리뷰 …
    display: str = ""  # 선택 UI/매칭에 쓰는 표시 이름(충돌 시에만 canonical 과 다름)


class NoMatchReason(str, Enum):
    """비교 매칭 실패 사유(화면 진단 문구의 단일 출처)."""

    NONE = "NONE"  # 매칭 성공
    NO_DIE_PHOTO = "NO_DIE_PHOTO"  # 같은 die 의 비교 사진 자체가 없음
    COORD_FAIL = "COORD_FAIL"  # 비교 사진은 있으나 좌표 추출 실패로 제외됨
    OVER_TOLERANCE = "OVER_TOLERANCE"  # 같은 die 후보는 있으나 허용오차 초과


@dataclass
class MatchResult:
    """기준 defect 1개에 대한 특정 비교 layer 의 매칭 결과."""

    compare_layer: str
    base: DefectRecord
    matched: Optional[DefectRecord] = None
    distance: Optional[float] = None
    # ---- 진단용(매칭 실패 사유 파악) ----
    nearest: Optional[DefectRecord] = None  # 허용오차 무시한 같은 die 최근접 후보
    nearest_distance: Optional[float] = None  # 그 거리
    die_candidates: int = 0  # 같은 (wafer,die) 의 좌표 OK 비교 record 수
    failed_in_die: int = 0  # 이 비교 layer·같은 wafer 의 좌표 추출 실패 수(근사)
    ambiguous: bool = False  # 허용오차 내 거의 동률 후보가 둘 이상(어느 것이 맞는지 모호)

    @property
    def is_match(self) -> bool:
        return self.matched is not None

    @property
    def reason(self) -> NoMatchReason:
        """매칭 실패 사유 분류."""
        if self.is_match:
            return NoMatchReason.NONE
        if self.die_candidates > 0:
            return NoMatchReason.OVER_TOLERANCE
        if self.failed_in_die > 0:
            return NoMatchReason.COORD_FAIL
        return NoMatchReason.NO_DIE_PHOTO


@dataclass
class BaseDefectMatches:
    """기준 defect 1개 + 모든 비교 layer 매칭 결과 묶음."""

    base: DefectRecord
    results: list[MatchResult] = field(default_factory=list)
    # 근접(<50) 중복 defect 을 대표 1개로 접었을 때, 묶인 전체 base 를 담는다(‘+n’ 표시용).
    # None 이면 단독(중복 없음). app.clustering.Cluster 타입(순환 import 회피 위해 미주석).
    base_cluster: object = None

    def for_layer(self, layer: str) -> Optional[MatchResult]:
        for r in self.results:
            if r.compare_layer == layer:
                return r
        return None


# =============================================================================
# app/safety.py   [#5]
# =============================================================================
"""원본 보호 게이트 (문서 Section 1 — 절대 안전 규칙).

본 프로그램의 최우선 원칙은 원본 데이터 훼손 방지이다. 모든 쓰기/캐시/Excel 경로는
반드시 이 모듈의 게이트를 통과해야 한다.

2중 보호:
  1차(로직) : 소스(LOT) 경로에 대해서는 write/delete/move/rename 을 절대 호출하지 않는다.
              이미지/INI/info 파일은 read_bytes()/open('rb') 읽기 전용으로만 접근한다.
  2차(경로) : 출력 경로가 어떤 소스 루트와 같거나 그 하위면 저장을 차단한다.
"""




class OriginalProtectionError(Exception):
    """원본 보호 규칙 위반 시 발생. 저장/쓰기를 차단한다."""


def _resolve(path: str | Path) -> Path:
    """심볼릭 링크/상대경로/대소문자 차이를 흡수해 비교 가능한 절대경로로 변환.

    네트워크 경로(UNC)나 존재하지 않는 출력 경로도 다룰 수 있어야 하므로
    strict=False 로 resolve 한다.
    """
    try:
        return Path(path).resolve(strict=False)
    except (OSError, RuntimeError):
        return Path(path).absolute()


def is_within(child: str | Path, parent: str | Path) -> bool:
    """child 가 parent 와 같거나 그 하위 경로이면 True."""
    c = _resolve(child)
    p = _resolve(parent)
    if c == p:
        return True
    try:
        c.relative_to(p)
        return True
    except ValueError:
        return False


def assert_output_safe(
    output_path: str | Path,
    source_roots: Iterable[str | Path],
) -> Path:
    """출력 경로가 어떤 소스 루트와 같거나 하위이면 차단한다 (2차 보호).

    Section 1.1: "원본 경로와 출력 경로가 같거나, 출력 경로가 원본 경로 하위이면 저장을 차단한다."

    Returns:
        검증을 통과한 절대 출력 경로.

    Raises:
        OriginalProtectionError: 출력 경로가 원본 폴더 내부일 때.
    """
    out = _resolve(output_path)
    for root in source_roots:
        if not str(root):
            continue
        if is_within(out, root):
            raise OriginalProtectionError(
                "출력 경로가 원본(LOT) 폴더 내부이거나 동일합니다. "
                "원본 보호 규칙에 따라 저장을 차단했습니다.\n"
                f"  출력 경로 : {out}\n"
                f"  원본 경로 : {_resolve(root)}\n"
                "원본 폴더 밖의 다른 폴더를 선택하세요."
            )
    return out


def is_output_safe(output_path: str | Path, source_roots: Iterable[str | Path]) -> bool:
    """예외 없이 안전 여부만 반환(경고 표시 등 비차단 검사용)."""
    out = _resolve(output_path)
    return not any(str(root) and is_within(out, root) for root in source_roots)


def conflicting_source(
    output_path: str | Path, source_roots: Iterable[str | Path]
) -> Path | None:
    """output_path 를 포함하는 원본 루트가 있으면 그 경로를, 없으면 None 을 반환."""
    out = _resolve(output_path)
    for root in source_roots:
        if str(root) and is_within(out, root):
            return _resolve(root)
    return None


def safe_makedirs(target_dir: str | Path, source_roots: Iterable[str | Path]) -> Path:
    """출력 디렉터리를 만들되, 원본 내부면 차단한다."""
    out = assert_output_safe(target_dir, source_roots)
    out.mkdir(parents=True, exist_ok=True)
    return out


def read_only_bytes(path: str | Path) -> bytes:
    """원본 파일을 읽기 전용으로만 읽는다 (1차 보호).

    명시적으로 'rb' 모드만 사용하여 원본 수정 가능성을 원천 차단한다.
    """
    with open(path, "rb") as fh:
        return fh.read()


# =============================================================================
# app/ui/flow_layout.py   [#6]
# =============================================================================
"""줄바꿈(Flow) 레이아웃.

위젯을 가로로 배치하되 폭이 부족하면 다음 줄로 넘긴다. 가로 스크롤 의존을 없애기 위해
비교 Layer 체크박스 영역 등에 사용한다(문서 Section 9 / 사용성 — 가로 휠 미사용).

Qt 공식 FlowLayout 예제를 본 프로젝트 스타일에 맞춰 정리한 구현.
"""





class FlowLayout(QLayout):
    """폭에 따라 자동 줄바꿈하는 레이아웃."""

    def __init__(
        self,
        parent: Optional[QWidget] = None,
        margin: int = 0,
        h_spacing: int = 8,
        v_spacing: int = 6,
    ):
        super().__init__(parent)
        self._items: list[QLayoutItem] = []
        self._h_space = h_spacing
        self._v_space = v_spacing
        self.setContentsMargins(QMargins(margin, margin, margin, margin))

    # ---- QLayout 필수 구현 ------------------------------------------------
    def addItem(self, item: QLayoutItem) -> None:  # noqa: N802
        self._items.append(item)

    def count(self) -> int:
        return len(self._items)

    def itemAt(self, index: int) -> Optional[QLayoutItem]:  # noqa: N802
        if 0 <= index < len(self._items):
            return self._items[index]
        return None

    def takeAt(self, index: int) -> Optional[QLayoutItem]:  # noqa: N802
        if 0 <= index < len(self._items):
            return self._items.pop(index)
        return None

    def expandingDirections(self) -> Qt.Orientations:  # noqa: N802
        return Qt.Orientation(0)

    def hasHeightForWidth(self) -> bool:  # noqa: N802
        return True

    def heightForWidth(self, width: int) -> int:  # noqa: N802
        return self._do_layout(QRect(0, 0, width, 0), test_only=True)

    def setGeometry(self, rect: QRect) -> None:  # noqa: N802
        super().setGeometry(rect)
        self._do_layout(rect, test_only=False)

    def sizeHint(self) -> QSize:  # noqa: N802
        return self.minimumSize()

    def minimumSize(self) -> QSize:  # noqa: N802
        size = QSize()
        for item in self._items:
            size = size.expandedTo(item.minimumSize())
        margins = self.contentsMargins()
        size += QSize(
            margins.left() + margins.right(), margins.top() + margins.bottom()
        )
        return size

    # ---- 배치 로직 --------------------------------------------------------
    def _do_layout(self, rect: QRect, test_only: bool) -> int:
        margins = self.contentsMargins()
        effective = rect.adjusted(
            margins.left(), margins.top(), -margins.right(), -margins.bottom()
        )
        x = effective.x()
        y = effective.y()
        line_height = 0

        for item in self._items:
            widget = item.widget()
            if widget is not None and not widget.isVisible():
                continue
            hint = item.sizeHint()
            next_x = x + hint.width() + self._h_space
            if next_x - self._h_space > effective.right() and line_height > 0:
                # 줄바꿈
                x = effective.x()
                y = y + line_height + self._v_space
                next_x = x + hint.width() + self._h_space
                line_height = 0
            if not test_only:
                item.setGeometry(QRect(QPoint(x, y), hint))
            x = next_x
            line_height = max(line_height, hint.height())

        return y + line_height - rect.y() + margins.bottom()


# =============================================================================
# app/ui/image_loader.py   [#7]
# =============================================================================
"""비동기 이미지 로더 (문서 Section 10 — UI 멈춤 최소화).

비교 그리드의 원본 이미지를 네트워크 경로에서 동기로 읽으면 탐색 시 UI 가 멈춘다.
이 로더는 QThreadPool 워커에서 이미지를 읽어 디스플레이 크기로 축소한 뒤, 메모리 LRU
캐시에 보관한다. 같은 이미지를 다시 볼 때(이전/다음 왕복)는 캐시에서 즉시 표시된다.

원본 이미지는 QImageReader 의 읽기 전용 접근으로만 읽으며 원본을 수정하지 않는다.
"""





class _LoadSignals(QObject):
    done = Signal(int, str, object)  # request_id, path, QImage|None


class _LoadTask(QRunnable):
    def __init__(self, request_id: int, path: str, max_dim: int):
        super().__init__()
        self.request_id = request_id
        self.path = path
        self.max_dim = max_dim
        self.signals = _LoadSignals()

    @Slot()
    def run(self) -> None:
        image: Optional[QImage] = None
        try:
            reader = QImageReader(self.path)
            reader.setAutoTransform(True)
            size = reader.size()
            # 큰 이미지는 디코드 단계에서 축소하여 메모리/시간을 절약한다.
            if size.isValid() and max(size.width(), size.height()) > self.max_dim:
                scaled = size.scaled(
                    QSize(self.max_dim, self.max_dim), Qt.KeepAspectRatio
                )
                reader.setScaledSize(scaled)
            img = reader.read()
            if not img.isNull():
                image = img
        except Exception:  # noqa: BLE001 - 로드 실패는 placeholder 로 처리
            image = None
        self.signals.done.emit(self.request_id, self.path, image)


class ImageLoader(QObject):
    """비동기 이미지 로더 + LRU 캐시."""

    loaded = Signal(int, object)  # request_id, QImage|None

    def __init__(
        self,
        max_dim: int = 600,
        cache_size: int = 128,
        parent: Optional[QObject] = None,
    ):
        super().__init__(parent)
        self._max_dim = max_dim
        self._cache_size = cache_size
        self._cache: "OrderedDict[str, QImage]" = OrderedDict()
        self._next_id = 0
        self._pool = QThreadPool.globalInstance()

    def request(self, path: str) -> int:
        """이미지 로드를 요청하고 request_id 를 반환. 결과는 loaded 시그널로 전달."""
        self._next_id += 1
        rid = self._next_id
        cached = self._cache.get(path)
        if cached is not None:
            self._cache.move_to_end(path)
            # 캐시 적중도 비동기로 통일 (호출 측이 동일 경로로 처리)
            QTimer.singleShot(0, lambda: self.loaded.emit(rid, cached))
            return rid
        task = _LoadTask(rid, path, self._max_dim)
        task.signals.done.connect(self._on_done)
        self._pool.start(task)
        return rid

    def prefetch(self, paths: list[str]) -> None:
        """탐색 체감 향상을 위해 인접 이미지를 미리 디코드해 캐시에 채운다.

        이미 캐시에 있으면 건너뛴다. 결과는 캐시에만 적재되고 loaded 시그널 소비자는
        request_id 로 자신과 무관한 결과를 무시하므로 UI 에 영향이 없다.
        """
        for path in paths:
            if not path or path in self._cache:
                continue
            task = _LoadTask(-1, path, self._max_dim)
            task.signals.done.connect(self._on_prefetched)
            self._pool.start(task)

    @Slot(int, str, object)
    def _on_prefetched(self, rid: int, path: str, image: object) -> None:
        if isinstance(image, QImage) and not image.isNull():
            self._cache[path] = image
            self._cache.move_to_end(path)
            while len(self._cache) > self._cache_size:
                self._cache.popitem(last=False)

    @Slot(int, str, object)
    def _on_done(self, rid: int, path: str, image: object) -> None:
        if isinstance(image, QImage) and not image.isNull():
            self._cache[path] = image
            self._cache.move_to_end(path)
            while len(self._cache) > self._cache_size:
                self._cache.popitem(last=False)
        self.loaded.emit(rid, image)

    def clear_cache(self) -> None:
        self._cache.clear()


# =============================================================================
# app/ui/theme.py   [#8]
# =============================================================================
"""다크 + 파란 네온 테마 (문서 Section 9).

어두운 바탕, 파란 네온 강조, 깔끔하고 복잡하지 않은 화면.
버튼은 hover/pressed 시 시각적 변화가 있어야 한다(QSS state 로 처리).
"""



# ---- UI 글자 크기(전역 스케일) ----
# 설정의 ui_font_size 값 → 배율. 보통 기준 크게=+30%.
FONT_SCALES = {"normal": 1.0, "large": 1.3}
# 인라인 스타일(px 직접 지정) 위젯이 참조하는 현재 배율. apply_theme 에서 갱신.
FONT_SCALE = 1.0


def scale_for(key: str | None) -> float:
    """설정 값(normal/large) → 글자 크기 배율."""
    return FONT_SCALES.get(key or "normal", 1.0)


def fpx(base: int) -> int:
    """현재 배율을 반영한 글자 크기(px). 인라인 스타일용."""
    return max(1, round(base * FONT_SCALE))


# 팔레트 — 저채도 슬레이트 다크 테마(부드러운 대비, 넓은 여백 지향)
BG = "#11151c"
BG_PANEL = "#171c26"
BG_ELEV = "#1f2632"
NEON = "#5b8db8"        # 강조: 저채도 슬레이트블루
NEON_DIM = "#456b8f"
NEON_SOFT = "#2c343f"   # 저채도 경계선
TEXT = "#dde3ec"
TEXT_DIM = "#8b95a4"
MATCH = "#6ec59a"
NOMATCH = "#d98a8a"
BASE_GLOW = "#7fa8cc"
WARN = "#d8b773"        # "허용오차 초과" 등 진단 강조
OVERLAY_BG = "rgba(17, 21, 28, 0.62)"  # 이미지 위 배지 반투명 배경

STYLESHEET = f"""
* {{
    font-family: 'Segoe UI', 'Malgun Gothic', sans-serif;
    color: {TEXT};
}}
QWidget#root, QMainWindow {{
    background-color: {BG};
}}
QFrame#panel {{
    background-color: {BG_PANEL};
    border: 1px solid {NEON_SOFT};
    border-radius: 14px;
}}
QFrame#sidebar {{
    background-color: {BG_PANEL};
    border: 1px solid {NEON_SOFT};
    border-radius: 14px;
}}
QLabel {{
    color: {TEXT};
    background: transparent;
}}
QLabel#dim {{ color: {TEXT_DIM}; }}
QLabel#title {{ font-size: 16px; font-weight: 700; color: {TEXT}; }}
QLabel#lotName {{ font-size: 13px; font-weight: 600; color: {BASE_GLOW}; }}
QLabel#section {{ font-size: 11px; font-weight: 700; color: {TEXT_DIM};
    letter-spacing: 1px; }}
/* 이미지 위 Layer 배지 + 진단 라벨 */
QLabel#layerBadge {{
    background-color: {OVERLAY_BG}; color: {TEXT};
    border-radius: 8px; padding: 3px 10px; font-weight: 700; font-size: 13px;
}}
QLabel#layerBadgeBase {{
    background-color: {OVERLAY_BG}; color: {BASE_GLOW};
    border-radius: 8px; padding: 3px 10px; font-weight: 700; font-size: 13px;
}}
QLabel#diag {{ color: {TEXT_DIM}; font-size: 10px; }}
QLabel#diagWarn {{ color: {WARN}; font-size: 10px; }}

/* ---- 버튼 ---- */
QPushButton {{
    background-color: {BG_ELEV};
    color: {TEXT};
    border: 1px solid {NEON_SOFT};
    border-radius: 10px;
    padding: 8px 16px;
    font-size: 12px;
}}
QPushButton:hover {{
    background-color: {NEON_SOFT};
    border: 1px solid {NEON};
    color: {TEXT};
}}
QPushButton:pressed {{
    background-color: {NEON_DIM};
    border: 1px solid {NEON};
}}
QPushButton:disabled {{
    color: {TEXT_DIM};
    border: 1px solid #232a33;
    background-color: #161b23;
}}
QPushButton#primary {{
    background-color: {NEON_DIM};
    border: 1px solid {NEON};
    font-weight: 700;
    color: {TEXT};
}}
QPushButton#primary:hover {{ background-color: {NEON}; color: {TEXT}; }}
QPushButton#primary:pressed {{ background-color: {NEON_DIM}; }}
/* 컴팩트 버튼(전체/해제·설정 등) */
QPushButton#mini {{
    padding: 3px 10px;
    font-size: 11px;
    border-radius: 8px;
}}
/* 토글(checkable) mini 버튼이 켜지면 네온 배경으로 확실히 구분 */
QPushButton#mini:checked {{
    background-color: {NEON_DIM};
    border: 1px solid {NEON};
    color: {TEXT};
    font-weight: 700;
}}
QPushButton#mini:checked:hover {{ background-color: {NEON}; }}

/* 스크롤 영역은 기본 흰 배경 대신 투명(뒤 패널이 비치게) */
QScrollArea {{ background: transparent; border: none; }}

/* ---- 입력(콤보/스핀/라인) ---- */
QComboBox, QSpinBox, QDoubleSpinBox, QLineEdit {{
    background-color: {BG_ELEV};
    border: 1px solid {NEON_SOFT};
    border-radius: 8px;
    padding: 6px 10px;
    min-height: 22px;
    color: {TEXT};
    selection-background-color: {NEON_DIM};
}}
QComboBox:hover, QSpinBox:hover, QDoubleSpinBox:hover, QLineEdit:hover {{
    border: 1px solid {NEON};
}}
QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus, QLineEdit:focus {{
    border: 1px solid {NEON};
    background-color: {BG_PANEL};
}}
QComboBox:disabled, QSpinBox:disabled, QDoubleSpinBox:disabled {{
    color: {TEXT_DIM};
    background-color: #161b23;
}}

/* 콤보 드롭다운 버튼 (화살표 이미지는 apply_theme 에서 주입) */
QComboBox::drop-down {{ width: 24px; border: none; }}
QComboBox QAbstractItemView {{
    background-color: {BG_ELEV};
    border: 1px solid {NEON};
    border-radius: 8px;
    padding: 4px;
    selection-background-color: {NEON_DIM};
    outline: none;
}}
QComboBox QAbstractItemView::item {{
    min-height: 26px; padding: 3px 8px; border-radius: 6px;
}}
QComboBox QAbstractItemView::item:hover {{ background: {NEON_SOFT}; }}

/* 스핀박스 ↑↓ 버튼 + 화살표 */
QAbstractSpinBox {{ padding-right: 22px; }}
/* 버튼 없는 입력(허용오차)은 일반 패딩 */
QDoubleSpinBox#tol {{ padding-right: 10px; }}
QDoubleSpinBox#tol::up-button, QDoubleSpinBox#tol::down-button {{ width: 0; border: none; }}
QAbstractSpinBox::up-button, QAbstractSpinBox::down-button {{
    subcontrol-origin: border;
    width: 20px;
    background-color: {BG_ELEV};
    border-left: 1px solid {NEON_SOFT};
}}
QAbstractSpinBox::up-button {{
    subcontrol-position: top right;
    border-top-right-radius: 8px;
}}
QAbstractSpinBox::down-button {{
    subcontrol-position: bottom right;
    border-bottom-right-radius: 8px;
    border-top: 1px solid {NEON_SOFT};
}}
QAbstractSpinBox::up-button:hover, QAbstractSpinBox::down-button:hover {{
    background-color: {NEON_SOFT};
}}
QAbstractSpinBox::up-button:pressed, QAbstractSpinBox::down-button:pressed {{
    background-color: {NEON_DIM};
}}
/* 스핀박스 화살표 이미지는 apply_theme 에서 주입 */

/* ---- 체크박스(비교 layer 선택) ---- */
QCheckBox {{ spacing: 6px; padding: 4px; }}
QCheckBox::indicator {{
    width: 16px; height: 16px;
    border: 1px solid {NEON_SOFT};
    border-radius: 5px;
    background: {BG_ELEV};
}}
QCheckBox::indicator:hover {{ border: 1px solid {NEON}; }}
QCheckBox::indicator:checked {{
    background: {NEON};
    border: 1px solid {NEON_DIM};
}}

/* ---- 다이얼로그 ---- */
QDialog {{ background-color: {BG}; }}

/* ---- 리스트(출력 선택) ---- */
QListWidget {{
    background-color: {BG_ELEV};
    border: 1px solid {NEON_SOFT};
    border-radius: 10px;
    outline: none;
}}
QListWidget::item {{ padding: 10px; border-radius: 8px; min-height: 30px; }}
QListWidget::item:hover {{ background: {NEON_SOFT}; }}
QListWidget::item:selected {{ background: {NEON_DIM}; color: {TEXT}; }}

/* ---- 우클릭/드롭다운 메뉴 — 어두운 테마 통일(흰 배경 방지) ---- */
QMenu {{
    background-color: {BG_ELEV};
    color: {TEXT};
    border: 1px solid {NEON_SOFT};
    border-radius: 8px;
    padding: 4px;
}}
QMenu::item {{
    padding: 6px 18px; border-radius: 6px; background: transparent;
}}
QMenu::item:selected {{ background: {NEON_DIM}; color: {TEXT}; }}
QMenu::item:disabled {{ color: {TEXT_DIM}; }}
QMenu::separator {{ height: 1px; background: {NEON_SOFT}; margin: 4px 8px; }}

/* ---- 아이템 뷰(트리/리스트/테이블) — 어두운 테마 통일(흰 배경 방지) ---- */
QTreeView, QListView, QTableView, QColumnView {{
    background-color: {BG_ELEV};
    alternate-background-color: {BG_ELEV};
    color: {TEXT};
    border: 1px solid {NEON_SOFT};
    border-radius: 8px;
    outline: none;
}}
QTreeView::item, QListView::item, QTableView::item {{
    padding: 3px 6px; border-radius: 6px;
}}
QTreeView::item:hover, QListView::item:hover, QTableView::item:hover {{
    background: {NEON_SOFT};
}}
QTreeView::item:selected, QListView::item:selected, QTableView::item:selected {{
    background: {NEON_DIM}; color: {TEXT};
}}
QTreeView::branch {{ background: transparent; }}
QHeaderView::section {{
    background-color: {BG_PANEL};
    color: {TEXT_DIM};
    border: none;
    border-bottom: 1px solid {NEON_SOFT};
    padding: 4px 6px;
}}

/* ---- 스플리터 손잡이(넓고 차분하게) ---- */
QSplitter::handle {{ background: transparent; }}
QSplitter::handle:horizontal {{ width: 10px; }}

/* ---- 스크롤바 ---- */
QScrollBar:horizontal, QScrollBar:vertical {{
    background: transparent; border: none;
}}
QScrollBar:horizontal {{ height: 10px; }}
QScrollBar:vertical {{ width: 10px; }}
QScrollBar::handle {{
    background: {NEON_SOFT}; border-radius: 5px; min-width: 30px; min-height: 30px;
}}
QScrollBar::handle:hover {{ background: {NEON_DIM}; }}
QScrollBar::add-line, QScrollBar::sub-line {{ width: 0; height: 0; }}
QScrollBar::add-page, QScrollBar::sub-page {{ background: transparent; }}

QProgressBar {{
    background-color: {BG_ELEV};
    border: 1px solid {NEON_SOFT};
    border-radius: 8px;
    text-align: center;
    height: 16px;
}}
QProgressBar::chunk {{
    background-color: {NEON_DIM};
    border-radius: 7px;
}}
QToolTip {{
    background-color: {BG_ELEV};
    color: {TEXT};
    border: 1px solid {NEON_SOFT};
    border-radius: 6px;
    padding: 5px;
}}
"""


def _make_arrow(direction: str, color: str, size: int = 12) -> str:
    """삼각형 화살표 PNG 를 생성하고 파일 경로(posix)를 반환한다.

    QSS 의 border-삼각형 트릭은 Qt 버전/플랫폼에 따라 깨지므로, 깔끔한 화살표를
    런타임에 그려 이미지로 주입한다(에셋 파일 불필요).
    """
    import tempfile
    from pathlib import Path

    from PySide6.QtCore import QPoint, Qt
    from PySide6.QtGui import QColor, QPainter, QPixmap, QPolygon

    pm = QPixmap(size, size)
    pm.fill(Qt.transparent)
    p = QPainter(pm)
    p.setRenderHint(QPainter.Antialiasing, True)
    p.setPen(Qt.NoPen)
    p.setBrush(QColor(color))
    m = 2  # 여백
    if direction == "down":
        pts = [QPoint(m, m + 1), QPoint(size - m, m + 1), QPoint(size // 2, size - m)]
    else:  # up
        pts = [QPoint(m, size - m - 1), QPoint(size - m, size - m - 1), QPoint(size // 2, m)]
    p.drawPolygon(QPolygon(pts))
    p.end()

    out_dir = Path(tempfile.gettempdir()) / "defect_tracker_theme"
    out_dir.mkdir(parents=True, exist_ok=True)
    safe = color.lstrip("#")
    path = out_dir / f"arrow_{direction}_{safe}_{size}.png"
    pm.save(str(path), "PNG")
    return path.as_posix()


_FONT_RE = re.compile(r"font-size:\s*(\d+)px")
_ORIG_PT: float | None = None


def _scaled_sheet(scale: float) -> str:
    """스타일시트의 모든 font-size(px)를 배율만큼 키운다.

    보통(1.0)은 현재와 완전히 동일. 크게일 때는 명시 크기 없는 위젯(콤보/입력/체크박스 등)의
    기본 글자 크기도 커지도록 `*` 규칙에 기본 font-size 를 함께 주입한다(구체 선택자가 우선).
    """
    if scale == 1.0:
        return STYLESHEET
    sheet = _FONT_RE.sub(
        lambda m: f"font-size: {max(8, round(int(m.group(1)) * scale))}px", STYLESHEET
    )
    return sheet + f"\n* {{ font-size: {round(12 * scale)}px; }}\n"


def apply_theme(app, scale: float = 1.0) -> None:
    global FONT_SCALE, _ORIG_PT
    FONT_SCALE = scale
    # 앱 기본 폰트 크기도 배율만큼(명시 크기 없는 네이티브 요소·메뉴 등 대응). 원본을 한 번만
    # 기억해 반복 적용해도 배율이 누적되지 않게 한다.
    if _ORIG_PT is None:
        f0 = app.font()
        _ORIG_PT = f0.pointSizeF() if f0.pointSizeF() > 0 else 9.0
    f = app.font()
    f.setPointSizeF(_ORIG_PT * scale)
    app.setFont(f)

    # 콤보/스핀 화살표를 런타임 이미지로 주입(테마색 삼각형).
    # 주의: `QComboBox:hover::down-arrow` 규칙은 Qt 에서 화살표가 두 번 그려지는
    # QSS 버그를 유발하므로 사용하지 않는다(화살표 색은 고정 TEXT_DIM 으로 충분).
    down = _make_arrow("down", TEXT_DIM)
    up = _make_arrow("up", TEXT_DIM)
    arrow_qss = f"""
QComboBox::down-arrow {{ image: url("{down}"); width: 12px; height: 12px; }}
QAbstractSpinBox::down-arrow {{ image: url("{down}"); width: 9px; height: 9px; }}
QAbstractSpinBox::up-arrow {{ image: url("{up}"); width: 9px; height: 9px; }}
"""
    app.setStyleSheet(_scaled_sheet(scale) + arrow_qss)


# =============================================================================
# app/wafermap_align.py   [#9]
# =============================================================================
"""웨이퍼 맵 die 정합(alignment) — 관측 die 와 디바이스 DB die_map 의 원점 맞추기.

배경: 디바이스 DB 의 die_map 은 Map 그리드의 (ci,ri)(좌상단 원점)인데, record 의
(col,row) 는 파서마다 다른 오프셋(KLA `+count//2`, Camtek INI `row_base-row`, 파일명
직접)을 거친다. 따라서 die_map 을 그대로 valid 로 쓰면 실제 모양과 어긋날 수 있다.

해법(translation voting): 관측 die o 와 die_map die d 의 모든 쌍에 대해 평행이동
s = o - d 에 투표하면, 가장 표를 많이 받은 s 가 두 좌표계를 가장 잘 겹치게 하는 이동이다.
그 s 로 die_map 을 관측 좌표계로 옮겨(valid = die_map + s) 그리면 실제 디바이스 모양과
관측 die 가 정렬된다. 겹침 비율(overlap)이 낮으면 정합 실패로 보고 호출 측이 폴백한다.
"""



# 좌표쌍 타입
Die = tuple[int, int]


@dataclass(frozen=True)
class Alignment:
    """관측→die_map 정합 결과."""

    dcol: int  # die_map 을 관측 좌표계로 옮기는 평행이동(col)
    drow: int
    overlap: float  # 관측 die 중 디바이스 모양 위에 놓인 비율(0~1)

    @property
    def ok(self) -> bool:
        return self.overlap > 0.0


def align_observed_to_diemap(
    observed: set[Die],
    die_map: frozenset[Die] | set[Die],
    *,
    max_samples: int = 120,
) -> Alignment:
    """관측 die 집합을 die_map 에 가장 잘 겹치게 하는 평행이동을 찾는다.

    반환 Alignment.(dcol,drow) 는 die_map 을 관측 좌표계로 옮기는 이동이며,
    valid = {(ci+dcol, ri+drow) for (ci,ri) in die_map} 로 쓰면 관측과 정렬된다.
    overlap 은 관측 die 중 옮긴 die_map 위에 놓인 비율이다.
    """
    if not observed or not die_map:
        return Alignment(0, 0, 0.0)

    # 표본으로 평행이동 투표(비용 제한). 정수 die 이므로 단순 차이로 충분.
    sample = list(observed)[:max_samples]
    votes: Counter[Die] = Counter()
    for oc, orow in sample:
        for dc, dr in die_map:
            votes[(oc - dc, orow - dr)] += 1

    # 동점(같은 최다 득표) translation 이 여럿일 수 있다 — 관측 die 가 성기면 작은
    # cluster 가 큰(dense) 디바이스 모양 안 여러 위치에 똑같이 맞아 overlap 이 같아진다.
    # 해시 순서로 임의 선택하면 윤곽이 defect 셀 옆으로 밀려 보이고, 관측 중심으로 맞추면
    # 오히려 이미 제자리인 die 를 밀어버린다(관측이 die_map 부분집합일 때 shift 0 이 정답인데
    # centroid 로 (-1,1) 등을 고름). 두 좌표계는 원점을 공유하도록 설계됐으므로, 최다 득표
    # 후보 중 **이동 크기(|dcol|+|drow|)가 가장 작은**(= 원 좌표를 최대한 보존하는) 것을
    # 결정론적으로 고른다. 이상적(부분집합) 경우 shift 0 을 되찾는다.
    best_votes = max(votes.values())
    candidates = [s for s, c in votes.items() if c == best_votes]
    sdc, sdr = min(candidates, key=lambda s: (abs(s[0]) + abs(s[1]), s))

    # 전체 관측 기준 겹침 비율 산정(표본이 아닌 전체로 신뢰도 측정).
    shifted = {(ci + sdc, ri + sdr) for ci, ri in die_map}
    hit = sum(1 for o in observed if o in shifted)
    return Alignment(sdc, sdr, hit / len(observed))


def shifted_die_map(die_map: frozenset[Die] | set[Die], align: Alignment) -> set[Die]:
    """die_map 을 정합 이동만큼 옮긴 valid 집합(관측 좌표계)."""
    return {(ci + align.dcol, ri + align.drow) for ci, ri in die_map}


# =============================================================================
# app/clustering.py   [#10]
# =============================================================================
"""defect 근접 클러스터링 + layer 간 교차 매칭 (순수 로직, UI 무관).

- `cluster_records`: 같은 wafer·같은 die 안에서 local 좌표 거리가 `CLUSTER_RADIUS` 미만인
  defect 들을 하나로 묶는다(대표 1개 + 나머지). 히트맵에서 근접 중복 defect 을 대표 1장 +
  "+n" 으로 접어 보여주기 위한 것.
- `cross_layer_groups`: 여러 layer 의 cluster 대표들을 layer 간에 매칭(그리디)해, 한 위치의
  defect 들을 "어느 layer 끼리 같은 defect 인지" 그룹으로 묶는다. 특정 기준 layer 에 종속되지
  않는다("전체 defect 보기"용).

매칭/거리 기준은 `matcher` 와 동일한 개념(같은 wafer·die ±1·local 좌표 거리)이며,
거리 계산은 `DefectRecord.distance_to`(app/models.py)를 재사용한다.
"""




# 같은 defect 으로 볼 local 좌표 거리 임계값(미만이면 하나로 묶음).
CLUSTER_RADIUS = 50.0
# 교차 매칭 시 die 이웃 탐색 범위(±DIE_TOL). matcher.DEFAULT_DIE_TOL 과 동일 개념.
_DIE_TOL = 1


def clustering__norm_wafer(wafer_id: str) -> str:
    return (wafer_id or "").strip().lower()


@dataclass
class Cluster:
    """근접 defect 묶음 — 대표 1개 + 전체 members(대표 포함)."""

    representative: DefectRecord
    members: list[DefectRecord] = field(default_factory=list)

    @property
    def extra_count(self) -> int:
        """대표 외 추가 defect 수('+n' 표기용)."""
        return max(0, len(self.members) - 1)


class _UnionFind:
    def __init__(self, n: int):
        self.parent = list(range(n))

    def find(self, a: int) -> int:
        while self.parent[a] != a:
            self.parent[a] = self.parent[self.parent[a]]
            a = self.parent[a]
        return a

    def union(self, a: int, b: int) -> None:
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[rb] = ra


def _rep_of(members: list[DefectRecord]) -> DefectRecord:
    """대표 선택 — image_path 이름순 첫 항목(결정론)."""
    return min(members, key=lambda r: str(r.image_path))


def cluster_records(
    records: Iterable[DefectRecord], radius: float = CLUSTER_RADIUS
) -> list[Cluster]:
    """같은 wafer·die 안에서 거리 < radius 인 record 를 union-find 로 묶는다.

    좌표가 없는(ok=False) record 는 각자 단독 cluster. 반환은 대표 image_path 이름순 정렬.
    """
    recs = list(records)
    clusters: list[Cluster] = []

    # 좌표 없는 record 는 단독.
    ok = [r for r in recs if r.ok]
    for r in recs:
        if not r.ok:
            clusters.append(Cluster(representative=r, members=[r]))

    if ok:
        uf = _UnionFind(len(ok))
        # 같은 (wafer, die) 버킷 안에서만 쌍 비교.
        buckets: dict[tuple[str, int, int], list[int]] = {}
        for i, r in enumerate(ok):
            buckets.setdefault((clustering__norm_wafer(r.wafer_id), int(r.col), int(r.row)), []).append(i)
        for idxs in buckets.values():
            for a in range(len(idxs)):
                for b in range(a + 1, len(idxs)):
                    ia, ib = idxs[a], idxs[b]
                    d = ok[ia].distance_to(ok[ib])
                    if d is not None and d < radius:
                        uf.union(ia, ib)
        comps: dict[int, list[DefectRecord]] = {}
        for i, r in enumerate(ok):
            comps.setdefault(uf.find(i), []).append(r)
        for members in comps.values():
            clusters.append(Cluster(representative=_rep_of(members), members=members))

    clusters.sort(key=lambda c: str(c.representative.image_path))
    return clusters


def _clusters_match(a: Cluster, b: Cluster, tolerance: float) -> Optional[float]:
    """두 cluster 대표가 매칭되면 거리, 아니면 None. 같은 wafer·die ±DIE_TOL·거리<=tol."""
    ra, rb = a.representative, b.representative
    if not ra.ok or not rb.ok:
        return None
    if clustering__norm_wafer(ra.wafer_id) != clustering__norm_wafer(rb.wafer_id):
        return None
    if abs(int(ra.col) - int(rb.col)) > _DIE_TOL or abs(int(ra.row) - int(rb.row)) > _DIE_TOL:
        return None
    d = ra.distance_to(rb)
    if d is None or d > tolerance:
        return None
    return d


def cross_layer_groups(
    layer_to_clusters: dict[str, list[Cluster]], tolerance: float
) -> list[dict[str, Cluster]]:
    """layer 별 cluster 들을 layer 간에 매칭해 그룹으로 묶는다(기준 layer 없음).

    거리 오름차순 그리디 union: 서로 다른 layer 의 cluster 두 개가 매칭되고, 두 그룹에
    겹치는 layer 가 없을 때만 합친다(그룹당 layer 최대 1개). 매칭 안 된 cluster 는 원소
    1개짜리 그룹으로 개별 반환. 반환 순서는 결정론(각 그룹 대표 image_path 이름순).
    """
    # 평탄화: (layer, cluster) 노드 목록.
    nodes: list[tuple[str, Cluster]] = []
    for layer, clusters in layer_to_clusters.items():
        for c in clusters:
            nodes.append((layer, c))
    n = len(nodes)
    if n == 0:
        return []

    uf = _UnionFind(n)
    group_layers: dict[int, set[str]] = {i: {nodes[i][0]} for i in range(n)}

    # 후보 간선(서로 다른 layer, 매칭) 을 거리 오름차순으로.
    edges: list[tuple[float, int, int]] = []
    for i in range(n):
        for j in range(i + 1, n):
            if nodes[i][0] == nodes[j][0]:
                continue
            d = _clusters_match(nodes[i][1], nodes[j][1], tolerance)
            if d is not None:
                edges.append((d, i, j))
    edges.sort(key=lambda e: (e[0], e[1], e[2]))

    for _d, i, j in edges:
        ri, rj = uf.find(i), uf.find(j)
        if ri == rj:
            continue
        if group_layers[ri] & group_layers[rj]:
            continue  # 같은 layer 중복 → 합치지 않음(그룹당 layer 1개 유지)
        uf.union(ri, rj)
        root = uf.find(ri)
        merged = group_layers[ri] | group_layers[rj]
        group_layers[root] = merged

    comps: dict[int, dict[str, Cluster]] = {}
    for i, (layer, cluster) in enumerate(nodes):
        comps.setdefault(uf.find(i), {})[layer] = cluster

    groups = list(comps.values())
    groups.sort(
        key=lambda g: min(str(c.representative.image_path) for c in g.values())
    )
    return groups


def _merge_member_results(members: list[BaseDefectMatches]) -> list[MatchResult]:
    """클러스터 멤버들의 비교 결과를 layer 별 best 매치로 병합한다.

    각 layer 에 대해 매칭된 것(is_match) 중 거리 최소를 우선, 없으면 첫 후보를 쓴다.
    → 대표가 개별로는 미매칭이어도 멤버가 매치했으면 그 매치가 반영된다.
    """
    order: list[str] = []
    by_layer: dict[str, list[MatchResult]] = {}
    for m in members:
        for r in m.results:
            if r.compare_layer not in by_layer:
                by_layer[r.compare_layer] = []
                order.append(r.compare_layer)
            by_layer[r.compare_layer].append(r)
    merged: list[MatchResult] = []
    for layer in order:
        cands = by_layer[layer]
        matched = [c for c in cands if c.is_match]
        if matched:
            best = min(matched, key=lambda c: (c.distance if c.distance is not None else 1e18))
        else:
            best = cands[0]
        merged.append(best)
    return merged


def collapse_matches(
    matches: list[BaseDefectMatches], radius: float = CLUSTER_RADIUS
) -> list[BaseDefectMatches]:
    """근접(같은 wafer·die·거리<radius) 기준 defect 을 대표 1개로 접는다.

    각 클러스터마다: 대표 base = `cluster.representative`(image_path 이름순, 결정론),
    결과는 멤버들의 layer 별 best 매치로 병합, `base_cluster` 에 전체 멤버를 담는다.
    반환 순서는 원래 등장 순서(클러스터 내 최소 index) 기준으로 안정 정렬.
    """
    if not matches:
        return []
    by_base = {id(m.base): m for m in matches}
    index_of = {id(m.base): i for i, m in enumerate(matches)}
    clusters = cluster_records([m.base for m in matches], radius)
    collapsed: list[tuple[int, BaseDefectMatches]] = []
    for cl in clusters:
        member_matches = [by_base[id(b)] for b in cl.members]
        results = _merge_member_results(member_matches)
        bdm = BaseDefectMatches(
            base=cl.representative,
            results=results,
            base_cluster=cl,
        )
        order_idx = min(index_of[id(b)] for b in cl.members)
        collapsed.append((order_idx, bdm))
    collapsed.sort(key=lambda t: t[0])
    return [bdm for _, bdm in collapsed]


# =============================================================================
# app/diagnostics.py   [#11]
# =============================================================================
"""좌표 추출 실패 진단 리포트(개발용) — 단일 markdown 파일로 관리.

스캔에서 좌표를 뽑지 못한 record 의 '왜'를 모아 한 파일로 남긴다. 매 스캔마다
**누적 추가(append)** 하여 이력을 보존한다. 원본이 아닌 워크스페이스에만 쓴다.
민감정보(좌표값 등)는 적지 않고 파일명/구조/카운트/사유만 기록한다.

핵심: 같은 '시도 트레일(note)'을 가진 실패끼리 **서명(signature) 클러스터링**해
대표 예시·카운트·처방 힌트를 보여줌으로써 대량 실패의 근본 원인을 빠르게 파악한다.
"""




_STATUS_LABEL = {
    ParseStatus.NOT_FOUND.value: "좌표/매칭 정보 없음",
    ParseStatus.INFO_FILE_NOT_FOUND.value: "info 파일 없음",
    ParseStatus.INVALID_INFO.value: "info 값 부족/오류",
}

# 서명(시도 트레일)에 특정 신호가 있으면 처방 힌트를 붙인다.
_HINTS = [
    ("KLA 원본", "파일명이 KLA 원본형 — 자재(material) 폴더가 아니라 KLA 원본 폴더를 "
                 "선택했을 가능성. 폴더 레벨을 확인하세요."),
    ("ColorImageGrabingInfo.ini 없음", "Camtek INI 가 없는 layer — KLA info(.001)로만 "
                                       "좌표가 나옵니다. info 파일 존재를 확인하세요."),
    ("TiffFileName", "KLA info 의 TiffFileName 목록과 이미지 파일명이 어긋남 — 확장자/명명 "
                     "규칙 불일치 가능."),
    ("DiePitchY 없음", "KLA info header 가 비표준 — DiePitchY 라인을 확인하세요."),
    ("die 위치 음수", "XINDEX/YINDEX 또는 제품 zero offset 이 맞지 않음 — 제품 프로파일을 "
                     "확인하세요."),
    ("필드 누락", "INI section 에 x/y/col/row 키가 부족 — INI 생성 설정을 확인하세요."),
    ("미분류(class 0) 후보 이미지", "KLA 가 찍었지만 정식 결함으로 분류/등록하지 않은 후보 "
                                   "이미지 — 실제 결함이 아니면 정상이며 무시해도 됩니다."),
]


# 미분류(class 0) 후보 — 원본 info(DefectList)에 정식 결함으로 등록되지 않아 좌표가 없는,
# '정상적으로 제외되는' 이미지. 진짜 실패(info 없음·불일치 등)와 구분해 따로 표기한다.
_UNCLASSIFIED_MARK = "미분류(class 0)"


def _is_unclassified(rec: DefectRecord) -> bool:
    return _UNCLASSIFIED_MARK in (rec.note or "")


def _signature(rec: DefectRecord) -> str:
    """동일 원인 묶음을 위한 서명(시도 트레일 문자열)."""
    return rec.note or rec.status.value


def _hint_for(signature: str) -> str:
    for needle, hint in _HINTS:
        if needle in signature:
            return hint
    return ""


def build_failure_report(lot_name: str, records: list[DefectRecord],
                         scan_errors: list[str] | None = None) -> str:
    """실패 진단 markdown 문자열을 만든다(파일 쓰기는 write_parse_failure_report)."""
    failed = [r for r in records if not r.ok]
    total = len(records)
    # 미분류(class 0) 후보와 '확인 필요' 실패를 분리한다.
    unclassified = [r for r in failed if _is_unclassified(r)]
    real_failed = [r for r in failed if not _is_unclassified(r)]

    lines: list[str] = []
    lines.append(f"# 좌표 추출 진단 — {lot_name}")
    lines.append("")
    lines.append(f"- 전체 이미지: **{total}개**")
    lines.append(f"- 좌표 OK: **{total - len(failed)}개**")
    if unclassified:
        lines.append(
            f"- 미분류(class 0) 후보: **{len(unclassified)}개** (정식 결함 아님 · 무시 가능)"
        )
    lines.append(f"- 실패: **{len(real_failed)}개**")
    lines.append("")

    if not failed:
        lines.append("이번 스캔에서 좌표 추출 실패가 없습니다. ✅")
        if scan_errors:
            lines.append("")
            lines.append("## 접근 실패 경로")
            for e in scan_errors[:50]:
                lines.append(f"- {e}")
        return "\n".join(lines) + "\n"

    # 미분류(class 0) 후보 — 무시 가능(간단 요약만, KLA info 덤프 없음)
    if unclassified:
        lines.extend(_unclassified_section(unclassified))

    # 확인 필요 실패 — 상태별 카운트 + 원인 클러스터(상세)
    if real_failed:
        lines.extend(_failure_clusters(real_failed))
    else:
        lines.append("확인이 필요한 실패는 없습니다 — 위 미분류(class 0) 후보만 있으며 정상입니다. ✅")
        lines.append("")

    if scan_errors:
        lines.append("## 접근 실패 경로")
        for e in scan_errors[:50]:
            lines.append(f"- {e}")
        lines.append("")
    return "\n".join(lines) + "\n"


def _unclassified_section(recs: list[DefectRecord]) -> list[str]:
    """미분류(class 0) 후보 요약 — layer/wafer 분포 + 예시(무거운 info 덤프 없음)."""
    lines: list[str] = []
    lines.append(f"## 미분류(class 0) 후보 — 무시 가능 ({len(recs)}개)")
    lines.append("")
    lines.append(
        "> KLA 가 캡처했지만 정식 결함으로 분류/등록하지 않은 후보 이미지입니다. 원본 "
        "info(DefectList)에 좌표 항목이 없어 정상적으로 제외됩니다(실제 결함 아님)."
    )
    by_layer = Counter(r.layer_folder for r in recs)
    lines.append(
        "- layer 분포: " + ", ".join(f"{k}×{v}" for k, v in by_layer.most_common(10))
    )
    by_wafer = Counter(r.wafer_id for r in recs)
    lines.append(
        "- wafer 분포: " + ", ".join(f"{k}×{v}" for k, v in by_wafer.most_common(10))
    )
    lines.append("- 예시 파일:")
    for r in recs[:5]:
        lines.append(f"  - `{Path(r.image_path).name}`")
    lines.append("")
    return lines


def _failure_clusters(recs: list[DefectRecord]) -> list[str]:
    """확인 필요 실패의 상태별 카운트 + 동일 사유 클러스터(상세 컨텍스트 포함)."""
    lines: list[str] = []
    by_status: Counter[str] = Counter(r.status.value for r in recs)
    lines.append("## 상태별 카운트")
    lines.append("")
    lines.append("| 상태 | 개수 |")
    lines.append("|---|---|")
    for status, n in by_status.most_common():
        lines.append(f"| {_STATUS_LABEL.get(status, status)} | {n} |")
    lines.append("")

    clusters: dict[str, list[DefectRecord]] = defaultdict(list)
    for r in recs:
        clusters[_signature(r)].append(r)
    lines.append("## 실패 원인 클러스터 (동일 사유끼리 묶음)")
    lines.append("")
    for sig, rs in sorted(clusters.items(), key=lambda kv: -len(kv[1])):
        lines.append(f"### ({len(rs)}개) {sig}")
        hint = _hint_for(sig)
        if hint:
            lines.append(f"> 처방: {hint}")
        by_layer = Counter(r.layer_folder for r in rs)
        dist = ", ".join(f"{k}×{v}" for k, v in by_layer.most_common(6))
        lines.append(f"- layer 분포: {dist}")
        lines.append("- 예시 파일:")
        for r in rs[:5]:
            lines.append(f"  - `{Path(r.image_path).resolve()}`")
        lines.append("")

        # 폴더별 진단 컨텍스트(같은 클러스터에서 고유 wafer_dir 기준)
        seen_dirs: set[str] = set()
        diag_count = 0
        for r in rs:
            if diag_count >= 3:
                break
            d = r.diag
            if not d:
                continue
            wdir = d.get("wafer_dir", "")
            if wdir in seen_dirs:
                continue
            seen_dirs.add(wdir)
            diag_count += 1
            lines.extend(_format_diag_context(r))
    return lines


def _format_diag_context(rec: DefectRecord) -> list[str]:
    """단일 실패 record 의 진단 컨텍스트를 markdown 줄 목록으로 포맷한다."""
    d = rec.diag
    if not d:
        return []
    lines: list[str] = []
    wdir = d.get("wafer_dir", str(Path(rec.image_path).parent))
    lines.append(f"#### 폴더 컨텍스트: `{wdir}`")
    lines.append("")

    # 좌표 출처·상태
    lines.append(f"- 좌표 출처(source): **{rec.source.value if hasattr(rec.source, 'value') else rec.source}**")
    lines.append(f"- 파싱 상태(status): **{rec.status.value if hasattr(rec.status, 'value') else rec.status}**")
    lines.append(f"- 시도 트레일(note): {rec.note or '(없음)'}")
    lines.append("")

    # 폴더 내 파일 목록
    all_files = d.get("files_in_folder", [])
    img_count = d.get("image_count", 0)
    lines.append(f"**폴더 내 파일 ({len(all_files)}개, 이미지 {img_count}개):**")
    lines.append("")
    for fname in all_files:
        lines.append(f"- `{fname}`")
    lines.append("")

    # Camtek INI 정보
    ini_files = d.get("ini_files", [])
    has_ini = d.get("has_ini_sections", False)
    if ini_files:
        lines.append(f"**Camtek INI 파일:** {', '.join(f'`{f}`' for f in ini_files)}")
        if has_ini:
            keys = d.get("ini_section_keys", [])
            lines.append(f"- INI section 수: {len(keys)}")
            if keys:
                preview = ", ".join(keys[:10])
                suffix = f" ... 외 {len(keys) - 10}개" if len(keys) > 10 else ""
                lines.append(f"- section 키 예시: `{preview}`{suffix}")
        lines.append("")
    else:
        lines.append("**Camtek INI 파일:** 없음")
        lines.append("")

    # KLA info 정보
    kla_file = d.get("kla_info_file")
    if kla_file:
        lines.append(f"**KLA info 파일:** `{kla_file}`")
        pitch_y = d.get("kla_die_pitch_y")
        tiff_count = d.get("kla_tiff_count", 0)
        all_defect_count = d.get("kla_all_defect_count", 0)
        lines.append(f"- DiePitchY: {pitch_y}")
        lines.append(f"- TiffFileName 매핑 수: {tiff_count}")
        lines.append(f"- 전체 DefectList 엔트리 수: {all_defect_count}")
        lines.append("")
        info_text = d.get("kla_info_text", "")
        if info_text:
            lines.append("<details><summary>KLA info 파일 내용 (펼치기)</summary>")
            lines.append("")
            lines.append("```")
            lines.append(info_text)
            lines.append("```")
            lines.append("")
            lines.append("</details>")
            lines.append("")
    else:
        lines.append("**KLA info 파일:** 없음 (선택된 info 파일 없음)")
        lines.append("")

    return lines


diagnostics__MAX_LOG_SIZE = 25 * 1024 * 1024  # 25 MB


def _pick_log_file(logs_dir: Path) -> Path:
    """현재 쓸 로그 파일을 반환한다. 25MB 초과 시 다음 번호 파일을 만든다."""
    base = logs_dir / "parse_failures.md"
    if not base.exists() or base.stat().st_size < diagnostics__MAX_LOG_SIZE:
        return base
    idx = 2
    while True:
        candidate = logs_dir / f"parse_failures_{idx}.md"
        if not candidate.exists() or candidate.stat().st_size < diagnostics__MAX_LOG_SIZE:
            return candidate
        idx += 1


def write_parse_failure_report(
    log_dir: Path, lot_name: str, records: list[DefectRecord],
    scan_errors: list[str] | None = None,
) -> Path:
    """진단 리포트를 log_dir/parse_failures*.md 에 **누적 추가**하고 경로를 반환한다."""
    logs = Path(log_dir)
    logs.mkdir(parents=True, exist_ok=True)
    out = _pick_log_file(logs)
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    separator = f"\n---\n\n> 스캔 시각: {stamp}\n\n"
    report = build_failure_report(lot_name, records, scan_errors)
    with open(out, "a", encoding="utf-8") as f:
        f.write(separator + report)
    return out


# =============================================================================
# app/layout.py   [#12]
# =============================================================================
"""Layer 폴더명 정규화 및 비교 그리드 배치 (문서 Section 8.2 / 8.4).

폴더명 예: "1. LYA4", "2. LYC3_재리뷰", "3. LYA4재리뷰"
  - 선행 순번 "N. " 제거
  - 접미 "재리뷰"(앞 언더바·공백은 있어도/없어도) 제거 후 is_re_review 플래그
  - 남은 토큰을 canonical layer 토큰으로 사용 (예: LYA4, LYC3)

비교 화면 그리드(Section 8.4)는 config.DEFAULT_LAYER_GRID 를 기본으로 하되,
실제 존재하는 layer 만 배치하고 나머지는 발견 순서대로 빈 칸을 채운다(graceful fallback).
"""




_ORDER_PREFIX_RE = re.compile(r"^\s*\d+\s*[.\-_)]\s*")
_RE_REVIEW_SUFFIXES = (
    "_재리뷰", "_재 리뷰", "재리뷰", "재 리뷰", "_rereview", "_re-review",
)
# 재리뷰 레벨: 접미가 (구분자) + ("재"×n) + "리뷰"(공백 허용). n = 재리뷰 깊이.
#   재리뷰/_재리뷰=1, 재재리뷰=2, 재재재리뷰=3 … (앞 언더바·공백은 있어도/없어도 인식)
_RE_REVIEW_LEVEL_RE = re.compile(r"[_\s]*((?:재\s*)+)리뷰\s*$")
# 영문 다중 re- 도 보조 지원: _re-review=1, _re-re-review=2 …
_EN_RE_REVIEW_RE = re.compile(r"_\s*((?:re-?)+)review\s*$", re.IGNORECASE)


def re_review_level(folder_name: str) -> int:
    """폴더명의 재리뷰 깊이를 센다(0 없음 / 1 재리뷰 / 2 재재리뷰 …).

    한국어 "_재리뷰"·"_재재리뷰"(반복 재) 및 영문 "_rereview"/"_re-re-review" 를 인식한다.
    """
    name = _ORDER_PREFIX_RE.sub("", folder_name.strip())
    m = _RE_REVIEW_LEVEL_RE.search(name)
    if m:
        return m.group(1).count("재")
    m = _EN_RE_REVIEW_RE.search(name)
    if m:
        # "re" 출현 횟수 = 레벨(re-review→1, re-re-review→2)
        return max(1, m.group(1).lower().count("re"))
    return 0


def normalize_layer(folder_name: str) -> tuple[str, bool]:
    """폴더명 -> (canonical 토큰, 재리뷰 여부).

    재리뷰 접미(재재리뷰 포함)는 모두 제거해 같은 canonical 로 정규화한다.
    """
    name = folder_name.strip()
    name = _ORDER_PREFIX_RE.sub("", name)

    level = re_review_level(folder_name)
    if level > 0:
        # 접미(재…리뷰 / re…review)를 통째로 제거
        name = _RE_REVIEW_LEVEL_RE.sub("", name)
        name = _EN_RE_REVIEW_RE.sub("", name)
    else:
        lowered = name.lower()
        for suffix in _RE_REVIEW_SUFFIXES:
            if lowered.endswith(suffix.lower()):
                name = name[: -len(suffix)]
                level = 1
                break

    canonical = name.strip().strip("._- ")
    return canonical, level > 0


def _canon_token(s: str) -> str:
    """배치 매칭용 단순화 토큰 (대문자/영숫자만)."""
    return re.sub(r"[^a-z0-9]", "", s.lower())


def build_grid(layers: list[str]) -> list[list[str | None]]:
    """주어진 layer 목록을 기본 그리드 배치에 채워 2열 그리드를 만든다.

    기본 배치(config.DEFAULT_LAYER_GRID)의 셀과 canonical 토큰이 일치하면 그 자리에 두고,
    배치에 없는 layer 는 남은 빈 칸/추가 행에 순서대로 채운다.
    """
    remaining = list(layers)
    grid: list[list[str | None]] = []

    # 1차: 기본 배치 위치에 일치하는 layer 부터 채운다.
    template = DEFAULT_LAYER_GRID
    for row_tpl in template:
        row: list[str | None] = []
        for cell in row_tpl:
            match = None
            for lyr in remaining:
                if _canon_token(lyr) == _canon_token(cell):
                    match = lyr
                    break
            if match is not None:
                remaining.remove(match)
                row.append(match)
            else:
                row.append(None)
        grid.append(row)

    # 2차: 배치에 없던 layer 들을 빈 칸 → 추가 행 순으로 채운다.
    def next_remaining() -> str | None:
        return remaining.pop(0) if remaining else None

    for row in grid:
        for c in range(len(row)):
            if row[c] is None and remaining:
                row[c] = next_remaining()

    while remaining:
        grid.append([next_remaining(), next_remaining()])

    # 완전히 빈 행 제거
    grid = [r for r in grid if any(cell is not None for cell in r)]
    return grid


# =============================================================================
# app/matcher.py   [#13]
# =============================================================================
"""매칭 엔진 (문서 Section 8.3 + 원본 AOI 도구 Module_Compare 알고리즘).

기준 layer 의 각 defect 에 대해 비교 layer 에서 다음 조건을 만족하는 defect 를 찾는다:
  1. wafer ID 동일
  2. die 위치 (col,row) 가 ±DIE_TOL(기본 1) 이내 — 경계 정렬 차이를 흡수
  3. local x,y 거리(아래 '정합오차 보정' 적용) <= tolerance

원본 AOI 도구의 비교 알고리즘에서 가져온 개선:
  - **die ±1 허용**: 두 layer 의 die index 가 1 칸 어긋나도 매칭.
  - **layer 간 전역 정합오차(median offset) 보정**: 1:1 로 분명히 매칭되는 쌍들로
    두 layer 사이의 계통적 이동량(중앙값 dx,dy)을 추정하고, 그 이동량을 뺀 잔차로
    게이팅·선택한다. 두 스캔 사이에 일정한 위치 오프셋이 있어도 매칭이 성립한다.
    단, 이 오프셋이 die pitch 급으로 비정상적으로 크면(_MAX_OFFSET_MAGNITUDE 초과)
    실제 정합오차가 아니라 die 라벨링 불일치로 보고 보정을 적용하지 않는다 — 그렇지
    않으면 서로 다른 die 를 매칭으로 잘못 보고하게 된다.

대량 이미지에서도 빠르도록 비교 layer record 를 (wafer, col, row) 키로 한 번만
인덱싱한다. 모든 계산은 메모리에서만 수행하며 원본 파일을 수정하지 않는다.
"""




# layer -> {(norm_wafer, col, row): [record, ...]}
DieIndex = dict[str, dict[tuple[str, int, int], list[DefectRecord]]]
# layer -> {norm_wafer: 좌표 추출 실패(not ok) record 수}
FailIndex = dict[str, dict[str, int]]

# die index 허용 오차(±). 0 이면 정확 일치.
DEFAULT_DIE_TOL = 1


@dataclass(frozen=True)
class LayerOffset:
    """비교 layer 의 기준 layer 대비 전역 정합오차(중앙값)와 표본 수."""

    dx: float = 0.0
    dy: float = 0.0
    count: int = 0  # 추정에 쓰인 1:1 매칭 쌍 수


def matcher__norm_wafer(wafer_id: str) -> str:
    """layer 간 wafer 폴더명 표기 차이(대소문자/공백)를 흡수해 매칭을 완화."""
    return wafer_id.strip().lower()


def _gather_candidates(
    base: DefectRecord,
    layer_bucket: dict,
    die_tol: int,
) -> list[DefectRecord]:
    """base 의 die 주변 ±die_tol 범위 비교 후보를 모은다(같은 wafer)."""
    if base.col is None or base.row is None:
        return []
    w = matcher__norm_wafer(base.wafer_id)
    out: list[DefectRecord] = []
    for dc in range(-die_tol, die_tol + 1):
        for dr in range(-die_tol, die_tol + 1):
            out.extend(layer_bucket.get((w, base.col + dc, base.row + dr), []))
    return out


# 정합오차를 신뢰하려면 최소 이만큼의 1:1 표본이 있어야 한다(단일 쌍 오적용 방지).
_MIN_OFFSET_SAMPLES = 3

# 정합오차(median offset) 크기 상한(µm) — 정답 도구 근거 문서상 KLA↔Camtek 실측
# 정합오차는 ~110~125µm 수준이다. die pitch 급(수만 µm)으로 "일관된" 오프셋이
# 나온다면 이는 실제 장비 정합오차가 아니라 die 라벨링/파싱 불일치를 정합오차로
# 오인한 것이다 — 그런 오프셋을 그대로 적용하면 서로 다른 die 를 "매칭"으로
# 잘못 보고하면서 실제 거리(raw distance)만 크게 표시되는 문제가 생긴다. 이 상한을
# 넘는 표본은 아무리 일관돼도(MAD 작아도) 보정하지 않고 미매칭으로 둔다.
_MAX_OFFSET_MAGNITUDE = 1000.0  # µm


def _mad(values: list[float], center: float) -> float:
    """median absolute deviation — 표본 일관성(흩어짐) 측정."""
    return statistics.median([abs(v - center) for v in values]) if values else 0.0


def _estimate_offset(
    dxs: list[float], dys: list[float], tolerance: float
) -> LayerOffset:
    """1:1 매칭 쌍의 dx,dy 표본에서 전역 정합오차(중앙값)를 추정한다.

    오적용을 막기 위해 (1) 표본 수 ≥ _MIN_OFFSET_SAMPLES, (2) 표본이 일관(MAD ≤
    tolerance), (3) 오프셋 크기가 _MAX_OFFSET_MAGNITUDE 이내일 때만 보정값을
    만든다. 그 외에는 보정 없음(LayerOffset()).
    """
    if len(dxs) >= _MIN_OFFSET_SAMPLES:
        mdx = statistics.median(dxs)
        mdy = statistics.median(dys)
        if (
            abs(mdx) <= _MAX_OFFSET_MAGNITUDE
            and abs(mdy) <= _MAX_OFFSET_MAGNITUDE
            and _mad(dxs, mdx) <= tolerance
            and _mad(dys, mdy) <= tolerance
        ):
            return LayerOffset(mdx, mdy, len(dxs))
    return LayerOffset()


def compute_layer_offsets(
    base_records: list[DefectRecord],
    compare_layers: list[str],
    index: DieIndex,
    tolerance: float,
    die_tol: int = DEFAULT_DIE_TOL,
) -> dict[str, LayerOffset]:
    """비교 layer 별 전역 정합오차(중앙값 dx,dy)를 die-단일 매칭 쌍으로 추정한다.

    die 주변(±die_tol)에 후보가 **정확히 1개**인 경우만 표본으로 사용한다(모호 배제).
    거리 게이트를 두지 않으므로 **허용오차보다 큰 계통적 shift 도 추정**할 수 있다.
    """
    offsets: dict[str, LayerOffset] = {}
    for layer in compare_layers:
        bucket = index.get(layer, {})
        dxs: list[float] = []
        dys: list[float] = []
        for base in base_records:
            if not base.ok:
                continue
            cands = [c for c in _gather_candidates(base, bucket, die_tol) if c.ok]
            if len(cands) == 1:
                dxs.append(base.x - cands[0].x)  # type: ignore[operator]
                dys.append(base.y - cands[0].y)  # type: ignore[operator]
        offsets[layer] = _estimate_offset(dxs, dys, tolerance)
    return offsets


def build_die_index(
    records_by_layer: dict[str, list[DefectRecord]],
    layers: list[str],
) -> DieIndex:
    """비교 layer 들을 (norm_wafer, col, row) 키로 인덱싱한다(좌표 OK 인 record 만)."""
    index: DieIndex = {}
    for layer in layers:
        bucket: dict[tuple[str, int, int], list[DefectRecord]] = defaultdict(list)
        for rec in records_by_layer.get(layer, []):
            if rec.ok:
                bucket[(matcher__norm_wafer(rec.wafer_id), rec.col, rec.row)].append(  # type: ignore[index]
                    rec
                )
        index[layer] = bucket
    return index


def build_fail_index(
    records_by_layer: dict[str, list[DefectRecord]],
    layers: list[str],
) -> FailIndex:
    """비교 layer·wafer 별 좌표 추출 실패(not ok) record 수를 센다.

    실패 record 는 col/row 가 None 이라 die 단위로 키를 만들 수 없으므로
    (layer, wafer) 단위로 근사 집계한다(진단 표시용).
    """
    index: FailIndex = {}
    for layer in layers:
        bucket: dict[str, int] = defaultdict(int)
        for rec in records_by_layer.get(layer, []):
            if not rec.ok:
                bucket[matcher__norm_wafer(rec.wafer_id)] += 1
        index[layer] = bucket
    return index


# 두 후보의 잔차 차가 이 값 미만이면 "동률(모호)"로 본다(µm).
_AMBIGUOUS_EPSILON = 1.0


def find_nearest(
    base: DefectRecord,
    candidates: list[DefectRecord],
) -> tuple[DefectRecord | None, float | None]:
    """허용오차와 무관하게 후보 중 최근접 record 를 반환(진단용, raw 거리)."""
    best: DefectRecord | None = None
    best_dist: float | None = None
    for cand in candidates:
        dist = base.distance_to(cand)
        if dist is None:
            continue
        if best_dist is None or dist < best_dist:
            best = cand
            best_dist = dist
    return best, best_dist


def _select_match(
    base: DefectRecord,
    candidates: list[DefectRecord],
    tolerance: float,
    offset: LayerOffset,
) -> tuple[DefectRecord | None, float | None, bool]:
    """후보 중 최적 매치를 고른다.

    잔차 = hypot(dx-offset.dx, dy-offset.dy) <= tolerance 인 후보 중 잔차 최소를 고른다
    (정합오차가 tolerance 보다 커도 일관되면 보정). 반환 distance 는 보정 전 실제
    거리(raw)로 보고한다. offset 이 (0,0) 이면 raw 거리 기준 최근접과 같다.
    """
    if not base.ok:
        return None, None, False
    best: DefectRecord | None = None
    best_resid: float | None = None
    best_raw: float | None = None
    resids: list[float] = []
    for c in candidates:
        if not c.ok:
            continue
        dx = base.x - c.x  # type: ignore[operator]
        dy = base.y - c.y  # type: ignore[operator]
        if math.hypot(dx - offset.dx, dy - offset.dy) > tolerance:
            continue
        resid = math.hypot(dx - offset.dx, dy - offset.dy)  # tie-break/모호 판정 공통
        resids.append(resid)
        if best_resid is None or resid < best_resid:
            best_resid = resid
            best = c
            best_raw = math.hypot(dx, dy)
    ambiguous = (
        best is not None
        and sum(1 for r in resids if abs(r - best_resid) < _AMBIGUOUS_EPSILON) >= 2
    )
    return best, best_raw, ambiguous


def _build_result(
    base: DefectRecord,
    layer: str,
    candidates: list[DefectRecord],
    tolerance: float,
    offset: LayerOffset,
    failed_count: int,
) -> MatchResult:
    """미리 수집한 후보(candidates)로 한 layer 의 매칭 결과를 만든다.

    후보 수집을 호출자가 책임지므로(중복 수집 제거) 매칭 로직만 담당한다.
    """
    matched: DefectRecord | None = None
    dist: float | None = None
    nearest: DefectRecord | None = None
    nearest_dist: float | None = None
    failed_in_die = 0
    ambiguous = False
    if base.ok:
        matched, dist, ambiguous = _select_match(
            base, candidates, tolerance, offset
        )
        if matched is None:
            nearest, nearest_dist = find_nearest(base, candidates)
            failed_in_die = failed_count
    return MatchResult(
        compare_layer=layer,
        base=base,
        matched=matched,
        distance=dist,
        nearest=nearest,
        nearest_distance=nearest_dist,
        die_candidates=len(candidates),
        failed_in_die=failed_in_die,
        ambiguous=ambiguous,
    )


def match_base_against_layers(
    base: DefectRecord,
    compare_layers: list[str],
    records_by_layer: dict[str, list[DefectRecord]],
    tolerance: float,
    *,
    index: DieIndex | None = None,
    fail_index: FailIndex | None = None,
    offsets: dict[str, LayerOffset] | None = None,
    die_tol: int = DEFAULT_DIE_TOL,
) -> BaseDefectMatches:
    """기준 defect 1개를 모든 비교 layer 와 매칭.

    index/fail_index/offsets 를 미리 만들어 넘기면 재사용한다(match_all 에서 사용).
    넘기지 않으면 이 호출에 한해 즉석에서 만든다(offsets 미지정 시 정합오차 보정 없음).
    """
    idx = index if index is not None else build_die_index(records_by_layer, compare_layers)
    fidx = (
        fail_index
        if fail_index is not None
        else build_fail_index(records_by_layer, compare_layers)
    )
    result = BaseDefectMatches(base=base)
    for layer in compare_layers:
        candidates = (
            _gather_candidates(base, idx.get(layer, {}), die_tol) if base.ok else []
        )
        offset = (offsets or {}).get(layer, LayerOffset())
        failed_count = fidx.get(layer, {}).get(matcher__norm_wafer(base.wafer_id), 0)
        result.results.append(
            _build_result(
                base, layer, candidates, tolerance, offset, failed_count,
            )
        )
    return result


def match_all_with_offsets(
    base_records: list[DefectRecord],
    compare_layers: list[str],
    records_by_layer: dict[str, list[DefectRecord]],
    tolerance: float,
    *,
    index: DieIndex | None = None,
    fail_index: FailIndex | None = None,
    die_tol: int = DEFAULT_DIE_TOL,
) -> tuple[list[BaseDefectMatches], dict[str, LayerOffset]]:
    """매칭 결과 목록과 비교 layer 별 전역 정합오차를 함께 반환한다.

    각 (base, layer) 후보를 **한 번만** 수집해 정합오차 추정과 매칭에 함께 쓴다
    (이전엔 두 패스에서 각각 수집 → 중복 제거로 대량 이미지에서 빨라진다).
    """
    idx = index if index is not None else build_die_index(records_by_layer, compare_layers)
    fidx = (
        fail_index
        if fail_index is not None
        else build_fail_index(records_by_layer, compare_layers)
    )
    buckets = {layer: idx.get(layer, {}) for layer in compare_layers}

    # 1패스: (base, layer) 후보 수집 + 1:1 쌍에서 정합오차 표본 적립.
    cand_table: list[dict[str, list[DefectRecord]]] = []
    samples: dict[str, tuple[list[float], list[float]]] = {
        layer: ([], []) for layer in compare_layers
    }
    for base in base_records:
        row: dict[str, list[DefectRecord]] = {}
        for layer in compare_layers:
            cands = _gather_candidates(base, buckets[layer], die_tol) if base.ok else []
            row[layer] = cands
            if base.ok:
                ok_cands = [c for c in cands if c.ok]
                if len(ok_cands) == 1:
                    dxs, dys = samples[layer]
                    dxs.append(base.x - ok_cands[0].x)  # type: ignore[operator]
                    dys.append(base.y - ok_cands[0].y)  # type: ignore[operator]
        cand_table.append(row)

    offsets = {
        layer: _estimate_offset(samples[layer][0], samples[layer][1], tolerance)
        for layer in compare_layers
    }

    # 2패스: 수집한 후보 재사용해 매칭 결과 구성.
    matches: list[BaseDefectMatches] = []
    for base, row in zip(base_records, cand_table):
        result = BaseDefectMatches(base=base)
        wafer = matcher__norm_wafer(base.wafer_id)
        for layer in compare_layers:
            failed_count = fidx.get(layer, {}).get(wafer, 0)
            result.results.append(
                _build_result(
                    base, layer, row[layer], tolerance, offsets[layer], failed_count,
                )
            )
        matches.append(result)
    return matches, offsets


def match_all(
    base_records: list[DefectRecord],
    compare_layers: list[str],
    records_by_layer: dict[str, list[DefectRecord]],
    tolerance: float,
) -> list[BaseDefectMatches]:
    """기준 layer 의 모든 defect 에 대해 매칭 결과 목록을 만든다(정합오차 보정 포함)."""
    matches, _ = match_all_with_offsets(
        base_records, compare_layers, records_by_layer, tolerance,
    )
    return matches


# =============================================================================
# app/parsers/camtek_filename.py   [#14]
# =============================================================================
"""Camtek 파일명에서 직접 좌표 추출 (문서 Section 6 / 13.3.4 + 실제 AOI 도구 스키마).

Camtek 스캔 이미지는 파일명에 die 위치(col,row)와 local 좌표(x,y)가 포함된다.
관찰된 레이아웃(원본 자료 Module_KLA 의 파일명 생성 규칙 포함):

  (A) Section 6      : R_..._{wafer}_{col}_{row}_{x}_{y}_{DefectName}.jpg
  (B) Section 13.3.4 : R_..._{wafer}_{col}_{row}_{DefectName}_{x}_{y}
  (C) AOI 변환 결과  : {PREFIX}_{LOT}_{wafer}_{col}_{row}_{DefectName}_{x}_{y}_{DXSize}_{DYSize}_{DArea}.jpg

핵심 규칙(세 레이아웃 공통):
  - col/row 는 die 위치를 나타내는 **연속한 두 정수 토큰**(파일명에서 처음 등장하는 쌍).
  - col/row 뒤에는 **반드시 defect 이름(영문자 포함 토큰)** 이 존재한다.
    (KLA 원본 이름 `wafer_0_1_7_1` 처럼 이름이 없는 건 Camtek 형식이 아님 → NOT_FOUND)
  - x/y 는 col/row 뒤에서 **처음 등장하는 두 수치 토큰**(정수/소수 모두 가능).
    (C) 처럼 그 뒤에 수치가 더 있으면 차례로 DXSize/DYSize/DArea 로 본다.
"""




_INT_RE = re.compile(r"^-?\d+$")
camtek_filename__NUM_RE = re.compile(r"^-?\d+(?:\.\d+)?$")  # 정수 또는 소수
_HAS_LETTER_RE = re.compile(r"[A-Za-z가-힣]")


@dataclass
class CamtekNameResult:
    status: ParseStatus
    col: Optional[int] = None
    row: Optional[int] = None
    x: Optional[float] = None
    y: Optional[float] = None
    defect_name: str = ""
    dx_size: Optional[float] = None
    dy_size: Optional[float] = None
    d_area: Optional[float] = None
    reason: str = ""  # 진단용: 실패 사유(성공이면 빈 문자열)


def _is_number(tok: str) -> bool:
    return bool(camtek_filename__NUM_RE.match(tok))


def parse_camtek_filename(filename: str) -> CamtekNameResult:
    """Camtek 파일명에서 col/row/x/y(+크기/면적)와 defect 이름을 추출한다.

    좌표 형식이 아니면 ParseStatus.NOT_FOUND 를 반환한다.
    """
    stem = re.sub(r"\.(jpg|jpeg|png|bmp|tif|tiff)$", "", filename, flags=re.IGNORECASE)
    tokens = stem.split("_")

    # col, row = 연속한 두 정수 토큰 중 첫 번째 쌍 (die 위치)
    col = row = col_idx = None
    for i in range(len(tokens) - 1):
        if _INT_RE.match(tokens[i]) and _INT_RE.match(tokens[i + 1]):
            col, row, col_idx = int(tokens[i]), int(tokens[i + 1]), i
            break
    if col is None:
        return CamtekNameResult(
            ParseStatus.NOT_FOUND,
            reason=f"파일명에서 연속한 두 정수(die col/row)를 찾지 못함 (토큰 {len(tokens)}개)",
        )

    after = tokens[col_idx + 2:]
    # col/row 뒤에 defect 이름(영문자 포함)이 없으면 Camtek 형식이 아님(KLA 원본 등).
    if not any(_HAS_LETTER_RE.search(t) for t in after):
        return CamtekNameResult(
            ParseStatus.NOT_FOUND,
            reason="col/row 뒤 defect 이름(영문자) 토큰 없음 → KLA 원본 파일명일 가능성",
        )

    # x/y = col/row 뒤 처음 등장하는 두 수치 토큰. 그 뒤 수치는 크기/면적.
    nums = [t for t in after if _is_number(t)]
    if len(nums) < 2:
        return CamtekNameResult(
            ParseStatus.NOT_FOUND,
            reason=f"col/row 뒤 x/y 수치 토큰 부족({len(nums)}개, 2개 필요)",
        )
    x = float(nums[0])
    y = float(nums[1])
    dx_size = float(nums[2]) if len(nums) >= 3 else None
    dy_size = float(nums[3]) if len(nums) >= 4 else None
    d_area = float(nums[4]) if len(nums) >= 5 else None

    defect_name = " ".join(t for t in after if not _is_number(t)).strip()

    return CamtekNameResult(
        status=ParseStatus.OK,
        col=col,
        row=row,
        x=x,
        y=y,
        defect_name=defect_name,
        dx_size=dx_size,
        dy_size=dy_size,
        d_area=d_area,
    )


# =============================================================================
# app/parsers/camtek_ini.py   [#15]
# =============================================================================
"""ColorImageGrabingInfo.ini 기반 Camtek 좌표 산출 (문서 Section 13.3).

원본 이미지 이름(예: 253715.91797.c.-1104740629.1)에 ".jpeg" 를 붙여 INI section 을 찾고,
해당 section 의 X/Y(없으면 FaultX/FaultY)와 Col/Row 로 위치 정보를 계산한다.

  col = Col - camtek_col_offset
  row = camtek_row_base - Row
  x   = X - Col * camtek_pitch_x
  y   = Y - Row * camtek_pitch_y

(상수는 활성 제품 프로파일(`config.active_product()`)에서 온다 — 기본 DEVA 은
col_offset=2, row_base=7, pitch_x=37170.0, pitch_y=44830.0.)

원본 INI 는 read-only 로만 읽는다.
"""




camtek_ini__log = logging.getLogger("defect_tracker.parsers.ini")


@dataclass
class CamtekIniResult:
    status: ParseStatus
    col: Optional[int] = None
    row: Optional[int] = None
    x: Optional[float] = None
    y: Optional[float] = None
    reason: str = ""  # 진단용: 실패 사유(성공이면 빈 문자열)


def _parse_ini_sections(text: str) -> dict[str, dict[str, str]]:
    """INI 텍스트를 {section_name(소문자): {key(소문자): value}} 로 파싱.

    configparser 대신 직접 파싱: section 이름에 '.'/'-' 등 특수문자가 많고
    중복/비표준 라인이 있을 수 있으므로 관대하게 처리한다.
    """
    sections: dict[str, dict[str, str]] = {}
    current: Optional[str] = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith((";", "#")):
            continue
        if line.startswith("[") and line.endswith("]"):
            current = line[1:-1].strip().lower()
            sections.setdefault(current, {})
        elif "=" in line and current is not None:
            key, _, value = line.partition("=")
            sections[current][key.strip().lower()] = value.strip()
    return sections


def camtek_ini__read_text(ini_path: Path) -> str:
    data = read_only_bytes(ini_path)
    for encoding in ("utf-8-sig", "utf-8", "cp949", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1", errors="replace")


def load_ini(ini_path: str | Path) -> dict[str, dict[str, str]]:
    """INI 파일을 파싱해 section 사전을 반환 (캐싱은 호출 측 책임)."""
    return _parse_ini_sections(camtek_ini__read_text(Path(ini_path)))


def _to_float(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _to_int(value: Optional[str]) -> Optional[int]:
    f = _to_float(value)
    if f is None:
        return None
    i = int(round(f))
    if abs(f - i) > 0.01:
        # Col/Row 는 정수여야 한다. 비정수면 데이터 이상 가능성 → 경고(반올림은 진행).
        camtek_ini__log.warning("INI Col/Row 비정수 값 %s → %d 로 반올림", value, i)
    return i


def convert_from_sections(
    sections: dict[str, dict[str, str]], original_name: str
) -> CamtekIniResult:
    """미리 파싱한 section 사전에서 원본 이름에 해당하는 좌표를 계산."""
    # INI section 키는 원본 이미지 파일명(확장자 포함)인데 저장된 확장자가
    # 제각각일 수 있으므로 후보를 순서대로 시도한다(.jpeg 우선으로 기존 동작 보존).
    section = None
    for ext in (".jpeg", ".jpg", ".png", ".bmp", ".tif", ".tiff", ""):
        section = sections.get(f"{original_name}{ext}".lower())
        if section is not None:
            break
    if section is None:
        return CamtekIniResult(
            ParseStatus.NOT_FOUND,
            reason=f"INI 에 이미지 '{original_name}' section 없음(section {len(sections)}개)",
        )

    # X/Y 우선, 없으면 FaultX/FaultY (Section 13.3.3)
    x_raw = _to_float(section.get("x"))
    if x_raw is None:
        x_raw = _to_float(section.get("faultx"))
    y_raw = _to_float(section.get("y"))
    if y_raw is None:
        y_raw = _to_float(section.get("faulty"))
    col_ini = _to_int(section.get("col"))
    row_ini = _to_int(section.get("row"))

    if x_raw is None or y_raw is None or col_ini is None or row_ini is None:
        missing = [
            n for n, v in (("x/faultx", x_raw), ("y/faulty", y_raw),
                           ("col", col_ini), ("row", row_ini)) if v is None
        ]
        return CamtekIniResult(
            ParseStatus.INVALID_INFO,
            reason=f"INI section 필드 누락: {', '.join(missing)}",
        )

    prod = active_product()
    col = col_ini - prod.camtek_col_offset
    row = prod.camtek_row_base - row_ini
    x = x_raw - col_ini * prod.camtek_pitch_x
    y = y_raw - row_ini * prod.camtek_pitch_y

    return CamtekIniResult(status=ParseStatus.OK, col=col, row=row, x=x, y=y)


def convert_camtek_ini(ini_path: str | Path, original_name: str) -> CamtekIniResult:
    """INI 파일을 읽어 원본 이름에 해당하는 col_row_x_y 위치 정보를 계산."""
    try:
        sections = load_ini(ini_path)
    except OSError as exc:
        return CamtekIniResult(
            ParseStatus.INFO_FILE_NOT_FOUND, reason=f"INI 파일 열기 실패: {exc}"
        )
    return convert_from_sections(sections, original_name)


# =============================================================================
# app/parsers/kla_info.py   [#16]
# =============================================================================
"""KLA info(.001) 기반 좌표 변환 (문서 Section 13.2 및 KLA 변환 보고서).

같은 wafer 폴더의 KLA info 파일에서 jpg 파일명과 동일한 TiffFileName 을 찾고,
그 아래 DefectList line 의 XREL/YREL/XINDEX/YINDEX 와 header 의 DiePitchY 로 변환한다.

  col = XINDEX + zeroX
  row = YINDEX + zeroY
  x   = Round(XREL, 0)
  y   = Round(DiePitchY - YREL, 0)   # Y 방향은 DiePitchY 기준으로 반전

zeroX/zeroY 는 info 파일의 `SampleTestPlan` 블록(그 lot/step 에서 실제 관측된
XINDEX/YINDEX 최소값)에서 우선 계산한다 — 제품 설정(PackageX/Y÷2)은 이 블록이
없을 때만 쓰는 폴백이다. 제품 설정값은 디바이스 DB 시트를 잘못 참조하는 등으로
실측과 어긋날 수 있어(정답 VBA 도구/AOIDeviceDB.xlsx 로 확인한 DEVA 사례 —
"DEVA"과 "DEVA Live" 두 시트가 있는데 후자가 실제 운영 값(PackageY=6, zeroY=3)),
info 파일 자신의 실측값이 더 신뢰할 수 있다.

원본 info/이미지 파일은 read-only 로만 읽는다.
"""




kla_info__log = logging.getLogger("defect_tracker.parsers.kla")

# DefectList 필드 순서(문서 Section 13.2.5.2):
# DEFECTID X Y XREL YREL XINDEX YINDEX XSIZE YSIZE ...
_XREL_IDX = 3
_YREL_IDX = 4
_XINDEX_IDX = 5
_YINDEX_IDX = 6
_MIN_FIELDS = 7


@dataclass
class KlaResult:
    status: ParseStatus
    col: Optional[int] = None
    row: Optional[int] = None
    x: Optional[float] = None
    y: Optional[float] = None
    reason: str = ""  # 진단용: 실패 사유(성공이면 빈 문자열)


def select_info_file(filenames: list[str]) -> Optional[str]:
    """wafer 폴더 파일 목록에서 KLA info 파일을 고른다 (문서 Section 13.2.3).

    1) 확장자 .001 우선
    2) 없으면 .pass 가 아닌 비-jpg 파일
    3) .pass / .jpg(.jpeg) 는 info 로 사용하지 않음
    """
    def ext(name: str) -> str:
        return Path(name).suffix.lower()

    dot001 = [f for f in filenames if ext(f) == ".001"]
    if dot001:
        return sorted(dot001)[0]

    _skip = frozenset((".pass", ".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"))
    candidates = [
        f
        for f in filenames
        if ext(f) not in _skip
    ]
    return sorted(candidates)[0] if candidates else None


def kla_info__read_text(path: Path) -> str:
    data = read_only_bytes(path)
    for encoding in ("utf-8", "cp949", "utf-16-le", "utf-16-be", "latin-1"):
        try:
            return data.decode(encoding)
        except (UnicodeDecodeError, ValueError):
            continue
    return data.decode("latin-1", errors="replace")


def _parse_die_pitch_y(lines: list[str]) -> Optional[float]:
    for line in lines:
        s = line.strip()
        if s.lower().startswith("diepitch"):
            nums = re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
            if len(nums) >= 2:
                try:
                    return float(nums[1])
                except ValueError:
                    return None
    return None


def _parse_sample_test_plan(lines: list[str]) -> Optional[tuple[int, int]]:
    """`SampleTestPlan N` 블록에서 실측 XINDEX/YINDEX 최소값 기준 zeroX/zeroY 를 구한다.

    블록은 `SampleTestPlan <개수>` 라인 다음에 `XINDEX YINDEX` 쌍이 이어지고,
    선언된 개수만큼 읽거나 `;`로 끝나는 라인을 만나면 끝난다.
    """
    n = len(lines)
    i = 0
    while i < n:
        m = re.match(r"(?i)^\s*SampleTestPlan\s+(\d+)\s*$", lines[i])
        if not m:
            i += 1
            continue
        count = int(m.group(1))
        xs: list[int] = []
        ys: list[int] = []
        i += 1
        while i < n and len(xs) < count:
            entry = lines[i].strip()
            i += 1
            if not entry:
                continue
            terminated = entry.endswith(";")
            nums = re.findall(r"-?\d+", entry.rstrip(";").strip())
            if len(nums) >= 2:
                xs.append(int(nums[0]))
                ys.append(int(nums[1]))
            if terminated:
                break
        if xs and ys:
            return (-min(xs), -min(ys))
        return None
    return None


def _parse_defect_fields(data_line: str) -> Optional[list[float]]:
    cleaned = data_line.strip().rstrip(";").strip()
    parts = cleaned.split()
    if len(parts) < _MIN_FIELDS:
        return None
    try:
        return [float(p) for p in parts[:_MIN_FIELDS]]
    except ValueError:
        return None


_KLA_FNAME_RE = re.compile(
    r"^.+?_(-?\d+)_(-?\d+)_(\d+)_(\d+)\.[a-zA-Z]+$"
)


@dataclass
class _ParsedInfo:
    die_pitch_y: Optional[float]
    # jpg 파일명(소문자) -> DefectList 필드
    defects: dict[str, list[float]]
    all_defects: list[list[float]] = field(default_factory=list)
    # SampleTestPlan 실측 기반 (zeroX, zeroY). 블록이 없으면 None(제품 설정값 폴백).
    sample_zero: Optional[tuple[int, int]] = None


def parse_info_text(text: str) -> _ParsedInfo:
    """KLA info 텍스트를 파싱해 DiePitchY 와 TiffFileName->DefectList 매핑을 만든다.

    TiffFileName 이 없는 DefectList 엔트리도 all_defects 에 수집하여,
    TiffFileName 매칭 실패 시 XINDEX/YINDEX 기반 폴백 검색에 쓴다.
    """
    lines = text.splitlines()
    die_pitch_y = _parse_die_pitch_y(lines)
    sample_zero = _parse_sample_test_plan(lines)
    defects: dict[str, list[float]] = {}
    all_defects: list[list[float]] = []

    i = 0
    n = len(lines)
    pending_tiff: Optional[str] = None
    in_defect_block = False
    while i < n:
        s = lines[i].strip()
        m = re.match(r"(?i)^TiffFileName\s+(.+?);?\s*$", s)
        if m:
            pending_tiff = m.group(1).strip().strip(";").strip().lower()
            in_defect_block = False
            i += 1
            continue
        if s.lower().startswith("defectlist"):
            rest = s[len("defectlist"):].strip()
            if rest:
                fields = _parse_defect_fields(rest)
                if fields is not None:
                    all_defects.append(fields)
                    if pending_tiff:
                        defects[pending_tiff] = fields
                        pending_tiff = None
            else:
                in_defect_block = True
            i += 1
            continue
        if (in_defect_block or pending_tiff) and s:
            fields = _parse_defect_fields(s)
            if fields is not None:
                all_defects.append(fields)
                if pending_tiff:
                    defects[pending_tiff] = fields
                    pending_tiff = None
            else:
                in_defect_block = False
                pending_tiff = None
        i += 1

    return _ParsedInfo(die_pitch_y=die_pitch_y, defects=defects,
                       all_defects=all_defects, sample_zero=sample_zero)


def load_info(info_path: str | Path) -> _ParsedInfo:
    return parse_info_text(kla_info__read_text(Path(info_path)))


def convert_from_parsed(parsed: _ParsedInfo, jpg_filename: str) -> KlaResult:
    """미리 파싱한 info 에서 jpg 파일명에 해당하는 좌표를 계산."""
    key = Path(jpg_filename).name.lower()
    fields = parsed.defects.get(key)
    if fields is None:
        # 정확 파일명 실패 시 stem(확장자 제외) 일치로 한 번 더 시도.
        # 예) 이미지 foo.jpg ↔ TiffFileName foo.tif 처럼 확장자만 다른 경우.
        stem = Path(jpg_filename).stem.lower()
        for k, v in parsed.defects.items():
            if Path(k).stem == stem:
                fields = v
                break
    fname_match = _KLA_FNAME_RE.match(key)
    if fields is None and parsed.all_defects and fname_match:
        want_xi = int(fname_match.group(1))
        want_yi = int(fname_match.group(2))
        want_id = int(fname_match.group(4))
        for entry in parsed.all_defects:
            if (len(entry) >= _MIN_FIELDS
                    and int(round(entry[_XINDEX_IDX])) == want_xi
                    and int(round(entry[_YINDEX_IDX])) == want_yi
                    and int(round(entry[0])) == want_id):
                fields = entry
                break
    if fields is None:
        # class(그룹 3)==0(Unclassified) 파일명은 KLARF 에 정식 결함으로 등록된 적
        # 없는 미분류 후보 이미지일 가능성이 높다(정상) — 파일명을 사유에 넣지 않는
        # 고정 문구로 반환해 진단 리포트에서 파일마다 따로 클러스터링되지 않게 한다.
        if fname_match and fname_match.group(3) == "0":
            return KlaResult(
                ParseStatus.NOT_FOUND,
                reason=(
                    "KLA: 미분류(class 0) 후보 이미지 — info DefectList 에 정식 결함으로 "
                    "등록되지 않음(정상, 무시 가능)"
                ),
            )
        sample = ", ".join(list(parsed.defects)[:3])
        return KlaResult(
            ParseStatus.NOT_FOUND,
            reason=(
                f"info 에 TiffFileName '{key}' 매칭 실패"
                f"(보유 {len(parsed.defects)}개{', 예: ' + sample if sample else ''}, "
                f"DefectList {len(parsed.all_defects)}건)"
            ),
        )
    if parsed.die_pitch_y is None:
        return KlaResult(
            ParseStatus.INVALID_INFO, reason="info header 에 DiePitchY 없음"
        )

    xrel = fields[_XREL_IDX]
    yrel = fields[_YREL_IDX]
    xindex = int(round(fields[_XINDEX_IDX]))
    yindex = int(round(fields[_YINDEX_IDX]))

    zero_x, zero_y = parsed.sample_zero or (kla_zero_x(), kla_zero_y())
    col = xindex + zero_x
    row = yindex + zero_y
    # 음수 die 위치는 비정상(잘못된 XINDEX/YINDEX) — 잘못 매칭되지 않도록 실패 처리.
    if col < 0 or row < 0:
        kla_info__log.warning(
            "KLA die 위치가 음수입니다(col=%s,row=%s) — %s", col, row, jpg_filename
        )
        return KlaResult(
            ParseStatus.INVALID_INFO,
            reason=f"die 위치 음수(col={col},row={row}) — XINDEX/YINDEX 비정상",
        )
    x = round(xrel)
    y = round(parsed.die_pitch_y - yrel)
    return KlaResult(status=ParseStatus.OK, col=col, row=row, x=float(x), y=float(y))


def convert_kla(info_path: str | Path, jpg_filename: str) -> KlaResult:
    """KLA info 파일을 읽어 jpg 에 해당하는 col_row_x_y 위치 정보를 계산."""
    try:
        parsed = load_info(info_path)
    except OSError as exc:
        return KlaResult(
            ParseStatus.INFO_FILE_NOT_FOUND, reason=f"info 파일 열기 실패: {exc}"
        )
    return convert_from_parsed(parsed, jpg_filename)


# =============================================================================
# app/thumbnails.py   [#17]
# =============================================================================
"""썸네일 생성/캐시 (문서 Section 8.6, 10).

상단 썸네일은 사진 중앙 일부 구간(기본 10%)만 잘라 확대한 형태로 보여준다.
생성된 썸네일은 output workspace 내 cache 폴더에만 저장하며, 원본 폴더에는 쓰지 않는다.
캐시 키는 (원본 경로 + 크기 + mtime) 해시로 만들어 원본 변경 시 자동 갱신된다.

원본 이미지는 read-only 로만 읽는다.
"""






def _cache_key(path: Path, size: int, center_ratio: float, full: bool) -> str:
    try:
        stat = path.stat()
        sig = f"{path}|{stat.st_size}|{int(stat.st_mtime)}|{size}|{center_ratio}|{full}"
    except OSError:
        sig = f"{path}|{size}|{center_ratio}|{full}"
    return hashlib.sha1(sig.encode("utf-8")).hexdigest()


def _center_crop(img: Image.Image, ratio: float) -> Image.Image:
    """중앙 ratio 비율 영역을 잘라낸다 (사진 중앙 N% 확대용)."""
    w, h = img.size
    cw = max(1, int(w * ratio))
    ch = max(1, int(h * ratio))
    left = (w - cw) // 2
    top = (h - ch) // 2
    return img.crop((left, top, left + cw, top + ch))


class ThumbnailCache:
    """썸네일을 디스크 캐시에 저장하고 재사용한다."""

    def __init__(self, cache_dir: Path, size: int = THUMBNAIL_SIZE):
        self.cache_dir = Path(cache_dir)
        self.size = size
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _cache_path(self, key: str) -> Path:
        return self.cache_dir / f"{key}.png"

    def get_center_thumbnail(
        self,
        image_path: str | Path,
        center_ratio: float = THUMBNAIL_CENTER_RATIO,
    ) -> Optional[Path]:
        """중앙 center_ratio 영역을 확대한 썸네일 경로를 반환(없으면 생성)."""
        return self._get(image_path, center_ratio=center_ratio, full=False)

    def get_full_thumbnail(
        self, image_path: str | Path, max_size: Optional[int] = None
    ) -> Optional[Path]:
        """이미지 전체를 축소한 미리보기 썸네일(Excel/그리드용)."""
        return self._get(image_path, center_ratio=1.0, full=True, max_size=max_size)

    def _get(
        self,
        image_path: str | Path,
        center_ratio: float,
        full: bool,
        max_size: Optional[int] = None,
    ) -> Optional[Path]:
        path = Path(image_path)
        target = max_size or self.size
        key = _cache_key(path, target, center_ratio, full)
        out = self._cache_path(key)
        if out.exists():
            return out
        # 원자적 쓰기: 임시 파일에 저장 후 os.replace 로 교체 → 동시 읽기가 부분 파일을 보지
        # 않게 한다(백그라운드 워밍 워커와 UI fill 이 같은 캐시 파일을 동시에 다룰 수 있음).
        tmp = out.with_name(f"{out.stem}.{os.getpid()}_{id(path)}.tmp")
        try:
            data = read_only_bytes(path)
            with Image.open(io.BytesIO(data)) as img:
                img = img.convert("RGB")
                if not full and center_ratio < 1.0:
                    img = _center_crop(img, center_ratio)
                img.thumbnail((target, target), Image.LANCZOS)
                img.save(tmp, format="PNG")
            os.replace(tmp, out)
            return out
        except (OSError, ValueError):
            try:
                tmp.unlink()
            except OSError:
                pass
            return None


# =============================================================================
# app/ui/busy_overlay.py   [#18]
# =============================================================================
"""로딩(작업 중) 오버레이 — 부모 위 반투명 막 + 중앙 카드(부드러운 스피너·메시지·진행바).

무거운 작업이 진행되는 동안 '멈춘 것'처럼 보이지 않도록, 부드럽게 회전하는 네온 링과
(가능하면) 진행도를 표시한다. 부모의 크기에 맞춰 자동으로 덮는다.
"""






class _SpinnerRing(QWidget):
    """부드럽게 회전하는 네온 원호(브라유 글리프보다 매끄럽게)."""

    def __init__(self, parent: Optional[QWidget] = None, size: int = 52):
        super().__init__(parent)
        self._angle = 0.0
        self.setFixedSize(size, size)

    def advance(self, deg: float) -> None:
        self._angle = (self._angle + deg) % 360.0
        self.update()

    def paintEvent(self, event):  # noqa: N802
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        m = 7
        rect = QRectF(m, m, self.width() - 2 * m, self.height() - 2 * m)
        # 바탕 트랙(희미한 전체 원)
        track = QPen(QColor(NEON_SOFT))
        track.setWidth(5)
        track.setCapStyle(Qt.RoundCap)
        p.setPen(track)
        p.drawArc(rect, 0, 360 * 16)
        # 밝은 회전 호(약 110°)
        arc = QPen(QColor(NEON))
        arc.setWidth(5)
        arc.setCapStyle(Qt.RoundCap)
        p.setPen(arc)
        p.drawArc(rect, int(-self._angle * 16), 110 * 16)
        p.end()


class BusyOverlay(QWidget):
    """부모를 덮는 로딩 오버레이. start/stop/set_progress 로 제어."""

    def __init__(self, parent: QWidget):
        super().__init__(parent)
        self._host = parent
        self.setVisible(False)
        self.setAttribute(Qt.WA_TransparentForMouseEvents, False)  # 클릭 삼켜 조작 방지

        lay = QVBoxLayout(self)
        lay.setAlignment(Qt.AlignCenter)
        card = QFrame()
        card.setObjectName("busyCard")
        card.setFixedWidth(320)
        # 은은한 네온 테두리 + 살짝 밝은 배경으로 카드를 또렷하게.
        card.setStyleSheet(
            "QFrame#busyCard {"
            f" background:{BG_ELEV};"
            f" border:1px solid {NEON_SOFT};"
            " border-radius:14px; }"
        )
        cl = QVBoxLayout(card)
        cl.setContentsMargins(26, 24, 26, 22)
        cl.setSpacing(14)
        cl.setAlignment(Qt.AlignCenter)

        self._ring = _SpinnerRing()
        cl.addWidget(self._ring, alignment=Qt.AlignHCenter)

        self._base_msg = "처리 중"
        self._msg = QLabel(self._base_msg)
        self._msg.setAlignment(Qt.AlignCenter)
        self._msg.setWordWrap(True)
        self._msg.setStyleSheet(
            f"color:{TEXT}; font-weight:700; font-size:14px; border:none;"
        )
        cl.addWidget(self._msg)

        self._sub = QLabel("잠시만 기다려 주세요")
        self._sub.setAlignment(Qt.AlignCenter)
        self._sub.setStyleSheet(
            f"color:{TEXT_DIM}; font-size:11px; border:none;"
        )
        cl.addWidget(self._sub)

        self._bar = QProgressBar()
        self._bar.setTextVisible(True)
        self._bar.setFixedHeight(14)
        self._bar.setVisible(False)
        cl.addWidget(self._bar)

        lay.addWidget(card)

        self._dot = 0
        self._frame = 0
        self._timer = QTimer(self)
        self._timer.setInterval(40)  # 부드러운 회전(≈25fps)
        self._timer.timeout.connect(self._tick)

        self._host.installEventFilter(self)

    # ---- 표시 제어 ---------------------------------------------------
    def start(self, message: str = "처리 중", determinate: bool = False) -> None:
        self._base_msg = message.rstrip("… .")
        self._msg.setText(self._base_msg)
        self._sub.setVisible(not determinate)
        self._bar.setVisible(determinate)
        if determinate:
            self._bar.setRange(0, 100)
            self._bar.setValue(0)
        self._reposition()
        self.setVisible(True)
        self.raise_()
        if not self._timer.isActive():
            self._timer.start()

    def set_message(self, message: str) -> None:
        self._base_msg = message.rstrip("… .")
        self._msg.setText(self._base_msg + "." * self._dot)

    def set_progress(self, cur: int, total: int) -> None:
        if total <= 0:
            return
        if not self._bar.isVisible():
            self._bar.setVisible(True)
        self._bar.setRange(0, 100)
        self._bar.setValue(int(round(min(cur, total) / total * 100)))

    def stop(self) -> None:
        self._timer.stop()
        self.setVisible(False)

    def pump(self) -> None:
        """무거운 메인 스레드 작업 중에도 스피너가 계속 회전하도록 이벤트 루프를 잠깐 돌린다.

        사용자 입력 이벤트는 제외해 작업 도중 재진입(레이어 재변경 등)을 막고,
        타이머·페인트 이벤트만 처리해 애니메이션 프레임을 갱신한다.
        """
        if not self.isVisible():
            return
        QApplication.processEvents(QEventLoop.ExcludeUserInputEvents)

    # ---- 내부 ---------------------------------------------------------
    def _tick(self) -> None:
        self._ring.advance(14)  # 회전
        self._frame += 1
        if self._frame % 9 == 0:  # 메시지 말줄임(…) 애니메이션은 느리게
            self._dot = (self._dot + 1) % 4
            self._msg.setText(self._base_msg + "." * self._dot)

    def _reposition(self) -> None:
        self.setGeometry(self._host.rect())

    def eventFilter(self, obj, event):  # noqa: N802
        if obj is self._host and event.type() == QEvent.Resize and self.isVisible():
            self._reposition()
        return False

    def paintEvent(self, event):  # noqa: N802
        painter = QPainter(self)
        painter.fillRect(self.rect(), QColor(8, 11, 16, 185))  # 반투명 어두운 막
        painter.end()


# =============================================================================
# app/ui/controls.py   [#19]
# =============================================================================
"""좌측 사이드바 컨트롤 및 하단 탐색 바 (문서 Section 8.1, 8.3, 8.5).

사이드바(세로): 자재 폴더 선택 / 자재명 / 기준 Layer / 허용 오차 /
비교 Layer(체크, 세로 스크롤) / 설정·업데이트·결과 출력하기
탐색 바: 이전 / 현재 index·전체 / 다음

비교 Layer 선택부는 세로 스크롤 영역에 한 줄에 하나씩 쌓는다.
layer 목록 설정 시 시그널을 차단해 재계산 폭주를 막는다.
"""






class NoScrollDoubleSpinBox(QDoubleSpinBox):
    """마우스 휠로 값이 바뀌지 않는 스핀박스.

    사이드바를 세로 스크롤하다 허용오차가 실수로 바뀌는 것을 막는다. 포커스가 있을 때
    키보드/직접 입력은 정상 동작한다.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setFocusPolicy(Qt.StrongFocus)  # 휠이 아니라 클릭/탭으로만 포커스

    def wheelEvent(self, event):  # noqa: N802
        event.ignore()  # 휠은 항상 무시(부모 스크롤로 전달)


class NoScrollComboBox(QComboBox):
    """마우스 휠로 항목이 바뀌지 않는 콤보박스(실수 변경 방지)."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setFocusPolicy(Qt.StrongFocus)

    def wheelEvent(self, event):  # noqa: N802
        event.ignore()


class SideBar(QFrame):
    """좌측 세로 컨트롤 사이드바.

    상단 컨트롤 바를 대체하지만 공개 API(시그널/메서드/속성)는 동일하게 유지한다.
    """

    open_folder = Signal()
    base_layer_changed = Signal(str)
    compare_layers_changed = Signal()
    tolerance_changed = Signal(float)
    export_requested = Signal()
    settings_requested = Signal()
    update_requested = Signal()

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setObjectName("sidebar")
        self.setMinimumWidth(200)
        self.setMaximumWidth(360)
        self._compare_checks: list[QCheckBox] = []
        self._rereview_set: set = set()  # '재리뷰' 버튼이 선택할 선호 재리뷰 집합
        self._build()

    @staticmethod
    def _section_label(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setObjectName("section")
        return lbl

    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 10)
        outer.setSpacing(8)

        # ── 헤더: 자재 폴더 선택 + 자재명
        self.btn_open = QPushButton("📁  자재 폴더 선택")
        self.btn_open.setToolTip("리뷰가 진행된 자재(LOT) 폴더를 선택 (Ctrl+O)")
        self.btn_open.clicked.connect(self.open_folder)
        self.lbl_lot = QLabel("선택된 자재 없음")
        self.lbl_lot.setObjectName("lotName")
        self.lbl_lot.setWordWrap(True)
        outer.addWidget(self.btn_open)
        outer.addWidget(self.lbl_lot)

        # ── 기준 Layer
        outer.addWidget(self._section_label("기준 LAYER"))
        self.cmb_base = NoScrollComboBox()
        self.cmb_base.setMinimumWidth(150)
        self.cmb_base.currentTextChanged.connect(self._on_base_changed)
        outer.addWidget(self.cmb_base)

        # ── 허용 오차
        outer.addWidget(self._section_label("허용 오차"))
        self.spn_tol = NoScrollDoubleSpinBox()
        self.spn_tol.setObjectName("tol")
        self.spn_tol.setButtonSymbols(QAbstractSpinBox.NoButtons)  # ↑↓ 버튼 제거(깔끔한 입력)
        self.spn_tol.setRange(0.0, 100000.0)
        self.spn_tol.setDecimals(1)
        self.spn_tol.setValue(DEFAULT_TOLERANCE)
        self.spn_tol.setSingleStep(10.0)
        self.spn_tol.setSuffix(" µm")
        self.spn_tol.setToolTip(
            "기준과 비교 defect 의 die 내 local 좌표 거리(µm) 허용값.\n"
            "작을수록 엄격, 클수록 느슨하게 매칭됩니다."
        )
        self.spn_tol.valueChanged.connect(self.tolerance_changed)
        outer.addWidget(self.spn_tol)

        # 실시간 매칭 요약(허용오차 튜닝 피드백)
        self.lbl_match = QLabel("")
        self.lbl_match.setObjectName("dim")
        self.lbl_match.setWordWrap(True)
        outer.addWidget(self.lbl_match)

        # ── 비교 Layer: 라벨 + (아래 줄) 재리뷰/전체/해제 — 좁은 폭에서 라벨이 잘리지 않도록 분리
        outer.addWidget(self._section_label("비교 LAYER"))
        cmp_btns = QHBoxLayout()
        cmp_btns.setSpacing(6)
        self.btn_rereview = QPushButton("재리뷰")
        self.btn_rereview.setObjectName("mini")
        self.btn_rereview.setToolTip(
            "재리뷰 layer 만 선택(같은 layer 에 재재리뷰가 있으면 재재리뷰 우선)"
        )
        self.btn_rereview.clicked.connect(self._set_rereview_compares)
        self.btn_all = QPushButton("전체")
        self.btn_all.setObjectName("mini")
        self.btn_all.setToolTip("선택 가능한 비교 layer 를 모두 선택")
        self.btn_all.clicked.connect(lambda: self._set_all_compares(True))
        self.btn_none = QPushButton("해제")
        self.btn_none.setObjectName("mini")
        self.btn_none.setToolTip("비교 layer 선택 모두 해제")
        self.btn_none.clicked.connect(lambda: self._set_all_compares(False))
        cmp_btns.addWidget(self.btn_rereview)
        cmp_btns.addWidget(self.btn_all)
        cmp_btns.addWidget(self.btn_none)
        outer.addLayout(cmp_btns)

        # 비교 Layer 체크박스 (세로 스크롤 영역에 한 줄씩)
        self._compare_scroll = QScrollArea()
        self._compare_scroll.setWidgetResizable(True)
        self._compare_scroll.setFrameShape(QFrame.NoFrame)
        self._compare_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        # viewport 기본 흰색 제거 → 사이드바 패널이 비쳐 글자가 보이게
        self._compare_scroll.setStyleSheet("background: transparent;")
        self._compare_scroll.viewport().setAutoFillBackground(False)
        self._compare_host = QWidget()
        self._compare_host.setAutoFillBackground(False)
        self._compare_host.setStyleSheet("background: transparent;")
        self._compare_box = QVBoxLayout(self._compare_host)
        self._compare_box.setContentsMargins(0, 0, 0, 0)
        self._compare_box.setSpacing(4)
        self._compare_box.addStretch()
        self._compare_scroll.setWidget(self._compare_host)
        outer.addWidget(self._compare_scroll, 1)

        # ── 푸터: 설정(작게) + 결과 출력  — 업데이트는 설정 안으로 이동
        self.btn_settings = QPushButton("⚙ 설정")
        self.btn_settings.setObjectName("mini")
        self.btn_settings.setToolTip("작업공간·출력 폴더·기본값·업데이트")
        self.btn_settings.clicked.connect(self.settings_requested)
        self.btn_settings.setMaximumHeight(30)

        self.btn_export = QPushButton("결과 출력")
        self.btn_export.setObjectName("primary")
        self.btn_export.setToolTip("선택한 기준 사진의 비교 결과를 Excel 로 출력 (Ctrl+E)")
        self.btn_export.clicked.connect(self.export_requested)
        self.btn_export.setEnabled(False)
        self.btn_export.setMaximumHeight(30)

        footer = QHBoxLayout()
        footer.setSpacing(6)
        footer.addWidget(self.btn_settings)
        footer.addWidget(self.btn_export, 1)
        outer.addLayout(footer)

        # 제작 크레딧(항상 보이는 사이드바 하단) — 두 줄로 표기.
        credit = QLabel(CREDITS.replace(", ", "\n"))
        credit.setObjectName("dim")
        credit.setWordWrap(True)
        credit.setStyleSheet(f"font-size:{fpx(11)}px;")  # 만든이 문구 +20%(9→11)
        credit.setAlignment(Qt.AlignHCenter)
        outer.addWidget(credit)

        # 업데이트 버튼은 사이드바에서 제거(설정 다이얼로그로 이동). 호환용 더미 참조.
        self.btn_update = None
        self._update_available = False

    # ---- API ----------------------------------------------------------
    def set_lot_name(self, name: str) -> None:
        self.lbl_lot.setText(f"자재: {name}")

    def set_layers(
        self,
        layers: list[str],
        base: Optional[str] = None,
        compares: Optional[list[str]] = None,
        rereview: Optional[set] = None,
    ) -> None:
        """layer 목록으로 기준 콤보 + 비교 체크박스를 채운다.

        기본값 설정 중에는 시그널을 차단하여 재계산이 0회가 되도록 한다(호출 측에서 1회만 재구성).
        base 가 None 이면 기준은 **빈칸**으로 두어 사용자가 직접 고르게 한다(자동 선택 안 함).
        compares 가 None 이면 비교 기본값은 rereview(선호 재리뷰 집합)만 체크한다.
        """
        self._rereview_set = set(rereview) if rereview else set()

        self.cmb_base.blockSignals(True)
        self.cmb_base.clear()
        self.cmb_base.addItems(layers)
        self.cmb_base.setPlaceholderText("기준 layer 선택")

        # 비교 체크박스 재구성 (세로 스택: 끝의 stretch 앞에 삽입)
        for cb in self._compare_checks:
            self._compare_box.removeWidget(cb)
            cb.setParent(None)
            cb.deleteLater()
        self._compare_checks.clear()
        for lyr in layers:
            cb = QCheckBox(lyr)
            cb.blockSignals(True)
            cb.stateChanged.connect(lambda _=0: self.compare_layers_changed.emit())
            self._compare_box.insertWidget(self._compare_box.count() - 1, cb)
            self._compare_checks.append(cb)

        # 기준 선택: base 가 주어지면 적용, 없으면 빈칸(-1)으로 두어 사용자 선택을 유도.
        chosen_base = base if (base and base in layers) else ""
        if chosen_base:
            self.cmb_base.setCurrentText(chosen_base)
        else:
            self.cmb_base.setCurrentIndex(-1)

        # 비교 선택 기본값:
        #  - 저장된 선택(compares)이 있으면 그것(+체크 유지용 기준)을 복원
        #  - 없으면 선호 재리뷰 집합을 체크. 재리뷰 layer 가 전혀 없는 자재는
        #    빈 선택(매칭 불가)이 되어 막다른 화면이 되므로 전체를 기본 체크한다(폴백).
        # 기준 layer 는 비교에서 자동 제외되지만(아래 compare_layers) 체크 상태는 유지한다.
        if compares is not None:
            compare_set = set(compares)
            if chosen_base:
                compare_set.add(chosen_base)
        elif self._rereview_set:
            compare_set = set(self._rereview_set)
        else:
            compare_set = set(layers)
        for cb in self._compare_checks:
            cb.setChecked(cb.text() in compare_set)

        self._sync_compare_enabled(chosen_base)

        # 시그널 복원
        self.cmb_base.blockSignals(False)
        for cb in self._compare_checks:
            cb.blockSignals(False)
        self.btn_export.setEnabled(bool(layers))
        self.btn_all.setEnabled(bool(layers))
        self.btn_none.setEnabled(bool(layers))
        self.btn_rereview.setEnabled(bool(self._rereview_set))

    def set_match_summary(self, text: str) -> None:
        self.lbl_match.setText(text)

    def set_tolerance(self, value: float) -> None:
        self.spn_tol.blockSignals(True)
        self.spn_tol.setValue(value)
        self.spn_tol.blockSignals(False)

    def set_update_available(self, available: bool) -> None:
        """업데이트 가용 시 설정 버튼에 표식(•)을 단다(업데이트는 설정 안에 있음)."""
        self._update_available = available
        self.btn_settings.setText("⚙ 설정 •" if available else "⚙ 설정")
        self.btn_settings.setToolTip(
            "업데이트 있음 — 설정에서 적용" if available
            else "작업공간·출력 폴더·기본값·업데이트"
        )

    def set_update_busy(self, busy: bool) -> None:
        self.btn_settings.setEnabled(not busy)
        self.btn_open.setEnabled(not busy)

    def _set_all_compares(self, checked: bool) -> None:
        """비교 layer 전체 선택/해제 — 한 번의 신호로 처리."""
        changed = False
        for cb in self._compare_checks:
            if cb.isEnabled() and cb.isChecked() != checked:
                cb.blockSignals(True)
                cb.setChecked(checked)
                cb.blockSignals(False)
                changed = True
        if changed:
            self.compare_layers_changed.emit()

    def _set_rereview_compares(self) -> None:
        """선호 재리뷰 집합만 체크(같은 layer 재재리뷰 우선). 그 외는 해제 — 한 번의 신호."""
        if not self._rereview_set:
            return
        changed = False
        for cb in self._compare_checks:
            want = cb.text() in self._rereview_set
            if cb.isEnabled() and cb.isChecked() != want:
                cb.blockSignals(True)
                cb.setChecked(want)
                cb.blockSignals(False)
                changed = True
        if changed:
            self.compare_layers_changed.emit()

    def _on_base_changed(self, base: str) -> None:
        self._sync_compare_enabled(base)
        self.base_layer_changed.emit(base)

    def _sync_compare_enabled(self, base: str) -> None:
        """기준 layer 의 체크박스는 비활성(토글 불가)하되 체크 상태는 보존한다.

        실제 비교에서는 compare_layers() 가 기준 layer 를 자동 제외한다. 기준을 바꾸면
        이전 기준 layer 는 다시 활성화되고, 보존된 체크 상태로 비교에 복귀한다.
        """
        for cb in self._compare_checks:
            cb.setEnabled(cb.text() != base)

    def base_layer(self) -> str:
        return self.cmb_base.currentText()

    def compare_layers(self) -> list[str]:
        """체크된 layer 중 기준 layer 를 제외한 목록(기준은 비교 대상에서 자동 제외)."""
        base = self.base_layer()
        return [
            cb.text() for cb in self._compare_checks
            if cb.isChecked() and cb.text() != base
        ]

    def tolerance(self) -> float:
        return self.spn_tol.value()


class NavBar(QFrame):
    """하단 탐색 바: 이전 / index·전체 / 다음 (문서 Section 8.5)."""

    prev_clicked = Signal()
    next_clicked = Signal()

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setObjectName("panel")
        lay = QHBoxLayout(self)
        lay.setContentsMargins(12, 8, 12, 8)

        self.btn_prev = QPushButton("◀  이전")
        self.btn_prev.setToolTip("이전 기준 사진 (← / PageUp)")
        self.btn_prev.clicked.connect(self.prev_clicked)
        self.btn_next = QPushButton("다음  ▶")
        self.btn_next.setToolTip("다음 기준 사진 (→ / PageDown)")
        self.btn_next.clicked.connect(self.next_clicked)

        self.lbl_index = QLabel("0 / 0")
        self.lbl_index.setAlignment(Qt.AlignCenter)
        self.lbl_index.setStyleSheet(f"font-size:{fpx(13)}px; font-weight:600;")
        self.lbl_index.setMinimumWidth(90)

        self.lbl_status = QLabel("")
        self.lbl_status.setObjectName("dim")

        lay.addWidget(self.btn_prev)
        lay.addStretch()
        lay.addWidget(self.lbl_index)
        lay.addStretch()
        lay.addWidget(self.lbl_status)
        lay.addStretch()
        lay.addWidget(self.btn_next)
        self._lay = lay
        self.set_enabled(False)

    def add_widget(self, widget: QWidget) -> None:
        """탐색 바 오른쪽(다음 버튼 앞)에 보조 위젯을 추가한다."""
        self._lay.insertWidget(self._lay.count() - 1, widget)

    def set_index(self, current: int, total: int) -> None:
        self.lbl_index.setText(f"{current} / {total}")

    def set_status(self, text: str) -> None:
        self.lbl_status.setText(text)

    def set_status_tooltip(self, text: str) -> None:
        self.lbl_status.setToolTip(text)

    def set_enabled(self, enabled: bool) -> None:
        self.btn_prev.setEnabled(enabled)
        self.btn_next.setEnabled(enabled)


# 하위 호환: 옛 이름으로 import 하던 코드 지원
TopBar = SideBar


# =============================================================================
# app/ui/help_dialog.py   [#20]
# =============================================================================
"""도움말 다이얼로그 — 단축키 + 기능 안내(섹션 구성, 스크롤)."""





# 단축키 그룹: (그룹명, [(키, 설명), ...])
_SHORTCUT_GROUPS = [
    ("탐색", [
        ("← / → · PageUp / PageDown", "이전 / 다음 기준 사진"),
        ("Home / End", "처음 / 끝 기준 사진으로"),
        ("U", "다음 '미매칭 포함' 기준으로 점프"),
        ("F5", "현재 자재 폴더 다시 스캔"),
    ]),
    ("선택", [
        ("Ctrl + A / Ctrl + D", "비교 Layer 전체 선택 / 모두 해제"),
    ]),
    ("출력", [
        ("A", "현재 기준 사진을 출력 트레이에 담기"),
        ("Ctrl + E", "Excel 결과 출력"),
    ]),
    ("파일 · 도움말", [
        ("Ctrl + O", "자재 폴더 열기 (우클릭: 최근 폴더)"),
        ("F1", "이 도움말 열기"),
        ("이미지 클릭", "원본 전체 해상도 확대 보기"),
        ("이미지 우클릭", "경로 복사 / 파일·폴더 열기"),
    ]),
]

# 기능 안내: (기능명, 설명)
_FEATURES = [
    ("자재 폴더 선택기",
     "브레드크럼·폴더 트리·최근/즐겨찾기로 탐색하고, 고른 폴더가 자재(LOT)인지 실시간으로 "
     "확인합니다. layer·wafer 폴더를 골라도 자재 폴더로 자동 보정됩니다."),
    ("Defect 히트맵",
     "웨이퍼맵에 defect 밀도를 색으로 표시합니다. 위치를 클릭하면 그 자리의 defect 이 "
     "오른쪽에 나열됩니다. 상단 캡션에서 현재 모드·wafer 를 확인할 수 있습니다."),
    ("매치만 / 전체 defect 모드",
     "기본은 '매치만'(매칭된 기준 defect). '전체 defect'를 켜면 특정 기준 없이 선택 layer 를 "
     "서로 교차 매칭해 모든 defect 을 보여주고, 매칭 안 된 것은 개별로 표시합니다. "
     "웨이퍼맵 색도 모드에 따라 바뀝니다."),
    ("여러 다이 선택 · 드래그 박스",
     "웨이퍼맵에서 드래그하면 사각형으로 여러 die 를 한 번에 선택합니다(항상 가능). "
     "'여러 다이 선택'을 켜면 클릭이 die 를 누적 토글합니다."),
    ("defect 클러스터링",
     "같은 layer 에서 거리 50 미만으로 붙은 defect 은 하나로 묶어 대표 1장만 보이고, "
     "좌하단 '+n' 을 누르면 묶인 나머지도 모두 볼 수 있습니다."),
    ("출력 트레이",
     "'담기'로 원하는 defect 을 모아 두었다가 'Excel 출력'으로 한 번에 리포트를 만듭니다. "
     "layer·자재를 바꿔도 담은 사진은 유지됩니다."),
]

_KEY_COL_WIDTH = 190


class ShortcutsDialog(QDialog):
    """단축키 + 기능 안내."""

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("도움말")
        self.setMinimumSize(560, 560)
        self._build()

    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        scroll.viewport().setAutoFillBackground(False)
        body = QWidget()
        body.setStyleSheet("background: transparent;")
        lay = QVBoxLayout(body)
        lay.setContentsMargins(20, 18, 20, 14)
        lay.setSpacing(14)

        lay.addWidget(self._title("단축키"))
        for name, rows in _SHORTCUT_GROUPS:
            lay.addWidget(self._group_header(name))
            lay.addWidget(self._shortcut_grid(rows))

        lay.addSpacing(4)
        lay.addWidget(self._title("기능 안내"))
        for name, desc in _FEATURES:
            lay.addWidget(self._feature_card(name, desc))

        lay.addStretch()
        scroll.setWidget(body)
        outer.addWidget(scroll, 1)

        footer = QVBoxLayout()
        footer.setContentsMargins(20, 8, 20, 12)
        btn = QPushButton("닫기")
        btn.setObjectName("primary")
        btn.setDefault(True)
        btn.clicked.connect(self.accept)
        footer.addWidget(btn, alignment=Qt.AlignRight)
        outer.addLayout(footer)

    @staticmethod
    def _title(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setObjectName("title")
        return lbl

    @staticmethod
    def _group_header(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            f"color:{NEON}; font-weight:700; font-size:12px;"
            " letter-spacing:1px; margin-top:2px;"
        )
        return lbl

    @staticmethod
    def _shortcut_grid(rows: list[tuple[str, str]]) -> QWidget:
        host = QWidget()
        grid = QGridLayout(host)
        grid.setContentsMargins(6, 0, 0, 0)
        grid.setHorizontalSpacing(16)
        grid.setVerticalSpacing(6)
        grid.setColumnMinimumWidth(0, _KEY_COL_WIDTH)
        grid.setColumnStretch(1, 1)
        for r, (key, desc) in enumerate(rows):
            k = QLabel(key)
            k.setStyleSheet(
                f"font-weight:700; color:{TEXT};"
                f" background:{BG_ELEV}; border:1px solid {NEON_SOFT};"
                " border-radius:5px; padding:2px 8px;"
            )
            k.setTextInteractionFlags(Qt.TextSelectableByMouse)
            d = QLabel(desc)
            d.setObjectName("dim")
            d.setWordWrap(True)
            grid.addWidget(k, r, 0, Qt.AlignTop | Qt.AlignLeft)
            grid.addWidget(d, r, 1, Qt.AlignVCenter)
        return host

    @staticmethod
    def _feature_card(name: str, desc: str) -> QWidget:
        card = QFrame()
        card.setObjectName("cell")
        card.setStyleSheet(
            f"QFrame#cell {{ background:{BG_ELEV};"
            f" border:1px solid {NEON_SOFT}; border-radius:8px; }}"
        )
        lay = QVBoxLayout(card)
        lay.setContentsMargins(12, 8, 12, 10)
        lay.setSpacing(3)
        title = QLabel(name)
        title.setStyleSheet(f"font-weight:700; color:{TEXT};")
        desc_lbl = QLabel(desc)
        desc_lbl.setObjectName("dim")
        desc_lbl.setWordWrap(True)
        lay.addWidget(title)
        lay.addWidget(desc_lbl)
        return card


# =============================================================================
# app/ui/image_viewer.py   [#21]
# =============================================================================
"""원본 확대 뷰어 (read-only) — defect review 사용성.

그리드의 이미지를 클릭하면 원본 전체 해상도를 별도 창에서 확대해 본다.
맞춤/실제(1:1) 토글, 마우스 휠 줌, 메타데이터 표시, Esc 닫기를 지원한다.
원본은 QImageReader 로 읽기 전용 접근만 하며 수정하지 않는다.
"""





_MIN_SCALE = 0.1
_MAX_SCALE = 8.0


class ImageViewerDialog(QDialog):
    """원본 전체 해상도 확대 뷰어."""

    def __init__(self, record: DefectRecord, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.record = record
        self.setWindowTitle(f"원본 보기 — {Path(record.image_path).name}")
        self.setMinimumSize(640, 520)
        self.resize(960, 760)
        self._scale = 1.0
        self._fit = True
        # 클릭-드래그 패닝 상태(항목 6)
        self._panning = False
        self._pan_start = None  # QPoint (뷰포트 좌표)
        self._pan_h0 = 0
        self._pan_v0 = 0

        self._image = self._load(record.image_path)
        self._build()
        self._apply_scale(fit=True)

        QShortcut(QKeySequence(Qt.Key_Escape), self, activated=self.accept)
        QShortcut(QKeySequence.ZoomIn, self, activated=lambda: self._zoom(1.25))
        QShortcut(QKeySequence.ZoomOut, self, activated=lambda: self._zoom(0.8))

    @staticmethod
    def _load(path) -> QImage:
        reader = QImageReader(str(path))
        reader.setAutoTransform(True)  # EXIF 회전 반영
        img = reader.read()
        return img  # null 이면 placeholder 처리

    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(8)

        rec = self.record
        meta = QLabel(
            f"<b>{rec.layer}</b> · wafer {rec.wafer_id} · die({rec.col},{rec.row}) · "
            f"pos {rec.position_key} · <span style='color:{TEXT_DIM}'>"
            f"{rec.image_path}</span>"
        )
        meta.setWordWrap(True)
        meta.setObjectName("title")
        # 정보 텍스트를 드래그 선택·복사 가능하게(Ctrl+C). 마우스 커서도 텍스트형으로.
        meta.setTextInteractionFlags(
            Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard
        )
        meta.setCursor(Qt.IBeamCursor)
        meta_row = QHBoxLayout()
        meta_row.addWidget(meta, 1)
        btn_copy = QPushButton("정보 복사")
        btn_copy.setObjectName("mini")
        btn_copy.setToolTip("이 사진의 layer·wafer·die·좌표·경로를 클립보드로 복사")
        btn_copy.clicked.connect(self._copy_info)
        meta_row.addWidget(btn_copy, 0, Qt.AlignTop)
        outer.addLayout(meta_row)

        bar = QHBoxLayout()
        self.btn_fit = QPushButton("실제 크기 (1:1)")
        self.btn_fit.setCheckable(False)
        self.btn_fit.clicked.connect(self._toggle_fit)
        btn_zoom_out = QPushButton("－")
        btn_zoom_out.setFixedWidth(40)
        btn_zoom_out.clicked.connect(lambda: self._zoom(0.8))
        btn_zoom_in = QPushButton("＋")
        btn_zoom_in.setFixedWidth(40)
        btn_zoom_in.clicked.connect(lambda: self._zoom(1.25))
        self.lbl_zoom = QLabel("100%")
        self.lbl_zoom.setObjectName("dim")
        self.lbl_zoom.setFixedWidth(56)
        self.lbl_zoom.setAlignment(Qt.AlignCenter)
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)

        bar.addWidget(self.btn_fit)
        bar.addWidget(btn_zoom_out)
        bar.addWidget(self.lbl_zoom)
        bar.addWidget(btn_zoom_in)
        bar.addStretch()
        bar.addWidget(btn_close)
        outer.addLayout(bar)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(False)
        self._scroll.setAlignment(Qt.AlignCenter)
        # 확대 시 스크롤바는 못 쓰므로(드래그로 이동) 숨긴다. value 이동은 계속 동작.
        self._scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._canvas = QLabel()
        self._canvas.setAlignment(Qt.AlignCenter)
        if self._image.isNull():
            self._canvas.setText("이미지를 불러올 수 없습니다.")
        else:
            # 클릭-드래그로 화면 이동: 캔버스에서 마우스 이벤트를 받는다.
            self._canvas.setCursor(Qt.OpenHandCursor)
            self._canvas.installEventFilter(self)
            # 휠은 스크롤이 아니라 항상 줌으로만 동작하도록 뷰포트/캔버스에서 가로챈다.
            self._scroll.viewport().installEventFilter(self)
        self._scroll.setWidget(self._canvas)
        outer.addWidget(self._scroll, 1)

    def _info_text(self) -> str:
        """클립보드 복사용 정돈된 정보 텍스트."""
        r = self.record
        parts = [
            f"layer: {r.layer}",
            f"wafer: {r.wafer_id}",
            f"die: ({r.col},{r.row})",
            f"pos: {r.position_key}",
        ]
        if r.defect_name:
            parts.append(f"defect: {r.defect_name}")
        if r.dx_size is not None or r.dy_size is not None or r.d_area is not None:
            parts.append(f"size: dx={r.dx_size}, dy={r.dy_size}, area={r.d_area}")
        parts.append(f"path: {r.image_path}")
        return "\n".join(parts)

    def _copy_info(self) -> None:
        QApplication.clipboard().setText(self._info_text())

    # ---- 줌/맞춤 ---------------------------------------------------------
    def _toggle_fit(self) -> None:
        if self._fit:
            self._apply_scale(scale=1.0)  # 실제 크기
        else:
            self._apply_scale(fit=True)

    def _zoom(self, factor: float) -> None:
        if self._image.isNull():
            return
        self._apply_scale(scale=max(_MIN_SCALE, min(_MAX_SCALE, self._scale * factor)))

    def _zoom_at_cursor(self, factor: float) -> None:
        """마우스 커서 아래 지점을 고정한 채 확대/축소(휠 줌)."""
        if self._image.isNull():
            return
        from PySide6.QtGui import QCursor
        vp = self._scroll.viewport()
        pos = vp.mapFromGlobal(QCursor.pos())
        hbar = self._scroll.horizontalScrollBar()
        vbar = self._scroll.verticalScrollBar()
        old = self._scale
        new = max(_MIN_SCALE, min(_MAX_SCALE, old * factor))
        if new == old:
            return
        # 스케일 전, 커서 아래 콘텐츠 좌표
        cx = hbar.value() + pos.x()
        cy = vbar.value() + pos.y()
        self._apply_scale(scale=new)
        ratio = new / old
        # 스케일 후 그 콘텐츠 좌표가 다시 커서 아래 오도록 스크롤 이동
        hbar.setValue(round(cx * ratio - pos.x()))
        vbar.setValue(round(cy * ratio - pos.y()))

    def _apply_scale(self, scale: Optional[float] = None, fit: bool = False) -> None:
        if self._image.isNull():
            return
        if fit:
            self._fit = True
            area = self._scroll.viewport().size()
            iw, ih = self._image.width(), self._image.height()
            if iw > 0 and ih > 0:
                self._scale = min(area.width() / iw, area.height() / ih, 1.0)
                self._scale = max(self._scale, _MIN_SCALE)
        else:
            self._fit = False
            self._scale = scale if scale is not None else self._scale

        w = max(1, int(self._image.width() * self._scale))
        h = max(1, int(self._image.height() * self._scale))
        pix = QPixmap.fromImage(self._image).scaled(
            w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation
        )
        self._canvas.setPixmap(pix)
        self._canvas.resize(pix.size())
        self.lbl_zoom.setText(f"{int(self._scale * 100)}%")
        self.btn_fit.setText("화면 맞춤" if not self._fit else "실제 크기 (1:1)")

    def wheelEvent(self, event):  # noqa: N802
        if self._image.isNull():
            return
        delta = event.angleDelta().y()
        if delta != 0:
            self._zoom_at_cursor(1.2 if delta > 0 else 1 / 1.2)
            event.accept()

    # ---- 클릭-드래그 패닝(항목 6) --------------------------------------
    def eventFilter(self, obj, event):  # noqa: N802
        """휠=줌 전용, 캔버스 좌클릭 드래그=화면 이동."""
        from PySide6.QtCore import QEvent

        # 휠은 스크롤 영역이 먹어 스크롤되지 않도록 가로채 항상 줌으로만 처리한다.
        if event.type() == QEvent.Wheel and not self._image.isNull():
            delta = event.angleDelta().y()
            if delta != 0:
                self._zoom_at_cursor(1.2 if delta > 0 else 1 / 1.2)
            return True

        if obj is self._canvas and not self._image.isNull():
            et = event.type()
            if et == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                self._panning = True
                self._pan_start = event.globalPosition().toPoint()
                self._pan_h0 = self._scroll.horizontalScrollBar().value()
                self._pan_v0 = self._scroll.verticalScrollBar().value()
                self._canvas.setCursor(Qt.ClosedHandCursor)
                return True
            if et == QEvent.MouseMove and self._panning and self._pan_start is not None:
                delta = event.globalPosition().toPoint() - self._pan_start
                self._scroll.horizontalScrollBar().setValue(self._pan_h0 - delta.x())
                self._scroll.verticalScrollBar().setValue(self._pan_v0 - delta.y())
                return True
            if et == QEvent.MouseButtonRelease and event.button() == Qt.LeftButton:
                self._panning = False
                self._pan_start = None
                self._canvas.setCursor(Qt.OpenHandCursor)
                return True
        return super().eventFilter(obj, event)

    def resizeEvent(self, event):  # noqa: N802
        super().resizeEvent(event)
        if self._fit:
            self._apply_scale(fit=True)


# =============================================================================
# app/ui/notifications.py   [#22]
# =============================================================================
"""비차단 알림 배너(토스트) — 매끄러운 오류/안내 처리 (문서 Section 9 / 사용성).

화면을 막는 모달 팝업 대신, 창 상단에 부드럽게 나타났다 사라지는 인라인 배너를 사용한다.
info/success/warn/error 레벨별 색상, 선택적 액션 버튼, 자동 소멸을 지원한다.
배너는 비차단이므로 작업 흐름이 끊기지 않는다.
"""





_LEVEL_COLORS = {
    "info": (NEON_DIM, "#ffffff"),
    "success": ("#15803d", "#ffffff"),
    "warn": ("#b45309", "#ffffff"),
    "error": ("#b00020", "#ffffff"),
}

_ICONS = {"info": "ℹ", "success": "✓", "warn": "⚠", "error": "✕"}


class NotificationBanner(QFrame):
    """창 상단의 비차단 알림 배너."""

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setObjectName("banner")
        self._action_cb: Optional[Callable[[], None]] = None
        self._build()

        self._effect = QGraphicsOpacityEffect(self)
        self._effect.setOpacity(0.0)
        self.setGraphicsEffect(self._effect)

        self._fade = QPropertyAnimation(self._effect, b"opacity", self)
        self._fade.setDuration(180)
        self._fade.setEasingCurve(QEasingCurve.InOutCubic)
        self._collapse = QPropertyAnimation(self, b"maximumHeight", self)
        self._collapse.setDuration(200)
        self._collapse.setEasingCurve(QEasingCurve.InOutCubic)
        self._group = QParallelAnimationGroup(self)
        self._group.addAnimation(self._fade)
        self._group.addAnimation(self._collapse)

        self._auto_hide = QTimer(self)
        self._auto_hide.setSingleShot(True)
        self._auto_hide.timeout.connect(self.dismiss)

        self.setMaximumHeight(0)
        self.setVisible(False)

    def _build(self) -> None:
        lay = QHBoxLayout(self)
        lay.setContentsMargins(14, 8, 10, 8)
        lay.setSpacing(10)
        self._icon = QLabel("")
        self._icon.setStyleSheet("font-weight:700; color:#ffffff; background:transparent;")
        self._label = QLabel("")
        self._label.setWordWrap(True)
        self._label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self._label.setStyleSheet("color:#ffffff; background:transparent;")
        self._action = QPushButton("")
        self._action.setCursor(Qt.PointingHandCursor)
        self._action.setVisible(False)
        self._action.clicked.connect(self._on_action)
        self._action.setStyleSheet(
            "QPushButton { color:#ffffff; background:rgba(255,255,255,0.18);"
            " border:1px solid rgba(255,255,255,0.45); border-radius:6px;"
            " padding:3px 10px; }"
            "QPushButton:hover { background:rgba(255,255,255,0.32); }"
        )
        self._close = QPushButton("✕")
        self._close.setCursor(Qt.PointingHandCursor)
        self._close.setFixedSize(22, 22)
        self._close.clicked.connect(self.dismiss)
        self._close.setStyleSheet(
            "QPushButton { color:#ffffff; background:transparent; border:none;"
            " font-size:13px; }"
            "QPushButton:hover { background:rgba(255,255,255,0.22); border-radius:11px; }"
        )

        lay.addWidget(self._icon)
        lay.addWidget(self._label, 1)
        lay.addWidget(self._action)
        lay.addWidget(self._close)

    def _on_action(self) -> None:
        cb = self._action_cb
        self.dismiss()
        if cb is not None:
            cb()

    def show_message(
        self,
        text: str,
        level: str = "info",
        *,
        action_text: Optional[str] = None,
        action: Optional[Callable[[], None]] = None,
        timeout_ms: int = 4500,
    ) -> None:
        """배너를 표시한다. timeout_ms<=0 이면 자동 소멸하지 않는다."""
        bg, _fg = _LEVEL_COLORS.get(level, _LEVEL_COLORS["info"])
        # 프레임 배경만 레벨별로 바꾸고, 자식(라벨/버튼) 스타일은 _build 에서 고정.
        self.setStyleSheet(
            f"QFrame#banner {{ background:{bg}; border:none; border-radius:8px; }}"
        )
        self._icon.setText(_ICONS.get(level, "ℹ"))
        self._label.setText(text)
        self._action_cb = action
        if action_text and action is not None:
            self._action.setText(action_text)
            self._action.setVisible(True)
        else:
            self._action.setVisible(False)

        self.setVisible(True)
        target = max(self.sizeHint().height(), 40)
        self._group.stop()
        self._fade.setStartValue(self._effect.opacity())
        self._fade.setEndValue(1.0)
        self._collapse.setStartValue(self.maximumHeight())
        self._collapse.setEndValue(target)
        self._group.start()
        # 오버레이로 부모 위에 떠서 표시 — 레이아웃을 밀지 않아 화면이 흔들리지 않는다.
        self.reposition()
        self.raise_()

        self._auto_hide.stop()
        if timeout_ms > 0:
            self._auto_hide.start(timeout_ms)

    def reposition(self) -> None:
        """부모 위 상단 중앙에 배너를 배치한다(오버레이). 부모 크기 변화 시 호출."""
        parent = self.parentWidget()
        if parent is None:
            return
        w = min(760, max(320, parent.width() - 40))
        self.setFixedWidth(w)
        self.move((parent.width() - w) // 2, 12)

    def dismiss(self) -> None:
        self._auto_hide.stop()
        if not self.isVisible():
            return
        self._group.stop()
        self._fade.setStartValue(self._effect.opacity())
        self._fade.setEndValue(0.0)
        self._collapse.setStartValue(self.maximumHeight())
        self._collapse.setEndValue(0)
        try:
            self._group.finished.disconnect()
        except (RuntimeError, TypeError):
            pass
        self._group.finished.connect(self._after_hide)
        self._group.start()

    def _after_hide(self) -> None:
        try:
            self._group.finished.disconnect(self._after_hide)
        except (RuntimeError, TypeError):
            pass
        if self._effect.opacity() <= 0.01:
            self.setVisible(False)


# =============================================================================
# app/ui/settings_dialog.py   [#23]
# =============================================================================
"""설정 다이얼로그 — 작업공간/출력 폴더/기본 허용오차/업데이트 확인 (사용성).

OK 시 AppSettings 를 갱신·저장한다. 작업공간이 현재 LOT 내부면 경고하고 차단한다.
"""






class SettingsDialog(QDialog):
    """설정 편집 다이얼로그."""

    update_requested = Signal()  # "지금 업데이트 확인" 클릭 시

    def __init__(
        self,
        settings: AppSettings,
        current_lot: Optional[str] = None,
        parent: Optional[QWidget] = None,
        update_available: bool = False,
    ):
        super().__init__(parent)
        self.setWindowTitle("설정")
        self.setMinimumWidth(560)
        self._settings = settings
        self._current_lot = current_lot
        self._update_available = update_available
        self._wants_update = False
        self._build()

    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(18, 18, 18, 14)
        outer.setSpacing(12)

        title = QLabel("설정")
        title.setObjectName("title")
        outer.addWidget(title)

        form = QFormLayout()
        form.setSpacing(10)

        self.ed_workspace = QLineEdit(self._settings.workspace)
        form.addRow("작업공간 폴더", self._with_browse(self.ed_workspace, self._pick_workspace))

        self.ed_output = QLineEdit(self._settings.output_folder)
        self.ed_output.setPlaceholderText("(비우면 작업공간/exports 사용)")
        form.addRow("출력 폴더", self._with_browse(self.ed_output, self._pick_output))

        # 디바이스 DB 파일(AOIDeviceDB.xlsx) — 있으면 제품 목록을 확장한다.
        self.ed_device_db = QLineEdit(self._settings.device_db_path)
        self.ed_device_db.setPlaceholderText("(선택) AOIDeviceDB.xlsx 경로")
        form.addRow(
            "디바이스 DB", self._with_browse(self.ed_device_db, self._pick_device_db)
        )

        # 제품 프로파일(좌표 변환 상수). 변경은 다음 스캔부터 적용.
        self.cmb_product = NoScrollComboBox()
        self._reload_products(select=self._settings.product)
        self.cmb_product.setToolTip("제품별 좌표 변환 상수 — 변경 후 다시 스캔(F5)하세요")
        form.addRow("제품 프로파일", self.cmb_product)

        # 경로를 '직접 입력/붙여넣기' 해도(찾아보기 없이) 제품 목록이 채워지도록,
        # 편집이 끝나면 그 경로의 DB 를 읽어 콤보를 갱신한다.
        self.ed_device_db.editingFinished.connect(self._on_device_db_edited)

        # 시작 시 DB 경로가 있으면 미리 로드해 제품 목록을 채운다.
        if self._settings.device_db_path:
            self._load_db(self._settings.device_db_path, select=self._settings.product)

        # defect 근접 클러스터링 거리 — 같은 die 안에서 이 값 미만이면 하나로 묶는다.
        self.spn_cluster = QDoubleSpinBox()
        self.spn_cluster.setRange(0.0, 100000.0)
        self.spn_cluster.setDecimals(1)
        self.spn_cluster.setSingleStep(5.0)
        self.spn_cluster.setValue(
            getattr(self._settings, "cluster_radius", DEFAULT_CLUSTER_RADIUS)
        )
        self.spn_cluster.setToolTip(
            "같은 die 안에서 이 거리(좌표 단위) 미만인 defect 을 하나로 묶어 대표 1장+‘+n’ 으로 봅니다."
        )
        form.addRow("defect 클러스터 거리", self.spn_cluster)

        # 전체 UI 글자 크기(보통/크게).
        self.cmb_font = NoScrollComboBox()
        self.cmb_font.addItem("보통", "normal")
        self.cmb_font.addItem("크게", "large")
        fi = self.cmb_font.findData(getattr(self._settings, "ui_font_size", "normal"))
        self.cmb_font.setCurrentIndex(fi if fi >= 0 else 0)
        self.cmb_font.setToolTip(
            "전체 UI 글자 크기입니다. 변경하면 대부분 즉시 적용되고, 다시 시작하면 완전히 적용됩니다."
        )
        form.addRow("글자 크기", self.cmb_font)

        self.chk_update = QCheckBox("시작할 때 업데이트 확인")
        self.chk_update.setChecked(self._settings.auto_update_check)
        form.addRow("자동 업데이트", self.chk_update)

        # 수동 업데이트(사이드바에서 이동): 확인/적용 버튼
        upd_host = QWidget()
        upd_lay = QHBoxLayout(upd_host)
        upd_lay.setContentsMargins(0, 0, 0, 0)
        upd_lay.setSpacing(8)
        self.btn_update = QPushButton(
            "지금 업데이트" if self._update_available else "업데이트 확인"
        )
        if self._update_available:
            self.btn_update.setObjectName("primary")
        self.btn_update.setToolTip("최신 버전(메인 브랜치)으로 업데이트")
        self.btn_update.clicked.connect(self._on_update_clicked)
        self.lbl_update = QLabel(
            "새 버전이 있습니다." if self._update_available else ""
        )
        self.lbl_update.setObjectName("dim")
        upd_lay.addWidget(self.btn_update)
        upd_lay.addWidget(self.lbl_update, 1)
        form.addRow("업데이트", upd_host)

        # 개발자 모드 토글 — 작은 켜짐/꺼짐 버튼. 켜면 아래 dev 섹션(로그 경로·로그 폴더)
        # 이 나타나고, 저장 시 settings.dev_mode 에 기록된다. 환경변수 DEFECT_TRACKER_DEV
        # 로 강제 켜진 경우엔 항상 켜짐으로 두고 토글을 비활성화한다.
        self._dev_env_forced = dev_mode()  # settings=None → 환경변수만 반영
        dev_on = self._dev_env_forced or bool(getattr(self._settings, "dev_mode", False))
        self.btn_dev = QPushButton("켜짐" if dev_on else "꺼짐")
        self.btn_dev.setObjectName("mini")
        self.btn_dev.setCheckable(True)
        self.btn_dev.setChecked(dev_on)
        self.btn_dev.setToolTip("파일 로그·진단 리포트·로그 경로 설정을 켭니다.")
        if self._dev_env_forced:
            self.btn_dev.setEnabled(False)
            self.btn_dev.setToolTip("환경변수 DEFECT_TRACKER_DEV 로 강제로 켜져 있습니다.")
        self.btn_dev.toggled.connect(self._on_dev_toggled)
        dev_host = QWidget()
        dev_hl = QHBoxLayout(dev_host)
        dev_hl.setContentsMargins(0, 0, 0, 0)
        dev_hl.addWidget(self.btn_dev)
        dev_hl.addStretch(1)
        form.addRow("개발자 모드", dev_host)

        # 단축키·도움말 보기(상단 밴드에서 이동).
        self.btn_help = QPushButton("단축키 · 도움말 보기")
        self.btn_help.clicked.connect(self._open_help)
        form.addRow("도움말", self.btn_help)

        outer.addLayout(form)

        # 개발자 섹션(로그 저장 경로 · 로그 폴더 열기) — 토글로 표시/숨김.
        self._dev_box = QWidget()
        dev_form = QFormLayout(self._dev_box)
        dev_form.setContentsMargins(0, 0, 0, 0)
        dev_form.setSpacing(10)
        self.ed_log_dir = QLineEdit(self._settings.log_dir)
        self.ed_log_dir.setPlaceholderText("(비우면 작업공간/logs 사용)")
        dev_form.addRow(
            "로그 저장 경로", self._with_browse(self.ed_log_dir, self._pick_log_dir)
        )
        self.btn_logs = QPushButton("로그 폴더 열기")
        self.btn_logs.setToolTip("좌표 추출 진단(parse_failures.md)과 실행 로그가 있는 폴더")
        self.btn_logs.clicked.connect(self._open_logs)
        dev_form.addRow("진단/로그", self.btn_logs)
        self._dev_box.setVisible(dev_on)
        outer.addWidget(self._dev_box)

        self.lbl_err = QLabel("")
        self.lbl_err.setStyleSheet("color:#f87171;")
        self.lbl_err.setWordWrap(True)
        self.lbl_err.setVisible(False)
        outer.addWidget(self.lbl_err)

        footer = QLabel(f"{APP_NAME}  ·  버전 {__version__}")
        footer.setObjectName("dim")
        outer.addWidget(footer)

        credit = QLabel(CREDITS)
        credit.setObjectName("dim")
        credit.setStyleSheet(f"font-size:{fpx(12)}px;")  # 만든이 문구 +20%
        outer.addWidget(credit)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("저장")
        buttons.button(QDialogButtonBox.Cancel).setText("취소")
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        outer.addWidget(buttons)

    def _on_dev_toggled(self, on: bool) -> None:
        """개발자 모드 토글 — 버튼 라벨과 dev 섹션 표시를 갱신한다."""
        self.btn_dev.setText("켜짐" if on else "꺼짐")
        self._dev_box.setVisible(on)
        self.adjustSize()

    def _open_help(self) -> None:
        ShortcutsDialog(self).exec()

    def _open_logs(self) -> None:
        """진단/로그 폴더(로그 저장 경로, 비어 있으면 workspace/logs)를 파일 탐색기로 연다."""
        from PySide6.QtCore import QUrl
        from PySide6.QtGui import QDesktopServices

        log_dir = self.ed_log_dir.text().strip()
        if log_dir:
            logs = Path(log_dir)
        else:
            base = self.ed_workspace.text().strip() or self._settings.workspace
            logs = Path(base) / "logs"
        try:
            logs.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(logs)))

    def _with_browse(self, line: QLineEdit, handler) -> QWidget:
        host = QWidget()
        lay = QHBoxLayout(host)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(6)
        btn = QPushButton("찾아보기")
        btn.clicked.connect(handler)
        lay.addWidget(line, 1)
        lay.addWidget(btn)
        return host

    def _reload_products(self, select: str | None = None) -> None:
        """config.PRODUCTS 로 제품 콤보를 다시 채운다.

        익명 기본 프로파일(DEFAULT_PRODUCT, 표시명 'DEVA Live')은 실제 디바이스가 아니라
        내부 폴백이므로 목록에 노출하지 않고, 대신 '(자동 인식)' 항목으로 대체한다. 이
        항목을 고르면 기본 프로파일이 유지되며 저장 시 LOT 경로로 디바이스를 자동 인식한다.
        """
        self.cmb_product.blockSignals(True)
        self.cmb_product.clear()
        self.cmb_product.addItem("(자동 인식)", DEFAULT_PRODUCT)
        for key, prod in PRODUCTS.items():
            if key == DEFAULT_PRODUCT:
                continue
            self.cmb_product.addItem(f"{prod.name} ({key})", key)
        if select:
            idx = self.cmb_product.findData(select)
            if idx >= 0:
                self.cmb_product.setCurrentIndex(idx)
        self.cmb_product.blockSignals(False)

    def _load_db(self, path: str, select: str | None = None) -> None:
        from pathlib import Path

        if not path or not Path(path).exists():
            return
        try:

            profiles = load_device_db(path)
            register_devices(profiles)
            self._reload_products(select=select or self.cmb_product.currentData())
            if hasattr(self, "lbl_err"):
                self.lbl_err.setStyleSheet("color:#6ec59a;")
                self.lbl_err.setText(f"디바이스 {len(profiles)}개 로드됨")
                self.lbl_err.setVisible(True)
        except Exception as exc:  # noqa: BLE001
            self._error(f"디바이스 DB 로드 실패: {exc}")

    def _on_device_db_edited(self) -> None:
        """디바이스 DB 경로를 직접 입력/붙여넣기로 바꾼 뒤(찾아보기 없이) 제품 목록 갱신."""
        path = self.ed_device_db.text().strip()
        if path:
            self._load_db(path)

    def _pick_device_db(self) -> None:
        from pathlib import Path

        start = self.ed_device_db.text() or self.ed_workspace.text() or str(Path.home())
        path, _ = QFileDialog.getOpenFileName(
            self, "디바이스 DB(AOIDeviceDB.xlsx) 선택", start, "Excel 파일 (*.xlsx)"
        )
        if path:
            self.ed_device_db.setText(path)
            self._load_db(path)

    def _pick_workspace(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "작업공간 폴더 선택", self.ed_workspace.text() or str(Path.home())
        )
        if folder:
            self.ed_workspace.setText(folder)

    def _pick_output(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "출력 폴더 선택", self.ed_output.text() or self.ed_workspace.text()
        )
        if folder:
            self.ed_output.setText(folder)

    def _pick_log_dir(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "로그 저장 경로 선택", self.ed_log_dir.text() or self.ed_workspace.text()
        )
        if folder:
            self.ed_log_dir.setText(folder)

    def _on_accept(self) -> None:
        workspace = self.ed_workspace.text().strip()
        output = self.ed_output.text().strip()
        log_dir = self.ed_log_dir.text().strip()
        if not workspace:
            self._error("작업공간 폴더를 지정하세요.")
            return
        # 원본 보호: 작업공간/출력/로그 경로가 현재 LOT 내부면 차단.
        if self._current_lot:
            targets = [
                (workspace, "작업공간"),
                (output or workspace, "출력"),
            ]
            if self.btn_dev.isChecked():  # 개발자 모드에서만 로그 경로 사용
                targets.append((log_dir or workspace, "로그"))
            for target, label in targets:
                if conflicting_source(target, [self._current_lot]) is not None:
                    self._error(
                        f"{label} 폴더가 현재 LOT 폴더 내부에 있습니다. 원본 보호를 위해 "
                        "원본 밖의 폴더를 선택하세요."
                    )
                    return
        self.accept()

    def _on_update_clicked(self) -> None:
        """현재 입력값을 먼저 저장 의도로 반영하고 업데이트를 요청하며 닫는다."""
        self._wants_update = True
        self.updated_settings()
        self.update_requested.emit()
        self.accept()

    def wants_update(self) -> bool:
        return self._wants_update

    def _error(self, msg: str) -> None:
        self.lbl_err.setText(msg)
        self.lbl_err.setVisible(True)

    def updated_settings(self) -> AppSettings:
        """다이얼로그 입력을 반영한 설정(저장은 호출 측)."""
        self._settings.workspace = self.ed_workspace.text().strip()
        self._settings.output_folder = self.ed_output.text().strip()
        self._settings.log_dir = self.ed_log_dir.text().strip()
        self._settings.dev_mode = self.btn_dev.isChecked()
        self._settings.cluster_radius = self.spn_cluster.value()
        self._settings.auto_update_check = self.chk_update.isChecked()
        self._settings.product = self.cmb_product.currentData() or DEFAULT_PRODUCT
        self._settings.device_db_path = self.ed_device_db.text().strip()
        self._settings.ui_font_size = self.cmb_font.currentData() or "normal"
        return self._settings


# =============================================================================
# app/ui/splash.py   [#24]
# =============================================================================
"""시작 스플래시 — 무거운 MainWindow 임포트/구성 전에 즉시 피드백을 준다.

PySide6 임포트(가장 큰 비용) 직후, QApplication 이 만들어지자마자 표시한다.
"""




_W, _H = 460, 240


def make_splash() -> QSplashScreen:
    """앱 이름/버전과 '로딩 중...' 을 담은 스플래시 화면을 만든다."""
    pm = QPixmap(_W, _H)
    pm.fill(QColor(BG_PANEL))
    p = QPainter(pm)
    p.setRenderHint(QPainter.Antialiasing, True)
    p.setPen(QColor(NEON_SOFT))
    p.drawRoundedRect(0, 0, _W - 1, _H - 1, 14, 14)

    p.setPen(QColor(BASE_GLOW))
    title = QFont("Segoe UI", 17)
    title.setBold(True)
    p.setFont(title)
    p.drawText(pm.rect().adjusted(0, -34, 0, -34), Qt.AlignCenter, APP_NAME)

    # 버전은 제목 바로 밑에
    p.setPen(QColor(TEXT_DIM))
    sub = QFont("Segoe UI", 10)
    p.setFont(sub)
    p.drawText(
        pm.rect().adjusted(0, -8, 0, -8),
        Qt.AlignCenter,
        f"v{__version__}",
    )
    # 제작 크레딧(시작 화면)
    credit = QFont("Segoe UI", 9)
    p.setFont(credit)
    p.drawText(
        pm.rect().adjusted(0, 58, 0, 58),
        Qt.AlignCenter,
        CREDITS,
    )
    p.end()

    splash = QSplashScreen(pm)
    splash.setWindowFlag(Qt.WindowStaysOnTopHint, True)
    return splash


def show_status(splash: QSplashScreen, text: str) -> None:
    """스플래시 하단에 진행 상태 문구를 표시한다."""
    splash.showMessage(
        text,
        Qt.AlignHCenter | Qt.AlignBottom,
        QColor(TEXT),
    )


# =============================================================================
# app/ui/wafer_map.py   [#25]
# =============================================================================
"""웨이퍼 맵 네비게이터 — 현재 wafer 의 die 격자를 매칭 상태로 색칠하고,

die 클릭 시 해당 기준 사진으로 점프한다. 리뷰 현황을 한눈에 본다.

상태 색:
  matched(매칭)=초록, none(전무)=빨강, 기준없음=빈칸.
"""





_CELL = 16
_GAP = 2

_STATE_COLORS = {
    "matched": MATCH,
    "none": NOMATCH,
}


class WaferMapWidget(QWidget):
    """현재 wafer 의 die 상태 격자."""

    die_clicked = Signal(int, int)  # (col, row)

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self._cols = 0
        self._rows = 0
        self._states: dict[tuple[int, int], str] = {}
        self._current: Optional[tuple[int, int]] = None
        self._valid: Optional[frozenset] = None  # 존재하는 die (None 이면 전체 사각)
        # 그리기 원점(실좌표) — 내용 bounding box 의 좌상단을 (0,0) 픽셀에 맞춘다.
        # states/valid/current 는 실좌표 그대로 두고 그리기·클릭만 이 오프셋을 적용해,
        # 좌표계 원점이 wafer 마다 달라도 맵이 떠 보이거나 잘리지 않게 한다.
        self._origin_col = 0
        self._origin_row = 0
        self.setToolTip("웨이퍼 맵 — die 클릭 시 해당 기준 사진으로 이동")
        self.setMinimumSize(40, 40)

    def set_data(
        self,
        cols: int,
        rows: int,
        states: dict[tuple[int, int], str],
        current: Optional[tuple[int, int]] = None,
        valid: Optional[frozenset] = None,
        origin: tuple[int, int] = (0, 0),
    ) -> None:
        self._cols = max(0, cols)
        self._rows = max(0, rows)
        self._states = states
        self._current = current
        self._valid = valid if valid else None
        self._origin_col, self._origin_row = origin
        self.setFixedSize(
            max(40, self._cols * (_CELL + _GAP) + _GAP),
            max(40, self._rows * (_CELL + _GAP) + _GAP),
        )
        self.update()

    def clear(self) -> None:
        self._states = {}
        self._current = None
        self.update()

    def _cell_rect(self, col: int, row: int) -> QRect:
        """실좌표(col,row)를 원점 오프셋을 적용한 픽셀 사각형으로."""
        x = _GAP + (col - self._origin_col) * (_CELL + _GAP)
        y = _GAP + (row - self._origin_row) * (_CELL + _GAP)
        return QRect(x, y, _CELL, _CELL)

    def paintEvent(self, event):  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, False)
        empty = QColor(BG_ELEV)
        border = QColor(NEON_SOFT)
        for dr in range(self._rows):
            row = dr + self._origin_row
            for dc in range(self._cols):
                col = dc + self._origin_col
                # 디바이스 die 배치(valid)가 주어지면 존재하는 die 만 그린다(실제 모양).
                if self._valid is not None and (col, row) not in self._valid:
                    continue
                rect = self._cell_rect(col, row)
                status = self._states.get((col, row))
                color = QColor(_STATE_COLORS.get(status, "")) if status else empty
                painter.fillRect(rect, color)
                painter.setPen(QPen(border, 1))
                painter.drawRect(rect)
        if self._current is not None:
            cc, cr = self._current
            if (self._origin_col <= cc < self._origin_col + self._cols
                    and self._origin_row <= cr < self._origin_row + self._rows):
                painter.setPen(QPen(QColor(BASE_GLOW), 2))
                painter.drawRect(self._cell_rect(cc, cr).adjusted(0, 0, -1, -1))
        painter.end()

    def mousePressEvent(self, event):  # noqa: N802
        if event.button() != Qt.LeftButton:
            return
        pos = event.position().toPoint()
        dc = (pos.x() - _GAP) // (_CELL + _GAP)
        dr = (pos.y() - _GAP) // (_CELL + _GAP)
        if 0 <= dc < self._cols and 0 <= dr < self._rows:
            col = int(dc) + self._origin_col
            row = int(dr) + self._origin_row
            if (col, row) in self._states:
                self.die_clicked.emit(col, row)


# =============================================================================
# app/ui/widgets.py   [#26]
# =============================================================================
"""재사용 위젯 및 애니메이션 헬퍼 (문서 Section 8.6, 9).

- FadeImageLabel: 이미지 교체 시 부드러운 fade(기준) / 빠른 fade(비교) 전환.
- ClickableThumb: 클릭 가능한 썸네일(현재 선택 강조).
- 모든 움직임은 부드럽게(QPropertyAnimation 이징).
"""






class FadeImageLabel(QLabel):
    """이미지를 표시하는 라벨.

    ImageLoader 가 주입되면 이미지를 비동기로 로드하여 UI 멈춤을 막는다(Section 10).
    주입되지 않으면 동기 로드로 폴백한다(테스트/단독 사용).

    주의: QScrollArea 안에서 QGraphicsOpacityEffect 를 쓰면 스크롤 시 위젯이
    엉뚱한 위치에 그려지거나 사라지는 Qt 렌더 버그가 있어, 그리드 이미지는
    그래픽 이펙트 fade 를 쓰지 않고 즉시 교체한다.
    """

    def __init__(self, parent: Optional[QWidget] = None, duration: int = 220):
        super().__init__(parent)
        self.setAlignment(Qt.AlignCenter)
        self.setMinimumSize(120, 120)
        self.setScaledContents(False)
        self._source_pixmap: Optional[QPixmap] = None
        self._placeholder = "이미지 없음"
        self._loader: Optional[ImageLoader] = None
        self._pending_id = -1
        self._pending_animated = True

    def set_loader(self, loader: ImageLoader) -> None:
        self._loader = loader
        loader.loaded.connect(self._on_loaded)

    def set_duration(self, ms: int) -> None:  # 호환용 no-op (fade 제거)
        pass

    def _scaled(self) -> Optional[QPixmap]:
        if self._source_pixmap is None or self._source_pixmap.isNull():
            return None
        return self._source_pixmap.scaled(
            self.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation
        )

    def resizeEvent(self, event):  # noqa: N802 (Qt naming)
        super().resizeEvent(event)
        sc = self._scaled()
        if sc is not None:
            super().setPixmap(sc)

    def show_path(self, path: Optional[str | Path], animated: bool = True) -> None:
        """이미지 경로를 표시. None 이면 placeholder."""
        if path is None:
            self._pending_id = -1
            self._apply(None, animated)
            return
        self._pending_animated = animated
        if self._loader is not None:
            # 비동기 로드: 결과 도착 시 _on_loaded 에서 적용
            self._pending_id = self._loader.request(str(path))
        else:
            p = QPixmap(str(path))
            self._apply(p if not p.isNull() else None, animated)

    def _on_loaded(self, request_id: int, image: object) -> None:
        if request_id != self._pending_id:
            return  # 빠른 탐색으로 인한 지난 요청 결과는 무시
        animated = self._pending_animated
        if isinstance(image, QImage) and not image.isNull():
            self._apply(QPixmap.fromImage(image), animated)
        else:
            self._apply(None, animated)

    def show_message(self, text: str) -> None:
        self._source_pixmap = None
        super().clear()
        self.setText(text)

    def _apply(self, pixmap: Optional[QPixmap], animated: bool) -> None:
        self._source_pixmap = pixmap
        if pixmap is None:
            self.show_message(self._placeholder)
            return
        self.setText("")
        sc = self._scaled()
        if sc is not None:
            super().setPixmap(sc)


class ClickableThumb(QFrame):
    """클릭 가능한 기준 썸네일 (현재 선택 강조)."""

    clicked = Signal(int)

    def __init__(self, index: int, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.index = index
        self._selected = False
        self.setObjectName("thumb")
        self.setCursor(Qt.PointingHandCursor)
        self.setFixedSize(96, 96)
        self._build()
        self._refresh_style()

    def _build(self) -> None:
        lay = QVBoxLayout(self)
        lay.setContentsMargins(4, 4, 4, 4)
        lay.setSpacing(2)
        self.img = QLabel()
        self.img.setAlignment(Qt.AlignCenter)
        self.img.setFixedSize(84, 62)
        self.img.setStyleSheet(
            f"background:{BG}; border-radius:6px; color:{TEXT_DIM};"
        )
        # 매칭 상태 점(미매칭=빨강, 매칭=숨김) — 트리아지 표식
        self.dot = QLabel(self.img)
        self.dot.setFixedSize(10, 10)
        self.dot.move(70, 4)
        self.dot.hide()
        self.caption = QLabel("")
        self.caption.setObjectName("dim")
        self.caption.setAlignment(Qt.AlignCenter)
        self.caption.setStyleSheet(f"font-size: {fpx(9)}px;")
        lay.addWidget(self.img)
        lay.addWidget(self.caption)

    def set_image(self, path: Optional[str | Path]) -> None:
        if path is not None:
            p = QPixmap(str(path))
            if not p.isNull():
                self.img.setPixmap(
                    p.scaled(84, 62, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                )
                return
        self.img.setText("?")

    def set_caption(self, text: str) -> None:
        self.caption.setText(text)

    def set_status(self, status: str) -> None:
        """매칭 상태 점 표시: 'none'(빨강 점) / 'matched'(점 없음)."""
        if status != "none":
            self.dot.hide()
            return
        self.dot.setStyleSheet(
            f"background:{NOMATCH}; border:1px solid {BG}; border-radius:5px;"
        )
        self.dot.show()

    def set_tooltip(self, text: str) -> None:
        # 자식 위젯은 부모 tooltip 을 상속하지 않으므로 모두 지정한다.
        self.setToolTip(text)
        self.img.setToolTip(text)
        self.caption.setToolTip(text)

    def set_selected(self, selected: bool) -> None:
        self._selected = selected
        self._refresh_style()

    def _refresh_style(self) -> None:
        if self._selected:
            self.setStyleSheet(
                f"QFrame#thumb {{ background: {NEON_DIM};"
                f" border: 2px solid {BASE_GLOW}; border-radius: 8px; }}"
            )
        else:
            self.setStyleSheet(
                f"QFrame#thumb {{ background: {BG_ELEV};"
                f" border: 1px solid {NEON_SOFT}; border-radius: 8px; }}"
                f"QFrame#thumb:hover {{ border: 1px solid {NEON}; }}"
            )

    def mousePressEvent(self, event):  # noqa: N802
        if event.button() == Qt.LeftButton:
            self.clicked.emit(self.index)
        super().mousePressEvent(event)


# =============================================================================
# app/updater.py   [#27]
# =============================================================================
"""자동 업데이트 — 메인 브랜치를 가져와 적용.

설치 형태에 따라 자동 감지:
  - git 체크아웃(.git 존재 + git 설치) → git fetch + reset --hard origin/main
  - 그 외 → GitHub codeload 에서 main ZIP 을 받아 설치 폴더에 덮어쓰기

순수 함수로 분리(네트워크/Qt 없이 테스트 가능)하고, Qt 워커가 이를 감싼다.
사용자 작업공간(LOCALAPPDATA)은 건드리지 않으며, 앱 설치 폴더만 갱신한다.
"""




updater__ProgressCb = Optional[Callable[[str], None]]

# ZIP 추출 시 덮어쓰지 않을 최상위 폴더
_SKIP_DIRS = {".git", "__pycache__", ".pytest_cache", ".venv", "venv"}
# 자동 업데이트로 받아오지 않을(로컬 유지) 파일 이름 — 개발 문서는 배포본에서 갱신하지 않는다.
_SKIP_FILES = {"CLAUDE.md", "README.md"}

_API = "https://api.github.com/repos/{owner}/{repo}/commits/{branch}"
_ZIP = "https://codeload.github.com/{owner}/{repo}/zip/refs/heads/{branch}"


@dataclass
class UpdateStatus:
    available: bool
    local: Optional[str]
    remote: Optional[str]
    method: str  # "git" | "zip"
    error: Optional[str] = None


# ----------------------------------------------------------------- 경로/버전
def app_root() -> Path:
    """설치 루트(이 파일의 상위의 상위)."""
    return Path(__file__).resolve().parent  # 단일 파일: 자기 디렉터리를 설치 루트로 본다


def is_frozen() -> bool:
    return bool(getattr(sys, "frozen", False))


def git_available() -> bool:
    return shutil.which("git") is not None


def is_git_checkout(root: Path) -> bool:
    return (root / ".git").exists() and git_available()


def version_file(root: Path) -> Path:
    return root / "version.json"


def read_version(root: Path) -> Optional[str]:
    vf = version_file(root)
    if vf.exists():
        try:
            return json.loads(vf.read_text(encoding="utf-8")).get("commit")
        except (json.JSONDecodeError, OSError):
            return None
    return None


def write_version(root: Path, sha: str) -> None:
    version_file(root).write_text(
        json.dumps({"commit": sha}, indent=2), encoding="utf-8"
    )


def current_sha(root: Optional[Path] = None) -> Optional[str]:
    """현재 설치본의 커밋 sha (git 우선, 없으면 version.json)."""
    root = root or app_root()
    if is_git_checkout(root):
        try:
            out = subprocess.run(
                ["git", "-C", str(root), "rev-parse", "HEAD"],
                capture_output=True, text=True, timeout=15,
            )
            if out.returncode == 0:
                return out.stdout.strip()
        except (OSError, subprocess.SubprocessError):
            pass
    return read_version(root)


# -------------------------------------------------------------------- 원격
def fetch_remote_sha(
    owner: str,
    repo: str,
    branch: str,
    token: str = "",
    opener=urlopen,
    timeout: float = 10.0,
) -> Optional[str]:
    """GitHub API 로 브랜치 최신 커밋 sha 를 가져온다(public 무인증)."""
    url = _API.format(owner=owner, repo=repo, branch=branch)
    headers = {"Accept": "application/vnd.github+json", "User-Agent": "Defect Tracker"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = Request(url, headers=headers)
    with opener(req, timeout=timeout) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    sha = data.get("sha")
    return sha if isinstance(sha, str) and sha else None


def check_update(
    root: Optional[Path] = None,
    owner: str = UPDATE_OWNER,
    repo: str = UPDATE_REPO,
    branch: str = UPDATE_BRANCH,
    token: str = "",
    opener=urlopen,
) -> UpdateStatus:
    """업데이트 가능 여부를 판정(백그라운드 호출용, 예외는 error 로 흡수)."""
    root = root or app_root()
    method = "git" if is_git_checkout(root) else "zip"
    local = current_sha(root)
    try:
        remote = fetch_remote_sha(owner, repo, branch, token, opener)
    except Exception as exc:  # noqa: BLE001 - 네트워크/파싱 오류는 graceful
        return UpdateStatus(False, local, None, method, error=str(exc))
    if remote is None:
        return UpdateStatus(False, local, None, method, error="원격 버전 확인 실패")
    available = local is None or (local[:40] != remote[:40])
    return UpdateStatus(available, local, remote, method)


# --------------------------------------------------------------- 적용(git)
def apply_via_git(
    root: Path, branch: str = UPDATE_BRANCH, progress: updater__ProgressCb = None
) -> tuple[bool, str]:
    """git fetch + reset --hard origin/branch 로 메인을 그대로 반영."""
    def run(args: list[str]) -> subprocess.CompletedProcess:
        return subprocess.run(
            ["git", "-C", str(root), *args],
            capture_output=True, text=True, timeout=180,
        )

    # 개발 문서(CLAUDE.md·README.md)는 업데이트로 받아오지 않는다 → reset 전 로컬 내용을
    # 저장해 두었다가 이후 그대로 복원한다(로컬에 없으면 없는 상태로 되돌린다).
    saved = {
        name: ((root / name).read_bytes() if (root / name).exists() else None)
        for name in _SKIP_FILES
    }

    if progress:
        progress("원격 변경사항 받는 중...")
    f = run(["fetch", "--depth", "1", "origin", branch])
    if f.returncode != 0:
        return False, f"git fetch 실패: {f.stderr.strip()}"
    if progress:
        progress("최신 버전 적용 중...")
    r = run(["reset", "--hard", f"origin/{branch}"])
    if r.returncode != 0:
        return False, f"git reset 실패: {r.stderr.strip()}"

    # 저장해 둔 개발 문서를 복원(업데이트가 이 파일들을 바꾸지 않도록).
    for name, content in saved.items():
        p = root / name
        try:
            if content is None:
                p.unlink(missing_ok=True)
            else:
                p.write_bytes(content)
        except OSError:
            pass
    return True, "git 으로 업데이트했습니다."


# --------------------------------------------------------------- 적용(ZIP)
def download_zip(url: str, dest: Path, opener=urlopen, timeout: float = 60.0) -> Path:
    req = Request(url, headers={"User-Agent": "Defect Tracker"})
    with opener(req, timeout=timeout) as resp, open(dest, "wb") as fh:
        shutil.copyfileobj(resp, fh)
    return dest


def _safe_join(target_root: Path, rel: str) -> Optional[Path]:
    """zip-slip 방지: target_root 밖으로 나가는 경로는 None."""
    dest = (target_root / rel).resolve()
    root = target_root.resolve()
    if dest == root or root in dest.parents:
        return dest
    return None


def extract_over(
    zip_path: Path, target_root: Path, skip: Optional[set[str]] = None
) -> int:
    """ZIP(최상위 단일 폴더)을 target_root 에 덮어쓴다(삭제는 하지 않음).

    Returns: 기록한 파일 수.
    """
    skip = skip if skip is not None else _SKIP_DIRS
    written = 0
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            parts = Path(info.filename).parts
            if len(parts) <= 1:
                continue  # 최상위 폴더 자체
            rel_parts = parts[1:]  # 'repo-branch/' 접두 제거
            if any(p in skip for p in rel_parts):
                continue
            if rel_parts[-1] in _SKIP_FILES:
                continue  # 개발 문서(CLAUDE.md·README.md)는 갱신하지 않음
            rel = "/".join(rel_parts)
            dest = _safe_join(target_root, rel)
            if dest is None:
                continue
            dest.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(info) as src, open(dest, "wb") as out:
                shutil.copyfileobj(src, out)
            written += 1
    return written


def apply_via_zip(
    target_root: Path,
    owner: str = UPDATE_OWNER,
    repo: str = UPDATE_REPO,
    branch: str = UPDATE_BRANCH,
    remote_sha: Optional[str] = None,
    opener=urlopen,
    progress: updater__ProgressCb = None,
) -> tuple[bool, str]:
    """ZIP 을 받아 설치 폴더에 덮어쓰고 version.json 을 기록."""
    url = _ZIP.format(owner=owner, repo=repo, branch=branch)
    tmpdir = Path(tempfile.mkdtemp(prefix="defect_tracker_update_"))
    try:
        if progress:
            progress("새 버전 내려받는 중...")
        zip_path = download_zip(url, tmpdir / "main.zip", opener)
        if progress:
            progress("설치 폴더에 적용 중...")
        count = extract_over(zip_path, target_root)
        if count == 0:
            return False, "받은 패키지에 적용할 파일이 없습니다."
        if remote_sha:
            write_version(target_root, remote_sha)
        return True, f"{count}개 파일을 갱신했습니다."
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def apply_update(
    status: UpdateStatus,
    root: Optional[Path] = None,
    owner: str = UPDATE_OWNER,
    repo: str = UPDATE_REPO,
    branch: str = UPDATE_BRANCH,
    opener=urlopen,
    progress: updater__ProgressCb = None,
) -> tuple[bool, str]:
    """감지된 방식(method)에 따라 업데이트를 적용."""
    root = root or app_root()
    if is_frozen():
        return False, "실행파일(exe) 버전은 자동 업데이트를 지원하지 않습니다. 새 버전을 내려받아 교체하세요."
    if status.method == "git" and is_git_checkout(root):
        return apply_via_git(root, branch, progress)
    return apply_via_zip(
        root, owner, repo, branch, remote_sha=status.remote, opener=opener, progress=progress
    )


# ----------------------------------------------------------------- Qt 워커
try:
    from PySide6.QtCore import QObject, QRunnable, Signal, Slot

    class UpdateCheckSignals(QObject):
        done = Signal(object)  # UpdateStatus

    class UpdateCheckWorker(QRunnable):
        """시작 시 백그라운드로 업데이트 여부 확인."""

        def __init__(self, token: str = ""):
            super().__init__()
            self.token = token
            self.signals = UpdateCheckSignals()

        @Slot()
        def run(self) -> None:
            status = check_update(token=self.token)
            self.signals.done.emit(status)

    class UpdateApplySignals(QObject):
        progress = Signal(str)
        finished = Signal(bool, str)  # ok, message

    class UpdateApplyWorker(QRunnable):
        """업데이트 적용(진행 상황 신호)."""

        def __init__(self, status: UpdateStatus, token: str = ""):
            super().__init__()
            self.status = status
            self.token = token
            self.signals = UpdateApplySignals()

        @Slot()
        def run(self) -> None:
            try:
                ok, msg = apply_update(
                    self.status, progress=self.signals.progress.emit
                )
            except Exception as exc:  # noqa: BLE001
                ok, msg = False, f"업데이트 중 오류: {exc}"
            self.signals.finished.emit(ok, msg)

except ImportError:  # PySide6 미설치 환경(부트스트랩 전)
    pass


# =============================================================================
# app/export/excel_report.py   [#28]
# =============================================================================
"""Excel 결과 출력 (문서 Section 8.7).

선택된 기준 사진들에 대해 기준 layer 사진과 비교 layer 매칭 결과를 깔끔한 Excel 로 출력한다.
포함 정보: LOT명, wafer ID, 기준/비교 layer, col_row_x_y 위치, 허용 오차, 매칭 여부,
이미지 썸네일, 원본 경로(추적용, 무수정).

저장 경로는 반드시 assert_output_safe 게이트를 통과해야 하며, 원본 폴더 내부면 차단된다.
원본 이미지는 read-only 로만 읽는다(썸네일 캐시를 통해).
"""





# 다크/네온 느낌과 어울리는 보고서 색 (Excel 은 밝은 배경이 가독성 좋아 절제해 사용)
_NAVY = "FF1B2A4A"
_NEON = "FF1E90FF"
_LIGHT = "FFF2F6FF"
_MATCH = "FF1F7A1F"
_NOMATCH = "FFB00020"
_GREY = "FF6B7280"

_THIN = Side(style="thin", color="FFB8C4DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_IMG_COL_WIDTH = 30  # Excel 폭 단위
_IMG_PX = 190  # 썸네일 픽셀
_IMG_ROW_HEIGHT = 150  # 포인트


def _set_cell(ws, row, col, value, *, bold=False, color=None, fill=None,
              size=10, align="left", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color or "FF1B2A4A", size=size)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    cell.border = _BORDER
    return cell


def _place_image(ws, thumb_cache, rec, row, col) -> None:
    """(row,col) 셀에 rec 썸네일을 앵커한다. 실패 시 안내 텍스트."""
    col_letter = get_column_letter(col)
    cell = ws.cell(row=row, column=col)
    cell.border = _BORDER
    thumb = thumb_cache.get_full_thumbnail(rec.image_path, max_size=_IMG_PX)
    if thumb is not None:
        try:
            xl = XLImage(str(thumb))
            xl.anchor = f"{col_letter}{row}"
            ws.add_image(xl)
            return
        except (OSError, ValueError):
            cell.value = "(이미지 로드 실패)"
            return
    cell.value = "(이미지 없음)"


def export_excel(
    output_path: str | Path,
    *,
    lot_name: str,
    base_layer: str,
    compare_layers: list[str],
    tolerance: float,
    selected: list[BaseDefectMatches],
    thumb_cache: ThumbnailCache,
    source_roots: Iterable[str | Path],
    progress: Optional[Callable[[int, int], None]] = None,
) -> Path:
    """선택된 기준 defect 들의 비교 결과를 Excel 로 저장한다.

    Returns:
        저장된 파일의 절대 경로.

    Raises:
        OriginalProtectionError: 출력 경로가 원본 폴더 내부일 때.
    """
    out = assert_output_safe(output_path, source_roots)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "비교 결과"
    ws.sheet_view.showGridLines = False

    # 열: A=라벨, B=기준, C..=비교 layer. 각 블록이 '자기' 기준/비교 layer 로 스스로
    # 라벨링하므로(여러 layer 를 한 번에 담아도 섞이지 않게), 열 수는 블록별 비교 layer
    # 최댓값으로 잡는다.
    max_cmp = max((len(item.results) for item in selected), default=len(compare_layers))
    n_data_cols = 1 + max_cmp  # 기준 + 비교(최대)
    n_cols = n_data_cols + 1   # + 라벨열(A)

    ws.column_dimensions["A"].width = 16
    for ci in range(n_data_cols):
        ws.column_dimensions[get_column_letter(2 + ci)].width = _IMG_COL_WIDTH

    # ---- 보고서 헤더 ----
    r = 1
    title = ws.cell(row=r, column=1, value="Defect Tracker 비교 결과 보고서")
    title.font = Font(bold=True, color="FFFFFFFF", size=14)
    title.fill = PatternFill("solid", fgColor=_NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    ws.row_dimensions[r].height = 26
    r += 1

    # 담긴 항목들의 실제 기준 layer(혼합일 수 있음)를 헤더에 표기.
    base_layers_present: list[str] = []
    for item in selected:
        bl = item.base.layer or base_layer
        if bl not in base_layers_present:
            base_layers_present.append(bl)
    if len(base_layers_present) <= 1:
        base_desc = base_layers_present[0] if base_layers_present else base_layer
    else:
        base_desc = "혼합(기준 없이): " + ", ".join(base_layers_present)

    meta = (
        f"LOT: {lot_name}    기준 Layer: {base_desc}    "
        f"허용 오차: {tolerance:g}    생성: {datetime.now():%Y-%m-%d %H:%M}    "
        f"※ 각 블록의 'Layer' 행이 실제 기준/비교 layer 를 표시"
    )
    mc = ws.cell(row=r, column=1, value=meta)
    mc.font = Font(color="FFFFFFFF", size=10)
    mc.fill = PatternFill("solid", fgColor=_NEON)
    mc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    ws.row_dimensions[r].height = 20
    r += 2

    # 상단 컬럼 헤더 행은 두지 않는다 — 블록마다 'Layer' 행으로 이미 표기하므로 중복이다(항목 1).

    # ---- 각 기준 defect 블록 (블록마다 '자기' 기준/비교 layer 로 표기) ----
    _total = len(selected)
    for idx, item in enumerate(selected, start=1):
        if progress is not None:
            progress(idx, _total)
        base = item.base
        base_layer_name = base.layer or base_layer
        results = list(item.results)

        # 블록 제목 — 기준 layer 를 함께 표기(여러 layer 혼합 대비).
        _set_cell(
            ws, r, 1,
            f"#{idx}", bold=True, color="FFFFFFFF", fill=_GREY, align="center",
        )
        head = (
            f"기준 {base_layer_name}   wafer {base.wafer_id}   "
            f"die ({base.col},{base.row})   pos {base.position_key}   "
            f"[{base.source.value}]"
        )
        _set_cell(ws, r, 2, head, bold=True, fill=_LIGHT, wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_cols)
        ws.row_dimensions[r].height = 18
        r += 1

        # Layer 이름 행 — 이 블록의 실제 기준/비교 layer(전역 기준에 종속되지 않음).
        _set_cell(ws, r, 1, "Layer", bold=True, align="center", fill=_LIGHT)
        _set_cell(
            ws, r, 2, "★ " + base_layer_name + " (기준)",
            bold=True, color="FFFFFFFF", fill=_NEON, align="center",
        )
        for ci, mr in enumerate(results):
            _set_cell(
                ws, r, 3 + ci, mr.compare_layer,
                bold=True, color="FFFFFFFF", fill=_NAVY, align="center",
            )
        for ci in range(len(results), max_cmp):
            _set_cell(ws, r, 3 + ci, "", fill=_LIGHT, align="center")
        ws.row_dimensions[r].height = 18
        r += 1

        # 이미지 행
        _set_cell(ws, r, 1, "이미지", bold=True, align="center", fill=_LIGHT)
        ws.row_dimensions[r].height = _IMG_ROW_HEIGHT
        _place_image(ws, thumb_cache, base, r, 2)
        for ci, mr in enumerate(results):
            rec = mr.matched
            if rec is not None:
                _place_image(ws, thumb_cache, rec, r, 3 + ci)
            else:
                _set_cell(ws, r, 3 + ci, "매칭 없음", color=_NOMATCH, align="center")
        for ci in range(len(results), max_cmp):
            ws.cell(row=r, column=3 + ci).border = _BORDER
        r += 1

        # 상세 정보 행
        _set_cell(ws, r, 1, "정보", bold=True, align="center", fill=_LIGHT)
        base_lines = ["기준 ★", f"위치 {base.position_key}", Path(base.image_path).name]
        extra = getattr(getattr(item, "base_cluster", None), "extra_count", 0) or 0
        if extra:
            base_lines.append(f"+{extra} 근접중복")
        _set_cell(ws, r, 2, "\n".join(base_lines), color=_NEON, wrap=True, size=9)
        for ci, mr in enumerate(results):
            rec = mr.matched
            if rec is not None:
                dist = mr.distance
                lines = [
                    f"매칭 O (거리 {dist:.1f})" if dist is not None else "매칭 O",
                    f"위치 {rec.position_key}",
                    Path(rec.image_path).name,
                ]
                color = _MATCH
            else:
                lines = ["매칭 X"]
                color = _NOMATCH
            _set_cell(ws, r, 3 + ci, "\n".join(lines), color=color, wrap=True, size=9)
        for ci in range(len(results), max_cmp):
            ws.cell(row=r, column=3 + ci).border = _BORDER
        ws.row_dimensions[r].height = 48
        r += 1

        # 원본 경로 행 (추적용)
        _set_cell(ws, r, 1, "원본경로", bold=True, align="center", fill=_LIGHT)
        _set_cell(ws, r, 2, str(base.image_path), color=_GREY, wrap=True, size=8)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_cols)
        ws.row_dimensions[r].height = 24
        r += 1
        r += 1  # 블록 간 간격

    wb.save(out)
    return out


# =============================================================================
# app/scanner.py   [#29]
# =============================================================================
"""LOT 폴더 스캔 및 인덱스 구축 (문서 Section 4, 8.1, 8.2).

구조: LOT 폴더 / layer 폴더 / wafer 폴더 / (defect 이미지 + info/ini 파일)

각 wafer 폴더의 이미지마다 좌표 출처를 판별해 col_row_x_y 위치 정보를 만든다:
  1) Camtek 파일명에 좌표가 있으면 그대로 추출
  2) ColorImageGrabingInfo.ini section 에서 산출
  3) KLA info(.001) 의 TiffFileName/DefectList 로 변환

원본은 read-only 로만 읽으며 어떤 파일도 생성하지 않는다.
"""




scanner__ProgressCb = Optional[Callable[[str, int, int], None]]
CancelCb = Optional[Callable[[], bool]]  # True 면 스캔을 협조적으로 중단

_INI_HINT = "colorimagegrabinginfo"
scanner__log = logging.getLogger("defect_tracker.scanner")

# wafer 스캔 병렬 워커 수(네트워크 I/O 바운드 → CPU 수보다 넉넉히).
_SCAN_WORKERS = max(4, min(16, (os.cpu_count() or 4) * 2))

# 디렉터리/파일 나열 중 접근 오류를 누적(스캔 1회 단위). 병렬 스캔(Phase 3) 대비 lock 사용.
_scan_errors: list[str] = []
_scan_errors_lock = threading.Lock()


def _record_scan_error(path: Path, exc: OSError) -> None:
    msg = f"{path}: {exc.__class__.__name__}: {exc}"
    with _scan_errors_lock:
        _scan_errors.append(msg)
    scanner__log.warning("접근 실패 — %s", msg)


@dataclass
class LotIndex:
    """스캔 결과. layer/wafer/record 인덱스를 보관."""

    lot_name: str
    lot_path: Path
    layers: list[LayerInfo] = field(default_factory=list)
    records: list[DefectRecord] = field(default_factory=list)
    scan_errors: list[str] = field(default_factory=list)  # 접근 불가 경로(권한/네트워크)

    def layer_canonicals(self) -> list[str]:
        """선택 UI 에 표시할 layer 이름 목록(폴더 순서, 유니크).

        canonical 이 충돌하지 않으면 canonical 그대로, 충돌하면 재리뷰 등으로
        구분된 display 이름을 사용한다(scan_lot 에서 display 산정).
        """
        seen: list[str] = []
        for lyr in self.layers:
            name = lyr.display or lyr.canonical
            if name not in seen:
                seen.append(name)
        return seen

    def records_for_layer(self, canonical: str) -> list[DefectRecord]:
        return [r for r in self.records if r.layer == canonical]

    def records_by_layer(self) -> dict[str, list[DefectRecord]]:
        out: dict[str, list[DefectRecord]] = {}
        for r in self.records:
            out.setdefault(r.layer, []).append(r)
        return out

    def wafers(self) -> list[str]:
        seen: list[str] = []
        for r in self.records:
            if r.wafer_id not in seen:
                seen.append(r.wafer_id)
        return seen


def _list_dirs(path: Path) -> list[Path]:
    try:
        return sorted([p for p in path.iterdir() if p.is_dir()], key=lambda p: p.name)
    except OSError as exc:
        _record_scan_error(path, exc)
        return []


def _list_files(path: Path) -> list[Path]:
    try:
        return sorted([p for p in path.iterdir() if p.is_file()], key=lambda p: p.name)
    except OSError as exc:
        _record_scan_error(path, exc)
        return []


def _is_image(name: str) -> bool:
    return Path(name).suffix.lower() in IMAGE_EXTENSIONS


def _dir_has_image(path: Path) -> bool:
    try:
        with os.scandir(path) as it:
            for e in it:
                if e.is_file() and _is_image(e.name):
                    return True
    except OSError:
        return False
    return False


def _image_depth(root: Path, max_depth: int = 4, breadth: int = 24) -> Optional[int]:
    """root 아래에서 이미지 파일을 직접 담은 디렉터리가 처음 나타나는 깊이를 반환.

    레벨별 BFS(레벨당 폴더 수 breadth 로 제한)로 가볍게 탐색한다(네트워크 경로 대비).
    못 찾으면 None.
    """
    level = [root]
    for depth in range(0, max_depth + 1):
        for d in level[:breadth]:
            if _dir_has_image(d):
                return depth
        nxt: list[Path] = []
        for d in level[:breadth]:
            try:
                with os.scandir(d) as it:
                    nxt.extend(Path(e.path) for e in it if e.is_dir())
            except OSError:
                continue
        if not nxt:
            break
        level = nxt
    return None


def _child_dirs(root: Path, breadth: int = 24) -> list[Path]:
    """root 의 바로 아래 하위 디렉터리 목록(breadth 로 제한, 네트워크 대비)."""
    try:
        with os.scandir(root) as it:
            dirs = [Path(e.path) for e in it if e.is_dir()]
    except OSError:
        return []
    return dirs[:breadth]


def _has_child_dir_with_images(root: Path, breadth: int = 24) -> bool:
    """root 의 바로 아래 하위 폴더(=wafer 후보) 중 하나라도 이미지를 직접 담고 있는가."""
    return any(_dir_has_image(c) for c in _child_dirs(root, breadth))


def _has_grandchild_dir_with_images(root: Path, breadth: int = 24) -> bool:
    """root/자식(layer)/손자(wafer)/이미지 구조가 있는가(= root 가 자재(LOT)인가).

    자식(layer) 폴더 중 하나라도 '이미지를 직접 담은 하위 폴더(wafer)'를 가지면 True.
    """
    return any(_has_child_dir_with_images(c, breadth) for c in _child_dirs(root, breadth))


def classify_selection(path: str | Path) -> tuple[str, Optional[Path]]:
    """선택한 폴더가 자재 구조(LOT/layer/wafer/사진)에서 어느 레벨인지 **구조로** 판별.

    LOT 폴더의 정의: `LOT/layer/wafer/사진` — 사진이 정확히 2단계 아래(wafer 폴더)에
    있어야 자재(LOT)로 인정한다. 얕은 위치(LOT·layer 폴더)에 흔히 섞여 있는 요약/맵
    이미지 같은 잡파일에 흔들리지 않도록 **가장 깊은 구조가 우선**하도록 판정한다.

    Returns:
        (kind, material_path) — kind 는
          'material'(LOT 정상) / 'layer' / 'wafer'(둘 다 자동으로 상위 자재로 보정) /
          'too_high'(device 등 상위, 재선택 필요) / 'unknown'(이미지 못 찾음).
        material_path 는 layer/wafer/material 일 때 추정 자재(LOT) 폴더, 그 외 None.
    """
    p = Path(path)
    # 깊은 구조 우선: LOT/layer/wafer/사진(손자에 이미지) → 자재(LOT).
    if _has_grandchild_dir_with_images(p):
        return ("material", p)
    # p/wafer/사진(자식이 이미지 직접 보유) → p 는 layer, 상위가 LOT.
    if _has_child_dir_with_images(p):
        return ("layer", p.parent)
    # p/사진(직접 보유) → p 는 wafer, 상위상위가 LOT.
    if _dir_has_image(p):
        return ("wafer", p.parent.parent)
    # 여기까지 아니지만 더 깊은 곳에 이미지가 있으면 상위(device 등) → 재선택 필요.
    if _image_depth(p) is not None:
        return ("too_high", None)
    return ("unknown", None)


def _merge_ini_sections(ini_paths: list[Path]) -> dict[str, dict[str, str]]:
    """폴더 내 모든 ColorImageGrabingInfo.ini section 을 하나로 합친다."""
    merged: dict[str, dict[str, str]] = {}
    for p in ini_paths:
        try:
            merged.update(load_ini(p))
        except OSError:
            continue
    return merged


def _build_record_for_image(
    image_path: Path,
    wafer_id: str,
    layer: str,
    layer_folder: str,
    ini_sections: dict[str, dict[str, str]],
    kla_parsed,  # Optional[_ParsedInfo]
    wafer_diag: dict,  # 폴더 레벨 진단 컨텍스트
) -> DefectRecord:
    name = image_path.name
    stem = image_path.stem
    # 각 파서 시도의 실패 사유를 모아(시도 트레일) 진단 리포트에 쓴다.
    trail: list[str] = []

    # 1) Camtek 파일명 직접 파싱
    fn = parse_camtek_filename(name)
    if fn.status == ParseStatus.OK:
        return DefectRecord(
            image_path=image_path,
            wafer_id=wafer_id,
            layer=layer,
            layer_folder=layer_folder,
            source=Source.CAMTEK_FILENAME,
            status=ParseStatus.OK,
            col=fn.col,
            row=fn.row,
            x=fn.x,
            y=fn.y,
            defect_name=fn.defect_name,
            dx_size=fn.dx_size,
            dy_size=fn.dy_size,
            d_area=fn.d_area,
        )
    trail.append(f"파일명: {fn.reason}")

    # 2) ColorImageGrabingInfo.ini
    if ini_sections:
        ini_res = convert_from_sections(ini_sections, stem)
        if ini_res.status == ParseStatus.OK:
            return DefectRecord(
                image_path=image_path,
                wafer_id=wafer_id,
                layer=layer,
                layer_folder=layer_folder,
                source=Source.CAMTEK_INI,
                status=ParseStatus.OK,
                col=ini_res.col,
                row=ini_res.row,
                x=ini_res.x,
                y=ini_res.y,
            )
        trail.append(f"INI: {ini_res.reason}")
    else:
        trail.append("INI: ColorImageGrabingInfo.ini 없음")

    # 3) KLA info
    if kla_parsed is not None:
        kla_res = convert_from_parsed(kla_parsed, name)
        if kla_res.status == ParseStatus.OK:
            return DefectRecord(
                image_path=image_path,
                wafer_id=wafer_id,
                layer=layer,
                layer_folder=layer_folder,
                source=Source.KLA,
                status=ParseStatus.OK,
                col=kla_res.col,
                row=kla_res.row,
                x=kla_res.x,
                y=kla_res.y,
            )
        trail.append(f"KLA: {kla_res.reason}")
        return DefectRecord(
            image_path=image_path,
            wafer_id=wafer_id,
            layer=layer,
            layer_folder=layer_folder,
            source=Source.KLA,
            status=kla_res.status,
            col=kla_res.col,
            row=kla_res.row,
            x=kla_res.x,
            y=kla_res.y,
            note=" | ".join(trail),
            diag=wafer_diag,
        )
    trail.append("KLA: info 파일 없음")

    # 어느 방법으로도 좌표를 찾지 못함
    return DefectRecord(
        image_path=image_path,
        wafer_id=wafer_id,
        layer=layer,
        layer_folder=layer_folder,
        source=Source.UNKNOWN,
        status=ParseStatus.NOT_FOUND,
        note=" | ".join(trail),
        diag=wafer_diag,
    )


def _read_info_text_safe(path: Path) -> str:
    """진단용: info 파일 텍스트 전문을 읽는다."""
    try:
        raw = read_only_bytes(path)
        for enc in ("utf-8", "cp949", "utf-16-le", "utf-16-be", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        else:
            text = raw.decode("latin-1", errors="replace")
        return text
    except OSError:
        return "(읽기 실패)"


def _scan_wafer_folder(
    wafer_dir: Path, layer: str, layer_folder: str
) -> list[DefectRecord]:
    files = _list_files(wafer_dir)
    filenames = [f.name for f in files]

    image_files = [f for f in files if _is_image(f.name)]
    ini_files = [
        f for f in files if _INI_HINT in f.name.lower() and f.suffix.lower() == ".ini"
    ]
    ini_sections = _merge_ini_sections(ini_files)

    # KLA info 파일 선택. Camtek INI 파일(.ini)은 KLA info 후보에서 제외하여
    # 혼재 폴더에서 오인하지 않도록 한다. 이 파싱은 .001/info 가 있는 KLA 폴더에서만 발생.
    kla_parsed = None
    kla_candidates = [n for n in filenames if Path(n).suffix.lower() != ".ini"]
    info_name = select_info_file(kla_candidates)
    info_text = ""
    if info_name:
        info_path = wafer_dir / info_name
        info_text = _read_info_text_safe(info_path)
        try:
            kla_parsed = load_info(info_path)
        except OSError:
            kla_parsed = None

    # 진단용 폴더 컨텍스트(실패 record 에만 첨부)
    wafer_diag: dict = {
        "wafer_dir": str(wafer_dir),
        "files_in_folder": filenames,
        "image_count": len(image_files),
        "ini_files": [f.name for f in ini_files],
        "has_ini_sections": bool(ini_sections),
        "ini_section_keys": sorted(ini_sections.keys())[:20] if ini_sections else [],
        "kla_info_file": info_name,
        "kla_info_text": info_text,
        "kla_tiff_count": len(kla_parsed.defects) if kla_parsed else 0,
        "kla_all_defect_count": len(kla_parsed.all_defects) if kla_parsed else 0,
        "kla_die_pitch_y": kla_parsed.die_pitch_y if kla_parsed else None,
    }

    wafer_id = wafer_dir.name
    return [
        _build_record_for_image(
            img, wafer_id, layer, layer_folder, ini_sections, kla_parsed,
            wafer_diag,
        )
        for img in image_files
    ]


def _assign_displays(infos: list[LayerInfo]) -> None:
    """canonical 이 충돌하는 경우에만 재리뷰 등으로 구분한 display 이름을 부여한다.

    충돌이 없으면 display = canonical (기존 동작 유지). 충돌 시 재리뷰 깊이에 따라
    "{canonical}_재리뷰"(레벨1) / "{canonical}_재재리뷰"(레벨2) … 로 구분하고, 일반은
    canonical, 그래도 겹치면 " (2)", " (3)" … 로 유일화.
    """
    counts: dict[str, int] = {}
    for inf in infos:
        counts[inf.canonical] = counts.get(inf.canonical, 0) + 1

    used: set[str] = set()
    for inf in infos:
        if counts.get(inf.canonical, 0) <= 1:
            name = inf.canonical
        elif inf.re_review_level >= 1:
            suffix = "재" * inf.re_review_level + "리뷰"
            name = f"{inf.canonical}_{suffix}"
        else:
            name = inf.canonical
        base = name
        k = 2
        while name in used:
            name = f"{base} ({k})"
            k += 1
        used.add(name)
        inf.display = name


def scan_lot(
    lot_path: str | Path,
    progress: scanner__ProgressCb = None,
    cancel_check: CancelCb = None,
) -> LotIndex:
    """LOT 폴더를 스캔해 LotIndex 를 만든다. 원본은 읽기 전용으로만 접근.

    cancel_check 가 True 를 반환하면 wafer 처리 루프를 협조적으로 중단한다(부분 결과 반환).
    """
    lot = Path(lot_path)
    index = LotIndex(lot_name=lot.name, lot_path=lot)

    with _scan_errors_lock:
        _scan_errors.clear()
    scanner__log.info("스캔 시작: %s", lot)

    layer_dirs = _list_dirs(lot)
    total = len(layer_dirs) or 1

    # 1차: layer 폴더별 정규화 정보 수집 후 표시 이름(display) 산정(충돌 시에만 구분).
    for layer_dir in layer_dirs:
        canonical, is_rr = normalize_layer(layer_dir.name)
        level = re_review_level(layer_dir.name)
        index.layers.append(
            LayerInfo(
                folder_name=layer_dir.name,
                canonical=canonical,
                path=layer_dir,
                is_re_review=is_rr,
                re_review_level=level,
            )
        )
    _assign_displays(index.layers)

    # 2차: (layer, wafer) 작업 목록을 만들고 wafer 단위로 병렬 스캔한다.
    # 네트워크 경로에서 디렉터리/info 파일 I/O latency 가 지배적이므로 병렬화가 효과적.
    # 결과는 입력 순서(layer 순 → wafer 정렬)대로 모아 직렬 스캔과 동일한 결정적 순서 유지.
    tasks: list[tuple[LayerInfo, Path]] = []
    for info in index.layers:
        for wafer_dir in _list_dirs(info.path):
            tasks.append((info, wafer_dir))
    wafer_total = len(tasks) or 1

    def _scan_task(task: tuple[LayerInfo, Path]) -> list[DefectRecord]:
        if cancel_check and cancel_check():  # 이미 제출된 작업도 빠르게 빠져나간다
            return []
        info, wafer_dir = task
        try:
            return _scan_wafer_folder(wafer_dir, info.display, info.folder_name)
        except Exception as exc:  # noqa: BLE001 - 한 wafer 실패가 전체 스캔을 막지 않게
            _record_scan_error(wafer_dir, exc if isinstance(exc, OSError) else OSError(str(exc)))
            return []

    if tasks:
        workers = min(_SCAN_WORKERS, len(tasks))
        with ThreadPoolExecutor(max_workers=workers) as ex:
            for done, recs in enumerate(ex.map(_scan_task, tasks), start=1):
                if cancel_check and cancel_check():
                    scanner__log.info("스캔 중단 요청 — %d/%d wafer 에서 멈춤", done, wafer_total)
                    break
                index.records.extend(recs)
                if progress:
                    info, wafer_dir = tasks[done - 1]
                    progress(
                        f"스캔: {info.folder_name}/{wafer_dir.name}", done, wafer_total
                    )

    with _scan_errors_lock:
        index.scan_errors = list(_scan_errors)
    scanner__log.info(
        "스캔 완료: layer %d · record %d · 접근오류 %d",
        len(index.layers), len(index.records), len(index.scan_errors),
    )
    if progress:
        progress("스캔 완료", total, total)
    return index


# =============================================================================
# app/ui/cluster_view.py   [#30]
# =============================================================================
"""클러스터 defect 표시용 공유 위젯 (히트맵·메인 매치 공통).

- `load_thumb_holder`: 썸네일 QLabel 생성.
- `ClickThumb`: 썸네일 클릭 시 원본 뷰어 열기.
- `ClusteredThumb`: 대표 썸네일 + layer 배지 + (클러스터면) 좌하단 '+n' 버튼.
- `ClusterMembersPopup`: 묶인 defect 전체를 가로(줄바꿈)로 보여주는 팝업.
"""






def attach_image_context_menu(widget: QWidget, path_getter) -> None:
    """위젯에 사진 우클릭 메뉴(경로 복사·파일 열기·폴더 열기)를 붙인다.

    path_getter() 는 표시 중인 이미지 경로를 반환한다(호출 시점 값 사용).
    메인 그리드(compare_grid)의 우클릭 메뉴와 동일한 동작을 히트맵 썸네일에도 제공한다.
    """
    widget.setContextMenuPolicy(Qt.CustomContextMenu)

    def _show(pos) -> None:
        path = str(path_getter())
        if not path:
            return
        menu = QMenu(widget)
        menu.addAction("경로 복사", lambda: QGuiApplication.clipboard().setText(path))
        menu.addAction(
            "파일 열기",
            lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)),
        )
        menu.addAction(
            "폴더 열기",
            lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(str(Path(path).parent))),
        )
        menu.exec(widget.mapToGlobal(pos))

    widget.customContextMenuRequested.connect(_show)


def _blank_holder(px: int) -> QLabel:
    holder = QLabel()
    holder.setAlignment(Qt.AlignCenter)
    holder.setFixedSize(px, int(px * 0.78))
    holder.setStyleSheet(
        f"background:{BG}; border:1px solid {NEON_SOFT};"
        f" border-radius:6px; color:{TEXT_DIM}; font-size:10px;"
    )
    return holder


def _start_loading_anim(holder: QLabel) -> None:
    """지연 로딩 placeholder 에 '로딩' 말줄임(…) 애니메이션을 건다(홀더 자식 타이머)."""
    holder._load_dots = 0  # type: ignore[attr-defined]
    holder.setText("로딩")
    timer = QTimer(holder)
    timer.setInterval(350)

    def _tick() -> None:
        holder._load_dots = (holder._load_dots + 1) % 4  # type: ignore[attr-defined]
        holder.setText("로딩" + "." * holder._load_dots)  # type: ignore[attr-defined]

    timer.timeout.connect(_tick)
    timer.start()
    holder._load_timer = timer  # type: ignore[attr-defined]


def _stop_loading_anim(holder: QLabel) -> None:
    timer = getattr(holder, "_load_timer", None)
    if timer is not None:
        timer.stop()
        holder._load_timer = None  # type: ignore[attr-defined]


def fill_holder(holder: QLabel, thumb_cache, image_path, px: int) -> None:
    """(UI 스레드) 캐시된 썸네일을 holder 에 그린다. 실패 시 '이미지 없음'."""
    _stop_loading_anim(holder)  # 로딩 애니메이션 정지 후 실제 이미지로 교체
    path = thumb_cache.get_full_thumbnail(image_path, max_size=px) \
        if thumb_cache is not None else None
    if path is not None:
        pix = QPixmap(str(path))
        if not pix.isNull():
            holder.setPixmap(
                pix.scaled(px, int(px * 0.78), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            )
            holder.setToolTip(str(image_path))
            return
    holder.setText("이미지 없음")


def load_thumb_holder(thumb_cache, image_path, px: int, defer: bool = False) -> QLabel:
    """썸네일 QLabel. defer=True 면 '로딩…' 애니메이션을 두고 나중에 fill_holder 로 채운다."""
    holder = _blank_holder(px)
    if defer:
        _start_loading_anim(holder)
        return holder
    fill_holder(holder, thumb_cache, image_path, px)
    return holder


class ClickThumb(QWidget):
    """썸네일 홀더를 감싸 클릭 시 원본 뷰어를 여는 래퍼."""

    def __init__(self, holder: QLabel, record, open_viewer, parent=None):
        super().__init__(parent)
        self._record = record
        self._open_viewer = open_viewer
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(holder)
        self.setCursor(Qt.PointingHandCursor)
        # 우클릭 → 경로 복사·파일/폴더 열기
        if record is not None:
            attach_image_context_menu(self, lambda: record.image_path)

    def mousePressEvent(self, event):  # noqa: N802
        if event.button() == Qt.LeftButton and self._record is not None:
            self._open_viewer(self._record)


class ClusterMembersPopup(QDialog):
    """클러스터에 묶인 defect 사진 전체를 가로(줄바꿈)로 보여주는 작은 팝업."""

    def __init__(self, records: list, layer: str, thumb_cache, open_viewer, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"{layer} — 묶인 defect {len(records)}개")
        self.setMinimumWidth(520)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        cap = QLabel("근접해 하나로 묶인 defect (클릭=원본)")
        cap.setObjectName("dim")
        outer.addWidget(cap)
        host = QWidget()
        flow = FlowLayout(host, margin=0, h_spacing=8, v_spacing=8)
        for rec in records:
            thumb = ClickThumb(load_thumb_holder(thumb_cache, rec.image_path, 150),
                               rec, open_viewer)
            flow.addWidget(thumb)
        outer.addWidget(host)


class ClusteredThumb(QWidget):
    """대표 썸네일 + layer 배지 + (클러스터면) 좌하단 '+n' 버튼.

    대표 클릭 → 원본 확대. '+n' 클릭 → 묶인 defect 전체 팝업.
    """

    def __init__(self, cluster: Cluster, layer: str, is_base: bool,
                 thumb_cache, open_viewer, px: int, parent=None, defer: bool = False):
        super().__init__(parent)
        self._cluster = cluster
        self._layer = layer
        self._thumb_cache = thumb_cache
        self._open_viewer = open_viewer
        rep = cluster.representative
        # 비동기(지연) 로딩용으로 holder·경로·크기를 노출한다.
        self.rep_path = rep.image_path
        self._px = px

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        holder = load_thumb_holder(thumb_cache, rep.image_path, px, defer=defer)
        self.holder = holder
        holder.setCursor(Qt.PointingHandCursor)
        # 대표 클릭 → 뷰어
        holder.mousePressEvent = self._on_rep_click  # type: ignore[assignment]
        # 우클릭 → 경로 복사·파일/폴더 열기(메인 그리드와 동일)
        attach_image_context_menu(
            holder, lambda: self._cluster.representative.image_path
        )
        # layer 배지
        badge = QLabel(("★ " + layer) if is_base else layer, holder)
        badge.setObjectName("layerBadgeBase" if is_base else "layerBadge")
        badge.adjustSize()
        badge.move(5, 5)
        badge.show()
        # '+n' 오버레이 버튼(클러스터 여분)
        if cluster.extra_count > 0:
            more = QPushButton(f"+{cluster.extra_count}", holder)
            more.setObjectName("mini")
            more.setToolTip("이 자리에 근접해 하나로 묶인 defect 을 모두 봅니다.")
            more.setCursor(Qt.PointingHandCursor)
            more.adjustSize()
            more.move(5, holder.height() - more.height() - 5)
            more.clicked.connect(self._on_more)
            more.show()
        lay.addWidget(holder)

    def fill(self) -> None:
        """지연 로딩(defer)한 대표 썸네일을 실제 이미지로 채운다(UI 스레드)."""
        fill_holder(self.holder, self._thumb_cache, self.rep_path, self._px)

    def _on_rep_click(self, event):
        if event.button() == Qt.LeftButton:
            self._open_viewer(self._cluster.representative)

    def _on_more(self) -> None:
        popup = ClusterMembersPopup(
            self._cluster.members, self._layer, self._thumb_cache, self._open_viewer, self
        )
        popup.exec()


# =============================================================================
# app/ui/compare_grid.py   [#31]
# =============================================================================
"""Layer 비교 그리드 (문서 Section 8.4).

layer별 이미지를 LYA4/LYB4 ... 형태로 배치한다. 기준 Layer 이미지는 강조 + "기준" 표기.
기준 사진 변경 시 비교 Layer 이미지들은 빠른 Fade 로 갱신된다.
"""







class LayerCell(QFrame):
    """단일 layer 이미지 셀 (제목 + 매칭 정보 + 이미지).

    이미지가 있을 때 클릭하면 record_clicked 로 현재 DefectRecord 를 알린다(원본 확대 보기).
    """

    record_clicked = Signal(object)  # DefectRecord
    cluster_clicked = Signal(object)  # 근접중복 '+n' 클릭 → 묶인 base 목록(list[DefectRecord])

    def __init__(
        self,
        layer: str,
        is_base: bool,
        loader: Optional[ImageLoader] = None,
        parent: Optional[QWidget] = None,
    ):
        super().__init__(parent)
        self.layer = layer
        self.is_base = is_base
        self._record: Optional[DefectRecord] = None
        self.setObjectName("cell")
        self._build()
        if loader is not None:
            self.image.set_loader(loader)
        self._apply_style(active=is_base)

    def _build(self) -> None:
        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 8, 10, 10)
        lay.setSpacing(6)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # 기준은 부드럽게(긴 fade), 비교는 빠른 fade
        self.image = FadeImageLabel(duration=320 if self.is_base else 180)
        self.image.setMinimumHeight(280)
        self.image.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Layer 이름 배지: 이미지 위에 floating (요구사항 4 — 사진 위 Layer명 표시)
        self.badge = QLabel(self.image)
        self.badge.setObjectName("layerBadgeBase" if self.is_base else "layerBadge")
        self.badge.setText(f"★ {self.layer} (기준)" if self.is_base else self.layer)
        self.badge.adjustSize()
        self.badge.move(10, 10)

        # 근접중복 '+n' 배지(기준 셀에서만) — 클릭 시 묶인 defect 전체 보기.
        self._cluster_members: list = []
        self.more_badge = QPushButton("", self.image)
        self.more_badge.setObjectName("mini")
        self.more_badge.setCursor(Qt.PointingHandCursor)
        self.more_badge.setToolTip("이 자리에 근접(<50)해 하나로 묶인 defect 을 모두 봅니다.")
        self.more_badge.clicked.connect(self._emit_cluster)
        self.more_badge.hide()

        # 이미지 아래 작은 진단/부가정보 줄(wafer명 노출 X — 상세는 tooltip)
        self.info = QLabel("")
        self.info.setObjectName("diag")
        self.info.setAlignment(Qt.AlignCenter)

        lay.addWidget(self.image, 1)
        lay.addWidget(self.info)

    def _apply_style(self, active: bool) -> None:
        if self.is_base:
            self.setStyleSheet(
                f"QFrame#cell {{ background:{BG_ELEV};"
                f" border:2px solid {BASE_GLOW}; border-radius:10px; }}"
            )
        elif active:
            self.setStyleSheet(
                f"QFrame#cell {{ background:{BG_PANEL};"
                f" border:1px solid {MATCH}; border-radius:10px; }}"
            )
        else:
            self.setStyleSheet(
                f"QFrame#cell {{ background:{BG_PANEL};"
                f" border:1px solid {NEON_SOFT}; border-radius:10px; }}"
            )

    def _set_info(self, text: str, *, warn: bool = False) -> None:
        """진단/부가정보 줄을 갱신하고 QSS objectName 을 다시 적용한다."""
        self.info.setText(text)
        self.info.setObjectName("diagWarn" if warn else "diag")
        self.info.style().unpolish(self.info)
        self.info.style().polish(self.info)

    def show_record(
        self, rec: Optional[DefectRecord], info: str, matched: bool, *, warn: bool = False
    ) -> None:
        self._set_record(rec)
        if rec is not None:
            self.image.show_path(rec.image_path, animated=not self.is_base)
        else:
            self.image.show_message("")
        self._set_info(info, warn=warn)
        self.badge.raise_()
        if not self.is_base:
            self._apply_style(active=matched)

    def show_base(self, rec: DefectRecord, extra: int = 0, members: Optional[list] = None) -> None:
        self._set_record(rec)
        self.image.show_path(rec.image_path, animated=True)
        # wafer명은 노출하지 않고 die 위치만 간단히(상세는 tooltip)
        info = f"die ({rec.col}, {rec.row})"
        if extra:
            info += f" · 근접중복 +{extra}"
        self._set_info(info)
        self.badge.raise_()
        # 근접중복 '+n' 배지 갱신(좌하단).
        self._cluster_members = list(members or [])
        if extra > 0:
            self.more_badge.setText(f"+{extra}")
            self.more_badge.adjustSize()
            self.more_badge.move(10, self.image.height() - self.more_badge.height() - 10)
            self.more_badge.show()
            self.more_badge.raise_()
        else:
            self.more_badge.hide()

    def _emit_cluster(self) -> None:
        if self._cluster_members:
            self.cluster_clicked.emit(self._cluster_members)

    def resizeEvent(self, event):  # noqa: N802
        super().resizeEvent(event)
        # 이미지 크기 변화에 맞춰 '+n' 배지를 좌하단에 유지.
        if self.more_badge.isVisible():
            self.more_badge.move(10, self.image.height() - self.more_badge.height() - 10)

    def _set_record(self, rec: Optional[DefectRecord]) -> None:
        self._record = rec
        if rec is not None:
            self.setCursor(Qt.PointingHandCursor)
            tip = (
                f"{self.layer} · wafer {rec.wafer_id} · die({rec.col},{rec.row})\n"
                f"pos {rec.position_key}"
            )
            if rec.defect_name:
                tip += f" · {rec.defect_name}"
            tip += f"\n{rec.image_path}\n\n클릭하면 원본을 크게 봅니다"
            self.setToolTip(tip)
            self.image.setToolTip(tip)
        else:
            self.setCursor(Qt.ArrowCursor)
            self.setToolTip("")
            self.image.setToolTip("")

    def mousePressEvent(self, event):  # noqa: N802
        if event.button() == Qt.LeftButton and self._record is not None:
            self.record_clicked.emit(self._record)
        super().mousePressEvent(event)

    def contextMenuEvent(self, event):  # noqa: N802
        if self._record is None:
            return
        rec = self._record
        path = str(rec.image_path)
        menu = QMenu(self)
        menu.addAction("경로 복사", lambda: QGuiApplication.clipboard().setText(path))
        menu.addAction(
            "파일 열기",
            lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)),
        )
        menu.addAction(
            "폴더 열기",
            lambda: QDesktopServices.openUrl(
                QUrl.fromLocalFile(str(_Path(path).parent))
            ),
        )
        menu.exec(event.globalPos())


class CompareGrid(QWidget):
    """layer 배치 그리드 컨테이너."""

    image_clicked = Signal(object)  # DefectRecord
    base_cluster_clicked = Signal(object)  # 근접중복 '+n' → 묶인 base 목록

    def __init__(
        self, loader: Optional[ImageLoader] = None, parent: Optional[QWidget] = None
    ):
        super().__init__(parent)
        self._grid = QGridLayout(self)
        self._grid.setContentsMargins(0, 0, 0, 0)
        self._grid.setSpacing(10)
        self._cells: dict[str, LayerCell] = {}
        self._layer_order: list[str] = []  # 셀 배치 순서(기준 우선, 이후 grid 순서)
        self._base_layer: str = ""
        self._loader = loader

    def build_layout(
        self, grid: list[list[Optional[str]]], base_layer: str
    ) -> None:
        """layer 배치(grid)에 따라 셀을 재구성하고 부드럽게 페이드 인한다."""
        # 기존 제거
        while self._grid.count():
            item = self._grid.takeAt(0)
            w = item.widget()
            if w is not None:
                w.setParent(None)
                w.deleteLater()
        self._cells.clear()
        self._layer_order = []
        self._base_layer = base_layer

        for r, row in enumerate(grid):
            for c, layer in enumerate(row):
                if not layer:
                    continue
                cell = LayerCell(
                    layer, is_base=(layer == base_layer), loader=self._loader
                )
                cell.record_clicked.connect(self.image_clicked)
                cell.cluster_clicked.connect(self.base_cluster_clicked)
                self._cells[layer] = cell
                self._layer_order.append(layer)
                self._grid.addWidget(cell, r, c)

    def _repack(self, visible_layers: list[str]) -> None:
        """보이는 layer 셀만 2열 그리드에 빈칸 없이 다시 배치하고 나머지는 숨긴다.

        위젯은 삭제하지 않고 재사용(위치만 이동)하므로 탐색 중 깜빡임이 적다.
        """
        while self._grid.count():
            self._grid.takeAt(0)
        visible_set = set(visible_layers)
        for layer, cell in self._cells.items():
            if layer not in visible_set:
                cell.setVisible(False)
        for i, layer in enumerate(visible_layers):
            cell = self._cells.get(layer)
            if cell is None:
                continue
            cell.setVisible(True)
            self._grid.addWidget(cell, i // 2, i % 2)

    def update_for_base(
        self, item: BaseDefectMatches, compare_layers: list[str]
    ) -> None:
        """기준 defect 변경 시 셀 갱신 — 기준 + 매칭된 비교 layer 셀만 압축 배치.

        매칭 없는 비교 layer 셀은 숨겨 빈칸을 없애고, 보이는 셀을 좌상단부터 2열로 채운다.
        """
        base = item.base
        visible: list[str] = []
        if self._base_layer in self._cells:
            cluster = getattr(item, "base_cluster", None)
            extra = getattr(cluster, "extra_count", 0) or 0
            members = list(getattr(cluster, "members", []) or [])
            self._cells[self._base_layer].show_base(base, extra=extra, members=members)
            visible.append(self._base_layer)

        ordered_compares = [l for l in self._layer_order if l != self._base_layer]
        for layer in ordered_compares:
            if layer not in compare_layers:
                continue
            cell = self._cells.get(layer)
            if cell is None:
                continue
            mr = item.for_layer(layer)
            if mr and mr.is_match and mr.matched is not None:
                info = f"매칭 O · 거리 {mr.distance:.1f} µm"
                if mr.ambiguous:
                    info += " · ⚠동률 후보"
                cell.show_record(mr.matched, info, matched=True, warn=mr.ambiguous)
                visible.append(layer)

        self._repack(visible)

    @staticmethod
    def _diag_text(mr) -> tuple[str, bool]:
        """매칭 실패 사유를 사용자 문구로 변환 (text, warn)."""
        reason = mr.reason
        if reason == NoMatchReason.COORD_FAIL:
            return f"좌표 추출 실패({mr.failed_in_die}장)", True
        if reason == NoMatchReason.OVER_TOLERANCE and mr.nearest_distance is not None:
            return f"허용오차 초과 · 최근접 {mr.nearest_distance:.1f}", True
        return "이 layer에 같은 die 사진 없음", False

    def show_empty(self, message: str) -> None:
        # 모든 셀을 다시 보이게 배치한 뒤 안내 메시지를 표시한다.
        self._repack(list(self._layer_order))
        for cell in self._cells.values():
            cell.image.show_message(message)
            cell.info.setText("")


# =============================================================================
# app/ui/export_dialog.py   [#32]
# =============================================================================
"""결과 출력 트레이 다이얼로그 (문서 Section 8.7 재설계).

'출력에 추가'로 담아 둔 기준 사진(트레이)들을 썸네일 카드로 확인하고, 개별 제거·전체
비우기 후 Excel 로 출력한다. 카드 그리드 패턴은 nomatch_gallery 와 동일한 톤으로 맞춘다.
"""





export_dialog__COLUMNS = 3
export_dialog__THUMB_PX = 180  # 카드 썸네일 크기(크게)


class ExportTrayDialog(QDialog):
    """출력 트레이(담긴 기준 사진)를 카드로 관리하고 Excel 출력을 확정하는 다이얼로그."""

    def __init__(
        self,
        entries: list[BaseDefectMatches],
        thumb_cache: Optional[ThumbnailCache] = None,
        all_matched: Optional[list[BaseDefectMatches]] = None,
        all_layers_provider: Optional[Callable[[], list[BaseDefectMatches]]] = None,
        parent: Optional[QWidget] = None,
    ):
        super().__init__(parent)
        # 담긴 항목(스냅샷). base image_path 로 중복 제거하며 유지한다.
        self._kept: list[BaseDefectMatches] = []
        self._keys: set[str] = set()
        for m in entries:
            self._add(m)
        # 이번 LOT 의 매칭 있는 기준 사진(전체 추가 버튼용).
        self._all_matched = list(all_matched or [])
        # 모든 layer 를 기준으로 매치를 합쳐 담는 공급자(느릴 수 있어 클릭 시 계산).
        self._all_layers_provider = all_layers_provider
        self._thumb_cache = thumb_cache
        self._wants_export = False  # True=Excel 출력, False=저장만(확인)
        self.setWindowTitle("결과 출력 — 담은 사진")
        self.setMinimumSize(620, 540)
        self._build()
        self._populate()
        # 무거운 '모든 매치(기준 없이)' 계산 동안 다이얼로그 위에 로딩+진행도 표시.
        self._busy = BusyOverlay(self)

    @staticmethod
    def _key(item: BaseDefectMatches) -> str:
        return str(item.base.image_path)

    def _add(self, item: BaseDefectMatches) -> bool:
        k = self._key(item)
        if k in self._keys:
            return False
        self._keys.add(k)
        self._kept.append(item)
        return True

    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 14, 16, 12)
        outer.setSpacing(10)

        head = QHBoxLayout()
        self.title = QLabel("")
        self.title.setObjectName("title")
        head.addWidget(self.title)
        head.addStretch()
        self.btn_add_all = QPushButton("기준 layer 매치 전체 추가")
        self.btn_add_all.setObjectName("mini")
        self.btn_add_all.setToolTip(
            "현재 선택된 기준 layer 기준으로, 매칭이 있는 기준 사진을 모두 담습니다."
        )
        self.btn_add_all.clicked.connect(self._add_all_matched)
        self.btn_add_all.setEnabled(bool(self._all_matched))
        head.addWidget(self.btn_add_all)
        if self._all_layers_provider is not None:
            self.btn_add_all_layers = QPushButton("모든 매치(기준 없이) 추가")
            self.btn_add_all_layers.setObjectName("mini")
            self.btn_add_all_layers.setToolTip(
                "모든 layer 를 각각 기준으로 매칭해, 어느 layer 에서든 매치된 defect 을 "
                "모두 담습니다(중복 제거). layer 수만큼 재계산해 잠시 걸릴 수 있습니다."
            )
            self.btn_add_all_layers.clicked.connect(self._add_all_layers)
            head.addWidget(self.btn_add_all_layers)
        self.btn_clear = QPushButton("전체 비우기")
        self.btn_clear.setObjectName("mini")
        self.btn_clear.setToolTip("담은 사진을 모두 뺍니다.")
        self.btn_clear.clicked.connect(self._clear_all)
        head.addWidget(self.btn_clear)
        outer.addLayout(head)

        hint = QLabel("담은 사진만 Excel 로 출력됩니다. 각 카드의 ✕ 로 개별 제거할 수 있습니다.")
        hint.setObjectName("dim")
        outer.addWidget(hint)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setFrameShape(QFrame.NoFrame)
        # 흰 배경 제거 → 다이얼로그(테마) 배경이 비치게.
        self._scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self._scroll.viewport().setAutoFillBackground(False)
        self._host = QWidget()
        self._host.setStyleSheet("background: transparent;")
        self._grid = QGridLayout(self._host)
        self._grid.setContentsMargins(4, 4, 4, 4)
        self._grid.setHorizontalSpacing(10)
        self._grid.setVerticalSpacing(10)
        self._grid.setAlignment(Qt.AlignTop)
        self._scroll.setWidget(self._host)
        outer.addWidget(self._scroll, 1)

        self._empty = QLabel("담은 사진이 없습니다. 위 '기준 layer 매치 전체 추가'로 담거나 창을 닫고 '＋ 출력에 추가'로 담아 주세요.")
        self._empty.setObjectName("dim")
        self._empty.setWordWrap(True)
        self._empty.setAlignment(Qt.AlignCenter)
        self._empty.setVisible(False)
        outer.addWidget(self._empty)

        # 하단 바 — 취소 + 확인(Excel 출력). QDialogButtonBox 는 빈 트레이일 때 OK 가
        # 흐려져 '없는 것'처럼 보였으므로, 명시적 버튼으로 확인을 항상 뚜렷하게 노출한다.
        bottom = QHBoxLayout()
        bottom.addStretch(1)
        btn_cancel = QPushButton("취소")
        btn_cancel.setToolTip("변경을 취소하고 닫습니다(담은 목록을 저장하지 않음).")
        btn_cancel.clicked.connect(self.reject)
        bottom.addWidget(btn_cancel)
        # 확인 = 담은 목록만 저장하고 닫는다(Excel 출력은 나중에).
        self.btn_ok = QPushButton("확인")
        self.btn_ok.setToolTip("담은 목록을 저장하고 닫습니다(지금 Excel 출력은 하지 않음).")
        self.btn_ok.clicked.connect(self._on_ok)
        bottom.addWidget(self.btn_ok)
        self.btn_export = QPushButton("Excel 출력")
        self.btn_export.setObjectName("primary")
        self.btn_export.setToolTip("담은 사진을 지금 Excel 파일로 출력합니다.")
        self.btn_export.setDefault(True)
        self.btn_export.clicked.connect(self._on_export)
        bottom.addWidget(self.btn_export)
        outer.addLayout(bottom)

    def _on_ok(self) -> None:
        """확인 — 담은 상태를 저장(트레이 반영)하고 닫는다. 출력은 하지 않음."""
        self._wants_export = False
        self.accept()

    def _on_export(self) -> None:
        """Excel 출력 — 저장 + 출력 흐름으로 진행."""
        self._wants_export = True
        self.accept()

    def wants_export(self) -> bool:
        """확인(False) vs Excel 출력(True) 구분."""
        return self._wants_export

    def _clear_grid(self) -> None:
        while self._grid.count():
            item = self._grid.takeAt(0)
            w = item.widget()
            if w is not None:
                w.setParent(None)
                w.deleteLater()

    def _populate(self) -> None:
        self._clear_grid()
        for i, item in enumerate(self._kept):
            self._grid.addWidget(self._make_card(item), i // export_dialog__COLUMNS, i % export_dialog__COLUMNS)
        shown = len(self._kept)
        self.title.setText(f"담은 사진 — 총 {shown}장")
        self._empty.setVisible(shown == 0)
        self.btn_export.setEnabled(shown > 0)
        self.btn_clear.setEnabled(shown > 0)

    def _make_card(self, item: BaseDefectMatches) -> QWidget:
        base = item.base
        px = export_dialog__THUMB_PX
        card = QFrame()
        card.setFixedWidth(px + 20)
        card.setObjectName("cell")
        card.setStyleSheet(
            f"QFrame#cell {{ background:{BG_ELEV};"
            f" border:1px solid {NEON_SOFT}; border-radius:8px; }}"
        )
        lay = QVBoxLayout(card)
        lay.setContentsMargins(6, 6, 6, 8)
        lay.setSpacing(4)

        # 큰 썸네일 + 그 위에 오버레이된 제거(✕) 버튼
        thumb = QLabel()
        thumb.setAlignment(Qt.AlignCenter)
        thumb.setFixedSize(px, int(px * 0.78))
        thumb.setStyleSheet(
            f"background:{BG}; border:1px solid {NEON_SOFT};"
            f" border-radius:6px; color:{TEXT_DIM}; font-size:10px;"
        )
        if self._thumb_cache is not None:
            path = self._thumb_cache.get_full_thumbnail(base.image_path, max_size=px)
            if path is not None:
                pix = QPixmap(str(path))
                if not pix.isNull():
                    thumb.setPixmap(pix.scaled(
                        px, int(px * 0.78), Qt.KeepAspectRatio, Qt.SmoothTransformation
                    ))
                else:
                    thumb.setText("이미지 없음")
            else:
                thumb.setText("이미지 없음")
        # ✕ 오버레이: 썸네일 우상단, 대비가 큰 반투명 어두운 배경 + 밝은 X.
        btn_x = QPushButton("✕", thumb)
        btn_x.setFixedSize(24, 24)
        btn_x.setCursor(Qt.PointingHandCursor)
        btn_x.setToolTip("이 사진을 출력 목록에서 뺍니다.")
        btn_x.setStyleSheet(
            "QPushButton { color:#ffffff; background:rgba(17,21,28,0.72);"
            " border:1px solid rgba(255,255,255,0.55); border-radius:12px;"
            " font-size:13px; font-weight:700; padding:0; }"
            "QPushButton:hover { background:#b00020; border:1px solid #ffffff; }"
        )
        btn_x.move(px - 28, 4)
        btn_x.clicked.connect(lambda _=0, k=self._key(item): self._remove(k))
        lay.addWidget(thumb, alignment=Qt.AlignHCenter)

        n_match = sum(1 for r in item.results if r.is_match)
        cap = QLabel(
            f"wafer {base.wafer_id} · die({base.col},{base.row})\n"
            f"매칭 {n_match}/{len(item.results)} · pos {base.position_key}"
        )
        cap.setObjectName("dim")
        cap.setStyleSheet("font-size:10px;")
        cap.setWordWrap(True)
        cap.setAlignment(Qt.AlignCenter)
        cap.setFixedWidth(px)
        lay.addWidget(cap, alignment=Qt.AlignHCenter)
        return card

    def _remove(self, key: str) -> None:
        self._kept = [m for m in self._kept if self._key(m) != key]
        self._keys.discard(key)
        self._populate()

    def _clear_all(self) -> None:
        self._kept = []
        self._keys = set()
        self._populate()

    def _add_all_matched(self) -> None:
        added = 0
        for m in self._all_matched:
            if self._add(m):
                added += 1
        self._populate()

    def _add_all_layers(self) -> None:
        """모든 layer 를 기준으로 한 매치를 공급자에서 받아 담는다(중복 제거).

        계산이 무거우므로 로딩 오버레이 + layer 단위 진행도를 표시한다.
        """
        if self._all_layers_provider is None:
            return
        from PySide6.QtWidgets import QApplication

        self._busy.start("모든 매치 계산 중", determinate=True)
        QApplication.processEvents()

        def _progress(cur: int, total: int) -> None:
            self._busy.set_message(f"모든 매치 계산 중  ({cur}/{total} layer)")
            self._busy.set_progress(cur, total)
            QApplication.processEvents()  # 동기 루프 중에도 진행바가 갱신되도록

        try:
            items = self._all_layers_provider(_progress) or []
        finally:
            self._busy.stop()
        for m in items:
            self._add(m)
        self._populate()

    def selected(self) -> list[BaseDefectMatches]:
        """최종 출력 대상 스냅샷 목록(담긴 순서 유지)."""
        return list(self._kept)


# =============================================================================
# app/ui/nomatch_gallery.py   [#33]
# =============================================================================
"""미매칭(기준 layer 의 defect 과 어떤 비교 layer 와도 매칭 안 된) 사진 갤러리.

후보에서 제외된 '매칭 없음' 기준 사진들을 한곳에 모아 썸네일 + 사유로 보여준다.
사유는 매칭 엔진이 이미 계산한 MatchResult.reason(NoMatchReason)을 그대로 재사용한다.
썸네일을 클릭하면 본문 탐색을 그 기준으로 이동한다(on_navigate 콜백).
"""





nomatch_gallery__COLUMNS = 4

# 사유별 (표시명, 색). 트리아지 우선순위는 _dominant 에서 별도로 정한다.
_REASON_META = {
    NoMatchReason.OVER_TOLERANCE: ("허용오차 초과", WARN),
    NoMatchReason.COORD_FAIL: ("좌표 추출 실패", NOMATCH),
    NoMatchReason.NO_DIE_PHOTO: ("같은 die 사진 없음", TEXT_DIM),
}
# 트리아지 우선순위(높을수록 먼저) — 거의 매칭된 것(허용오차 초과)을 가장 눈에 띄게.
_PRIORITY = [
    NoMatchReason.OVER_TOLERANCE,
    NoMatchReason.COORD_FAIL,
    NoMatchReason.NO_DIE_PHOTO,
]


def _layer_diag(mr) -> str:
    """한 비교 layer 의 미매칭 사유 짧은 문구(그리드와 동일 표현 재사용)."""

    return CompareGrid._diag_text(mr)[0]


def _dominant(item) -> Optional[NoMatchReason]:
    """기준 1개의 대표 미매칭 사유(우선순위가 가장 높은 것)."""
    present = {r.reason for r in item.results if not r.is_match}
    for reason in _PRIORITY:
        if reason in present:
            return reason
    return None


class NoMatchGalleryDialog(QDialog):
    """미매칭 기준 사진 갤러리(사유 표기 + 사유별 필터 + 클릭 이동)."""

    def __init__(
        self,
        entries: list[tuple[int, object]],
        thumb_cache,
        on_navigate: Callable[[int], None],
        parent: Optional[QWidget] = None,
    ):
        super().__init__(parent)
        self._entries = entries  # [(base_index, BaseDefectMatches), ...]
        self._thumb_cache = thumb_cache
        self._on_navigate = on_navigate
        self.setWindowTitle("매칭 없는 기준 사진")
        self.setMinimumSize(560, 460)
        self._build()
        self._populate()

    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 14, 16, 12)
        outer.setSpacing(10)

        head = QHBoxLayout()
        title = QLabel(f"매칭 없는 기준 사진 — 총 {len(self._entries)}장")
        title.setObjectName("title")
        head.addWidget(title)
        head.addStretch()
        head.addWidget(QLabel("사유:"))
        self.cmb_reason = QComboBox()
        self.cmb_reason.addItem("전체", None)
        for reason in _PRIORITY:
            self.cmb_reason.addItem(_REASON_META[reason][0], reason.value)
        self.cmb_reason.currentIndexChanged.connect(self._populate)
        head.addWidget(self.cmb_reason)
        outer.addLayout(head)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setFrameShape(QFrame.NoFrame)
        self._host = QWidget()
        self._grid = QGridLayout(self._host)
        self._grid.setContentsMargins(4, 4, 4, 4)
        self._grid.setHorizontalSpacing(10)
        self._grid.setVerticalSpacing(10)
        self._grid.setAlignment(Qt.AlignTop)
        self._scroll.setWidget(self._host)
        outer.addWidget(self._scroll, 1)

        self._empty = QLabel("해당하는 미매칭 기준 사진이 없습니다.")
        self._empty.setObjectName("dim")
        self._empty.setAlignment(Qt.AlignCenter)
        self._empty.setVisible(False)
        outer.addWidget(self._empty)

        btn = QPushButton("닫기")
        btn.setObjectName("primary")
        btn.clicked.connect(self.accept)
        outer.addWidget(btn, alignment=Qt.AlignRight)

    def _clear_grid(self) -> None:
        while self._grid.count():
            item = self._grid.takeAt(0)
            w = item.widget()
            if w is not None:
                w.setParent(None)
                w.deleteLater()

    def _populate(self) -> None:
        self._clear_grid()
        want = self.cmb_reason.currentData()
        shown = 0
        for idx, item in self._entries:
            present = {r.reason.value for r in item.results if not r.is_match}
            if want is not None and want not in present:
                continue
            self._grid.addWidget(self._make_cell(idx, item), shown // nomatch_gallery__COLUMNS, shown % nomatch_gallery__COLUMNS)
            shown += 1
        self._empty.setVisible(shown == 0)

    def _make_cell(self, idx: int, item) -> QWidget:
        cell = QFrame()
        cell.setFixedWidth(128)
        lay = QVBoxLayout(cell)
        lay.setContentsMargins(2, 2, 2, 6)
        lay.setSpacing(5)

        thumb = ClickableThumb(idx)
        base = item.base
        thumb.set_caption(f"{base.wafer_id}\n({base.col},{base.row})")
        path = self._thumb_cache.get_full_thumbnail(base.image_path, max_size=120)
        thumb.set_image(str(path) if path else None)
        thumb.set_status("none")
        thumb.clicked.connect(self._on_thumb_clicked)
        lay.addWidget(thumb, alignment=Qt.AlignHCenter)

        # 대표 사유 색 태그
        dom = _dominant(item)
        if dom is not None:
            label, color = _REASON_META[dom]
            tag = QLabel(label)
            tag.setAlignment(Qt.AlignCenter)
            tag.setStyleSheet(f"color:{color}; font-size:10px; font-weight:700;")
            lay.addWidget(tag)

        # layer 별 사유 요약(툴팁 + 작은 라벨)
        per_layer = [f"{r.compare_layer}: {_layer_diag(r)}" for r in item.results if not r.is_match]
        summary = QLabel(" · ".join(per_layer))
        summary.setObjectName("dim")
        summary.setStyleSheet("font-size:9px;")
        summary.setWordWrap(True)
        summary.setAlignment(Qt.AlignCenter)
        summary.setFixedWidth(120)
        if per_layer:
            summary.setToolTip("\n".join(per_layer))
        lay.addWidget(summary)
        return cell

    def _on_thumb_clicked(self, index: int) -> None:
        self._on_navigate(index)


# =============================================================================
# app/ui/thumbnail_strip.py   [#34]
# =============================================================================
"""상단 기준 썸네일 스트립 (문서 Section 8.6).

기준 Layer 사진들의 (중앙 10% 확대) 썸네일을 가로로 나열한다.
클릭 시 해당 사진을 기준 defect 로 설정하고, 현재 선택 썸네일을 강조한다.

가로 휠을 쓰지 않도록 **세로 휠 → 가로 스크롤** 로 매핑한다(사용성).
"""






class ThumbnailStrip(QScrollArea):
    """기준 사진 썸네일 가로 스트립."""

    thumb_clicked = Signal(int)

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setWidgetResizable(True)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setFixedHeight(120)
        self.setToolTip("세로 휠로 좌우 스크롤 · 클릭하면 기준 사진 변경")
        # viewport 기본 흰색 제거 → 뒤의 패널(BG_PANEL)이 비치게
        self.setFrameShape(QScrollArea.NoFrame)
        self.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self.viewport().setAutoFillBackground(False)
        self._container = QWidget()
        self._container.setObjectName("stripHost")
        self._container.setAutoFillBackground(False)
        self._container.setStyleSheet("#stripHost { background: transparent; }")
        self._layout = QHBoxLayout(self._container)
        self._layout.setContentsMargins(8, 8, 8, 8)
        self._layout.setSpacing(8)
        self._layout.addStretch()
        self.setWidget(self._container)
        self._thumbs: list[ClickableThumb] = []
        self._current = -1
        # 부드러운 가로 스크롤 애니메이션
        self._scroll_anim = QPropertyAnimation(self.horizontalScrollBar(), b"value", self)
        self._scroll_anim.setDuration(220)
        self._scroll_anim.setEasingCurve(QEasingCurve.OutCubic)

    def _animate_scroll_to(self, target: int) -> None:
        """수평 스크롤바를 target 값으로 부드럽게 이동."""
        bar = self.horizontalScrollBar()
        target = max(bar.minimum(), min(bar.maximum(), target))
        if target == bar.value():
            return
        self._scroll_anim.stop()
        self._scroll_anim.setStartValue(bar.value())
        self._scroll_anim.setEndValue(target)
        self._scroll_anim.start()

    def clear(self) -> None:
        for t in self._thumbs:
            t.setParent(None)
            t.deleteLater()
        self._thumbs.clear()
        self._current = -1

    def set_items(
        self,
        captions: list[str],
        tooltips: Optional[list[str]] = None,
        on_progress=None,
    ) -> None:
        """기준 record 개수만큼 썸네일 placeholder 를 만든다.

        on_progress 가 주어지면 일정 개수마다 호출해(예: 로딩 스피너 pump) 많은
        썸네일을 만드는 동안에도 UI 가 멈춘 것처럼 보이지 않게 한다.
        """
        self.clear()
        for i, cap in enumerate(captions):
            thumb = ClickableThumb(i)
            thumb.set_caption(cap)
            if tooltips and i < len(tooltips):
                thumb.set_tooltip(tooltips[i])
            thumb.clicked.connect(self.thumb_clicked)
            # stretch 앞에 삽입
            self._layout.insertWidget(self._layout.count() - 1, thumb)
            self._thumbs.append(thumb)
            # 썸네일이 많을 때 주기적으로 이벤트 루프에 양보 → 스피너 애니메이션 유지
            if on_progress is not None and (i & 15) == 15:
                on_progress()

    def set_thumbnail(self, index: int, path: str) -> None:
        if 0 <= index < len(self._thumbs):
            self._thumbs[index].set_image(path)

    def set_status_marks(self, statuses: list[str]) -> None:
        """각 썸네일에 매칭 상태 점을 표시(matched/none)."""
        for i, t in enumerate(self._thumbs):
            t.set_status(statuses[i] if i < len(statuses) else "matched")

    def set_visible_set(self, indices: Optional[list[int]]) -> None:
        """주어진 인덱스의 썸네일만 보이고 나머지는 숨긴다(후보 제외 반영).

        indices 가 None 이면 전부 표시. 인덱스는 기준 record 전체 기준(불변)이라
        클릭 시 emit 되는 index 도 그대로 유효하다.
        """
        if indices is None:
            for t in self._thumbs:
                t.setVisible(True)
            return
        sel = set(indices)
        for i, t in enumerate(self._thumbs):
            t.setVisible(i in sel)

    def set_current(self, index: int) -> None:
        if not (0 <= index < len(self._thumbs)):
            return
        for i, t in enumerate(self._thumbs):
            t.set_selected(i == index)
        self._current = index
        self._ensure_visible(index)

    def _ensure_visible(self, index: int) -> None:
        """선택 썸네일이 보이도록 부드럽게 가로 스크롤."""
        if not (0 <= index < len(self._thumbs)):
            return
        thumb = self._thumbs[index]
        bar = self.horizontalScrollBar()
        left = thumb.x()
        right = left + thumb.width()
        view_w = self.viewport().width()
        margin = 60
        cur = bar.value()
        if left - margin < cur:
            self._animate_scroll_to(left - margin)
        elif right + margin > cur + view_w:
            self._animate_scroll_to(right + margin - view_w)

    def wheelEvent(self, event):  # noqa: N802
        """세로 휠을 가로 스크롤로 변환 (가로 휠 불필요), 부드럽게 이동."""
        bar = self.horizontalScrollBar()
        delta = event.angleDelta().y()
        if delta == 0:
            delta = event.angleDelta().x()
        if delta != 0 and bar.maximum() > 0:
            # 진행 중 애니메이션의 목표값을 기준으로 누적 → 휠 연타도 매끄럽게
            anim_running = self._scroll_anim.state() == QPropertyAnimation.Running
            base = self._scroll_anim.endValue() if anim_running else bar.value()
            self._animate_scroll_to(int(base) - delta)
            event.accept()
        else:
            super().wheelEvent(event)


# =============================================================================
# app/ui/folder_picker.py   [#35]
# =============================================================================
"""자재(LOT) 폴더 선택 다이얼로그 — 브레드크럼 + 한 단계 목록 + 사이드바.

성능: 현재 디렉터리의 하위 폴더 한 단계만 os.scandir 로 나열한다(트리·워처·재귀·셸 아이콘
없음 → 네트워크 드라이브에서도 즉시). 명확함: 고른 폴더가 자재/layer/wafer 중 무엇인지,
유효한지(layer·wafer 개수)를 하단 배너에 실시간 표시한다(scanner.classify_selection 을
백그라운드로 실행). layer/wafer 를 골라도 선택 시 상위 자재 폴더로 보정한다.

편의: 최근 폴더·즐겨찾기·드라이브 사이드바, 이름 타이핑 필터. 원본은 읽기만 한다.
"""



folder_picker__NUM_RE = re.compile(r"(\d+)")


def natural_key(name: str) -> list:
    """자연 정렬 키 — 숫자 부분을 정수로 비교해 '2.'가 '10.'보다 앞에 오게 한다.

    예) 1., 2., …, 10., 11., …, 21. (사전식 1., 10., 11., 2. 방지)
    """
    return [int(t) if t.isdigit() else t.lower() for t in folder_picker__NUM_RE.split(name)]



# 검증 배너 종류별 색(테마 팔레트 재사용).
_BANNER_COLORS = {
    "material": (MATCH, "#10241b"),
    "layerwafer": (NEON, "#101a24"),
    "too_high": (WARN, "#241f10"),
    "unknown": (TEXT_DIM, BG_ELEV),
    "busy": (TEXT_DIM, BG_ELEV),
    "none": (TEXT_DIM, BG_ELEV),
}


def _subdir_count(path: Path) -> int:
    """path 바로 아래 폴더 수(싸게 os.scandir 1회)."""
    n = 0
    try:
        with os.scandir(path) as it:
            for e in it:
                if e.is_dir():
                    n += 1
    except OSError:
        return 0
    return n


def _first_subdir(path: Path) -> Optional[Path]:
    try:
        with os.scandir(path) as it:
            names = sorted((e.name for e in it if e.is_dir()), key=natural_key)
    except OSError:
        return None
    return path / names[0] if names else None


class _ValidateSignals(QObject):
    # token, kind, material_path, layer_count, wafer_count
    done = Signal(int, str, str, int, int)


class _ValidateWorker(QRunnable):
    """후보 폴더를 백그라운드로 판별한다(classify_selection + 개수 계산).

    UI 스레드를 막지 않도록 무거운 깊이 BFS 를 여기서 수행한다. 오래된 요청은 token 으로
    UI 에서 무시한다.
    """

    def __init__(self, token: int, path: str):
        super().__init__()
        self.token = token
        self.path = path
        self.signals = _ValidateSignals()

    @Slot()
    def run(self) -> None:
        kind, material = classify_selection(self.path)
        layers = wafers = 0
        try:
            if material is not None and kind in ("material", "layer", "wafer"):
                layers = _subdir_count(material)
                first = _first_subdir(material)
                if first is not None:
                    wafers = _subdir_count(first)
        except Exception:  # noqa: BLE001 - 개수는 부가 정보, 실패해도 판별은 전달
            layers = wafers = 0
        mat_str = str(material) if material is not None else ""
        self.signals.done.emit(self.token, kind, mat_str, layers, wafers)


class FolderPickerDialog(QDialog):
    """자재 폴더 선택기. `settings` 로 최근·즐겨찾기를 읽고 고정 토글 시 저장한다."""

    def __init__(self, settings, start_path: str, parent=None):
        super().__init__(parent)
        self.settings = settings
        self.setWindowTitle("자재(LOT) 폴더 선택")
        self.setMinimumSize(860, 560)
        self.resize(980, 640)

        self._pool = QThreadPool.globalInstance()
        self._cur: Path = self._safe_dir(start_path)
        self._history: list[Path] = []
        self._candidate: Optional[Path] = None
        # 마지막 검증 결과 캐시(candidate 경로 기준).
        self._valid_for: Optional[Path] = None
        self._valid_kind: str = ""
        self._valid_material: str = ""
        self._token = 0

        # '확인 중…' 배너 애니메이션 타이머(빌드 중 _set_banner 가 참조하므로 먼저 만든다).
        self._busy_base = ""
        self._busy_dots = 0
        self._busy_timer = QTimer(self)
        self._busy_timer.setInterval(350)
        self._busy_timer.timeout.connect(self._tick_busy)

        self._build_ui()
        self._reload_sidebar()
        self._go_to(self._cur, push=False)

    # ----------------------------------------------------------- helpers
    @staticmethod
    def _safe_dir(path: str) -> Path:
        try:
            p = Path(path)
            if p.exists() and p.is_dir():
                return p
        except OSError:
            pass
        return Path.home()

    def _list_subdirs(self, path: Path) -> list[str]:
        """한 단계 하위 폴더 이름만 나열(숨김 제외, 이름순). 네트워크에서도 가볍다."""
        try:
            with os.scandir(path) as it:
                names = [e.name for e in it if e.is_dir() and not e.name.startswith(".")]
        except OSError:
            return []
        return sorted(names, key=natural_key)

    # ----------------------------------------------------------- UI build
    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # 상단: 뒤로/위로 + 브레드크럼 + 경로 입력
        topbar = QHBoxLayout()
        topbar.setSpacing(6)
        self.btn_back = QPushButton("‹ 뒤로")
        self.btn_back.setFixedWidth(72)
        self.btn_back.clicked.connect(self._go_back)
        self.btn_up = QPushButton("↑ 위로")
        self.btn_up.setFixedWidth(72)
        self.btn_up.clicked.connect(self._go_up)
        self.btn_explorer = QPushButton("🗂 기본 탐색기")
        self.btn_explorer.setToolTip(
            "OS 기본 폴더 탐색기로 선택합니다.\n"
            "(네트워크 공유 등 트리에 아직 안 보이는 위치를 여기서 바로 고를 수 있습니다.)"
        )
        self.btn_explorer.clicked.connect(self._open_native_explorer)
        topbar.addWidget(self.btn_back)
        topbar.addWidget(self.btn_up)
        topbar.addWidget(self.btn_explorer)

        self._crumbs = QHBoxLayout()
        self._crumbs.setSpacing(2)
        crumb_host = QWidget()
        crumb_host.setLayout(self._crumbs)
        topbar.addWidget(crumb_host, 1)
        root.addLayout(topbar)

        self.ed_path = QLineEdit()
        self.ed_path.setPlaceholderText("경로를 붙여넣고 Enter — 예: \\\\server\\share\\LOT")
        self.ed_path.returnPressed.connect(self._on_path_entered)
        root.addWidget(self.ed_path)

        # 본문: 좌 사이드바 / 우 (필터 + 목록)
        body = QHBoxLayout()
        body.setSpacing(10)

        # 좌측: [Conder Scan 위치 지정 칸] + 폴더 구조 트리(지연 로딩) + 접이식 최근/즐겨찾기.
        left = QVBoxLayout()
        left.setSpacing(4)

        scan_box = QFrame()
        scan_box.setObjectName("panel")
        sb = QVBoxLayout(scan_box)
        sb.setContentsMargins(8, 6, 8, 8)
        sb.setSpacing(4)
        cap = QLabel(f"📌 {getattr(self.settings, 'scan_root_name', '') or '스캔 데이터'} 위치")
        cap.setStyleSheet(f"font-size:11px; font-weight:700; color:{NEON};")
        sb.addWidget(cap)
        row = QHBoxLayout()
        row.setSpacing(4)
        self.ed_scan_root = QLineEdit(getattr(self.settings, "scan_root_path", "") or "")
        self.ed_scan_root.setPlaceholderText("스캔 데이터 폴더 경로 — 설정하면 트리 맨 위에 고정")
        self.ed_scan_root.returnPressed.connect(self._apply_scan_root)
        row.addWidget(self.ed_scan_root, 1)
        btn_scan = QPushButton("찾기")
        btn_scan.setObjectName("mini")
        btn_scan.setFixedWidth(48)
        btn_scan.clicked.connect(self._pick_scan_root)
        row.addWidget(btn_scan)
        sb.addLayout(row)
        left.addWidget(scan_box)

        self.sidebar = QTreeWidget()
        self.sidebar.setHeaderHidden(True)
        self.sidebar.setFixedWidth(260)
        self.sidebar.setIndentation(12)  # 기본(~20)은 하위로 갈수록 여백 과다 → 축소
        self.sidebar.setRootIsDecorated(True)
        self.sidebar.itemClicked.connect(self._on_tree_clicked)
        self.sidebar.itemExpanded.connect(self._on_tree_expanded)
        left.addWidget(self.sidebar, 1)

        left_host = QWidget()
        left_host.setFixedWidth(260)
        left_host.setLayout(left)
        body.addWidget(left_host)

        right = QVBoxLayout()
        right.setSpacing(6)
        self.ed_filter = QLineEdit()
        self.ed_filter.setPlaceholderText("이 폴더 안에서 이름으로 찾기…")
        self.ed_filter.setClearButtonEnabled(True)
        self.ed_filter.textChanged.connect(self._apply_filter)
        right.addWidget(self.ed_filter)

        self.listw = QListWidget()
        self.listw.itemClicked.connect(self._on_item_clicked)
        self.listw.itemActivated.connect(self._on_item_activated)
        self.listw.itemDoubleClicked.connect(self._on_item_activated)
        right.addWidget(self.listw, 1)
        body.addLayout(right, 1)
        root.addLayout(body, 1)

        # 하단: 검증 배너 + ★ 고정 + 취소/선택
        self.banner = QLabel("폴더를 고르면 자재 여부를 여기서 확인합니다.")
        self.banner.setWordWrap(True)
        self.banner.setMinimumHeight(40)
        self.banner.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.banner.setContentsMargins(10, 6, 10, 6)
        self._set_banner("none", "폴더를 고르면 자재 여부를 여기서 확인합니다.")
        root.addWidget(self.banner)

        bottom = QHBoxLayout()
        bottom.setSpacing(8)
        self.btn_pin = QPushButton("☆ 즐겨찾기")
        self.btn_pin.setFixedWidth(120)
        self.btn_pin.clicked.connect(self._toggle_favorite)
        bottom.addWidget(self.btn_pin)
        bottom.addStretch(1)
        btn_cancel = QPushButton("취소")
        btn_cancel.clicked.connect(self.reject)
        self.btn_ok = QPushButton("이 폴더 선택")
        self.btn_ok.setDefault(True)
        self.btn_ok.clicked.connect(self.accept)
        bottom.addWidget(btn_cancel)
        bottom.addWidget(self.btn_ok)
        root.addLayout(bottom)

        # 검증 디바운스 타이머
        self._debounce = QTimer(self)
        self._debounce.setSingleShot(True)
        self._debounce.setInterval(150)
        self._debounce.timeout.connect(self._run_validation)

    # ----------------------------------------------------------- sidebar (폴더 트리)
    _LOADED = Qt.UserRole + 1  # 지연 로딩 완료 플래그

    def _make_dir_node(self, parent, label: str, path: str):
        """지연 확장 폴더 노드(펼치면 그때 한 단계만 os.scandir). 더미 자식으로 화살표 표시."""
        node = QTreeWidgetItem(parent, [label])
        node.setData(0, Qt.UserRole, path)
        node.setData(0, self._LOADED, False)
        node.setToolTip(0, path)
        QTreeWidgetItem(node, ["…"])  # placeholder → 펼침 화살표
        return node

    def _make_group_node(self, label: str):
        """접이식 그룹 헤더(경로 없음). 자식은 폴더 노드."""
        node = QTreeWidgetItem(self.sidebar, [label])
        node.setData(0, Qt.UserRole, None)
        node.setData(0, self._LOADED, True)  # 그룹은 지연 로딩 대상 아님
        node.setFlags(Qt.ItemIsEnabled)
        return node

    @staticmethod
    def _unc_anchor(s: str) -> str:
        """경로의 루트(anchor). UNC(\\\\server\\share)는 플랫폼과 무관하게 공유 루트를 돌려준다."""
        s = str(s)
        if s.startswith("\\\\") or s.startswith("//"):
            sep = "\\" if s[0] == "\\" else "/"
            parts = s.replace("/", "\\").split("\\")  # ['', '', server, share, ...]
            if len(parts) >= 4 and parts[2] and parts[3]:
                return f"{sep}{sep}{parts[2]}{sep}{parts[3]}{sep}"
            return s
        try:
            return Path(s).anchor
        except Exception:  # noqa: BLE001
            return ""

    @staticmethod
    def _drive_label(p: str) -> str:
        """드라이브 표시명 — 네트워크 드라이브는 볼륨 이름·UNC 대상까지 함께 보여준다."""
        drive = p.rstrip("/\\") or p
        name = ""
        dev = ""
        try:
            si = QStorageInfo(p)
            name = si.name() or ""
            raw = si.device()
            try:
                dev = bytes(raw).decode("utf-8", "ignore")
            except Exception:  # noqa: BLE001
                dev = str(raw)
        except Exception:  # noqa: BLE001
            pass
        is_net = dev.startswith("\\\\") or dev.startswith("//")
        extra = []
        if name:
            extra.append(name)
        if is_net and dev:
            extra.append(dev)
        label = drive + (f"  ({' · '.join(extra)})" if extra else "")
        return ("🌐 " if is_net else "💾 ") + label

    def _ensure_root_for(self, path) -> None:
        """path 를 포함하는 최상위 루트가 없으면(예: UNC 네트워크 폴더) 그 루트를 추가한다."""
        anchor = self._unc_anchor(path)
        if not anchor:
            return
        na = anchor.rstrip("/\\")
        for rp, _ in self._root_nodes:
            if str(rp).rstrip("/\\") == na:
                return
        is_net = anchor.startswith("\\\\") or anchor.startswith("//")
        label = ("🌐 " + anchor) if is_net else self._drive_label(anchor)
        node = self._make_dir_node(self.sidebar.invisibleRootItem(), label, anchor)
        self._root_nodes.append((anchor, node))

    def _open_native_explorer(self) -> None:
        """OS 기본 폴더 선택 대화상자로 폴더를 고른다(트리에 안 보이는 네트워크 공유 등)."""
        start = str(self._cur) if self._cur.exists() else str(Path.home())
        path = QFileDialog.getExistingDirectory(self, "폴더 선택 (기본 탐색기)", start)
        if path:
            self._go_to(self._safe_dir(path))

    def _apply_scan_root(self) -> None:
        """지정 칸의 스캔 데이터 폴더 경로를 저장·고정하고 그 폴더로 이동."""
        path = self.ed_scan_root.text().strip()
        self.settings.scan_root_path = path
        try:
            self.settings.save()
        except Exception:  # noqa: BLE001 - 저장 실패해도 세션 내 반영은 유지
            pass
        self._reload_sidebar()
        if path and Path(path).is_dir():
            self._go_to(self._safe_dir(path))

    def _pick_scan_root(self) -> None:
        start = self.ed_scan_root.text().strip() or str(self._cur)
        if not Path(start).exists():
            start = str(Path.home())
        path = QFileDialog.getExistingDirectory(self, "스캔 데이터 폴더 선택", start)
        if path:
            self.ed_scan_root.setText(path)
            self._apply_scan_root()

    def _pin_scan_roots(self, search_roots: list[str]) -> None:
        """스캔 데이터 폴더를 최상위 '📌' 고정 노드로 추가.

        1) 명시 지정 경로(scan_root_path)가 있으면 그것을 먼저(맨 위). 2) 이후 각 루트 바로
        아래에서 scan_root_name 폴더를 탐지해 추가(경로 중복 제거).
        """
        root = self.sidebar.invisibleRootItem()
        seen: set[str] = set()
        # 1) 명시 경로
        explicit = (getattr(self.settings, "scan_root_path", "") or "").strip()
        if explicit:
            try:
                ok = Path(explicit).is_dir()
            except OSError:
                ok = False
            if ok:
                key = str(Path(explicit)).rstrip("/\\")
                seen.add(key)
                node = self._make_dir_node(root, f"📌 {Path(explicit).name or explicit}", explicit)
                self._root_nodes.append((explicit, node))
        # 2) 이름 기반 자동 탐지
        name = (getattr(self.settings, "scan_root_name", "") or "").strip()
        if not name:
            return
        for r in search_roots:
            try:
                p = Path(r) / name
                if not p.is_dir():
                    continue
            except OSError:
                continue  # 연결 끊긴 드라이브/네트워크 → 스킵
            key = str(p).rstrip("/\\")
            if key in seen:
                continue
            seen.add(key)
            disp = r.rstrip("/\\") or r
            node = self._make_dir_node(root, f"📌 {name}  ({disp})", str(p))
            self._root_nodes.append((str(p), node))

    def _reload_sidebar(self) -> None:
        self.sidebar.clear()
        self._root_nodes = []  # (path, node) — 현재 위치 트리 동기화용 루트
        root = self.sidebar.invisibleRootItem()
        favs = [f for f in getattr(self.settings, "favorite_folders", []) if Path(f).exists()]
        recents = [f for f in getattr(self.settings, "recent_folders", []) if Path(f).exists()]
        # 검색 대상 루트(홈·드라이브·현재/즐겨찾기/최근의 네트워크 앵커).
        search_roots = [str(Path.home())] + [d.absoluteFilePath() for d in QDir.drives()]
        for cand in [str(self._cur), *favs, *recents]:
            a = self._unc_anchor(cand)
            if a and a not in search_roots:
                search_roots.append(a)
        # 1) 스캔 데이터 폴더(scan_root_name)가 있는 위치를 최상위에 📌 고정.
        self._pin_scan_roots(search_roots)
        # 2) 폴더 트리 위주: 홈 + 드라이브를 최상위 폴더 노드로 노출(탐색기처럼).
        home = self._make_dir_node(root, "🏠 홈", str(Path.home()))
        self._root_nodes.append((str(Path.home()), home))
        for d in QDir.drives():
            p = d.absoluteFilePath()
            self._root_nodes.append((p, self._make_dir_node(root, self._drive_label(p), p)))
        # 3) 네트워크(UNC) 위치는 드라이브 목록에 안 나오므로 앵커 루트를 추가.
        for cand in [str(self._cur), *favs, *recents]:
            self._ensure_root_for(cand)
        # 즐겨찾기·최근은 맨 아래 접이식 소형 그룹(기본 접힘).
        if favs:
            grp = self._make_group_node("★ 즐겨찾기")
            for f in favs:
                self._make_dir_node(grp, "★ " + (Path(f).name or f), f)
            grp.setExpanded(False)
        if recents:
            grp = self._make_group_node("↻ 최근")
            for f in recents:
                self._make_dir_node(grp, "📁 " + (Path(f).name or f), f)
            grp.setExpanded(False)
        self._reveal_in_tree(self._cur)

    def _on_tree_expanded(self, node: QTreeWidgetItem) -> None:
        if node.data(0, self._LOADED):
            return
        path = node.data(0, Qt.UserRole)
        node.takeChildren()  # 더미 제거
        if path:
            for name in self._list_subdirs(Path(path)):
                self._make_dir_node(node, "📁 " + name, str(Path(path) / name))
        node.setData(0, self._LOADED, True)

    def _reveal_in_tree(self, path: Path) -> None:
        """현재 경로를 포함하는 루트를 찾아 세그먼트마다 지연 확장하며 그 노드를 선택·스크롤."""
        roots = getattr(self, "_root_nodes", None)
        if roots is None:
            return
        # 네트워크(UNC) 등 기존 루트에 없는 위치면 루트를 즉석에서 추가.
        self._ensure_root_for(path)
        roots = self._root_nodes
        target = Path(path)
        best = None  # 가장 깊은(구체적인) 접두 루트
        for rp, node in roots:
            rpp = Path(rp)
            if target == rpp or rpp in target.parents:
                if best is None or len(str(rpp)) > len(str(Path(best[0]))):
                    best = (rp, node)
        if best is None:
            return
        rp, node = best
        node.setExpanded(True)  # itemExpanded → 지연 로딩(동기)
        try:
            rel = target.relative_to(Path(rp))
        except ValueError:
            rel = Path()
        cur = node
        for part in rel.parts:
            child = self._find_child_by_name(cur, part)
            if child is None:
                break
            cur = child
            cur.setExpanded(True)
        self.sidebar.setCurrentItem(cur)
        self.sidebar.scrollToItem(cur)

    @staticmethod
    def _find_child_by_name(parent: QTreeWidgetItem, name: str) -> Optional[QTreeWidgetItem]:
        for i in range(parent.childCount()):
            ch = parent.child(i)
            p = ch.data(0, Qt.UserRole)
            if p and Path(p).name == name:
                return ch
        return None

    def _on_tree_clicked(self, node: QTreeWidgetItem, _col: int = 0) -> None:
        path = node.data(0, Qt.UserRole)
        if path:
            self._go_to(self._safe_dir(path))

    # ----------------------------------------------------------- navigation
    def _go_to(self, path: Path, push: bool = True) -> None:
        path = self._safe_dir(str(path))
        if push and path != self._cur:
            self._history.append(self._cur)
        self._cur = path
        self.ed_path.setText(str(path))
        self.ed_filter.clear()
        self._populate_list()
        self._rebuild_crumbs()
        self.btn_back.setEnabled(bool(self._history))
        self.btn_up.setEnabled(path.parent != path)
        self._reveal_in_tree(path)  # 좌측 폴더 트리를 현재 위치로 확장·강조
        # 현재 폴더 자체를 후보로 삼아 자동 검증(자재로 바로 들어오면 즉시 확인).
        self._set_candidate(path)

    def _go_up(self) -> None:
        if self._cur.parent != self._cur:
            self._go_to(self._cur.parent)

    def _go_back(self) -> None:
        if self._history:
            prev = self._history.pop()
            self._go_to(prev, push=False)

    def _on_path_entered(self) -> None:
        text = self.ed_path.text().strip()
        if not text:
            return
        p = Path(text)
        if p.exists() and p.is_dir():
            self._go_to(p)
        else:
            self._set_banner("too_high", f"경로를 찾을 수 없습니다: {text}")

    def _rebuild_crumbs(self) -> None:
        while self._crumbs.count():
            w = self._crumbs.takeAt(0).widget()
            if w is not None:
                w.deleteLater()
        parts = list(self._cur.parts)
        acc = Path(parts[0]) if parts else self._cur
        for i, part in enumerate(parts):
            if i > 0:
                acc = acc / part
            label = part if part not in ("/", "\\") else "/"
            btn = QPushButton(label)
            btn.setFlat(True)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet(
                f"QPushButton {{ border: none; padding: 2px 6px; color: {NEON}; }}"
                f"QPushButton:hover {{ color: {TEXT}; text-decoration: underline; }}"
            )
            target = acc
            btn.clicked.connect(lambda _=False, t=target: self._go_to(t))
            self._crumbs.addWidget(btn)
            if i < len(parts) - 1:
                sep = QLabel("›")
                sep.setStyleSheet(f"color: {TEXT_DIM};")
                self._crumbs.addWidget(sep)
        self._crumbs.addStretch(1)

    # ----------------------------------------------------------- list
    def _populate_list(self) -> None:
        self.listw.clear()
        for name in self._list_subdirs(self._cur):
            it = QListWidgetItem("📁 " + name)
            it.setData(Qt.UserRole, name)
            self.listw.addItem(it)
        if self.listw.count() == 0:
            it = QListWidgetItem("(하위 폴더 없음 — 이 폴더가 자재일 수 있습니다)")
            it.setFlags(Qt.NoItemFlags)
            it.setForeground(Qt.gray)
            self.listw.addItem(it)

    def _apply_filter(self, text: str) -> None:
        needle = text.strip().lower()
        for i in range(self.listw.count()):
            it = self.listw.item(i)
            name = it.data(Qt.UserRole)
            if name is None:  # 안내/빈 항목은 그대로 둔다
                continue
            it.setHidden(bool(needle) and needle not in name.lower())

    def _on_item_clicked(self, item: QListWidgetItem) -> None:
        name = item.data(Qt.UserRole)
        if name is None:
            return
        self._set_candidate(self._cur / name)

    def _on_item_activated(self, item: QListWidgetItem) -> None:
        name = item.data(Qt.UserRole)
        if name is None:
            return
        self._go_to(self._cur / name)

    # ----------------------------------------------------------- validation
    def _set_candidate(self, path: Path) -> None:
        self._candidate = path
        self._update_pin_button()
        self._set_banner("busy", f"‘{path.name or path}’ 확인 중…")
        self._debounce.start()

    def _run_validation(self) -> None:
        if self._candidate is None:
            return
        self._token += 1
        worker = _ValidateWorker(self._token, str(self._candidate))
        worker.signals.done.connect(self._on_validated)
        self._pool.start(worker)

    @Slot(int, str, str, int, int)
    def _on_validated(
        self, token: int, kind: str, material: str, layers: int, wafers: int
    ) -> None:
        if token != self._token or self._candidate is None:
            return  # 오래된 결과 무시
        self._valid_for = self._candidate
        self._valid_kind = kind
        self._valid_material = material
        name = self._candidate.name or str(self._candidate)
        if kind == "material":
            self._set_banner(
                "material",
                f"✓ 자재(LOT) 폴더 · layer {layers}개 · wafer {wafers}개",
            )
            self.btn_ok.setEnabled(True)
        elif kind in ("layer", "wafer"):
            mat_name = Path(material).name if material else "?"
            self._set_banner(
                "layerwafer",
                f"{kind} 폴더 · 선택 시 자재 ‘{mat_name}’ 로 이동합니다"
                f" (layer {layers}개 · wafer {wafers}개)",
            )
            self.btn_ok.setEnabled(True)
        elif kind == "too_high":
            self._set_banner(
                "too_high",
                f"‘{name}’ 는 상위(device) 폴더입니다 · 자재 폴더로 들어가세요",
            )
            self.btn_ok.setEnabled(False)
        else:  # unknown
            self._set_banner(
                "unknown",
                f"‘{name}’ 에서 이미지를 찾지 못함 · 그래도 선택할 수 있습니다",
            )
            self.btn_ok.setEnabled(True)

    def _set_banner(self, kind: str, text: str) -> None:
        fg, bg = _BANNER_COLORS.get(kind, _BANNER_COLORS["none"])
        if kind == "busy":
            # 로딩(확인 중) — 말줄임 애니메이션으로 '움직이는' 로딩 표시.
            self._busy_base = text.rstrip("… .")
            self._busy_dots = 0
            self.banner.setText(self._busy_base)
            if not self._busy_timer.isActive():
                self._busy_timer.start()
        else:
            self._busy_timer.stop()
            self.banner.setText(text)
        self.banner.setStyleSheet(
            f"QLabel {{ background-color: {bg}; color: {fg};"
            f" border: 1px solid {fg}; border-radius: 6px; padding: 8px 10px; }}"
        )

    def _tick_busy(self) -> None:
        self._busy_dots = (self._busy_dots + 1) % 4
        self.banner.setText(self._busy_base + "." * self._busy_dots)

    # ----------------------------------------------------------- favorites
    def _update_pin_button(self) -> None:
        target = self._candidate or self._cur
        favs = list(getattr(self.settings, "favorite_folders", []))
        if str(target) in favs:
            self.btn_pin.setText("★ 고정됨")
        else:
            self.btn_pin.setText("☆ 즐겨찾기")

    def _toggle_favorite(self) -> None:
        target = str(self._candidate or self._cur)
        favs = list(getattr(self.settings, "favorite_folders", []))
        if target in favs:
            favs.remove(target)
        else:
            favs.insert(0, target)
        self.settings.favorite_folders = favs[:10]
        try:
            self.settings.save()
        except Exception:  # noqa: BLE001 - 저장 실패해도 세션 내 반영은 유지
            pass
        self._reload_sidebar()
        self._update_pin_button()

    # ----------------------------------------------------------- result
    def selected_path(self) -> str:
        """선택 확정 시 반환할 자재 경로(보정 포함). 취소/부적합이면 빈 문자열."""
        target = self._candidate or self._cur
        # 마지막 검증이 현재 후보에 대한 것이면 캐시 사용, 아니면 동기 재판정.
        if self._valid_for == target and self._valid_kind:
            kind, material = self._valid_kind, self._valid_material
        else:
            k, m = classify_selection(target)
            kind, material = k, (str(m) if m is not None else "")
        if kind == "material":
            return str(target)
        if kind in ("layer", "wafer") and material:
            return material
        if kind == "unknown":
            return str(target)
        return ""  # too_high / none

    def _resolved_kind(self) -> str:
        """마지막 후보의 구조 판정(material/layer/wafer/too_high/unknown)."""
        target = self._candidate or self._cur
        if self._valid_for == target and self._valid_kind:
            return self._valid_kind
        kind, _ = classify_selection(target)
        return kind

    def selected_wafer_folder(self) -> str:
        """wafer 폴더를 직접 골랐다면 그 폴더 경로, 아니면 빈 문자열.

        호출 측이 '개별 wafer 만 볼지' 를 물어보고, 아니면 selected_path()(상위 LOT)로
        회귀할 수 있게 한다.
        """
        target = self._candidate or self._cur
        return str(target) if self._resolved_kind() == "wafer" else ""


# =============================================================================
# app/ui/heatmap_dialog.py   [#36]
# =============================================================================
"""Defect 히트맵 팝업 (항목 4·5).

좌측 웨이퍼맵에 defect 밀도를 색으로 표시하고, 위치(die/하위셀)를 클릭하면 우측에 그
위치의 defect 들을 세로로 나열한다. 각 defect 행에는 기준 사진과 **매칭된 비교 layer
사진만** 가로로 나열하며(매칭 없는 칸은 표시하지 않음), 각 썸네일 좌상단에 layer 배지를
붙인다. 각 행은 '담기'로 공유 출력 트레이에 담을 수 있다.

die 개수가 50개 미만이면 각 die 를 5×5(25) 하위셀로 나눠 die 내부 위치를 구분한다.
wafer 선택에서 '전체'를 고르면 모든 wafer 의 defect 을 한 장에 밀도 합산해 보여준다.

순수 집계/분할 로직은 app.heatmap 에 있고 여기서는 시각화/상호작용만 담당한다.
"""





_ALL_WAFERS = "전체"
heatmap_dialog__THUMB_PX = 150  # 상세 목록 썸네일 크기(고정)


# 히트맵 die 색(배경 theme.BG 위에서 잘 보이도록 밝게). 배경/빈 die 와 구분된다.
_HEAT_LO = QColor("#6fb0e0")   # 낮은 밀도(밝은 슬레이트블루)
_HEAT_HI = QColor("#ff8a5c")   # 높은 밀도(밝은 코랄)


def _heat_color(count: int, max_count: int) -> QColor:
    """defect 개수를 밀도 색으로 변환(0=빈 die 톤, 많을수록 밝은 코랄)."""
    if count <= 0:
        return QColor(BG_ELEV)  # 빈 die: 배경(BG)보다 밝은 중간 톤
    lo = _HEAT_LO
    hi = _HEAT_HI
    if max_count <= 1:
        t = 1.0
    else:
        t = (count - 1) / (max_count - 1)
    t = max(0.0, min(1.0, t))
    r = int(lo.red() + (hi.red() - lo.red()) * t)
    g = int(lo.green() + (hi.green() - lo.green()) * t)
    b = int(lo.blue() + (hi.blue() - lo.blue()) * t)
    return QColor(r, g, b)


class HeatmapWaferMap(QWidget):
    """웨이퍼맵에 defect 밀도를 색으로 그리고 위치 클릭을 알린다.

    die 개수가 적으면(subdivide=True) 각 die 를 5×5 하위셀로 나눠 그린다.
    """

    selection_changed = Signal(object)  # list[HeatKey]

    _DIE_PX = 20          # 미분할 die 한 변
    _SUBCELL_PX = 9       # 하위셀 한 변
    _GAP = 3              # die 간 간격

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self._cols = 0
        self._rows = 0
        self._valid: Optional[frozenset] = None
        self._subdivide = False
        self._density: dict[HeatKey, int] = {}
        self._max_count = 1
        self._selected_keys: set[HeatKey] = set()
        # 여러 die 선택 모드 + 러버밴드 상태. 드래그 사각 선택은 _multi 와 무관하게 항상 가능;
        # _multi 는 '클릭=토글/박스=합집합'(on) vs '클릭·박스=교체'(off) 만 좌우한다.
        self._multi = False
        self._rubber_origin = None
        self._rubber_cur = None
        self._dragging = False
        # 그리기 원점(실좌표) — 내용 bounding box 좌상단을 (0,0) 픽셀에 맞춘다(맵 떠보임/잘림 방지).
        self._origin_col = 0
        self._origin_row = 0
        self.setToolTip("defect 밀도 — 색이 진할수록 defect 이 많음. 위치를 클릭하세요.")

    def set_multi(self, on: bool) -> None:
        self._multi = on
        self._selected_keys = set()
        self._rubber_origin = None
        self._rubber_cur = None
        self.update()
        self.selection_changed.emit([])

    def _die_w(self) -> int:
        return SUB_COLS * self._SUBCELL_PX if self._subdivide else self._DIE_PX

    def _die_h(self) -> int:
        return SUB_ROWS * self._SUBCELL_PX if self._subdivide else self._DIE_PX

    def set_data(
        self,
        cols: int,
        rows: int,
        valid: Optional[frozenset],
        subdivide: bool,
        density: dict[HeatKey, int],
        origin: tuple[int, int] = (0, 0),
    ) -> None:
        self._cols = max(0, cols)
        self._rows = max(0, rows)
        self._valid = frozenset(valid) if valid else None
        self._subdivide = subdivide
        self._density = density
        self._max_count = max(density.values(), default=1)
        self._selected_keys = set()
        self._rubber_origin = None
        self._rubber_cur = None
        self._origin_col, self._origin_row = origin
        dw, dh = self._die_w(), self._die_h()
        self.setFixedSize(
            max(40, self._cols * (dw + self._GAP) + self._GAP),
            max(40, self._rows * (dh + self._GAP) + self._GAP),
        )
        self.update()

    def _die_origin(self, col: int, row: int) -> tuple[int, int]:
        dw, dh = self._die_w(), self._die_h()
        return (self._GAP + (col - self._origin_col) * (dw + self._GAP),
                self._GAP + (row - self._origin_row) * (dh + self._GAP))

    def paintEvent(self, event):  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, False)
        # 위젯 전체를 UI 어두운 배경으로 채운다(흰 배경 제거, die 색과 구분).
        painter.fillRect(self.rect(), QColor(BG))
        border = QColor(NEON_SOFT)
        die_border = QColor(BASE_GLOW)
        for dr in range(self._rows):
            row = dr + self._origin_row
            for dc in range(self._cols):
                col = dc + self._origin_col
                if self._valid is not None and (col, row) not in self._valid:
                    continue
                ox, oy = self._die_origin(col, row)
                if self._subdivide:
                    self._paint_die_sub(painter, col, row, ox, oy, border)
                    # die 경계선(하위셀 위에 굵게) — die 구분 유지
                    painter.setPen(QPen(die_border, 1))
                    painter.drawRect(QRect(ox, oy, self._die_w(), self._die_h()))
                else:
                    count = self._density.get(HeatKey(col, row), 0)
                    rect = QRect(ox, oy, self._DIE_PX, self._DIE_PX)
                    painter.fillRect(rect, _heat_color(count, self._max_count))
                    painter.setPen(QPen(border, 1))
                    painter.drawRect(rect)
        self._paint_selection(painter)
        painter.end()

    def _paint_die_sub(self, painter, col, row, ox, oy, border) -> None:
        for sr in range(SUB_ROWS):
            for sc in range(SUB_COLS):
                count = self._density.get(HeatKey(col, row, sc, sr), 0)
                x = ox + sc * self._SUBCELL_PX
                y = oy + sr * self._SUBCELL_PX
                rect = QRect(x, y, self._SUBCELL_PX, self._SUBCELL_PX)
                painter.fillRect(rect, _heat_color(count, self._max_count))
                painter.setPen(QPen(border, 1))
                painter.drawRect(rect)

    def _key_rect(self, key: HeatKey) -> QRect:
        ox, oy = self._die_origin(key.col, key.row)
        if self._subdivide and key.subdivided:
            return QRect(ox + key.sub_col * self._SUBCELL_PX,
                         oy + key.sub_row * self._SUBCELL_PX,
                         self._SUBCELL_PX, self._SUBCELL_PX)
        return QRect(ox, oy, self._die_w(), self._die_h())

    def _paint_selection(self, painter) -> None:
        # 선택 die: 얇은(1px) 네온 녹색 외곽선만(채움 없음).
        green = QColor("#39ff14")
        painter.setPen(QPen(green, 1))
        for key in self._selected_keys:
            if not (self._origin_col <= key.col < self._origin_col + self._cols
                    and self._origin_row <= key.row < self._origin_row + self._rows):
                continue
            painter.drawRect(self._key_rect(key).adjusted(0, 0, -1, -1))
        # 드래그 러버밴드
        if self._rubber_origin is not None and self._rubber_cur is not None:
            rb = QRect(self._rubber_origin, self._rubber_cur).normalized()
            painter.setPen(QPen(green, 1, Qt.DashLine))
            painter.drawRect(rb)

    def _key_at(self, pos) -> Optional[HeatKey]:
        dw, dh = self._die_w(), self._die_h()
        dc = (pos.x() - self._GAP) // (dw + self._GAP)
        dr = (pos.y() - self._GAP) // (dh + self._GAP)
        if not (0 <= dc < self._cols and 0 <= dr < self._rows):
            return None
        col = int(dc) + self._origin_col
        row = int(dr) + self._origin_row
        if self._subdivide:
            ox, oy = self._die_origin(col, row)
            sc = min(SUB_COLS - 1, max(0, (pos.x() - ox) // self._SUBCELL_PX))
            sr = min(SUB_ROWS - 1, max(0, (pos.y() - oy) // self._SUBCELL_PX))
            key = HeatKey(col, row, int(sc), int(sr))
        else:
            key = HeatKey(col, row)
        return key if self._density.get(key, 0) > 0 else None

    def _keys_in_rect(self, rect: QRect) -> list[HeatKey]:
        out = []
        for key, cnt in self._density.items():
            if cnt > 0 and self._key_rect(key).intersects(rect):
                out.append(key)
        return out

    def _emit_selection(self) -> None:
        keys = sorted(self._selected_keys,
                      key=lambda k: (k.row, k.col, k.sub_row, k.sub_col))
        self.selection_changed.emit(keys)

    def mousePressEvent(self, event):  # noqa: N802
        if event.button() != Qt.LeftButton:
            return
        # 드래그 사각 선택은 항상 시작 가능(멀티 모드 여부와 무관).
        pos = event.position().toPoint()
        self._rubber_origin = pos
        self._rubber_cur = pos
        self._dragging = False

    def mouseMoveEvent(self, event):  # noqa: N802
        if self._rubber_origin is None:
            return
        pos = event.position().toPoint()
        if (pos - self._rubber_origin).manhattanLength() > 4:
            self._dragging = True
        self._rubber_cur = pos
        if self._dragging:
            self.update()

    def mouseReleaseEvent(self, event):  # noqa: N802
        if event.button() != Qt.LeftButton or self._rubber_origin is None:
            return
        origin = self._rubber_origin
        cur = self._rubber_cur or origin
        if self._dragging:
            # 사각 드래그: 사각 내 die 선택. 멀티면 합집합, 아니면 교체.
            keys = set(self._keys_in_rect(QRect(origin, cur).normalized()))
            self._selected_keys = (self._selected_keys | keys) if self._multi else keys
        else:
            # 클릭: 멀티면 토글, 아니면 단일 교체.
            key = self._key_at(origin)
            if self._multi:
                if key is not None:
                    self._selected_keys.discard(key) if key in self._selected_keys \
                        else self._selected_keys.add(key)
            else:
                self._selected_keys = {key} if key is not None else set()
        self._rubber_origin = None
        self._rubber_cur = None
        self._dragging = False
        self.update()
        self._emit_selection()


# 클러스터 표시 위젯은 공유 모듈로 이동(메인 매치와 공용). 하위 호환 별칭 유지.
_ClusteredThumb = ClusteredThumb
_ClusterMembersPopup = ClusterMembersPopup
_ClickThumb = ClickThumb
_load_thumb_holder = load_thumb_holder


class HeatmapDialog(QDialog):
    """defect 히트맵 팝업(좌 웨이퍼맵 · 우 defect 목록)."""

    def __init__(
        self,
        matches: list[BaseDefectMatches],
        base_layer: str,
        compare_layers: list[str],
        thumb_cache: ThumbnailCache,
        on_add_to_export: Callable[[list[int]], None],
        settings,
        current_wafer: Optional[str] = None,
        records_by_layer: Optional[dict] = None,
        parent: Optional[QWidget] = None,
    ):
        super().__init__(parent)
        self.matches = matches
        self._base_layer = base_layer
        self._compare_layers = list(compare_layers)
        self._thumb_cache = thumb_cache
        self._on_add = on_add_to_export
        self._settings = settings
        self._records_by_layer = records_by_layer or {}
        # 교차 매칭(전체 defect 모드)·매칭 판정용 허용오차.
        self._tolerance = getattr(settings, "tolerance", None) or DEFAULT_TOLERANCE
        # defect 근접 클러스터링 거리(설정에서 조절).
        self._cluster_radius = (
            getattr(settings, "cluster_radius", None) or DEFAULT_CLUSTER_RADIUS
        )
        self.setWindowTitle("Defect 히트맵")
        self.setMinimumSize(900, 600)

        self._groups: dict[HeatKey, list[int]] = {}
        self._selected_keys: list[HeatKey] = []  # 현재 선택된 위치(단일/다중)
        self._add_targets: list[int] = []  # '이 위치 출력에 넣기' 대상(매칭된 기준)
        self._pending_thumbs: list = []  # 상세 지연 로딩 썸네일 위젯
        self._thumb_token = 0            # stale 썸네일 로딩 무시
        self._active_thumb_workers: set = set()  # 실행 중 워커 참조 유지(GC 방지)
        self._hm_align_cache: dict = {}  # observed→alignment 캐시(매 refresh 재계산 방지)
        self._align_shift = (0, 0)  # 관측→die_map 좌표 정합 이동(밀도·선택 키를 die_map 좌표로)
        self._xr = (0.0, 1.0)   # die 내부 local 좌표 범위(subcell 판정용)
        self._yr = (0.0, 1.0)
        self._subdivide = False

        # 기본은 '전체' wafer(사용자 요청). current_wafer 는 무시하고 전체로 시작.
        self._current_wafer = _ALL_WAFERS

        self._build()
        # 시작 시 전체화면(최대화). exec 모달에서도 유지된다.
        self.setWindowState(Qt.WindowMaximized)
        self._refresh_map()

    # ---- 데이터 헬퍼 -------------------------------------------------
    def _wafers(self) -> list[str]:
        seen: list[str] = []
        for m in self.matches:
            w = m.base.wafer_id
            if w not in seen:
                seen.append(w)
        return seen

    def _available_layers(self) -> list[str]:
        """표시(필터) 가능한 비교 layer: matches 에 등장하는 모든 비교 layer(순서 유지)."""
        layers: list[str] = []
        for m in self.matches:
            for r in m.results:
                if r.compare_layer not in layers:
                    layers.append(r.compare_layer)
        return layers

    def _all_layers(self) -> list[str]:
        """조사 대상 전체 layer — 메인 기준 + 비교 layer(중복 제거, 순서 유지)."""
        out: list[str] = []
        for lyr in [self._base_layer, *self._available_layers()]:
            if lyr and lyr not in out:
                out.append(lyr)
        return out

    def _selected_layers(self) -> list[str]:
        """조사할(체크된) layer 목록. 기준 특별취급 없음."""
        return [lyr for lyr, cb in self._col_checks.items() if cb.isChecked()]

    def _wafer_ok(self, rec) -> bool:
        return self._current_wafer == _ALL_WAFERS or rec.wafer_id == self._current_wafer

    def _base_entries(self) -> list[tuple[int, object]]:
        """기준 defect 전체를 (base_index, base_record) 로. wafer 필터 적용.

        매치만 모드의 맵 density·선택→index 매핑(_groups) 공통. (미매칭 숨김은 상세 목록에서만.)
        """
        out: list[tuple[int, object]] = []
        for i, m in enumerate(self.matches):
            b = m.base
            if b.col is None or b.row is None or b.col < 0 or b.row < 0:
                continue
            if not self._wafer_ok(b):
                continue
            out.append((i, b))
        return out

    def _is_matched(self, bi: int) -> bool:
        """메인 비교 layer 중 하나 이상에서 매치되면 True('출력에 넣기' 대상 판정)."""
        m = self.matches[bi]
        return any((r := m.for_layer(lyr)) and r.is_match for lyr in self._compare_layers)

    def _all_defect_entries(self) -> list[tuple[int, object]]:
        """체크된 layer 의 좌표 OK defect(맵 density 용). 기준 특별취급 없음.

        근접 중복(같은 wafer·die·거리<cluster_radius) defect 은 한 클러스터로 묶어
        밀도에는 1개(대표)로만 계산한다.
        """
        out: list[tuple[int, object]] = []
        k = 0
        for lyr in self._selected_layers():
            recs = [
                rec for rec in self._records_by_layer.get(lyr, [])
                if getattr(rec, "ok", False) and self._wafer_ok(rec)
            ]
            for cl in cluster_records(recs, self._cluster_radius):
                out.append((k, cl.representative))
                k += 1
        return out

    def _map_entries(self) -> list[tuple[int, object]]:
        """맵 density 계산용 entries — 상시 조사 모드(체크된 layer 전체 defect)."""
        return self._all_defect_entries()

    # ---- UI 구성 -----------------------------------------------------
    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 12, 14, 12)
        outer.setSpacing(8)

        # 상단 컨트롤 바: 제목 / wafer 선택
        bar = QHBoxLayout()
        title = QLabel("Defect 히트맵")
        title.setObjectName("title")
        bar.addWidget(title)
        bar.addStretch()
        bar.addWidget(QLabel("Wafer"))
        self.cmb_wafer = QComboBox()
        self.cmb_wafer.addItem(_ALL_WAFERS)  # '전체' 맨 앞
        self.cmb_wafer.addItems(self._wafers())
        if self._current_wafer:
            self.cmb_wafer.setCurrentText(self._current_wafer)
        self.cmb_wafer.currentTextChanged.connect(self._on_wafer_changed)
        bar.addWidget(self.cmb_wafer)
        outer.addLayout(bar)

        # 조사할 layer(기준 개념 없음): 모든 layer 를 체크박스로, 체크된 것들의 defect 을 교차 조사.
        col_row = QHBoxLayout()
        col_row.setSpacing(8)
        col_row.addWidget(QLabel("조사할 layer:"))
        self._col_checks: dict[str, QCheckBox] = {}
        for lyr in self._all_layers():
            cb = QCheckBox(lyr)
            cb.setChecked(True)  # 기본 전체 체크
            cb.stateChanged.connect(lambda _=0: self._on_layers_changed())
            self._col_checks[lyr] = cb
            col_row.addWidget(cb)
        col_row.addStretch()
        outer.addLayout(col_row)

        # 본문 스플리터: 좌 맵 | 우 목록 (고정 배치)
        self._splitter = QSplitter(Qt.Horizontal)
        self._splitter.addWidget(self._build_map_panel())
        self._splitter.addWidget(self._build_list_panel())
        self._splitter.setStretchFactor(0, 0)
        self._splitter.setStretchFactor(1, 1)
        self._splitter.setSizes([360, 900])
        outer.addWidget(self._splitter, 1)

        btn_close = QPushButton("닫기")
        btn_close.setObjectName("primary")
        btn_close.clicked.connect(self.accept)
        outer.addWidget(btn_close, alignment=Qt.AlignRight)

    def _transparent_scroll(self) -> QScrollArea:
        """테마 배경이 비치는(흰 배경 없는) 스크롤 영역."""
        sc = QScrollArea()
        sc.setFrameShape(QFrame.NoFrame)
        sc.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        sc.viewport().setAutoFillBackground(False)
        return sc

    def _build_map_panel(self) -> QWidget:
        panel = QFrame()
        panel.setObjectName("panel")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(6)
        cap = QLabel("defect 밀도 (색이 진할수록 많음)")
        cap.setObjectName("dim")
        lay.addWidget(cap)
        self._map = HeatmapWaferMap()
        map_scroll = self._transparent_scroll()
        map_scroll.setWidgetResizable(False)
        map_scroll.setAlignment(Qt.AlignCenter)
        map_scroll.setWidget(self._map)
        self._map.selection_changed.connect(self._on_selection_changed)
        lay.addWidget(map_scroll, 1)
        self.lbl_map = QLabel("")
        self.lbl_map.setObjectName("dim")
        self.lbl_map.setStyleSheet("font-size:10px;")
        self.lbl_map.setWordWrap(True)
        lay.addWidget(self.lbl_map)
        # 여러 다이 선택 토글 — 웨이퍼 맵 바로 아래에 둔다(드래그 사각 선택은 이 토글과
        # 무관하게 항상 가능).
        self.btn_multi = QPushButton("여러 다이 선택: OFF")
        self.btn_multi.setObjectName("mini")
        self.btn_multi.setCheckable(True)
        self.btn_multi.setToolTip(
            "켜면 클릭이 여러 die 를 누적 토글합니다.\n"
            "(끄면 클릭=한 개 선택. 드래그 사각 다중 선택은 항상 가능합니다.)"
        )
        self.btn_multi.toggled.connect(self._on_multi_toggled)
        multi_row = QHBoxLayout()
        multi_row.setContentsMargins(0, 0, 0, 0)
        multi_row.addWidget(self.btn_multi, 0)
        multi_row.addStretch(1)
        lay.addLayout(multi_row)
        return panel

    def _build_list_panel(self) -> QWidget:
        panel = QFrame()
        panel.setObjectName("panel")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(6)
        head = QHBoxLayout()
        self.lbl_detail = QLabel("위치를 클릭하면 그 자리의 defect 이 여기에 나열됩니다.")
        self.lbl_detail.setObjectName("dim")
        head.addWidget(self.lbl_detail, 1)
        self.btn_add_all = QPushButton("이 위치 출력에 넣기")
        self.btn_add_all.setObjectName("mini")
        self.btn_add_all.setToolTip("선택 위치의 매칭된 기준 defect 을 출력 트레이에 담습니다.")
        self.btn_add_all.clicked.connect(self._add_all_current)
        self.btn_add_all.setEnabled(False)
        head.addWidget(self.btn_add_all, 0)
        lay.addLayout(head)

        self._detail_scroll = self._transparent_scroll()
        self._detail_scroll.setWidgetResizable(True)
        self._detail_host = QWidget()
        self._detail_host.setStyleSheet("background: transparent;")
        self._detail_box = QVBoxLayout(self._detail_host)
        self._detail_box.setContentsMargins(2, 2, 2, 2)
        self._detail_box.setSpacing(8)
        self._detail_box.addStretch()
        self._detail_scroll.setWidget(self._detail_host)
        lay.addWidget(self._detail_scroll, 1)
        return panel

    # ---- 웨이퍼맵 / 상호작용 ----------------------------------------
    def _on_wafer_changed(self, wafer: str) -> None:
        self._current_wafer = wafer
        self._selected_keys = []
        self._refresh_map()
        self._rebuild_detail()

    def _on_multi_toggled(self, on: bool) -> None:
        self.btn_multi.setText(f"여러 다이 선택: {'ON' if on else 'OFF'}")
        self._map.set_multi(on)  # set_multi 가 selection_changed([]) 를 emit → 상세 초기화

    def _on_layers_changed(self) -> None:
        # layer 체크 변경 → 맵(선택 layer 반영)·상세 실시간 갱신, 선택 위치 유지.
        self._refresh_map(preserve_selection=True)
        self._rebuild_detail()

    def _refresh_map(self, preserve_selection: bool = False) -> None:
        keep = list(self._selected_keys) if preserve_selection else []
        entries = self._map_entries()
        observed = {
            (r.col, r.row) for _, r in entries
            if r.col is not None and r.row is not None
        }
        die_count = len(observed)
        subdivide = should_subdivide(die_count)
        xr, yr = local_ranges([r for _, r in entries])
        self._xr, self._yr, self._subdivide = xr, yr, subdivide

        # 웨이퍼 '모양'과 defect '정합'을 분리한다. 모양은 항상 die_map 자체(고정 좌표)로
        # 그려 정합 오차가 모양에 구멍/잘림을 만들지 못하게 하고, defect 밀도는 정합 이동
        # (shift)만큼 die_map 좌표계로 옮겨 올린다.
        prod = active_product()
        valid = None
        shift = (0, 0)
        caption = prod.name if prod.source == "db" else ""  # 제품명만(‘모양 정합’ 미표기)
        if prod.die_map:
            if observed:
                # 정합은 관측 die 집합·제품이 같으면 재계산 불필요 → 캐시(매 refresh 비용 절감).
                ckey = (prod.name, frozenset(observed))
                align = self._hm_align_cache.get(ckey)
                if align is None:
                    align = align_observed_to_diemap(observed, prod.die_map)
                    self._hm_align_cache[ckey] = align
                shift = (align.dcol, align.drow)
            valid = set(prod.die_map)  # 모양 = die_map 그대로(항상 온전한 wafer 형태)
        self._align_shift = shift

        # 밀도/그룹 키를 die_map 좌표계로 맞춘다(관측 좌표 − shift).
        density_groups = group_defects(entries, subdivide, xr, yr)
        density = {self._shift_key(k): len(v) for k, v in density_groups.items()}
        # 매치만 상세용 그룹(base index)은 항상 기준 defect 전체로, 동일 subdivide/range 로 키 정합.
        base_groups = group_defects(self._base_entries(), subdivide, xr, yr)
        self._groups = {self._shift_key(k): v for k, v in base_groups.items()}
        self._selected_keys = [k for k in keep if k in density]

        observed_dm = {(c - shift[0], r - shift[1]) for c, r in observed}
        paint_valid = (valid | observed_dm) if valid else None
        if paint_valid is not None:
            # 내용 bounding box 로 정규화(맵이 여백에 떠 보이거나 잘리지 않게).
            content = set(paint_valid) | observed_dm
            min_col = min(c for c, _ in content)
            min_row = min(r for _, r in content)
            cols = max(c for c, _ in content) - min_col + 1
            rows = max(r for _, r in content) - min_row + 1
            origin = (min_col, min_row)
        else:
            max_col = max((c for c, _ in observed), default=0)
            max_row = max((r for _, r in observed), default=0)
            cols = max(prod.kla_package_x_count, max_col + 1)
            rows = max(prod.kla_package_y_count, max_row + 1)
            origin = (0, 0)
        self._map.set_data(cols, rows, paint_valid, subdivide, density, origin=origin)
        # set_data 가 선택을 비우므로, 그 뒤에 유지할 선택을 복원한다.
        if self._selected_keys:
            self._map._selected_keys = set(self._selected_keys)
            self._map.update()

        sub_txt = " · die 5×5 분할" if subdivide else ""
        n_def = sum(density.values())
        wafer_txt = "전체 wafer" if self._current_wafer == _ALL_WAFERS else f"wafer {self._current_wafer}"
        self.lbl_map.setText(
            f"{wafer_txt} · die {die_count}개 · defect {n_def}개{sub_txt}"
            + (f"\n{caption}" if caption else "")
        )

    def _clear_detail(self) -> None:
        # 마지막 stretch 를 제외한 위젯 제거
        while self._detail_box.count() > 1:
            item = self._detail_box.takeAt(0)
            w = item.widget()
            if w is not None:
                w.setParent(None)
                w.deleteLater()

    def _on_selection_changed(self, keys) -> None:
        self._selected_keys = list(keys)
        self._rebuild_detail()

    def _union_indices(self) -> list[int]:
        """선택된 위치(단일/다중)의 base defect index 합집합(순서 유지, 중복 제거)."""
        seen: set[int] = set()
        out: list[int] = []
        for k in self._selected_keys:
            for bi in self._groups.get(k, []):
                if bi not in seen:
                    seen.add(bi)
                    out.append(bi)
        return out

    def _shift_key(self, k: HeatKey) -> HeatKey:
        """관측 좌표 HeatKey 를 die_map 좌표계로 옮긴다(정합 shift 만큼)."""
        dc, dr = self._align_shift
        if dc == 0 and dr == 0:
            return k
        return HeatKey(k.col - dc, k.row - dr, k.sub_col, k.sub_row)

    def _key_for_record(self, rec) -> HeatKey:
        # 밀도/그룹 키와 같은 die_map 좌표계로 맞춘다(관측 좌표 − shift).
        dc, dr = self._align_shift
        col = int(rec.col) - dc
        row = int(rec.row) - dr
        if self._subdivide and rec.x is not None and rec.y is not None:
            sc, sr = subcell_of(rec.x, rec.y, self._xr, self._yr)
            return HeatKey(col, row, sc, sr)
        return HeatKey(col, row)

    def _records_at_selection(self) -> list[tuple[str, object]]:
        """선택 위치의 체크된 layer 모든 defect record — (layer, record)."""
        layers = self._selected_layers()
        wafer = self._current_wafer
        key_set = set(self._selected_keys)
        out: list[tuple[str, object]] = []
        for lyr in layers:
            for rec in self._records_by_layer.get(lyr, []):
                if not getattr(rec, "ok", False):
                    continue
                if wafer != _ALL_WAFERS and rec.wafer_id != wafer:
                    continue
                if rec.col is None or rec.row is None:
                    continue
                if self._key_for_record(rec) in key_set:
                    out.append((lyr, rec))
        return out

    def _open_viewer(self, record) -> None:
        if isinstance(record, DefectRecord):
            ImageViewerDialog(record, self).exec()

    def _rebuild_detail(self) -> None:
        self._clear_detail()
        self._pending_thumbs = []  # 이번 상세의 지연 썸네일 모음(비동기 로딩)
        # 상시 조사 모드: 체크된 layer 를 교차 매칭해 표시. '출력에 넣기' 대상은 선택 위치의
        # 매칭된 메인 기준 defect(출력은 기준 layer 기반).
        self._add_targets = [bi for bi in self._union_indices() if self._is_matched(bi)]
        self.btn_add_all.setEnabled(bool(self._add_targets))
        if not self._selected_keys:
            self.lbl_detail.setText("위치를 클릭하면 그 자리의 defect 이 여기에 나열됩니다.")
            return
        n_loc = len(self._selected_keys)
        loc = (self._key_label(self._selected_keys[0]) if n_loc == 1 else f"{n_loc}개 위치")
        self._build_all_detail(loc)
        # 썸네일은 백그라운드로 캐시를 구운 뒤 채워, 클릭 즉시 목록이 뜨고 멈추지 않게 한다.
        self._start_detail_thumbs()

    def _build_all_detail(self, loc: str) -> None:
        """조사 모드 — 체크된 layer 를 layer 간 교차 매칭(기준 종속 아님)해 그룹으로.

        교차매치 그룹(≥2 layer)은 전폭 행으로, 개별(미매칭, 1 layer)은 한 섹션에 가로(FlowLayout)
        나열해 빈칸을 줄인다.
        """
        by_layer: dict[str, list] = defaultdict(list)
        for lyr, rec in self._records_at_selection():
            by_layer[lyr].append(rec)
        layer_to_clusters = {
            lyr: cluster_records(recs, self._cluster_radius)
            for lyr, recs in by_layer.items()
        }
        groups = cross_layer_groups(layer_to_clusters, self._tolerance)
        matched = [g for g in groups if len(g) >= 2]
        individual = [g for g in groups if len(g) == 1]
        for g in matched:
            self._detail_box.insertWidget(
                self._detail_box.count() - 1, self._make_group_row(g)
            )
        if individual:
            self._detail_box.insertWidget(
                self._detail_box.count() - 1, self._make_individual_section(individual)
            )
        self.lbl_detail.setText(
            f"{loc} — 교차매치 {len(matched)}그룹 · 개별(미매칭) {len(individual)}개"
        )

    def _make_individual_section(self, groups: list[dict]) -> QWidget:
        """개별(미매칭) 단일-layer 그룹들을 layer 배지+대표(+n) 컴팩트 카드로 가로 나열."""

        box = QFrame()
        box.setObjectName("cell")
        box.setStyleSheet(
            f"QFrame#cell {{ background:{BG_ELEV};"
            f" border:1px solid {NEON_SOFT}; border-radius:8px; }}"
        )
        outer = QVBoxLayout(box)
        outer.setContentsMargins(8, 6, 8, 8)
        outer.setSpacing(4)
        cap = QLabel(f"개별(미매칭) — {len(groups)}개")
        cap.setStyleSheet(f"font-size:10px; font-weight:700; color:{NOMATCH};")
        outer.addWidget(cap)
        host = QWidget()
        flow = FlowLayout(host, margin=0, h_spacing=8, v_spacing=8)
        for g in groups:
            (layer, cluster), = g.items()
            # 조사 모드에는 기준(★)이 없다 — 항상 일반 배지로 표시.
            flow.addWidget(self._clustered_thumb(cluster, layer, False))
        outer.addWidget(host)
        return box

    @staticmethod
    def _key_label(key: HeatKey) -> str:
        return (f"die({key.col},{key.row})"
                + (f" · 하위셀({key.sub_col},{key.sub_row})" if key.subdivided else ""))

    def _row_frame(self):
        row = QFrame()
        row.setObjectName("cell")
        row.setStyleSheet(
            f"QFrame#cell {{ background:{BG_ELEV};"
            f" border:1px solid {NEON_SOFT}; border-radius:8px; }}"
        )
        lay = QHBoxLayout(row)
        lay.setContentsMargins(8, 6, 8, 6)
        lay.setSpacing(8)
        return row, lay

    def _clustered_thumb(self, cluster: Cluster, layer: str, is_base: bool) -> QWidget:
        # 지연 로딩: 위젯은 즉시 만들고, 썸네일은 백그라운드로 캐시를 구운 뒤 채운다.
        w = _ClusteredThumb(cluster, layer, is_base, self._thumb_cache,
                            self._open_viewer, heatmap_dialog__THUMB_PX, defer=True)
        self._pending_thumbs.append(w)
        return w

    def _start_detail_thumbs(self) -> None:
        """상세에 만든 지연 썸네일들을 백그라운드로 캐시 굽고 준비되는 대로 채운다."""
        pending = self._pending_thumbs
        if not pending:
            return
        self._thumb_token += 1
        token = self._thumb_token
        items = [(i, str(w.rep_path), w._px) for i, w in enumerate(pending)]
        worker = FullThumbWorker(self._thumb_cache, items)

        def _fill(i, t=token, snap=pending):
            if t == self._thumb_token and 0 <= i < len(snap):
                snap[i].fill()

        worker.signals.ready.connect(_fill)
        # 워커가 끝날 때까지 참조를 유지한다(안 그러면 GC 로 signals 가 사라져 ready 가
        # 도착 안 해 썸네일이 '로딩…'에 멈춘다 — 이 버그의 근본 원인).
        self._active_thumb_workers.add(worker)
        worker.signals.done.connect(
            lambda w=worker: self._active_thumb_workers.discard(w)
        )
        QThreadPool.globalInstance().start(worker)

    def _make_group_row(self, group: dict) -> QWidget:
        """조사 행 — 그룹의 layer 별 대표(+n). 단독이면 개별(미매칭)."""
        row, lay = self._row_frame()
        is_matched = len(group) >= 2
        rep0 = next(iter(group.values())).representative
        head = QVBoxLayout()
        head.setSpacing(4)
        tag = QLabel("교차매치" if is_matched else "개별(미매칭)")
        tag.setStyleSheet(
            f"font-size:10px; font-weight:700;"
            f" color:{MATCH if is_matched else NOMATCH};"
        )
        head.addWidget(tag)
        info = QLabel(f"wafer {rep0.wafer_id}\ndie({rep0.col},{rep0.row})")
        info.setObjectName("dim")
        info.setStyleSheet("font-size:10px;")
        head.addWidget(info)
        head.addStretch()
        head_host = QWidget()
        head_host.setFixedWidth(120)
        head_host.setLayout(head)
        lay.addWidget(head_host)

        for lyr in self._selected_layers():
            if lyr in group:
                # 조사 모드에는 기준(★)이 없다 — 항상 일반 배지로 표시.
                lay.addWidget(self._clustered_thumb(group[lyr], lyr, False))
        lay.addStretch()
        return row

    # ---- 출력 트레이 담기 -------------------------------------------
    def _add_all_current(self) -> None:
        if self._add_targets:
            self._on_add(list(self._add_targets))


# =============================================================================
# app/workers.py   [#37]
# =============================================================================
"""백그라운드 작업 (문서 Section 10 성능 요구사항).

스캔/썸네일 생성은 QThreadPool 워커에서 수행하여 UI 멈춤을 최소화한다.
"""





workers__log = logging.getLogger("defect_tracker.workers")

# 썸네일 생성 병렬 워커 수(이미지 디코드+I/O 혼합).
_THUMB_WORKERS = max(2, min(8, (os.cpu_count() or 4)))


class ScanSignals(QObject):
    progress = Signal(str, int, int)  # message, current, total
    finished = Signal(object)  # LotIndex
    error = Signal(str)


class ScanWorker(QRunnable):
    """LOT 폴더를 백그라운드에서 스캔한다."""

    def __init__(self, lot_path: str | Path):
        super().__init__()
        self.lot_path = lot_path
        self.signals = ScanSignals()
        self._cancelled = False

    def cancel(self) -> None:
        """협조적 취소 — 다음 wafer 처리 지점에서 스캔 루프가 멈춘다."""
        self._cancelled = True

    @Slot()
    def run(self) -> None:
        try:
            def cb(msg: str, cur: int, total: int) -> None:
                self.signals.progress.emit(msg, cur, total)

            index: LotIndex = scan_lot(
                self.lot_path, progress=cb, cancel_check=lambda: self._cancelled
            )
            if self._cancelled:
                return  # 중단된 결과는 UI 로 보내지 않는다(토큰 게이트와 이중 안전)
            self.signals.finished.emit(index)
        except Exception as exc:  # noqa: BLE001 - 워커는 모든 예외를 UI 로 전달
            workers__log.exception("스캔 워커 실패: %s", self.lot_path)
            self.signals.error.emit(str(exc))


class ThumbnailSignals(QObject):
    ready = Signal(int, str)  # index, thumbnail path
    done = Signal()


class ThumbnailWorker(QRunnable):
    """기준 record 들의 중앙 crop 썸네일을 백그라운드에서 생성한다(확대율은 center_ratio)."""

    def __init__(self, cache, items: list[tuple[int, Path]], center_ratio: float | None = None):
        super().__init__()
        self.cache = cache
        self.items = items
        self.center_ratio = center_ratio
        self.signals = ThumbnailSignals()
        self._cancelled = False

    def cancel(self) -> None:
        self._cancelled = True

    @Slot()
    def run(self) -> None:
        def _make(item: tuple[int, Path]):
            index, path = item
            if self._cancelled:
                return None
            try:
                if self.center_ratio is not None:
                    thumb = self.cache.get_center_thumbnail(path, self.center_ratio)
                else:
                    thumb = self.cache.get_center_thumbnail(path)
            except Exception:  # noqa: BLE001 - 개별 실패는 건너뛴다
                workers__log.exception("썸네일 생성 실패: %s", path)
                return None
            return (index, thumb)

        if self.items:
            workers = min(_THUMB_WORKERS, len(self.items))
            with ThreadPoolExecutor(max_workers=workers) as ex:
                for result in ex.map(_make, self.items):
                    if self._cancelled:
                        break
                    if result is not None and result[1] is not None:
                        self.signals.ready.emit(result[0], str(result[1]))
        self.signals.done.emit()


class MatchSignals(QObject):
    finished = Signal(object, object)  # matches(list[BaseDefectMatches]), offsets(dict)
    error = Signal(str)


class MatchWorker(QRunnable):
    """매칭(match_all_with_offsets + collapse_matches)을 백그라운드에서 수행한다.

    UI 스레드가 멈추지 않도록 무거운 2-패스 매칭 + 근접 클러스터링을 여기서 실행한다.
    die 인덱스(index/fail_index)는 UI 에서 만들어(캐시 재사용) 넘겨받는다.
    """

    def __init__(self, base_records, compare_layers, records_by_layer, tolerance,
                 index=None, fail_index=None, cluster_radius=None):
        super().__init__()
        self.base_records = base_records
        self.compare_layers = compare_layers
        self.records_by_layer = records_by_layer
        self.tolerance = tolerance
        self.index = index
        self.fail_index = fail_index
        self.cluster_radius = cluster_radius
        self.signals = MatchSignals()

    @Slot()
    def run(self) -> None:
        try:

            all_matches, offsets = match_all_with_offsets(
                self.base_records, self.compare_layers, self.records_by_layer,
                self.tolerance, index=self.index, fail_index=self.fail_index,
            )
            if self.cluster_radius is not None:
                matches = collapse_matches(all_matches, self.cluster_radius)
            else:
                matches = collapse_matches(all_matches)
            self.signals.finished.emit(matches, offsets)
        except Exception as exc:  # noqa: BLE001 - 워커 예외는 UI 로 전달
            workers__log.exception("매칭 워커 실패")
            self.signals.error.emit(str(exc))


class ExportSignals(QObject):
    progress = Signal(int, int)  # cur, total
    finished = Signal(str)       # output path
    error = Signal(str)


class ExportWorker(QRunnable):
    """Excel 출력을 백그라운드에서 수행한다(진행도 콜백 → progress 시그널)."""

    def __init__(self, kwargs: dict):
        super().__init__()
        self.kwargs = kwargs
        self.signals = ExportSignals()

    @Slot()
    def run(self) -> None:
        try:

            path = export_excel(
                progress=lambda c, t: self.signals.progress.emit(c, t), **self.kwargs
            )
            self.signals.finished.emit(str(path))
        except Exception as exc:  # noqa: BLE001 - 워커 예외는 UI 로 전달
            workers__log.exception("Excel 출력 워커 실패")
            self.signals.error.emit(str(exc))


class FullThumbSignals(QObject):
    ready = Signal(int)  # index (해당 썸네일 캐시 준비됨)
    done = Signal()


class FullThumbWorker(QRunnable):
    """전체 썸네일 캐시를 백그라운드에서 미리 굽는다(히트맵 상세 등 지연 로딩용).

    실제 QPixmap 세팅은 UI 스레드에서(ready 시그널 후), 여기서는 디코드+저장만 한다.
    """

    def __init__(self, thumb_cache, items):  # items: list[(index, image_path, px)]
        super().__init__()
        self.thumb_cache = thumb_cache
        self.items = items
        self.signals = FullThumbSignals()
        self._cancelled = False

    def cancel(self) -> None:
        self._cancelled = True

    @Slot()
    def run(self) -> None:
        def _warm(item):
            i, path, px = item
            if self._cancelled:
                return None
            try:
                self.thumb_cache.get_full_thumbnail(path, max_size=px)
            except Exception:  # noqa: BLE001 - 개별 실패는 건너뛴다
                pass
            return i

        if self.items and self.thumb_cache is not None:
            workers = min(_THUMB_WORKERS, len(self.items))
            with ThreadPoolExecutor(max_workers=workers) as ex:
                for i in ex.map(_warm, self.items):
                    if self._cancelled:
                        break
                    if i is not None:
                        self.signals.ready.emit(i)
        self.signals.done.emit()


# =============================================================================
# app/ui/main_window.py   [#38]
# =============================================================================
"""메인 윈도우 — 전체 workflow 조립 (문서 Section 8 전체).

폴더 선택 → 스캔 → 기준/비교 layer 선택 → 매칭 → 탐색/비교 → 결과 출력.
모든 원본 접근은 read-only, 결과는 output workspace 에만 저장한다.

상태 갱신 원칙(사용성):
  - 기준 layer 변경/새 LOT  → 전체 재구성(_rebuild_all), 인덱스 0.
  - 비교 layer 토글         → 그리드 컬럼만 재구성(_rematch, rebuild_grid=True), 현재 인덱스 유지.
  - 허용 오차 변경          → 재매칭만(_rematch, rebuild_grid=False), 그리드/썸네일 유지, 인덱스 유지.
오류는 비차단 배너로 안내하여 작업 흐름을 끊지 않는다.
"""






class MainWindow(QMainWindow):
    def __init__(self, settings: Optional[AppSettings] = None):
        super().__init__()
        self._base_title = f"{APP_NAME}  ·  v{__version__}"
        self.setWindowTitle(self._base_title)

        self.settings = settings or AppSettings.load()
        self.settings.ensure_workspace()
        self.thumb_cache = ThumbnailCache(self.settings.cache_path)
        self.image_loader = ImageLoader(max_dim=self._target_image_dim())
        self.pool = QThreadPool.globalInstance()

        self.lot_index: Optional[LotIndex] = None
        self.base_records: list[DefectRecord] = []  # 대표(접힌) 기준 목록
        self._base_records_raw: list[DefectRecord] = []  # 접기 전 raw 기준 목록
        self.matches: list[BaseDefectMatches] = []
        self.current = -1
        self._thumb_worker: Optional[ThumbnailWorker] = None
        self._scan_worker: Optional[ScanWorker] = None  # 진행 중 스캔(중단용)
        self._scan_token = 0  # stale 스캔/썸네일 결과 무시용
        self._match_token = 0  # stale 매칭(중첩 요청) 결과 무시용
        self._wafer_filter: Optional[str] = None  # 특정 wafer 만 보기(wafer 폴더 선택 시)
        self._exporting = False  # Excel 출력 진행 중(중복 방지)
        # 매칭 인덱스 캐시(비교 layer 집합이 같으면 허용오차만 바뀔 때 재사용)
        self._match_sig: object = None
        self._match_idx = None
        self._match_fail = None
        self._layer_offsets: dict = {}  # 비교 layer 별 전역 정합오차(median)
        # 보기 필터는 '매칭만' 고정(드롭다운 제거) — 매칭 0인 후보는 항상 후보에서 제외.
        self._filter = "matched"
        # 출력 담기 트레이: 담은 BaseDefectMatches 스냅샷 목록(base image_path 로 중복 제거).
        # 스냅샷이라 기준 layer·자재(LOT)를 바꿔도 담은 것이 그대로 유지된다.
        self._export_tray: list = []
        self._view_cache: Optional[list[int]] = None  # _view_indices 캐시
        self._align_cache: dict = {}  # (lot_id, wafer, product) -> Alignment (웨이퍼 맵 정합)
        # 실행 중 워커는 풀 스레드에서 도는 동안 GC 되지 않도록 참조를 유지한다.
        self._active_workers: set = set()

        self._update_status: Optional[UpdateStatus] = None
        self._updating = False

        self._restore_geometry()
        self._build_ui()
        # 무거운 작업(매칭·Excel 출력) 중 표시할 로딩 오버레이.
        self.busy = BusyOverlay(self)
        # 허용오차 스핀박스는 연속 변경되므로 재매칭을 디바운스한다.
        self._tol_timer = QTimer(self)
        self._tol_timer.setSingleShot(True)
        self._tol_timer.setInterval(250)
        self._tol_timer.timeout.connect(lambda: self._rematch(rebuild_grid=False))
        self._install_shortcuts()
        self._apply_saved_prefs()
        self._maybe_check_update()

    # ----------------------------------------------------- 화면/DPI 보조
    @staticmethod
    def _target_image_dim() -> int:
        """그리드 이미지 로딩 해상도를 화면 크기·DPR 기준으로 산정(고DPI 선명도)."""
        screen = QGuiApplication.primaryScreen()
        if screen is None:
            return 700
        geo = screen.availableGeometry()
        dpr = screen.devicePixelRatio()
        # 셀은 화면 절반 폭 이하 → 그 정도 해상도면 충분히 선명
        return int(max(560, min(1600, (geo.width() // 2) * dpr)))

    def _restore_geometry(self) -> None:
        screen = QGuiApplication.primaryScreen()
        if self.settings.window_geometry:
            try:
                x, y, w, h = (int(v) for v in self.settings.window_geometry.split(","))
                self.setGeometry(x, y, w, h)
                return
            except (ValueError, TypeError):
                pass
        if screen is not None:
            avail = screen.availableGeometry()
            w = int(avail.width() * 0.80)
            h = int(avail.height() * 0.84)
            self.resize(max(1100, w), max(720, h))
            self.move(avail.center().x() - self.width() // 2,
                      avail.center().y() - self.height() // 2)
        else:
            self.resize(1280, 860)
        self.setMinimumSize(1024, 680)

    def show_initial(self) -> None:
        """초기 표시 — 기본 최대화(설정). 최대화를 끈 적이 있으면 저장된 창 크기로 연다.

        _restore_geometry 가 normal 상태의 창 크기/위치를 미리 설정해 두므로, 최대화를
        해제하면 그 크기로 복원된다.
        """
        if self.settings.window_maximized:
            self.showMaximized()
        else:
            self.show()

    def maybe_prompt_device_db(self) -> None:
        """디바이스 DB(AOIDeviceDB.xlsx)가 없으면 설정을 안내하는 팝업을 띄운다.

        DB 가 로드되지 않으면(등록된 db 제품 0개) 웨이퍼 맵 die 배치·좌표 변환이 정확히
        표시되지 않으므로, 시작 시 한 번 설정을 권한다. '지금 설정'을 누르면 설정 창을 연다.
        """
        has_db = any(
            getattr(p, "source", "") == "db" for p in PRODUCTS.values()
        )
        if has_db:
            return
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Warning)
        box.setWindowTitle("디바이스 DB 설정 필요")
        box.setText("디바이스 DB(AOIDeviceDB.xlsx)가 설정되어 있지 않습니다.")
        box.setInformativeText(
            "웨이퍼 맵의 die 배치와 좌표 변환을 정확히 표시하려면 디바이스 DB 를 "
            "설정해야 합니다.\n지금 설정하시겠습니까?"
        )
        btn_set = box.addButton("지금 설정", QMessageBox.AcceptRole)
        box.addButton("나중에", QMessageBox.RejectRole)
        box.setDefaultButton(btn_set)
        box.exec()
        if box.clickedButton() is btn_set:
            self._open_settings()

    # ------------------------------------------------------------------ UI
    def _build_ui(self) -> None:
        root = QWidget()
        root.setObjectName("root")
        main = QVBoxLayout(root)
        main.setContentsMargins(10, 10, 10, 10)
        main.setSpacing(8)

        # 알림 배너는 레이아웃에 넣지 않고 창 위 오버레이로 띄운다(표시 시 UI 가 밀리지 않음).
        self.banner = NotificationBanner(root)
        self.banner.hide()

        # 좌측 사이드바 | 우측(짧은 상단 + 큰 그리드) — 수평 스플리터
        self.splitter = QSplitter(Qt.Horizontal)

        # ── 좌측: 컨트롤 사이드바
        self.top = SideBar()
        self.top.open_folder.connect(self._choose_folder)
        self.top.base_layer_changed.connect(lambda _: self._rebuild_all())
        self.top.compare_layers_changed.connect(lambda: self._rematch(rebuild_grid=True))
        self.top.tolerance_changed.connect(lambda _: self._tol_timer.start())
        self.top.export_requested.connect(self._export)
        self.top.settings_requested.connect(self._open_settings)
        # 업데이트는 설정 다이얼로그로 이동(_open_settings 에서 연결)
        # 자재 폴더 버튼: 우클릭 시 최근 폴더 메뉴
        self.top.btn_open.setContextMenuPolicy(Qt.CustomContextMenu)
        self.top.btn_open.customContextMenuRequested.connect(self._show_recent_menu)
        self.top.btn_open.setToolTip(
            "리뷰가 진행된 자재(LOT) 폴더를 선택 (Ctrl+O) · 우클릭: 최근 폴더"
        )
        self.splitter.addWidget(self.top)

        # ── 우측: 짧은 상단(썸네일 + 탐색) + 큰 비교 그리드
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(8)

        top_band = QFrame()
        top_band.setObjectName("panel")
        band_layout = QVBoxLayout(top_band)
        band_layout.setContentsMargins(10, 8, 10, 8)
        band_layout.setSpacing(6)
        self.strip = ThumbnailStrip()
        self.strip.thumb_clicked.connect(self._goto)
        # 썸네일 + 웨이퍼 맵을 한 줄에(맵은 현재 wafer 의 die 현황)
        strip_row = QHBoxLayout()
        strip_row.setContentsMargins(0, 0, 0, 0)
        strip_row.setSpacing(8)
        strip_row.addWidget(self.strip, 1)
        # defect 히트맵 보기(항목 4) — 웨이퍼맵에 defect 밀도를 표시하고 위치별 비교.
        self.btn_heatmap = QPushButton("히트맵\n보기")
        self.btn_heatmap.setFixedSize(96, 96)
        self.btn_heatmap.setToolTip(
            "defect 밀도 히트맵을 새 창으로 엽니다. 위치를 클릭하면 그 자리의 defect 들을 "
            "layer 별로 나란히 비교하고 출력에 담을 수 있습니다."
        )
        self.btn_heatmap.clicked.connect(self._open_heatmap)
        self.btn_heatmap.setEnabled(False)
        strip_row.addWidget(self.btn_heatmap, 0, Qt.AlignVCenter)
        # 웨이퍼 맵 + 캡션(디바이스/정합 안내)을 세로로 묶는다.
        wafer_box = QVBoxLayout()
        wafer_box.setContentsMargins(0, 0, 0, 0)
        wafer_box.setSpacing(2)
        self.wafer_map = WaferMapWidget()
        self.wafer_map.die_clicked.connect(self._jump_to_die)
        wafer_box.addWidget(self.wafer_map, 0, Qt.AlignHCenter)
        self.lbl_wafer = QLabel("")
        self.lbl_wafer.setObjectName("dim")
        self.lbl_wafer.setStyleSheet("font-size:9px;")
        self.lbl_wafer.setAlignment(Qt.AlignCenter)
        self.lbl_wafer.setWordWrap(True)
        self.lbl_wafer.setFixedWidth(140)
        wafer_box.addWidget(self.lbl_wafer, 0, Qt.AlignHCenter)
        strip_row.addLayout(wafer_box)
        band_layout.addLayout(strip_row)
        self.nav = NavBar()
        self.nav.prev_clicked.connect(self._prev)
        self.nav.next_clicked.connect(self._next)
        # 보기 필터는 '매칭만' 고정(드롭다운 제거) — 어떤 비교 layer 와도 매칭 안 된
        # 기준 사진은 항상 후보에서 제외한다.
        # 항목 9 에서 비운 자리에 '출력에 추가'(트레이 담기) 버튼을 둔다(항목 1).
        self.btn_add_export = QPushButton("＋ 출력에 추가")
        self.btn_add_export.setObjectName("mini")
        self.btn_add_export.setToolTip(
            "현재 기준 사진을 출력 목록(트레이)에 담습니다. (A)\n"
            "담은 것들은 '결과 출력' 시 함께 Excel 로 나옵니다."
        )
        self.btn_add_export.clicked.connect(self._add_current_to_export)
        self.btn_add_export.setEnabled(False)
        self.nav.add_widget(self.btn_add_export)
        self.lbl_view = QLabel("")
        self.lbl_view.setObjectName("dim")
        self.nav.add_widget(self.lbl_view)
        band_layout.addWidget(self.nav)
        right_layout.addWidget(top_band)

        # 진행바 + 중단 버튼(스캔 중에만 표시)
        progress_row = QHBoxLayout()
        progress_row.setContentsMargins(0, 0, 0, 0)
        progress_row.setSpacing(8)
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setTextVisible(False)
        progress_row.addWidget(self.progress, 1)
        self.btn_stop = QPushButton("■ 중단")
        self.btn_stop.setObjectName("mini")
        self.btn_stop.setToolTip("진행 중인 스캔을 중단합니다.")
        self.btn_stop.clicked.connect(self._stop_scan)
        self.btn_stop.setVisible(False)
        progress_row.addWidget(self.btn_stop, 0)
        right_layout.addLayout(progress_row)

        # 비교 그리드 (스크롤 가능) — 큰 메인 영역
        grid_scroll = QScrollArea()
        grid_scroll.setWidgetResizable(True)
        grid_scroll.setFrameShape(QFrame.NoFrame)
        grid_host = QFrame()
        grid_host.setObjectName("panel")
        grid_host_layout = QVBoxLayout(grid_host)
        grid_host_layout.setContentsMargins(12, 12, 12, 12)
        self.grid = CompareGrid(loader=self.image_loader)
        self.grid.image_clicked.connect(self._open_viewer)
        self.grid.base_cluster_clicked.connect(self._show_cluster_members)
        grid_host_layout.addWidget(self.grid)
        self._empty_label = QLabel("자재 폴더를 선택하면 비교 화면이 표시됩니다.")
        self._empty_label.setObjectName("dim")
        self._empty_label.setAlignment(Qt.AlignCenter)
        self._empty_label.setMinimumHeight(200)
        grid_host_layout.addWidget(self._empty_label)
        grid_host_layout.addStretch()
        grid_scroll.setWidget(grid_host)
        right_layout.addWidget(grid_scroll, 1)

        self.splitter.addWidget(right)
        self.splitter.setStretchFactor(0, 0)
        self.splitter.setStretchFactor(1, 1)
        self.splitter.setCollapsible(0, False)
        sw = max(180, int(self.settings.sidebar_width))
        self.splitter.setSizes([sw, max(600, self.width() - sw)])
        main.addWidget(self.splitter, 1)

        self.setCentralWidget(root)

    def resizeEvent(self, event):  # noqa: N802
        super().resizeEvent(event)
        # 오버레이 배너를 창 크기에 맞춰 상단 중앙에 유지한다.
        if getattr(self, "banner", None) is not None:
            self.banner.reposition()
            self.banner.raise_()

    def _install_shortcuts(self) -> None:
        QShortcut(QKeySequence(Qt.Key_Right), self, activated=self._next)
        QShortcut(QKeySequence(Qt.Key_Left), self, activated=self._prev)
        QShortcut(QKeySequence(Qt.Key_PageDown), self, activated=self._next)
        QShortcut(QKeySequence(Qt.Key_PageUp), self, activated=self._prev)
        QShortcut(QKeySequence(Qt.Key_Home), self,
                  activated=lambda: self._goto_view_edge(False))
        QShortcut(QKeySequence(Qt.Key_End), self,
                  activated=lambda: self._goto_view_edge(True))
        QShortcut(QKeySequence("Ctrl+O"), self, activated=self._choose_folder)
        QShortcut(QKeySequence("Ctrl+E"), self, activated=self._export)
        QShortcut(QKeySequence(Qt.Key_F5), self, activated=self._rescan)
        # 비교 layer 전체/해제, 다음 미매칭 점프
        QShortcut(QKeySequence("Ctrl+A"), self,
                  activated=lambda: self.top._set_all_compares(True))
        QShortcut(QKeySequence("Ctrl+D"), self,
                  activated=lambda: self.top._set_all_compares(False))
        QShortcut(QKeySequence(Qt.Key_U), self, activated=self._jump_unmatched)
        QShortcut(QKeySequence(Qt.Key_A), self, activated=self._add_current_to_export)
        QShortcut(QKeySequence(Qt.Key_F1), self, activated=self._open_help)

    def _apply_saved_prefs(self) -> None:
        # 0.0(정확 일치)도 유효한 사용자 설정이므로 falsy 검사로 떨어뜨리지 않는다.
        if self.settings.tolerance is not None:
            self.top.set_tolerance(self.settings.tolerance)

    # ----------------------------------------------------------- 폴더/스캔
    def _choose_folder(self) -> None:
        last = self.settings.last_lot_folder
        start = str(Path(last).parent) if last and Path(last).exists() else str(Path.home())
        # 네이티브 탐색기 대신 앱 내 커스텀 폴더 트리 선택기를 사용한다.
        dlg = FolderPickerDialog(self.settings, start, self)
        if dlg.exec():
            folder = dlg.selected_path()
            wafer_folder = dlg.selected_wafer_folder()
            if folder:
                if wafer_folder:
                    self._open_wafer_selection(wafer_folder, folder)
                else:
                    self._open_folder(folder)

    def _open_wafer_selection(self, wafer_folder: str, lot: str) -> None:
        """wafer 폴더를 골랐을 때: 개별 wafer 만 볼지 물어보고, 아니면 LOT 로 회귀한다."""
        from PySide6.QtWidgets import QMessageBox

        wafer_id = Path(wafer_folder).name
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Question)
        box.setWindowTitle("wafer 폴더 선택")
        box.setText(f"wafer 폴더 ‘{wafer_id}’ 를 선택했습니다.\n개별 wafer 의 매칭 정보를 볼까요?")
        box.setInformativeText("‘전체 LOT 보기’를 누르면 상위 LOT 폴더로 이동합니다.")
        btn_wafer = box.addButton("이 wafer 만 보기", QMessageBox.YesRole)
        box.addButton("전체 LOT 보기", QMessageBox.NoRole)
        box.setDefaultButton(btn_wafer)
        box.exec()
        if box.clickedButton() is btn_wafer:
            self.load_lot(lot, wafer_filter=wafer_id)
        else:
            self.load_lot(lot)  # 잘못 고른 경우 → LOT 자동 회귀

    def _open_folder(self, folder: str) -> None:
        """선택 폴더의 구조 레벨을 판별해 자재 폴더로 보정하거나 재선택을 안내한다.

        모든 안내는 비차단 배너로(팝업 없음). 원본 read-only.
        """
        kind, material = classify_selection(folder)
        if kind == "material":
            self.load_lot(folder)
        elif kind in ("layer", "wafer") and material is not None:
            label = "layer" if kind == "layer" else "wafer"
            self.banner.show_message(
                f"{label} 폴더가 선택되었으니 자재 폴더로 자동 이동하여 탐색합니다.",
                "info",
            )
            self.load_lot(str(material))
        elif kind == "too_high":
            self.banner.show_message(
                "상위(device) 폴더가 선택되었습니다. 자재 폴더를 선택해 주세요.",
                "warn",
                action_text="자재 폴더 선택",
                action=self._choose_folder,
                timeout_ms=0,
            )
        else:  # unknown — 그대로 시도(스캔에서 layer 없음 경고로 처리)
            self.load_lot(folder)

    def _rescan(self) -> None:
        """현재 LOT 폴더를 다시 스캔한다(F5). 데이터가 갱신됐을 때 사용."""
        last = self.settings.last_lot_folder
        if last and Path(last).exists():
            self.load_lot(last)

    def _show_recent_menu(self) -> None:
        recents = [f for f in self.settings.recent_folders if Path(f).exists()]
        if not recents:
            self.banner.show_message("최근 연 자재 폴더가 없습니다.", "info")
            return
        menu = QMenu(self)
        for folder in recents:
            menu.addAction(folder, lambda f=folder: self._open_folder(f))
        menu.exec(self.top.btn_open.mapToGlobal(self.top.btn_open.rect().bottomLeft()))

    def _push_recent(self, folder: str) -> None:
        recents = [f for f in self.settings.recent_folders if f != folder]
        recents.insert(0, folder)
        self.settings.recent_folders = recents[:5]

    def _auto_select_product(self, folder: str) -> None:
        """자재 경로에서 디바이스(제품)를 자동 인식해 활성화한다(스캔 전 호출).

        인식되면 좌표 변환·웨이퍼 맵 die 배치가 그 제품 기준으로 적용된다. 실패하면
        설정의 제품(settings.product)을 그대로 둔다.
        """
        try:
            key, score = match_product_for_path(folder)
        except Exception:  # noqa: BLE001 - 인식 실패는 치명적이지 않음
            return
        if key and key != _active_product:
            set_active_product(key)
            # 빌트인(die_map 없음)으로 인식됐으면 같은 크기의 DB die_map 제품으로 승격.
            ensure_die_map_product()
            prod = active_product()
            if prod.source == "db":
                self.banner.show_message(
                    f"디바이스 자동 인식: {prod.name}", "info", timeout_ms=2500
                )

    def load_lot(
        self,
        folder: str,
        wafer_filter: Optional[str] = None,
        auto_detect: bool = True,
    ) -> None:
        # 2차 원본 보호(Section 1.1): 캐시/결과 작업공간이 이 LOT 내부면 차단한다.
        if not self._verify_workspace_outside(folder):
            return
        # 특정 wafer 만 보기(wafer 폴더를 골라 '이 wafer 만 보기' 선택 시). 없으면 전체.
        self._wafer_filter = wafer_filter

        # 디바이스 자동 인식(스캔 전): 좌표 변환·die 배치를 올바른 제품으로 맞춘다.
        # 설정에서 사용자가 직접 제품을 고른 재스캔(auto_detect=False)에서는 건너뛴다.
        if auto_detect:
            self._auto_select_product(folder)

        self.settings.last_lot_folder = folder
        self._push_recent(folder)
        self._scan_token += 1
        token = self._scan_token

        self.progress.setVisible(True)
        self.btn_stop.setVisible(True)
        # 실제 진행(총 wafer 수)이 오기 전까지는 불확정(busy) 애니메이션으로 움직이게 한다
        # (0% 에 멈춰 '로딩중'만 뜨는 것 방지). _on_scan_progress 가 오면 0~100%로 전환.
        self.progress.setRange(0, 0)
        self.progress.setTextVisible(True)
        self.progress.setFormat("스캔 준비 중...")
        self.nav.set_status("스캔 중...")
        self.top.set_lot_name(Path(folder).name)
        self.setWindowTitle(f"{self._base_title}  —  {Path(folder).name}")

        worker = ScanWorker(folder)
        self._scan_worker = worker
        worker.signals.progress.connect(self._on_scan_progress)
        worker.signals.finished.connect(lambda idx, t=token: self._on_scan_finished(idx, t))
        worker.signals.error.connect(lambda msg, t=token: self._on_scan_error(msg, t))
        self._track_worker(worker, worker.signals.finished, worker.signals.error)
        self.pool.start(worker)

    def _track_worker(self, worker, *terminal_signals) -> None:
        """워커가 끝날 때까지 참조를 유지(실행 중 GC 로 인한 시그널 삭제 방지)."""
        self._active_workers.add(worker)
        for sig in terminal_signals:
            sig.connect(lambda *_, w=worker: self._active_workers.discard(w))

    def _stop_scan(self) -> None:
        """진행 중인 스캔을 중단한다(협조적 취소 + stale 토큰으로 결과 폐기)."""
        if self._scan_worker is not None:
            self._scan_worker.cancel()
            self._scan_worker = None
        self._scan_token += 1  # 늦게 도착하는 finished/progress 결과를 무시
        self.progress.setVisible(False)
        self.btn_stop.setVisible(False)
        self.nav.set_status("스캔 중단됨")
        self.banner.show_message("스캔을 중단했습니다.", "info")

    def _verify_workspace_outside(self, folder: str) -> bool:
        """캐시/내보내기 작업공간이 선택한 LOT 폴더 내부면 안내하고 차단한다."""
        for target in (self.settings.cache_path, self.settings.exports_path):
            conflict = conflicting_source(target, [folder])
            if conflict is not None:
                self.banner.show_message(
                    "작업공간(캐시/결과)이 원본 LOT 폴더 내부에 있어 차단했습니다. "
                    "원본 보호를 위해 다른 폴더를 선택하세요.",
                    "error",
                    action_text="작업공간 변경",
                    action=self._change_workspace,
                    timeout_ms=0,
                )
                return False
        return True

    def _change_workspace(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "작업공간(캐시/결과) 폴더 선택", str(Path.home())
        )
        if folder:
            self.settings.workspace = folder
            self.settings.output_folder = ""
            self.settings.ensure_workspace()
            self.thumb_cache = ThumbnailCache(self.settings.cache_path)
            self.settings.save()
            self.banner.show_message("작업공간을 변경했습니다.", "success")

    def _on_scan_progress(self, msg: str, cur: int, total: int) -> None:
        self.nav.set_status(msg)
        if total > 0:
            # layer 단위 진행을 0~100%로 환산해 진행바에 표시
            pct = int(round(min(cur, total) / total * 100))
            self.progress.setRange(0, 100)
            self.progress.setValue(pct)
            self.progress.setFormat(f"{msg}  %p%")

    def _on_scan_error(self, message: str, token: int) -> None:
        if token != self._scan_token:
            return
        self.progress.setVisible(False)
        self.btn_stop.setVisible(False)
        self._scan_worker = None
        self.nav.set_status("스캔 오류")
        self.banner.show_message(f"폴더 스캔 중 오류: {message}", "error", timeout_ms=0)

    def _on_scan_finished(self, index: LotIndex, token: int = -1) -> None:
        if token != -1 and token != self._scan_token:
            return  # 오래된(stale) 스캔 결과 무시
        self.progress.setVisible(False)
        self.btn_stop.setVisible(False)
        self._scan_worker = None
        self.lot_index = index
        # 새 LOT: 웨이퍼 맵 정합 캐시를 비운다(id(lot_index) 재사용으로 인한 stale 방지).
        self._align_cache.clear()
        layers = index.layer_canonicals()
        if not layers:
            self.banner.show_message(
                "선택한 폴더에서 layer 를 찾지 못했습니다. 자재 폴더를 확인하세요.",
                "warn", timeout_ms=0,
            )
            self.nav.set_status("layer 없음")
            self._empty_label.setVisible(True)
            self.grid.build_layout([], "")
            self.nav.set_enabled(False)
            return

        # 기준 layer 는 빈칸으로 시작(사용자가 직접 선택), 비교 기본값은 선호 재리뷰 집합만.
        # (자재 폴더를 바꿀 때마다 재리뷰만 선택되도록 저장값을 자동 복원하지 않는다.)
        rereview = self._preferred_rereview(index)
        self.top.set_layers(layers, base=None, compares=None, rereview=rereview)
        self.settings.save()

        ok = sum(1 for r in index.records if r.ok)
        failed = [r for r in index.records if not r.ok]
        status = (
            f"layer {len(layers)}개 · wafer {len(index.wafers())}개 · "
            f"이미지 {len(index.records)}개(좌표 OK {ok}개"
        )
        if failed:
            status += f", 실패 {len(failed)}개"
        status += ")"
        self.nav.set_status(status)
        self.nav.set_status_tooltip(self._failure_summary(failed))
        # 좌표 추출 실패 진단 리포트는 개발자 모드(환경변수 또는 설정)에서만 파일로 남긴다.
        report_path = self._write_diag_report(index) if dev_mode(self.settings) else None
        if failed:
            self.banner.show_message(
                f"{len(failed)}개 이미지의 좌표를 추출하지 못했습니다(상태표시줄에 상세).",
                "warn",
                action_text="진단 로그 열기" if report_path else None,
                action=(lambda p=report_path: QDesktopServices.openUrl(
                    QUrl.fromLocalFile(str(p)))) if report_path else None,
            )
        # 접근 불가(권한/네트워크) 경로가 있으면 조용히 누락되지 않도록 알린다.
        errors = getattr(index, "scan_errors", [])
        if errors:
            preview = "\n".join(errors[:8])
            if len(errors) > 8:
                preview += f"\n… 외 {len(errors) - 8}개"
            self.banner.show_message(
                f"{len(errors)}개 경로를 읽지 못해 일부 layer/wafer 가 누락됐을 수 있습니다.",
                "warn",
                timeout_ms=0,
            )
            self.nav.set_status_tooltip(
                self._failure_summary(failed) + "\n\n[접근 실패 경로]\n" + preview
            )
        self._rebuild_all()

    def _write_diag_report(self, index):
        """좌표 추출 실패 진단 리포트를 단일 md 에 누적 추가한다(실패 시 None)."""
        try:
            return write_parse_failure_report(
                self.settings.log_dir_path,
                index.lot_name,
                index.records,
                getattr(index, "scan_errors", []),
            )
        except OSError:
            return None

    @staticmethod
    def _failure_summary(failed: list[DefectRecord]) -> str:
        """좌표 파싱 실패 항목 요약(상태 표시줄 tooltip)."""
        if not failed:
            return "모든 이미지의 좌표를 정상 추출했습니다."
        by_status: dict[str, int] = {}
        for r in failed:
            by_status[r.status.value] = by_status.get(r.status.value, 0) + 1
        lines = ["좌표 추출 실패 항목:"]
        labels = {
            ParseStatus.NOT_FOUND.value: "좌표/매칭 정보 없음",
            ParseStatus.INFO_FILE_NOT_FOUND.value: "info 파일 없음",
            ParseStatus.INVALID_INFO.value: "info 값 부족/오류",
        }
        for code, n in sorted(by_status.items(), key=lambda kv: -kv[1]):
            lines.append(f"  · {labels.get(code, code)}: {n}개")
        for r in failed[:8]:
            lines.append(f"    - {r.layer}/{r.wafer_id}/{Path(r.image_path).name}")
        if len(failed) > 8:
            lines.append(f"    … 외 {len(failed) - 8}개")
        return "\n".join(lines)

    def _recommended_base(self) -> str:
        """그리드 기본 배치 순서에서 현재 LOT 에 존재하는 첫 layer 를 추천(자동선택 X)."""
        if self.lot_index is None:
            return ""
        layers = self.lot_index.layer_canonicals()
        present = {_canon_token(l): l for l in layers}
        for row in DEFAULT_LAYER_GRID:
            for cell in row:
                if _canon_token(cell) in present:
                    return present[_canon_token(cell)]
        return layers[0] if layers else ""

    def _show_base_prompt(self) -> None:
        """기준 layer 미선택 시 대기 화면: 매칭·탐색·스트립·맵을 비우고 선택을 안내한다."""
        self.base_records = []
        self.matches = []
        self._view_cache = None
        self.current = -1
        self.strip.set_items([], [])
        self.wafer_map.clear()
        self.lbl_wafer.setText("")
        self.nav.set_enabled(False)
        self.nav.set_index(0, 0)
        self.top.set_match_summary("")
        self._update_add_export_button()  # 트레이는 유지, 버튼 상태만 갱신
        rec = self._recommended_base()
        hint = f"  (추천: {rec})" if rec else ""
        self.grid.show_empty(f"기준 layer 를 선택하세요.{hint}")
        self._empty_label.setVisible(False)

    # ------------------------------------------------------- 재계산(비동기)
    def _rebuild_all(self) -> None:
        """새 LOT·기준 layer 변경: base 목록·썸네일·그리드 전체 재구성, 인덱스 0."""
        if self.lot_index is None:
            return
        base_layer = self.top.base_layer()
        if not base_layer:
            # 기준 layer 미선택: 매칭/탐색을 비우고 선택을 유도(대기 상태).
            self._show_base_prompt()
            return
        self._save_prefs()
        # raw 기준 목록(근접 중복 접기 전) — 매칭 입력·재매칭에 사용.
        self._base_records_raw = [
            r for r in self.lot_index.records_for_layer(base_layer) if r.ok
        ]
        # 특정 wafer 만 보기: 기준 defect 을 그 wafer 로 한정(매칭은 같은 wafer 안에서만
        # 일어나므로 비교 layer 는 그대로 두어도 그 wafer 매치만 나온다).
        if self._wafer_filter:
            only = [r for r in self._base_records_raw if r.wafer_id == self._wafer_filter]
            if only:
                self._base_records_raw = only
                self.banner.show_message(
                    f"개별 wafer ‘{self._wafer_filter}’ 만 표시합니다.", "info"
                )
            else:
                # 그 wafer 에 기준 defect 이 없으면 전체로 회귀.
                self.banner.show_message(
                    f"wafer ‘{self._wafer_filter}’ 에 기준 defect 이 없어 전체 LOT 를 표시합니다.",
                    "warn",
                )
                self._wafer_filter = None
        if not self._base_records_raw:
            self.matches, self.base_records, self._view_cache = [], [], None
            self._update_match_summary()
            self._after_rebuild()
            return
        # 무거운 매칭은 백그라운드에서(로딩 오버레이), 완료되면 화면을 재구성한다.
        self._start_match(self._after_rebuild)

    def _after_rebuild(self) -> None:
        """매칭 완료 후: 스트립/썸네일/그리드/탐색 재구성(기준 layer·새 LOT)."""
        captions, tooltips = [], []
        for m in self.matches:
            r = m.base
            extra = getattr(m.base_cluster, "extra_count", 0) or 0
            cap = f"{r.wafer_id}\n({r.col},{r.row})"
            if extra:
                cap += f"  +{extra}"
            captions.append(cap)
            tt = f"wafer {r.wafer_id} · die({r.col},{r.row}) · pos {r.position_key}"
            if r.defect_name:
                tt += f" · {r.defect_name}"
            if extra:
                tt += f" · 근접중복 +{extra}"
            tooltips.append(tt)
        self.strip.set_items(captions, tooltips, on_progress=self.busy.pump)
        self._start_thumbnails()
        self.busy.pump()

        self._rebuild_grid()
        self.busy.pump()
        self._empty_label.setVisible(not self.base_records)
        self.nav.set_enabled(bool(self.base_records))
        if self.base_records:
            view = self._view_indices()
            self._goto(view[0] if view else 0)
        else:
            self.nav.set_index(0, 0)
            self.grid.show_empty("기준 layer 에 좌표 OK 인 사진이 없습니다.")
            self.wafer_map.clear()
        self._refresh_strip_marks()
        self._update_add_export_button()

    def _rematch(self, rebuild_grid: bool) -> None:
        """비교 토글/허용오차 변경: 재매칭(비동기). 현재 인덱스는 유지(범위 clamp)."""
        if self.lot_index is None or not self._base_records_raw:
            return
        self._save_prefs()
        self._start_match(lambda: self._after_rematch(rebuild_grid))

    def _after_rematch(self, rebuild_grid: bool) -> None:
        if rebuild_grid:
            self._rebuild_grid()
        if self.matches:
            self.current = max(0, min(self.current, len(self.matches) - 1))
            self._goto(self.current)
        else:
            self.nav.set_index(0, 0)
        self._refresh_strip_marks()

    def _start_match(self, after) -> None:
        """매칭을 백그라운드 워커로 실행하고, 완료되면 after() 를 UI 스레드에서 호출."""
        compare_layers = self.top.compare_layers()
        tolerance = self.top.tolerance()
        rbl = self.lot_index.records_by_layer()
        idx, fidx = self._get_match_indices(compare_layers, rbl)
        self._match_token += 1
        token = self._match_token
        self.busy.start("매칭 중…")
        worker = MatchWorker(
            self._base_records_raw, compare_layers, rbl, tolerance,
            index=idx, fail_index=fidx,
            cluster_radius=getattr(self.settings, "cluster_radius", None),
        )
        worker.signals.finished.connect(
            lambda ms, offs, t=token, cb=after: self._on_match_done(ms, offs, t, cb)
        )
        worker.signals.error.connect(lambda msg, t=token: self._on_match_error(msg, t))
        self._track_worker(worker, worker.signals.finished, worker.signals.error)
        self.pool.start(worker)

    def _on_match_done(self, matches, offsets, token: int, after) -> None:
        if token != self._match_token:
            return  # 오래된(중첩 요청) 결과 무시
        self.matches = matches
        self._layer_offsets = offsets
        self.base_records = [m.base for m in self.matches]
        self._view_cache = None
        self._update_match_summary()
        # 오버레이를 먼저 끄지 않고(스피너 정지·화면 멈춤 방지), 무거운 재구성이
        # 끝난 뒤에 끈다. 재구성 도중에는 busy.pump() 로 스피너가 계속 돌게 한다.
        try:
            if after is not None:
                after()
        finally:
            self.busy.stop()

    def _on_match_error(self, msg: str, token: int) -> None:
        if token != self._match_token:
            return
        self.busy.stop()
        self.banner.show_message(f"매칭 실패: {msg}", "error")

    def _open_heatmap(self) -> None:
        if not self.matches:
            self.banner.show_message("먼저 자재 폴더와 기준 layer 를 선택하세요.", "info")
            return
        current_wafer = None
        if 0 <= self.current < len(self.matches):
            current_wafer = self.matches[self.current].base.wafer_id
        records_by_layer = (
            self.lot_index.records_by_layer() if self.lot_index else {}
        )
        dlg = HeatmapDialog(
            self.matches,
            self.top.base_layer(),
            self.top.compare_layers(),
            self.thumb_cache,
            self._add_indices_to_export,
            self.settings,
            current_wafer=current_wafer,
            records_by_layer=records_by_layer,
            parent=self,
        )
        dlg.exec()

    def _get_match_indices(self, compare_layers, rbl):
        """(lot, 비교 layer 집합) 기준으로 die/실패 인덱스를 캐시·재사용한다."""
        sig = (id(self.lot_index), tuple(compare_layers))
        if sig != self._match_sig:
            self._match_idx = build_die_index(rbl, compare_layers)
            self._match_fail = build_fail_index(rbl, compare_layers)
            self._match_sig = sig
        return self._match_idx, self._match_fail

    def _update_match_summary(self) -> None:
        """사이드바에 실시간 매칭 요약을 표시(허용오차 튜닝 피드백)."""
        if not self.matches:
            self.top.set_match_summary("")
            return
        total_pairs = sum(len(m.results) for m in self.matches)
        matched_pairs = sum(1 for m in self.matches for r in m.results if r.is_match)
        bases_matched = sum(
            1 for m in self.matches if any(r.is_match for r in m.results)
        )
        self.top.set_match_summary(
            f"매칭 {matched_pairs}/{total_pairs} 쌍 · "
            f"기준 {bases_matched}/{len(self.matches)}장"
        )
        # layer 간 전역 정합오차(median)를 tooltip 으로 안내(레지스트레이션 shift 파악)
        lines = ["[layer 간 정합오차(중앙값)]"]
        for layer, off in self._layer_offsets.items():
            if off.count:
                lines.append(
                    f"  {layer}: dx {off.dx:+.1f}, dy {off.dy:+.1f} µm "
                    f"(1:1 {off.count}쌍 기준, 보정 적용)"
                )
        if len(lines) > 1:
            self.top.lbl_match.setToolTip("\n".join(lines))
        else:
            self.top.lbl_match.setToolTip("")

    def _rebuild_grid(self) -> None:
        base_layer = self.top.base_layer()
        compare_layers = self.top.compare_layers()
        grid = build_grid([base_layer] + compare_layers)
        self.grid.build_layout(grid, base_layer)

    def _start_thumbnails(self) -> None:
        if self._thumb_worker is not None:
            self._thumb_worker.cancel()
        items = [(i, r.image_path) for i, r in enumerate(self.base_records)]
        if not items:
            return
        worker = ThumbnailWorker(
            self.thumb_cache, items, center_ratio=self._thumbnail_center_ratio()
        )
        token = self._scan_token
        worker.signals.ready.connect(
            lambda i, p, t=token: self.strip.set_thumbnail(i, p)
            if t == self._scan_token else None
        )
        self._track_worker(worker, worker.signals.done)
        self._thumb_worker = worker
        self.pool.start(worker)

    def _save_prefs(self) -> None:
        self.settings.tolerance = self.top.tolerance()
        self.settings.base_layer = self.top.base_layer()
        self.settings.compare_layers = self.top.compare_layers()

    # ------------------------------------------------- 썸네일 확대율
    @staticmethod
    def _thumbnail_center_ratio() -> float:
        """상단 썸네일 중앙 crop 비율 — 5× 고정(사진 중앙 20%)."""
        return THUMBNAIL_CENTER_RATIO

    # ------------------------------------------------- 출력 담기 트레이(항목 1)
    def _add_current_to_export(self) -> None:
        if not self.matches or not (0 <= self.current < len(self.matches)):
            self.banner.show_message("담을 기준 사진이 없습니다.", "info")
            return
        self._add_indices_to_export([self.current])

    def _tray_keys(self) -> set:
        return {str(m.base.image_path) for m in self._export_tray}

    def _add_indices_to_export(self, indices: list[int]) -> None:
        """주어진 base index 들의 매칭 스냅샷을 출력 트레이에 담는다(중복 무시)."""
        keys = self._tray_keys()
        added = 0
        for i in indices:
            if 0 <= i < len(self.matches):
                m = self.matches[i]
                k = str(m.base.image_path)
                if k not in keys:
                    self._export_tray.append(m)
                    keys.add(k)
                    added += 1
        self._update_add_export_button()
        if added:
            self.banner.show_message(
                f"출력 목록에 {added}장 담았습니다. (현재 {len(self._export_tray)}장)",
                "success", timeout_ms=2000,
            )
        else:
            self.banner.show_message("이미 담긴 사진입니다.", "info", timeout_ms=1500)

    def _clear_export_tray(self) -> None:
        self._export_tray = []
        self._update_add_export_button()

    def _update_add_export_button(self) -> None:
        n = len(self._export_tray)
        self.btn_add_export.setText(f"＋ 출력에 추가 ({n})" if n else "＋ 출력에 추가")
        self.btn_add_export.setEnabled(bool(self.matches))
        self.btn_heatmap.setEnabled(bool(self.matches))

    # ------------------------------------------------------------ 탐색
    def _goto(self, index: int) -> None:
        if not self.matches or not (0 <= index < len(self.matches)):
            return
        self.current = index
        item = self.matches[index]
        self.grid.update_for_base(item, self.top.compare_layers())
        self.strip.set_current(index)
        # 탐색 번호는 현재 보기(필터 후보) 기준으로 표시(제외된 후보는 세지 않음)
        view = self._view_indices()
        if index in view:
            self.nav.set_index(view.index(index) + 1, len(view))
        else:
            self.nav.set_index(index + 1, len(self.matches))
        self._prefetch_neighbors(index)
        self._update_wafer_map(item)

    def _prefetch_neighbors(self, index: int) -> None:
        """인접 기준의 이미지(기준+매칭 비교)를 미리 로드해 탐색 체감을 높인다."""
        paths: list[str] = []
        for j in (index + 1, index - 1, index + 2):
            if 0 <= j < len(self.matches):
                m = self.matches[j]
                paths.append(str(m.base.image_path))
                for r in m.results:
                    if r.is_match and r.matched is not None:
                        paths.append(str(r.matched.image_path))
        if paths:
            self.image_loader.prefetch(paths)

    @staticmethod
    def _preferred_rereview(index) -> set:
        """선호 재리뷰 layer 집합: canonical 별 최대 재리뷰 레벨(≥1)의 display 만.

        같은 canonical 에 재리뷰·재재리뷰가 모두 있으면 더 깊은(재재리뷰) 것만 고른다.
        """
        best_level: dict[str, int] = {}
        for layer in index.layers:
            lv = getattr(layer, "re_review_level", 0)
            if lv >= 1:
                best_level[layer.canonical] = max(best_level.get(layer.canonical, 0), lv)
        chosen = set()
        for layer in index.layers:
            lv = getattr(layer, "re_review_level", 0)
            if lv >= 1 and lv == best_level.get(layer.canonical):
                chosen.add(layer.display or layer.canonical)
        return chosen

    @staticmethod
    def _match_status(item) -> str:
        """기준 1개의 매칭 상태: matched(하나라도 매칭) / none(전무)."""
        return "matched" if any(r.is_match for r in item.results) else "none"

    def _passes_filter(self, item) -> bool:
        if self._filter == "all":
            return True
        if self._filter == "matched":
            return self._match_status(item) != "none"
        return True

    def _view_indices(self) -> list[int]:
        if self._view_cache is None:
            idxs = [i for i, m in enumerate(self.matches) if self._passes_filter(m)]
            # 기본 '매칭만' 필터가 모든 후보를 제외하면(예: 비교 layer 미선택) 빈 화면
            # 대신 전체를 보인다(혼란 방지). 사용자가 고른 명시 필터는 그대로 둔다.
            if not idxs and self.matches and self._filter == "matched":
                idxs = list(range(len(self.matches)))
            self._view_cache = idxs
        return self._view_cache

    def _step(self, delta: int) -> None:
        if not self.matches:
            return
        view = self._view_indices()
        if not view:
            self.banner.show_message("필터에 해당하는 기준 사진이 없습니다.", "info")
            return
        if self.current in view:
            pos = view.index(self.current)
            nxt = view[(pos + delta) % len(view)]
        else:
            nxt = view[0]
        self._goto(nxt)

    def _prev(self) -> None:
        self._step(-1)

    def _next(self) -> None:
        self._step(1)

    def _goto_view_edge(self, last: bool) -> None:
        view = self._view_indices()
        if view:
            self._goto(view[-1] if last else view[0])

    def _jump_unmatched(self) -> None:
        """현재 다음에 위치한 '미매칭 포함' 기준으로 점프(트리아지).

        현재 보기(필터)에 포함된 후보 중에서만 점프한다(제외된 후보는 건너뜀).
        """
        if not self.matches:
            return
        view = self._view_indices()
        targets = [i for i in view if self._match_status(self.matches[i]) == "none"]
        if not targets:
            self.banner.show_message("미매칭이 있는 기준 사진이 없습니다.", "success")
            return
        for i in targets:
            if i > self.current:
                self._goto(i)
                return
        self._goto(targets[0])  # 끝까지 없으면 처음으로 순환

    def _refresh_view_count(self) -> None:
        if self._filter == "all" or not self.matches:
            self.lbl_view.setText("")
            return
        n_view = len(self._view_indices())
        excluded = len(self.matches) - n_view
        if self._filter == "matched" and excluded > 0:
            self.lbl_view.setText(f"({n_view}개 · 제외 {excluded})")
        else:
            self.lbl_view.setText(f"({n_view}개)")

    def _refresh_strip_marks(self) -> None:
        """썸네일에 매칭 상태 점을 반영한다."""
        if not self.matches:
            return
        self.strip.set_status_marks([self._match_status(m) for m in self.matches])
        # 매칭 0인 기준 사진은 후보(썸네일)에서도 제외해 보이도록 반영
        self.strip.set_visible_set(self._view_indices())
        self._refresh_view_count()

    def _open_viewer(self, record: object) -> None:
        if isinstance(record, DefectRecord):
            dlg = ImageViewerDialog(record, self)
            dlg.exec()

    def _show_cluster_members(self, members: list) -> None:
        """기준 셀의 '+n' 클릭 — 근접(<50)으로 접힌 defect 전체를 팝업으로 보여준다."""
        if not members:
            return
        ClusterMembersPopup(
            members, self.top.base_layer(), self.thumb_cache, self._open_viewer, self
        ).exec()

    def _open_help(self) -> None:
        ShortcutsDialog(self).exec()

    # ---- 웨이퍼 맵 ----
    _ALIGN_MIN_OVERLAP = 0.6  # 이 비율 이상 겹쳐야 디바이스 모양을 신뢰

    def _update_wafer_map(self, item) -> None:
        """현재 wafer 의 die 격자를 매칭 상태로 갱신한다.

        디바이스 DB die_map 이 있으면 관측 die 와 **정합(평행이동)** 시켜 실제 모양으로
        그린다. 정합 신뢰도가 낮으면 사각 전체로 폴백하고 캡션·로그로 알린다.
        """

        wafer = item.base.wafer_id
        states: dict[tuple[int, int], str] = {}
        for m in self.matches:
            b = m.base
            if b.wafer_id != wafer or b.col is None or b.row is None:
                continue
            if b.col < 0 or b.row < 0:
                continue
            states[(b.col, b.row)] = self._match_status(m)
        observed = set(states.keys())

        prod = active_product()
        valid: Optional[set] = None
        # 캡션은 제품명만 표기(‘모양 정합 %’ 등은 노출하지 않음).
        caption = prod.name if prod.source == "db" else ""
        if prod.die_map and observed:
            align = self._get_alignment(wafer, prod, observed)
            if align.overlap >= self._ALIGN_MIN_OVERLAP:
                valid = shifted_die_map(prod.die_map, align)

        current = (item.base.col, item.base.row)
        # 실제 관측(매칭)된 die 는 DB 고정 모양(valid) 밖이어도 항상 그린다 — 그렇지 않으면
        # 정합 후 모양 밖으로 나온 새 die 가 격자만 커지고 색칠 없이 사라져 보인다.
        paint_valid = (valid | observed) if valid is not None else None

        if paint_valid is not None:
            # 디바이스 모양: 실제로 그려지는 셀(paint_valid = valid∪observed)의 bounding box
            # 로 격자를 정규화한다. 좌표계 원점이 wafer 마다 달라도 맵이 여백에 떠 보이거나
            # 좌·상단이 잘리지 않는다. (current 는 여기서 제외 — 음수 좌표로 걸러진 die 가
            # 헛여백을 만들지 않도록. 유효한 current 는 이미 paint_valid 안에 있다.)
            content = set(paint_valid)
            min_col = min(c for c, _ in content)
            min_row = min(r for _, r in content)
            max_col = max(c for c, _ in content)
            max_row = max(r for _, r in content)
            cols = max_col - min_col + 1
            rows = max_row - min_row + 1
            origin = (min_col, min_row)
        else:
            # 사각 폴백: 원점 (0,0) + 패키지 크기(관측 max 로 확장).
            max_col = max((c for c, _ in observed), default=0)
            max_row = max((r for _, r in observed), default=0)
            cols = max(prod.kla_package_x_count, max_col + 1)
            rows = max(prod.kla_package_y_count, max_row + 1)
            origin = (0, 0)

        self.wafer_map.set_data(cols, rows, states, current, valid=paint_valid, origin=origin)
        self.lbl_wafer.setText(caption)
        self.wafer_map.setToolTip(
            "웨이퍼 맵 — die 클릭 시 해당 기준 사진으로 이동"
            + (f"\n{caption}" if caption else "")
        )

    def _get_alignment(self, wafer: str, prod, observed: set):
        """(lot, wafer, product) 단위로 정합 결과를 캐시·재사용한다."""

        key = (id(self.lot_index), wafer, prod.key)
        cached = self._align_cache.get(key)
        if cached is None:
            cached = align_observed_to_diemap(observed, prod.die_map)
            self._align_cache[key] = cached
            if cached.overlap < self._ALIGN_MIN_OVERLAP:
                import logging
                logging.getLogger("defect_tracker.wafermap").info(
                    "die 정합 신뢰도 낮음 — wafer=%s product=%s overlap=%.2f",
                    wafer, prod.key, cached.overlap,
                )
        return cached

    def _jump_to_die(self, col: int, row: int) -> None:
        if not self.matches:
            return
        wafer = self.matches[self.current].base.wafer_id
        for i, m in enumerate(self.matches):
            b = m.base
            if b.wafer_id == wafer and b.col == col and b.row == row:
                self._goto(i)
                return

    # ------------------------------------------------------------ 설정
    def _open_settings(self) -> None:
        current_lot = str(self.lot_index.lot_path) if self.lot_index else None
        old_workspace = self.settings.workspace
        old_output = self.settings.output_folder
        old_cluster_radius = getattr(self.settings, "cluster_radius", None)
        old_font = getattr(self.settings, "ui_font_size", "normal")
        update_available = bool(self._update_status and self._update_status.available)
        dlg = SettingsDialog(
            self.settings, current_lot, self, update_available=update_available
        )
        accepted = dlg.exec()
        # "지금 업데이트/업데이트 확인" 클릭 시: 설정 저장 후 기존 비동기 흐름 재사용
        if dlg.wants_update():
            try:
                dlg.updated_settings().save()
            except OSError:
                pass
            self._manual_update()
            return
        if not accepted:
            return
        s = dlg.updated_settings()
        # 디바이스 DB 를 다시 읽어 제품 목록을 갱신한 뒤 활성 제품 적용.
        # 경로가 지정되면 그 경로를, 비면 번들 DB 를 자동으로 읽는다(하드코딩 금지).
        try:
            from pathlib import Path as _P

            db_path = _P(s.device_db_path) if s.device_db_path else None
            if db_path is None or not db_path.exists():
                db_path = bundled_device_db_path()
            if db_path is not None:
                register_devices(load_device_db(db_path))
        except Exception:  # noqa: BLE001
            self.banner.show_message("디바이스 DB 로드 실패(설정 확인).", "warn")
        # 디바이스 DB 를 새로 반영한 뒤, 아직 빌트인 기본 제품(DEFAULT_PRODUCT)에 머물러
        # 있으면 현재 LOT 경로에서 제품(디바이스)을 DB 시트명 기준으로 자동 인식해 프로파일을
        # 맞춘다("DB 저장하면 알아서 읽도록"). 사용자가 특정 디바이스를 직접 고른 경우엔
        # 그 값(≠기본)이 유지되므로 자동 인식이 덮어쓰지 않는다.
        if s.product == DEFAULT_PRODUCT and self.lot_index is not None:
            try:
                detected, _ = match_product_for_path(str(self.lot_index.lot_path))
            except Exception:  # noqa: BLE001 - 인식 실패는 치명적이지 않음
                detected = None
            if detected and detected in PRODUCTS:
                s.product = detected
                prod = PRODUCTS[detected]
                if getattr(prod, "source", "") == "db":
                    self.banner.show_message(
                        f"디바이스 자동 인식: {prod.name}", "info", timeout_ms=2500
                    )
        old_active = _active_product
        set_active_product(s.product)
        ensure_die_map_product()
        # 제품/DB 가 바뀌면 die_map 이 달라지므로 웨이퍼 맵 정합 캐시를 무효화한다.
        self._align_cache.clear()
        # 디바이스(제품)가 바뀌면 좌표 변환(col/row/x/y)이 '파싱 시점'에 계산되므로
        # 단순 재그리기로는 좌표가 갱신되지 않는다. 현재 LOT 을 다시 스캔해 새 pitch/offset
        # 으로 좌표를 재계산한다(사용자가 직접 고른 제품이므로 자동 인식은 건너뜀).
        if (
            _active_product != old_active
            and self.settings.last_lot_folder
            and self.lot_index is not None
        ):
            self.load_lot(
                self.settings.last_lot_folder,
                wafer_filter=self._wafer_filter,
                auto_detect=False,
            )
        elif self.matches:
            # 다음 네비게이션까지 기다리지 않고 지금 바로 새 제품 기준으로 다시 그린다.
            self._goto(self.current)
        # 작업공간/출력 폴더가 바뀌면 캐시를 재생성한다(원본 밖 보장은 다이얼로그에서 검증).
        if s.workspace != old_workspace or s.output_folder != old_output:
            s.ensure_workspace()
            self.thumb_cache = ThumbnailCache(s.cache_path)
        # 기본 허용오차를 스핀박스에도 반영(현재 매칭은 사용자가 바꾼 값 유지).
        try:
            s.save()
        except OSError as exc:
            self.banner.show_message(f"설정 저장 실패: {exc}", "error", timeout_ms=0)
            return
        # 개발자 모드를 방금 켰다면 재시작 없이 파일 로그를 바로 시작한다(setup_logging 은
        # 중복 핸들러를 막으므로 반복 호출해도 안전). 끄는 것은 다음 실행부터 반영된다.
        if dev_mode(s):
            setup_logging(s.log_dir_path)
        # defect 클러스터 거리가 바뀌면 근접 묶음이 달라지므로 재매칭한다.
        if getattr(s, "cluster_radius", None) != old_cluster_radius and self._base_records_raw:
            self._rematch(rebuild_grid=True)
        # 글자 크기(보통/크게)가 바뀌면 테마를 다시 적용해 대부분 즉시 반영한다.
        if getattr(s, "ui_font_size", "normal") != old_font:
            from PySide6.QtWidgets import QApplication

            apply_theme(QApplication.instance(), scale_for(s.ui_font_size))
        self.banner.show_message("설정을 저장했습니다.", "success")

    # ------------------------------------------------------------ 업데이트
    def _maybe_check_update(self) -> None:
        if not self.settings.auto_update_check:
            return
        self._start_update_check(manual=False)

    def _manual_update(self) -> None:
        if self._updating:
            return
        self.nav.set_status("업데이트 확인 중...")
        self._start_update_check(manual=True)

    def _start_update_check(self, manual: bool) -> None:
        worker = UpdateCheckWorker(token=self.settings.update_token)
        worker.signals.done.connect(lambda st, m=manual: self._on_update_checked(st, m))
        self._track_worker(worker, worker.signals.done)
        self.pool.start(worker)

    def _on_update_checked(self, status, manual: bool) -> None:
        self._update_status = status
        if status.available:
            self.top.set_update_available(True)
            answer = QMessageBox.question(
                self,
                "업데이트",
                "새 버전이 있습니다. 지금 업데이트할까요?\n"
                "업데이트 후 프로그램이 종료되며, 다시 시작하면 적용됩니다.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if answer == QMessageBox.Yes:
                self._do_update(status)
            elif not manual:
                self.banner.show_message(
                    "상단 '업데이트' 버튼으로 언제든 업데이트할 수 있습니다.", "info"
                )
        else:
            self.top.set_update_available(False)
            if manual:
                if status.error:
                    self.banner.show_message(
                        f"업데이트 확인 실패: {status.error}", "warn", timeout_ms=6000
                    )
                else:
                    self.banner.show_message("이미 최신 버전입니다.", "success")

    def _do_update(self, status) -> None:
        self._updating = True
        self.top.set_update_busy(True)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        self.nav.set_status("업데이트 준비 중...")

        worker = UpdateApplyWorker(status, token=self.settings.update_token)
        worker.signals.progress.connect(self.nav.set_status)
        worker.signals.finished.connect(self._on_update_finished)
        self._track_worker(worker, worker.signals.finished)
        self.pool.start(worker)

    def _on_update_finished(self, ok: bool, message: str) -> None:
        self._updating = False
        self.progress.setVisible(False)
        self.top.set_update_busy(False)
        if ok:
            QMessageBox.information(
                self,
                "업데이트 완료",
                f"{message}\n\n프로그램을 종료합니다. 다시 시작해 주세요.",
            )
            self.close()
        else:
            self.nav.set_status("업데이트 실패")
            self.banner.show_message(f"업데이트 실패: {message}", "error", timeout_ms=0)

    @staticmethod
    def _safe_filename(name: str) -> str:
        """Windows 금지문자/끝 공백·점 제거."""
        for ch in '<>:"/\\|?*':
            name = name.replace(ch, "_")
        name = name.rstrip(" .")
        return name or "compare"

    # ------------------------------------------------------------ 출력
    def _compute_all_layers_matched(self, progress=None) -> list[BaseDefectMatches]:
        """모든 layer 를 각각 기준으로 매칭해, 어느 layer 에서든 매치된 defect 을 합친다.

        기준 layer 종속 없이 '전체 매치'를 담기 위한 것. base image_path 로 중복 제거하고,
        현재 wafer 필터가 걸려 있으면 그 wafer 로 한정한다. (layer 수만큼 매칭을 다시 돌림)
        진행도는 progress(cur, total) 콜백으로 알린다(layer 단위).
        """

        if self.lot_index is None:
            return []
        layers = self.lot_index.layer_canonicals()
        rbl = self.lot_index.records_by_layer()
        tolerance = self.top.tolerance()
        total = len(layers)
        seen: set[str] = set()
        out: list[BaseDefectMatches] = []
        for i, base_layer in enumerate(layers):
            if progress is not None:
                progress(i, total)  # layer i 처리 시작(진행 표시)
            base_records = [
                r for r in self.lot_index.records_for_layer(base_layer) if r.ok
            ]
            if self._wafer_filter:
                base_records = [r for r in base_records if r.wafer_id == self._wafer_filter]
            if base_records:
                compare_layers = [lyr for lyr in layers if lyr != base_layer]
                matches, _ = match_all_with_offsets(
                    base_records, compare_layers, rbl, tolerance
                )
                for m in collapse_matches(matches):
                    if self._match_status(m) != "none":
                        k = str(m.base.image_path)
                        if k not in seen:
                            seen.add(k)
                            out.append(m)
        if progress is not None:
            progress(total, total)
        return out

    def _provide_all_layers_matched(self, progress=None) -> list[BaseDefectMatches]:
        """다이얼로그의 '모든 매치(기준 없이)' 버튼 공급자(진행 콜백 지원)."""
        return self._compute_all_layers_matched(progress)

    def _export(self) -> None:
        if not self.matches:
            self.banner.show_message("먼저 자재 폴더를 불러오세요.", "info")
            return
        # 이번 LOT 에서 매칭 있는 기준 사진(스냅샷) — 다이얼로그의 '전체 추가' 버튼용.
        all_matched = [
            m for m in self.matches if self._match_status(m) != "none"
        ]
        # 트레이가 비어 있어도 다이얼로그를 열어(전체 추가 버튼 사용) 담을 수 있게 한다.
        dlg = ExportTrayDialog(
            list(self._export_tray), self.thumb_cache,
            all_matched=all_matched,
            all_layers_provider=self._provide_all_layers_matched,
            parent=self,
        )
        if not dlg.exec():
            return
        selected = dlg.selected()  # list[BaseDefectMatches]
        # 다이얼로그에서 편집한 결과를 트레이에 반영(다음 출력에도 유지).
        self._export_tray = list(selected)
        self._update_add_export_button()
        # '확인'(저장만) → 트레이 상태만 저장하고 닫는다. Excel 출력은 나중에.
        if not dlg.wants_export():
            self.banner.show_message(
                f"출력 목록을 저장했습니다 ({len(selected)}장).", "success"
            )
            return
        if not selected:
            self.banner.show_message("출력할 사진이 없습니다.", "info")
            return

        # 컬럼(compare layer)은 담긴 스냅샷들에 등장하는 비교 layer 의 합집합(등장 순서).
        compare_union: list[str] = []
        for m in selected:
            for r in m.results:
                if r.compare_layer not in compare_union:
                    compare_union.append(r.compare_layer)

        default_dir = str(self.settings.exports_path)
        Path(default_dir).mkdir(parents=True, exist_ok=True)
        default_name = self._safe_filename(f"{self.lot_index.lot_name}_compare") + ".xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel 저장 위치", str(Path(default_dir) / default_name),
            "Excel 파일 (*.xlsx)",
        )
        if not path:
            return

        # Excel 출력은 이미지 디코드+저장이 무거우므로 백그라운드에서(진행도 오버레이).
        if self._exporting:
            self.banner.show_message("이미 Excel 출력 중입니다.", "info")
            return
        self._exporting = True
        self.busy.start("Excel 출력 준비 중…", determinate=True)
        kwargs = dict(
            output_path=path,
            lot_name=self.lot_index.lot_name,
            base_layer=self.top.base_layer(),
            compare_layers=compare_union or self.top.compare_layers(),
            tolerance=self.top.tolerance(),
            selected=selected,
            thumb_cache=self.thumb_cache,
            source_roots=[self.lot_index.lot_path],
        )
        worker = ExportWorker(kwargs)
        worker.signals.progress.connect(self._on_export_progress)
        worker.signals.finished.connect(self._on_export_done)
        worker.signals.error.connect(self._on_export_error)
        self._track_worker(worker, worker.signals.finished, worker.signals.error)
        self.pool.start(worker)

    def _on_export_progress(self, cur: int, total: int) -> None:
        self.busy.set_message(f"Excel 출력 중…  {cur}/{total}")
        self.busy.set_progress(cur, total)

    def _on_export_done(self, out: str) -> None:
        self._exporting = False
        self.busy.stop()
        if self.settings.output_folder == "":
            self.settings.output_folder = str(Path(out).parent)
        self.settings.save()
        self.banner.show_message(
            f"결과를 저장했습니다: {Path(out).name}",
            "success",
            action_text="폴더 열기",
            action=lambda p=out: QDesktopServices.openUrl(
                QUrl.fromLocalFile(str(Path(p).parent))
            ),
            timeout_ms=7000,
        )

    def _on_export_error(self, msg: str) -> None:
        self._exporting = False
        self.busy.stop()
        self.banner.show_message(f"Excel 출력 중 오류: {msg.splitlines()[0]}", "error", timeout_ms=0)

    # ------------------------------------------------------------ 종료
    def closeEvent(self, event):  # noqa: N802
        # 최대화 상태를 기억하고, 창 크기는 normal(복원) 기하로 저장한다.
        self.settings.window_maximized = self.isMaximized()
        geo = self.normalGeometry()
        self.settings.window_geometry = f"{geo.x()},{geo.y()},{geo.width()},{geo.height()}"
        try:
            self.settings.sidebar_width = self.splitter.sizes()[0]
        except (AttributeError, IndexError):
            pass
        self._save_prefs()
        try:
            self.settings.save()
        except OSError:
            pass
        super().closeEvent(event)


# =============================================================================
# main.py   [진입점]
# =============================================================================
"""Defect Tracker — 진입점.

원본 데이터를 절대 훼손하지 않는(read-only) defect 이미지 비교 뷰어.
실행: python main.py

필요 라이브러리가 없으면 GUI 를 띄우는 대신 친절한 안내를 출력한다.
의존성 자동 설치: python bootstrap.py
"""



# import 이름 -> 안내용 표기
_REQUIRED = {
    "PySide6": "PySide6",
    "PIL": "Pillow",
    "openpyxl": "openpyxl",
}


def _check_dependencies() -> list[str]:
    return [name for name in _REQUIRED if importlib.util.find_spec(name) is None]


def _load_device_db(settings) -> None:
    """디바이스 DB 를 읽어 제품 목록에 병합한다(실패 무시).

    설정에 경로가 있으면 그 경로를, 없으면 앱과 함께 배포되는 번들 DB
    (data/AOIDeviceDB.xlsx)를 자동으로 읽는다. DB 를 하드코딩하지 않고 항상 파일에서
    읽어 와 웨이퍼맵 die 위치가 DB 기준으로 표시되도록 한다.
    """
    from pathlib import Path


    path = getattr(settings, "device_db_path", "")
    db_path = Path(path) if path else None
    if db_path is None or not db_path.exists():
        db_path = bundled_device_db_path()  # 번들 DB 자동 로드
    if db_path is None:
        return
    try:

        register_devices(load_device_db(db_path))
    except Exception:  # noqa: BLE001 - DB 로드 실패는 치명적이지 않음

        get_logger().exception("디바이스 DB 로드 실패: %s", db_path)


def main() -> int:
    missing = _check_dependencies()
    if missing:
        names = ", ".join(_REQUIRED[m] for m in missing)
        print(
            "필요한 라이브러리가 설치되어 있지 않습니다: " + names + "\n"
            "다음 명령으로 설치하세요:\n"
            "    python bootstrap.py\n"
            "또는:\n"
            "    pip install -r requirements.txt",
            file=sys.stderr,
        )
        return 1

    # 터미널 실행 시: 무거운 라이브러리(PySide6) 로딩 동안 즉시 안내(콘솔).
    # windowed(.exe) 모드에서는 stderr 가 없을 수 있으므로 안전하게 처리.
    if sys.stderr is not None:
        try:
            print("Defect Tracker 시작 중... (라이브러리 로딩, 잠시만 기다려 주세요)",
                  file=sys.stderr, flush=True)
        except (OSError, ValueError):
            pass

    # 의존성 확인 후에 import (누락 시 깔끔한 메시지를 위해 함수 내부에서 import).
    # 무거운 모듈(MainWindow 트리: openpyxl/PIL 등)은 스플래시를 띄운 뒤 임포트한다.
    from PySide6.QtCore import Qt
    from PySide6.QtGui import QGuiApplication
    from PySide6.QtWidgets import QApplication


    # 고DPI: 분수 배율을 그대로 통과시켜 다양한 모니터에서 또렷하게(반올림 깨짐 방지).
    QGuiApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough
    )
    app = QApplication(sys.argv)
    app.setApplicationName("Defect Tracker")
    # 설정을 먼저 읽어 글자 크기(보통/크게)를 테마에 반영한 뒤 스플래시를 띄운다.
    settings = AppSettings.load()
    apply_theme(app, scale_for(settings.ui_font_size))

    # Qt 준비 직후 즉시 스플래시 표시 → 무거운 구성 동안 "로딩 중" 피드백을 보여준다.

    splash = make_splash()
    splash.show()
    show_status(splash, "로딩 중...")
    app.processEvents()

    _load_device_db(settings)
    set_active_product(settings.product)
    # 빌트인 폴백(die_map 없음) 대신 같은 패키지 크기의 DB die_map 제품으로 승격 →
    # 웨이퍼맵이 기본적으로 실제 die 모양으로 표시되도록.
    ensure_die_map_product()
    # 개발자 모드(환경변수 또는 설정)일 때만 파일 로그를 남긴다. 일반 사용자에겐 로그를 만들지 않음.
    if dev_mode(settings):
        setup_logging(settings.log_dir_path)
        get_logger().info("애플리케이션 시작 (제품=%s)", settings.product)

    show_status(splash, "화면 구성 중...")
    app.processEvents()


    window = MainWindow(settings)
    window.show_initial()  # 기본 최대화(설정), 해제 이력이 있으면 저장된 창 크기
    splash.finish(window)
    # 창이 뜬 뒤(이벤트 루프 진입 후) 디바이스 DB 미설정 안내 팝업을 띄운다.
    from PySide6.QtCore import QTimer

    QTimer.singleShot(0, window.maybe_prompt_device_db)
    return app.exec()


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# 단일 파일 업데이트 정책 (생성기가 추가 — 단일 파일 전용, app/ 소스에는 없음)
# =============================================================================
# 이 배포본은 defect_tracker.py 한 파일이므로, 업데이트는 레포 전체 ZIP 전개 대신
# GitHub main 의 single_file/defect_tracker.py 를 받아 "자기 자신"만 원자적으로 교체한다.
# 버전 파일(version.json)은 항상 "이 py 파일이 있는 폴더"에만 둔다.
_SINGLE_FILE_DIR = Path(__file__).resolve().parent  # 이 py 파일이 있는 폴더


def app_root() -> Path:
    """단일 파일: 버전/업데이트 기준 폴더 = 이 파일이 있는 폴더.

    맨 끝에서 재정의하므로 check_update·current_sha·read_version(읽기)과 apply_update(쓰기)가
    모두 이 폴더의 version.json 을 쓴다. 상위 폴더나 CWD 로 새지 않는다.
    """
    return _SINGLE_FILE_DIR


_SINGLE_FILE_RAW_URL = (
    "https://raw.githubusercontent.com/"
    "{owner}/{repo}/{branch}/single_file/defect_tracker.py"
).format(owner=UPDATE_OWNER, repo=UPDATE_REPO, branch=UPDATE_BRANCH)


def apply_update(status, root=None, owner=UPDATE_OWNER, repo=UPDATE_REPO,
                 branch=UPDATE_BRANCH, opener=urlopen, progress=None):
    """단일 파일 자기 교체 업데이트(레포 트리 전개 대체). 이름을 덮어써 이 정의가 쓰인다."""
    if is_frozen():
        return False, "실행파일(exe) 버전은 자동 업데이트를 지원하지 않습니다. 새 파일을 받아 교체하세요."
    target = _SINGLE_FILE_DIR / Path(__file__).name
    try:
        if progress:
            progress("새 단일 파일 내려받는 중...")
        req = Request(_SINGLE_FILE_RAW_URL, headers={"User-Agent": "Defect Tracker"})
        with opener(req, timeout=60) as resp:
            data = resp.read()
        if not data.strip():
            return False, "받은 파일이 비어 있습니다."
        if progress:
            progress("적용 중...")
        tmp = target.with_name(target.name + ".new")
        tmp.write_bytes(data)
        os.replace(tmp, target)  # 원자적 교체(같은 볼륨)
        if status.remote:
            write_version(_SINGLE_FILE_DIR, status.remote)  # version.json 은 이 폴더에만
        return True, "단일 파일을 최신으로 교체했습니다."
    except Exception as exc:  # noqa: BLE001 - 네트워크/IO 오류 graceful
        return False, "업데이트 실패: {}".format(exc)
